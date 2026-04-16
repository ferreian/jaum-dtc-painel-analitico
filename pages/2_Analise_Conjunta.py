"""
pages/2_Analise_Conjunta.py — Análise Conjunta de Produtividade
"""
import numpy as np
import pandas as pd
import streamlit as st
from scipy import stats
import plotly.graph_objects as go_plt

# ── Cores globais por status ──────────────────────────────────────────────────
COR_STATUS_PLOT = {
    "CHECK":    "#F4B184",
    "STINE":    "#2976B6",
    "LINHAGEM": "#00FF01",
    "DP2":      "#C4DFB4",
}
COR_BORDA = {
    "CHECK":    "#C46A3A",
    "STINE":    "#1A4F7A",
    "LINHAGEM": "#009900",
    "DP2":      "#7AAF6A",
}

# ── Escala de cores global para todos os heatmaps de produção relativa ─────────
# vermelho → salmão → creme neutro (85%) → amarelo (90%) → verde escuro (100%)
COLORSCALE_PERF = [
    [0.00, "#d73027"],
    [0.70, "#f89374"],
    [0.80, "#fff5cc"],
    [0.87, "#fee08b"],
    [0.90, "#a6d96a"],
    [1.00, "#1a9850"],
]

from utils.theme import aplicar_tema, page_header, secao_titulo
from utils.loader import carregar_2023, carregar_2024, carregar_2025
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, JsCode, ColumnsAutoSizeMode

st.set_page_config(
    page_title="Análise Conjunta · JAUM DTC",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

aplicar_tema()

# ── Helper AgGrid — configuração restaurada do v0 (menu funcionando) ──────────
def ag_table(df, height=400):
    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_default_column(
        resizable=True,
        sortable=True,
        filter=True,
        suppressMenu=False,
        menuTabs=["generalMenuTab", "filterMenuTab", "columnsMenuTab"],
        cellStyle={"fontSize": "14px", "color": "#000000", "fontFamily": "Helvetica Neue, sans-serif"},
    )
    gb.configure_grid_options(
        headerHeight=36,
        rowHeight=32,
        domLayout="normal",
        suppressMenuHide=True,              # mantém ícone de menu sempre visível
        suppressColumnVirtualisation=True,
        suppressContextMenu=False,
        enableRangeSelection=True,
    )
    go = gb.build()
    go["defaultColDef"]["headerClass"] = "ag-header-black"
    go["onFirstDataRendered"] = JsCode("function(params) { params.api.sizeColumnsToFit(); }")
    AgGrid(
        df,
        gridOptions=go,
        height=height,
        update_mode=GridUpdateMode.NO_UPDATE,   # evita re-render que bloqueia menu
        fit_columns_on_grid_load=False,
        columns_auto_size_mode=2,
        allow_unsafe_jscode=True,
        enable_enterprise_modules=True,          # habilita módulos de filtro/menu avançados
        custom_css={
            ".ag-header":                      {"background-color": "#4A4A4A !important"},
            ".ag-header-row":                  {"background-color": "#4A4A4A !important"},
            ".ag-header-cell":                 {"background-color": "#4A4A4A !important"},
            ".ag-header-cell-label":           {"color": "#FFFFFF !important", "font-weight": "700"},
            ".ag-header-cell-text":            {"color": "#FFFFFF !important", "font-size": "13px !important", "font-weight": "700 !important"},
            ".ag-icon":                        {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-header-icon":                 {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-header-cell-menu-button":     {"opacity": "1 !important", "visibility": "visible !important"},
            ".ag-header-cell-menu-button span":{"color": "#FFFFFF !important"},
            ".ag-icon-menu":                   {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-icon-filter":                 {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-cell":                        {"font-size": "13px !important", "color": "#000000 !important"},
            ".ag-row":                         {"font-size": "13px !important"},
        },
        theme="streamlit",          # "streamlit" não conflita com custom_css de header escuro
        use_container_width=True,
    )


# ── Helper exportar Excel genérico (restaurado do v0) ────────────────────────
def exportar_excel(df, nome_arquivo="tabela.xlsx", label="⬇️ Exportar Excel", key=None):
    """Botão para exportar um DataFrame como Excel com formatação básica."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active

    df = df.reset_index(drop=True)
    df = df.loc[:, ~df.columns.str.startswith("::")].copy()

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=str(col))
        cell.font      = Font(bold=True, name="Arial", size=10, color="1A1A1A")
        cell.fill      = PatternFill("solid", start_color="F2F2F2")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(12, len(str(col)) + 2)
    ws.row_dimensions[1].height = 28

    for ri, row_data in enumerate(df.itertuples(index=False), start=2):
        for ci, val in enumerate(row_data, 1):
            try:
                if val is None or (isinstance(val, float) and np.isnan(val)):
                    val = None
                elif hasattr(val, '__class__') and type(val).__name__ in ('NAType', 'NaTType'):
                    val = None
                else:
                    if pd.isna(val):
                        val = None
            except (TypeError, ValueError):
                pass
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="left" if ci == 1 else "center", vertical="center")
            cell.border    = border

    wb.save(buf)
    buf.seek(0)
    st.download_button(
        label=label,
        data=buf,
        file_name=nome_arquivo,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=key,
    )


# CSS para fonte nas tabelas nativas do Streamlit
st.markdown("""
<style>
[data-testid="stDataFrame"] td,
[data-testid="stDataFrame"] th,
[data-testid="stDataFrame"] [role="columnheader"],
[data-testid="stDataFrame"] [role="columnheader"] span,
[data-testid="stDataFrame"] [role="columnheader"] div {
    font-size: 14px !important;
    font-weight: 700 !important;
    color: #000000 !important;
    opacity: 1 !important;
}
/* Escurece captions globalmente */
[data-testid="stCaptionContainer"] p,
[data-testid="stCaptionContainer"] {
    color: #374151 !important;
    opacity: 1 !important;
}
</style>
""", unsafe_allow_html=True)

page_header(
    "Análise Conjunta",
    "Avalie o desempenho produtivo dos materiais em faixa considerando múltiplos locais e safras. Identifique cultivares com alta média, baixo CV e boa produção relativa frente às testemunhas — esses são os candidatos ao avanço.",
    imagem="Business mission-amico.png",
)

# ── Carregamento ──────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def carregar_concat():
    frames = []
    for loader in [carregar_2023, carregar_2024, carregar_2025]:
        d = loader()
        if d.get("ok") and d.get("ta_faixa") is not None:
            frames.append(d["ta_faixa"])
    if not frames:
        return pd.DataFrame()
    df = pd.concat(frames, ignore_index=True)
    df["gm"] = pd.to_numeric(df["gm"], errors="coerce")
    df["gm_cat"] = (df["gm"] / 10).round(1)
    return df

with st.spinner("Carregando dados..."):
    ta_raw = carregar_concat()

if ta_raw.empty:
    st.error("❌ Nenhum dado disponível. Verifique a página de Diagnóstico.")
    st.stop()

# Garante que filtro de cultivar inicia vazio em sessões novas
if "_cult_initialized" not in st.session_state:
    st.session_state["_cult_sel"] = set()
    st.session_state["_cult_initialized"] = True

# Normalizar GM_visual: divide por 10 se mediana > 10 (ex: 80.9 → 8.09)
if "GM_visual" in ta_raw.columns:
    med_gm = ta_raw["GM_visual"].dropna()
    med_gm = med_gm[med_gm > 0]
    if len(med_gm) > 0 and med_gm.median() > 10:
        ta_raw["GM_visual"] = (ta_raw["GM_visual"] / 10).round(1)

# Fallback população: usa pop_plantas_ha quando pop_plantasFinal_ha estiver nula (ex: safra 2024/25)
if "pop_plantasFinal_ha" in ta_raw.columns and "pop_plantas_ha" in ta_raw.columns:
    mask = ta_raw["pop_plantasFinal_ha"].isna() & ta_raw["pop_plantas_ha"].notna()
    ta_raw.loc[mask, "pop_plantasFinal_ha"] = ta_raw.loc[mask, "pop_plantas_ha"]
elif "pop_plantas_ha" in ta_raw.columns and "pop_plantasFinal_ha" not in ta_raw.columns:
    ta_raw["pop_plantasFinal_ha"] = ta_raw["pop_plantas_ha"]

# ── Sidebar — Filtros encadeados ──────────────────────────────────────────────
with st.sidebar:
    st.markdown(
        '<p style="font-size:11px;font-weight:600;color:#6B7280;text-transform:uppercase;'
        'letter-spacing:0.05em;padding: 0.5rem;">Filtros</p>',
        unsafe_allow_html=True,
    )

    if st.button("🔄 Limpar filtros", use_container_width=True):
        for key in list(st.session_state.keys()):
            if any(key.startswith(p) for p in ["safra_","macro_","micro_","estado_","cidade_",
                                                "fazenda_","resp_","status_","cult_"]):
                del st.session_state[key]
        for key in ["_cidade_sel", "_cult_sel", "busca_cidade", "busca_cult"]:
            if key in st.session_state:
                del st.session_state[key]
        st.rerun()

    def checkboxes(label, opcoes, default_all=True, defaults=None, prefix=""):
        sel = []
        for o in opcoes:
            checked = (o in defaults) if defaults is not None else default_all
            if st.checkbox(str(o), value=checked, key=f"{prefix}_{o}"):
                sel.append(o)
        return sel

    # ── 1. Safra — padrão só 2025/26 ──────────────────────────────────────────
    with st.expander("📅 Safra", expanded=True):
        safras_all = sorted(ta_raw["safra"].dropna().unique().tolist())
        safra_default = [s for s in safras_all if "2025" in str(s)] or safras_all[-1:]
        safras_sel = checkboxes("Safra", safras_all, defaults=safra_default, prefix="safra")

    ta_f1 = ta_raw[ta_raw["safra"].isin(safras_sel)] if safras_sel else ta_raw.iloc[0:0]

    # ── 2. Região Macro ────────────────────────────────────────────────────────
    with st.expander("🗺️ Região Macro", expanded=False):
        macros_all = sorted(ta_f1["regiao_macro"].dropna().unique().tolist())
        macros_sel = checkboxes("Macro", macros_all, prefix="macro")

    ta_f2 = ta_f1[ta_f1["regiao_macro"].isin(macros_sel)] if macros_sel else ta_f1.iloc[0:0]

    # ── 3. Região Micro ────────────────────────────────────────────────────────
    with st.expander("📍 Região Micro", expanded=False):
        micros_all = sorted(ta_f2["regiao_micro"].dropna().unique().tolist())
        micros_sel = checkboxes("Micro", micros_all, prefix="micro")

    ta_f3 = ta_f2[ta_f2["regiao_micro"].isin(micros_sel)] if micros_sel else ta_f2.iloc[0:0]

    # ── 4. Estado ──────────────────────────────────────────────────────────────
    with st.expander("🏛️ Estado", expanded=False):
        estados_all = sorted(ta_f3["estado_sigla"].dropna().unique().tolist())
        estados_sel = checkboxes("Estado", estados_all, prefix="estado")

    ta_f4 = ta_f3[ta_f3["estado_sigla"].isin(estados_sel)] if estados_sel else ta_f3.iloc[0:0]

    # ── 5. Cidade ──────────────────────────────────────────────────────────────
    with st.expander("🏙️ Cidade", expanded=False):
        cidades_all = sorted(ta_f4["cidade_nome"].dropna().unique().tolist())
        if "_cidade_sel" not in st.session_state:
            st.session_state["_cidade_sel"] = set(cidades_all)
        # remove cidades que saíram do escopo
        st.session_state["_cidade_sel"] &= set(cidades_all)
        busca_cidade = st.text_input("🔍 Buscar cidade", value="", key="busca_cidade",
                                     placeholder="Digite parte do nome...")
        cidades_filtradas = (
            [c for c in cidades_all if busca_cidade.strip().lower() in c.lower()]
            if busca_cidade.strip() else cidades_all
        )
        hidden_sel_cid = [c for c in st.session_state["_cidade_sel"] if c not in cidades_filtradas]
        if hidden_sel_cid:
            st.caption(f"✓ {len(hidden_sel_cid)} selecionada(s) fora da busca")
        if busca_cidade.strip() and not cidades_filtradas:
            st.caption("Nenhuma cidade encontrada.")
        for c in cidades_filtradas:
            val = st.checkbox(c, value=(c in st.session_state["_cidade_sel"]), key=f"cidade_{c}")
            if val:
                st.session_state["_cidade_sel"].add(c)
            else:
                st.session_state["_cidade_sel"].discard(c)
        cidades_sel = list(st.session_state["_cidade_sel"])

    ta_f5 = ta_f4[ta_f4["cidade_nome"].isin(cidades_sel)] if cidades_sel else ta_f4.iloc[0:0]

    # ── 6. Fazenda ─────────────────────────────────────────────────────────────
    with st.expander("🏡 Fazenda", expanded=False):
        fazendas_all = sorted(ta_f5["nomeFazenda"].dropna().unique().tolist())
        fazendas_sel = checkboxes("Fazenda", fazendas_all, prefix="fazenda")

    ta_f6 = ta_f5[ta_f5["nomeFazenda"].isin(fazendas_sel)] if fazendas_sel else ta_f5.iloc[0:0]

    # ── 7. Responsável ─────────────────────────────────────────────────────────
    with st.expander("👤 Responsável", expanded=False):
        resps_all = sorted(ta_f6["nomeResponsavel"].dropna().unique().tolist())
        resps_sel = checkboxes("Resp", resps_all, prefix="resp")

    ta_f7 = ta_f6[ta_f6["nomeResponsavel"].isin(resps_sel)] if resps_sel else ta_f6.iloc[0:0]

    # ── 8. Status do material ──────────────────────────────────────────────────
    with st.expander("🏷️ Status do Material", expanded=False):
        status_all = sorted(ta_f7["status_material"].dropna().unique().tolist())
        status_sel = checkboxes("Status", status_all, prefix="status")

    ta_f8 = ta_f7[ta_f7["status_material"].isin(status_sel)] if status_sel else ta_f7.iloc[0:0]

    # ── 9. Cultivar (dePara) ──────────────────────────────────────────────────
    with st.expander("🌱 Cultivar", expanded=False):
        cultivares_all = sorted(ta_f8["dePara"].dropna().unique().tolist())
        if "_cult_sel" not in st.session_state:
            st.session_state["_cult_sel"] = set()
        # remove cultivares que saíram do escopo
        st.session_state["_cult_sel"] &= set(cultivares_all)
        busca_cult = st.text_input("🔍 Buscar cultivar", value="", key="busca_cult",
                                   placeholder="Digite parte do nome...")
        cultivares_filtrados = (
            [c for c in cultivares_all if busca_cult.strip().lower() in c.lower()]
            if busca_cult.strip() else cultivares_all
        )
        hidden_sel_cult = [c for c in st.session_state["_cult_sel"] if c not in cultivares_filtrados]
        if hidden_sel_cult:
            st.caption(f"✓ {len(hidden_sel_cult)} selecionado(s) fora da busca")
        if busca_cult.strip() and not cultivares_filtrados:
            st.caption("Nenhum cultivar encontrado.")
        for c in cultivares_filtrados:
            val = st.checkbox(c, value=(c in st.session_state["_cult_sel"]), key=f"cult_{c}")
            if val:
                st.session_state["_cult_sel"].add(c)
            else:
                st.session_state["_cult_sel"].discard(c)
        cultivares_sel = list(st.session_state["_cult_sel"])

    # Sem seleção = sem filtro (todos passam) — permite selecionar só os que quer comparar
    ta_f9 = ta_f8[ta_f8["dePara"].isin(cultivares_sel)] if cultivares_sel else ta_f8

    # ── 10. GM — slider ────────────────────────────────────────────────────────
    with st.expander("🎯 Grupo de Maturidade", expanded=False):
        gm_min = float(round(ta_f9["gm_cat"].min(), 1)) if not ta_f9.empty else 5.0
        gm_max = float(round(ta_f9["gm_cat"].max(), 1)) if not ta_f9.empty else 9.0
        if gm_min >= gm_max:
            gm_max = round(gm_min + 0.1, 1)
        gm_range = st.slider("GM", min_value=gm_min, max_value=gm_max,
                             value=(gm_min, gm_max), step=0.1, format="%.1f")

    ta_filtrado = ta_f9[ta_f9["gm_cat"].between(gm_range[0], gm_range[1])]

# ── Aviso se sem dados ────────────────────────────────────────────────────────
if ta_filtrado.empty:
    st.warning("⚠️ Nenhum dado para os filtros selecionados.")
    st.stop()


# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 1 — TABELA DE DADOS BRUTOS
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Auditoria",
    "Quais são os dados por ensaio?",
    "Visão individual de cada observação. Use para auditoria e conferência dos dados antes da análise.",
)

# ── Seletor de Produção Relativa ──────────────────────────────────────────────
col_ref, col_test, _ = st.columns([2, 2, 3])

with col_ref:
    base_rel = st.selectbox(
        "Base da Produção Relativa",
        options=["Média geral do ensaio", "Maior produtividade", "Testemunha"],
        index=0,
    )

with col_test:
    if base_rel == "Testemunha":
        testemunhas = sorted(
            ta_filtrado[ta_filtrado["status_material"].isin(["CHECK", "STINE"])]["dePara"].dropna().unique().tolist()
        )
        if testemunhas:
            testemunha_sel = st.selectbox("Selecione a testemunha", options=testemunhas)
        else:
            st.warning("Nenhuma testemunha disponível nos filtros atuais.")
            testemunha_sel = None
    else:
        testemunha_sel = None

# ── Calcular produção relativa ────────────────────────────────────────────────
df_tabela = ta_filtrado.copy()

if base_rel == "Média geral do ensaio":
    ref_valor = df_tabela["kg_ha"].mean()
elif base_rel == "Maior produtividade":
    ref_valor = df_tabela["kg_ha"].max()
elif base_rel == "Testemunha" and testemunha_sel:
    ref_valor = df_tabela[df_tabela["dePara"] == testemunha_sel]["kg_ha"].mean()
else:
    ref_valor = df_tabela["kg_ha"].mean()

df_tabela["prod_relativa_pct"] = ((df_tabela["kg_ha"] / ref_valor) * 100).round(1) if ref_valor else np.nan

# ── Colunas para exibir ───────────────────────────────────────────────────────
col_map = {
    "safra":               "Safra",
    "cod_fazenda":         "Cód. Local",
    "nomeFazenda":         "Fazenda",
    "cidade_nome":         "Cidade",
    "estado_sigla":        "Estado",
    "regiao_macro":        "Região Macro",
    "regiao_micro":        "Região Micro",
    "nomeResponsavel":     "Responsável",
    "dePara":              "Cultivar",
    "status_material":     "Status",
    "gm_cat":              "GM",
    "GM_visual":           "GM Visual",
    "kg_ha":               "kg/ha",
    "sc_ha":               "sc/ha",
    "prod_relativa_pct":   "Prod. Relativa (%)",
    "umidadeParcela":      "Umidade (%)",
    "pop_plantasFinal_ha": "Pop. Final (pl/ha)",
    "media_AIV":           "AIV (cm)",
    "media_ALT":           "ALP (cm)",
    "notaAC":              "Acamamento",
    "dias_ate_DMF":        "Ciclo (dias)",
}

cols_disp = [c for c in col_map.keys() if c in df_tabela.columns]
df_show = df_tabela[cols_disp].rename(columns=col_map)

ag_table(df_show, height=min(400, 36 + 32 * len(df_show) + 20))
exportar_excel(df_show, nome_arquivo="auditoria.xlsx", label="⬇️ Exportar Auditoria", key="exp_auditoria")

st.divider()

# ── Calcular LSD via QMR da ANOVA (cultivar × local) ─────────────────────────
def calcular_lsd(df, col="kg_ha", fator="dePara", bloco="cod_fazenda", alpha=0.05):
    """
    LSD para análise conjunta via OLS numpy (rápido, sem dependências extras).
    Funciona com delineamentos balanceados e desbalanceados.
    Modelo: y = µ + cultivar + local + erro
    """
    try:
        d = df[[col, fator, bloco]].dropna().copy()
        d = d[d[col] > 0].reset_index(drop=True)
        if d.empty or d[fator].nunique() < 2 or d[bloco].nunique() < 2:
            return np.nan
        y       = d[col].values.astype(float)
        X_cult  = pd.get_dummies(d[fator],  drop_first=True).values.astype(float)
        X_local = pd.get_dummies(d[bloco],  drop_first=True).values.astype(float)
        X       = np.hstack([np.ones((len(y), 1)), X_cult, X_local])
        beta, _, rank, _ = np.linalg.lstsq(X, y, rcond=None)
        ss_res  = np.sum((y - X @ beta) ** 2)
        gl_res  = len(y) - rank
        if gl_res <= 0:
            return np.nan
        qmr     = ss_res / gl_res
        n_bloco = d[bloco].nunique()
        t_crit  = stats.t.ppf(1 - alpha / 2, df=gl_res)
        return round(t_crit * np.sqrt(2 * qmr / n_bloco), 1)
    except Exception:
        return np.nan


# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 2 — DESCRITIVA GERAL DO CONJUNTO
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Conjunto",
    "Como está o experimento como um todo?",
    "Estatísticas descritivas por variável considerando todos os ensaios filtrados.",
)

vars_desc = {
    "kg_ha":               "kg/ha",
    "sc_ha":               "sc/ha",
    "umidadeParcela":      "Umidade (%)",
    "pop_plantasFinal_ha": "Pop. Final",
    "media_AIV":           "AIV (cm)",
    "media_ALT":           "ALP (cm)",
    "notaAC":              "AC",
    "dias_ate_DMF":        "Ciclo",
}

medidas = ["Total de Observações", "Média", "Desvio Padrão", "Mínimo",
           "1º Quartil", "Mediana", "3º Quartil", "Máximo", "CV (%)", "LSD (5%)", "Locais"]

rows_geral = {m: {} for m in medidas}

for col, label in vars_desc.items():
    if col not in ta_filtrado.columns:
        continue
    serie = ta_filtrado[col].dropna()
    if col in ["kg_ha", "sc_ha", "pop_plantasFinal_ha", "notaAC", "media_AIV", "media_ALT", "dias_ate_DMF", "umidadeParcela"]:
        serie = serie[serie > 0]
    if len(serie) == 0:
        for m in medidas:
            rows_geral[m][label] = "—"
        continue

    media = serie.mean()
    dp    = serie.std()
    cv    = round(dp / media * 100, 2) if media > 0 else np.nan
    q1, q2, q3 = serie.quantile([0.25, 0.50, 0.75])
    lsd_kg = calcular_lsd(ta_filtrado)
    if col == "kg_ha":
        lsd = round(lsd_kg, 1) if isinstance(lsd_kg, (int, float)) and not np.isnan(lsd_kg) else "—"
    elif col == "sc_ha":
        lsd = round(lsd_kg / 60, 2) if isinstance(lsd_kg, (int, float)) and not np.isnan(lsd_kg) else "—"
    else:
        lsd = "—"
    locais = ta_filtrado["cod_fazenda"].nunique()

    rows_geral["Total de Observações"][label] = int(len(serie))
    rows_geral["Média"][label]                = round(media, 2)
    rows_geral["Desvio Padrão"][label]        = round(dp, 2)
    rows_geral["Mínimo"][label]               = round(serie.min(), 2)
    rows_geral["1º Quartil"][label]           = round(q1, 2)
    rows_geral["Mediana"][label]              = round(q2, 2)
    rows_geral["3º Quartil"][label]           = round(q3, 2)
    rows_geral["Máximo"][label]               = round(serie.max(), 2)
    rows_geral["CV (%)"][label]               = round(cv, 2) if not np.isnan(cv) else "—"
    rows_geral["LSD (5%)"][label]             = lsd
    rows_geral["Locais"][label]               = locais

df_geral = pd.DataFrame(rows_geral).T.reset_index().rename(columns={"index": "Medida"})
ag_table(df_geral, height=425)
exportar_excel(df_geral, nome_arquivo="descritiva_conjunto.xlsx", label="⬇️ Exportar Descritiva Geral", key="exp_geral")
st.caption(
    "ℹ️ **CV (%) desta tabela** = Desvio Padrão ÷ Média × 100, calculado sobre todas as observações brutas. "
    "Inclui a variação entre cultivares, entre locais e o erro experimental — por isso tende a ser maior. "
    "Para avaliar a qualidade do experimento, use o **CV da ANOVA** exibido na Tabela de Apresentação, "
    "que desconta os efeitos de cultivar e local, retendo apenas o erro residual."
)
st.caption(
    "ℹ️ **LSD (5%)** — Se a diferença de produtividade entre dois cultivares for maior que este valor, "
    "ela é real e não fruto do acaso (95% de confiança). "
    "Tecnicamente: Diferença Mínima Significativa = t(α/2, gl_resíduo) × √(2 × QMR / nº de locais), "
    "onde QMR é o Quadrado Médio do Resíduo da ANOVA conjunta (modelo: y = μ + cultivar + local + erro)."
)

st.divider()

# ════════════════════════════════════════════════════════════════════════════════

st.divider()

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 3 — GM × PRODUTIVIDADE
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "GM × Produtividade",
    "Como cada material se posiciona dentro do seu grupo de maturação?",
    "Compare cultivares no mesmo GM e entre grupos — identifique quais materiais se destacam dentro do seu nicho de ciclo.",
)

with st.popover("ℹ️ Como interpretar · GM × Produtividade", use_container_width=False):
    st.markdown("""
**📌 O que este gráfico mostra**

Cada ponto é um cultivar. A posição horizontal representa o **Grupo de Maturação (GM)** — quanto mais à direita, mais tardio o material. A posição vertical representa a **produtividade média (sc/ha)** no conjunto filtrado.

---

**📐 Como ler**

- **Eixo X → GM** — grupo de maturação do cultivar (ex: 6.5, 7.0, 7.5...)
- **Eixo Y → sc/ha** — média de produtividade no conjunto de locais filtrado
- **Cor do ponto** → status do material (CHECK, STINE, DP2)
- **Nome** → identificação do cultivar

---

**💡 O que observar**

- **Tendência ascendente** → materiais mais tardios produzem mais neste ambiente — o GM mais alto está sendo bem aproveitado
- **Tendência plana ou descendente** → o GM não explica a produtividade — outros fatores dominam
- **Pontos isolados acima da tendência** → cultivares que superam o esperado para o seu GM — candidatos a destaque
- **Pontos abaixo da tendência** → cultivares que ficaram aquém do esperado para seu grupo de maturação
""")

# ── Cálculo ──────────────────────────────────────────────────────────────────
df_gm = (
    ta_filtrado[ta_filtrado["sc_ha"] > 0]
    .groupby("dePara")
    .agg(
        media_sc  = ("sc_ha",    "mean"),
        gm        = ("gm_cat",   "median"),
        status    = ("status_material", "first"),
    )
    .reset_index()
    .dropna(subset=["gm"])
)
df_gm["media_sc"] = df_gm["media_sc"].round(1)
df_gm["gm"]       = df_gm["gm"].round(1)

fig_gm = go_plt.Figure()

for status, cor in COR_STATUS_PLOT.items():
    df_s = df_gm[df_gm["status"] == status]
    if df_s.empty: continue
    fig_gm.add_trace(go_plt.Scatter(
        x=df_s["gm"],
        y=df_s["media_sc"],
        mode="markers+text",
        name=status,
        text=df_s["dePara"],
        textposition="top center",
        textfont=dict(size=13, color="#333333", weight="bold"),
        marker=dict(
            color=cor,
            size=14,
            line=dict(color=COR_BORDA.get(status, "#888"), width=1.5),
            opacity=0.9,
        ),
        hovertemplate=(
            "<b>%{text}</b><br>"
            "GM: %{x}<br>"
            "Média: %{y:.1f} sc/ha<extra></extra>"
        ),
    ))

# Linha de tendência geral (regressão linear simples)
if len(df_gm) >= 3:
    x_tr = df_gm["gm"].values
    y_tr = df_gm["media_sc"].values
    z    = np.polyfit(x_tr, y_tr, 1)
    p    = np.poly1d(z)
    x_line = np.linspace(x_tr.min(), x_tr.max(), 100)
    fig_gm.add_trace(go_plt.Scatter(
        x=x_line, y=p(x_line),
        mode="lines",
        name="Tendência",
        line=dict(color="#AAAAAA", width=1.5, dash="dash"),
        hoverinfo="skip",
    ))

fig_gm.update_layout(
    height=500,
    plot_bgcolor="#F5F5F5", paper_bgcolor="#FFFFFF",
    font=dict(family="Helvetica Neue, sans-serif", size=13, color="#111111"),
    xaxis=dict(
        title=dict(text="<b>Grupo de Maturação (GM)</b>", font=dict(size=14, color="#111111")),
        tickfont=dict(size=12, color="#111111", weight="bold"),
        gridcolor="#FFFFFF", gridwidth=1.5,
        zeroline=False,
        dtick=0.1,
        range=[
            round(df_gm["gm"].min() - 0.2, 1),
            round(df_gm["gm"].max() + 0.2, 1),
        ],
        showline=True, linecolor="#CCCCCC", linewidth=1,
    ),
    yaxis=dict(
        title=dict(text="<b>Produtividade média (sc/ha)</b>", font=dict(size=14, color="#111111")),
        tickfont=dict(size=12, color="#111111", weight="bold"),
        gridcolor="#FFFFFF", gridwidth=1.5,
        zeroline=False,
        showline=True, linecolor="#CCCCCC", linewidth=1,
    ),
    legend=dict(
        title=dict(text="<b>Status</b>", font=dict(size=13, color="#111111")),
        orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0,
        font=dict(size=13, color="#111111", weight="bold"),
    ),
    margin=dict(t=60, b=60, l=70, r=40),
)

st.plotly_chart(fig_gm, use_container_width=True)
st.caption(
    "ℹ️ Cada ponto = média de produtividade do cultivar no conjunto filtrado. "
    "Linha tracejada = tendência linear entre GM e produtividade. "
    "Pontos acima da linha superam o esperado para o seu GM."
)

# SEÇÃO 4 — ESTATÍSTICAS DESCRITIVAS POR CULTIVAR
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Desempenho",
    "Como cada material se comportou no conjunto?",
    "Estatísticas descritivas agrupadas por cultivar. Avalie consistência, variabilidade e desempenho relativo entre locais.",
)

with st.popover("ℹ️ Como interpretar · Legenda", use_container_width=False):
    st.markdown("""
**📌 Como interpretar esta tabela**

Cada linha representa um cultivar com suas estatísticas calculadas sobre **todos os locais nos filtros ativos**.
Os materiais são ordenados por **Média (kg/ha) decrescente**.

---

**📋 Glossário das colunas**
""")
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("""
| Coluna | Descrição |
|---|---|
| **Status** | Categoria do material (LINHAGEM, DP2, STINE, CHECK) |
| **GM** | Grupo de Maturidade (escala 5.5–8.5) |
| **Locais** | Nº de locais onde o cultivar foi avaliado |
| **N** | Nº total de observações válidas (kg/ha > 0) |
| **Média (kg/ha)** | Média aritmética de produtividade entre os locais |
| **Prod. Relativa (%)** | Quanto o cultivar produziu em relação à base escolhida no seletor acima da tabela de auditoria — 100% = igual à base, acima de 100% = superou, abaixo de 100% = ficou aquém |
""")
    with col2:
        st.markdown("""
| Coluna | Descrição |
|---|---|
| **DP** | Desvio Padrão (kg/ha) — dispersão absoluta entre locais |
| **CV (%)** | Coeficiente de Variação — dispersão relativa (DP ÷ Média × 100) |
| **Mínimo / Máximo** | Menor e maior valor observado entre os locais |
| **Q1 / Mediana / Q3** | Faixa de produtividade: Q1 = piso (75% dos locais produziram acima), Mediana = valor central, Q3 = teto (25% dos locais produziram acima) |
""")
    st.markdown("""
---

**📐 Entendendo cada medida**

> ⚠️ **Média e Mediana são medidas diferentes e podem contar histórias distintas sobre o mesmo cultivar.** Quando estiverem próximas, o desempenho é equilibrado. Quando divergirem, vale investigar quais ambientes estão puxando a média para cima ou para baixo.

- **Média:** é a soma de todas as produtividades dividida pelo número de locais. É o valor mais conhecido, mas pode ser "puxada" para cima por uma área excepcional ou para baixo por uma área muito ruim — por isso nem sempre representa bem o comportamento típico do cultivar.

- **Mediana (Q2):** é o valor "do meio" quando todos os resultados estão em ordem. Em campo: é a produtividade que representa o "típico" do cultivar — metade das áreas produz menos que isso e metade produz mais. **Diferente da média, não sofre distorção por áreas muito ruins ou muito boas.**

- **Q1 (primeiro quartil):** é o ponto que separa os 25% piores resultados dos 75% restantes. Em campo: é o **teto inferior** — abaixo de Q1 estão as situações em que o cultivar costuma ir mal. Um Q1 alto significa que mesmo nos piores ambientes o cultivar ainda entrega uma produtividade razoável — sinal de segurança na recomendação.

- **Q3 (terceiro quartil):** é o ponto que separa os 75% melhores resultados dos 25% mais altos. Em campo: acima de Q3 estão as áreas onde o cultivar performa entre os melhores resultados que ele consegue entregar.

---

**🎯 O que observar**

- **CV baixo (< 15%)** → cultivar consistente entre locais — menor risco de recomendação
- **CV alto (> 25%)** → adaptação específica — bom em alguns locais, ruim em outros
- **DP baixo** → pouca variação absoluta de produtividade entre locais — comportamento previsível independente do ambiente
- **Mediana próxima da média** → distribuição equilibrada, sem locais muito discrepantes
- **Q1 alto** → piso de produtividade elevado — cultivar seguro mesmo nos piores ambientes
""")

desc_rows = []

for cultivar, grp in ta_filtrado.groupby("dePara", dropna=True):
    kg = grp["kg_ha"].dropna()
    kg = kg[kg > 0]
    if len(kg) == 0:
        continue

    q1, q2, q3 = kg.quantile([0.25, 0.50, 0.75])
    media = kg.mean()
    dp    = kg.std()
    cv    = (dp / media * 100) if media > 0 else np.nan
    locais = grp["cod_fazenda"].nunique()

    prod_rel = ((media / ref_valor) * 100) if ref_valor else np.nan

    desc_rows.append({
        "Cultivar":           cultivar,
        "Status":             grp["status_material"].mode()[0] if not grp["status_material"].mode().empty else "",
        "GM":                 grp["gm_cat"].mode()[0] if not grp["gm_cat"].mode().empty else "",
        "Locais":             locais,
        "N":                  int(len(kg)),
        "Média (kg/ha)":      round(media, 1),
        "Prod. Relativa (%)": round(prod_rel, 1) if not np.isnan(prod_rel) else np.nan,
        "DP":                 round(dp, 1) if not np.isnan(dp) else "—",
        "CV (%)":             round(cv, 1) if not np.isnan(cv) else "—",
        "Mínimo":             round(kg.min(), 1),
        "Q1":                 round(q1, 1),
        "Mediana":            round(q2, 1),
        "Q3":                 round(q3, 1),
        "Máximo":             round(kg.max(), 1),
    })

df_desc = pd.DataFrame(desc_rows).sort_values("Média (kg/ha)", ascending=False).reset_index(drop=True)

ag_table(df_desc, height=min(600, 36 + 32 * len(df_desc) + 20))
exportar_excel(df_desc, nome_arquivo="desempenho_cultivares.xlsx", label="⬇️ Exportar Desempenho", key="exp_desc")

st.divider()

# ── Pré-calcular df_apres (usado no lollipop e na tabela de apresentação) ─────
_cols_apres = {
    "kg_ha":               "kg/ha",
    "sc_ha":               "sc/ha",
    "umidadeParcela":      "Umidade (%)",
    "pop_plantasFinal_ha": "Pop. Final",
    "media_AIV":           "AIV (cm)",
    "media_ALT":           "ALP (cm)",
    "notaAC":              "AC",
    "dias_ate_DMF":        "Ciclo",
    "GM_visual":           "GM Visual",
}
_apres_rows = []
for _cultivar, _grp in ta_filtrado.groupby("dePara", dropna=True):
    _row = {
        "Cultivar": _cultivar,
        "status_material": _grp["status_material"].mode()[0] if not _grp["status_material"].mode().empty else "",
    }
    for _col, _label in _cols_apres.items():
        if _col not in _grp.columns:
            _row[_label] = None
            continue
        _serie = _grp[_col].dropna()
        if _col in ["kg_ha","sc_ha","pop_plantasFinal_ha","notaAC","media_AIV","media_ALT","dias_ate_DMF","umidadeParcela","GM_visual"]:
            _serie = _serie[_serie > 0]
        if _col == "pop_plantasFinal_ha":
            _row[_label] = int(round(_serie.mean(), 0)) if len(_serie) > 0 else None
        else:
            _row[_label] = round(_serie.mean(), 1) if len(_serie) > 0 else None
    _apres_rows.append(_row)

df_apres = pd.DataFrame(_apres_rows)
if not df_apres.empty:
    df_apres = df_apres.sort_values("sc/ha", ascending=False).reset_index(drop=True)

# ── LSD pré-calculado (usado no lollipop e na tabela) ─────────────────────────
_lsd_apres = calcular_lsd(ta_filtrado, col="kg_ha")
lsd_sc = round(_lsd_apres / 60, 2) if isinstance(_lsd_apres, (int, float)) and not np.isnan(_lsd_apres) else None

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 6 — RANKING E LOLLIPOP
# ════════════════════════════════════════════════════════════════════════════════

# Subtítulo com filtros ativos
_all_safras  = sorted(ta_raw["safra"].dropna().unique().tolist())
_all_macros  = sorted(ta_raw["regiao_macro"].dropna().unique().tolist()) if "regiao_macro" in ta_raw.columns else []
_all_micros  = sorted(ta_raw["regiao_micro"].dropna().unique().tolist()) if "regiao_micro" in ta_raw.columns else []
_all_estados = sorted(ta_raw["estado"].dropna().unique().tolist()) if "estado" in ta_raw.columns else []
_all_cidades = sorted(ta_raw["municipio"].dropna().unique().tolist()) if "municipio" in ta_raw.columns else []

filtros_ativos = []
if safras_sel and set(safras_sel) != set(_all_safras):
    filtros_ativos.append(" / ".join(str(s) for s in safras_sel))
if macros_sel and set(macros_sel) != set(_all_macros):
    filtros_ativos.append("Macro: " + ", ".join(macros_sel))
if micros_sel and set(micros_sel) != set(_all_micros):
    filtros_ativos.append("Micro: " + ", ".join(micros_sel))
if estados_sel and set(estados_sel) != set(_all_estados):
    filtros_ativos.append(", ".join(estados_sel))
if cidades_sel and set(cidades_sel) != set(_all_cidades):
    if len(cidades_sel) <= 3:
        filtros_ativos.append(", ".join(cidades_sel))
contexto_str = "  ·  ".join(filtros_ativos) if filtros_ativos else "Todos os dados"

# Adiciona contagem de fazendas ao contexto
n_fazendas_ctx = ta_filtrado["cod_fazenda"].nunique()
n_cidades_ctx  = ta_filtrado["cidade_nome"].nunique()
contexto_str   = contexto_str + f"  ·  {n_cidades_ctx} cidades · {n_fazendas_ctx} locais"

# ── GRÁFICO 1 — LOLLIPOP ─────────────────────────────────────────────────────
secao_titulo(
    "Visualização",
    "Quem lidera o ranking de produtividade?",
    contexto_str,
)

with st.popover("ℹ️ Como interpretar · Ranking", use_container_width=False):
    st.markdown("""
**📌 Como ler este gráfico**

Cada ponto representa a **média de sc/ha** de um cultivar considerando todos os locais nos filtros ativos.
Os materiais são ordenados do mais produtivo (topo) ao menos produtivo (base).

---

**🔲 Elementos do gráfico**
""")
    col1, col2, col3 = st.columns(3)
    col1.markdown('<div style="background:#fff;border:1px solid #ccc;padding:8px;border-radius:4px;text-align:center;font-size:13px;"><b>Ponto colorido</b><br>Média sc/ha do cultivar<br>(cor = status do material)</div>', unsafe_allow_html=True)
    col2.markdown('<div style="background:#fff;border:1px solid #ccc;padding:8px;border-radius:4px;text-align:center;font-size:13px;"><b>Linha tracejada cinza</b><br>Média geral do conjunto<br>(referência visual)</div>', unsafe_allow_html=True)
    col3.markdown('<div style="background:#fff;border:1px solid #ccc;padding:8px;border-radius:4px;text-align:center;font-size:13px;"><b>Linha vermelha pontilhada</b><br>Corte pelo LSD (5%)<br>separa grupos estatísticos</div>', unsafe_allow_html=True)
    st.markdown("""
---

**🎨 Legenda de cores**
""")
    col1, col2, col3 = st.columns(3)
    col1.markdown('<div style="background:#C4DFB4;padding:6px 10px;border-radius:4px;text-align:center;font-size:13px;"><b>LINHAGEM / DP2</b><br>materiais em avaliação</div>', unsafe_allow_html=True)
    col2.markdown('<div style="background:#F4B184;padding:6px 10px;border-radius:4px;text-align:center;font-size:13px;"><b>CHECK</b><br>testemunhas externas</div>', unsafe_allow_html=True)
    col3.markdown('<div style="background:#2976B6;color:#fff;padding:6px 10px;border-radius:4px;text-align:center;font-size:13px;"><b>STINE</b><br>materiais comerciais Stine</div>', unsafe_allow_html=True)
    st.markdown("""
---

**📊 Interpretação prática**

- Materiais **acima da linha vermelha** não diferem estatisticamente do melhor — são os **candidatos ao avanço**
- Materiais **entre linhas vermelhas** formam um grupo intermediário
- Materiais **abaixo da última linha vermelha** são estatisticamente inferiores ao grupo de elite
- Quando **não há linha vermelha**, não há diferença estatística significativa entre os materiais
""")

if df_apres.empty:
    st.warning("⚠️ Nenhum dado para exibir.")
else:
    df_plot = df_apres[["Cultivar", "sc/ha", "status_material"]].dropna(subset=["sc/ha"]).copy()
    df_plot = df_plot.sort_values("sc/ha", ascending=False).reset_index(drop=True)

    # Adicionar n de observações válidas por cultivar
    n_obs = (
        ta_filtrado[ta_filtrado["sc_ha"] > 0]
        .groupby("dePara")["sc_ha"]
        .count()
        .rename("n_obs")
    )
    df_plot = df_plot.merge(n_obs, left_on="Cultivar", right_index=True, how="left")
    df_plot["n_obs"] = df_plot["n_obs"].fillna(0).astype(int)

    COR_STATUS_PLOT = {
        "CHECK":    "#F4B184",
        "STINE":    "#2976B6",
        "LINHAGEM": "#00FF01",
        "DP2":      "#C4DFB4",
    }
    COR_BORDA = {
        "CHECK":    "#C46A3A",
        "STINE":    "#1A4F7A",
        "LINHAGEM": "#009900",
        "DP2":      "#7AAF6A",
    }

    media_plot = df_plot["sc/ha"].mean()

    # Recalcular cortes para o plot
    sc_desc = df_apres["sc/ha"].dropna().tolist()
    if lsd_sc:
        lider_p = sc_desc[0]
        for i in range(1, len(sc_desc)):
            v = sc_desc[i]
            if v is not None and (lider_p - v) > lsd_sc:
                lider_p = v

    fig = go_plt.Figure()

    for _, row in df_plot.iterrows():
        fig.add_shape(
            type="line",
            x0=0, x1=row["sc/ha"],
            y0=row["Cultivar"], y1=row["Cultivar"],
            line=dict(color="#DDDDDD", width=1.5),
        )

    cores_pt  = [COR_STATUS_PLOT.get(s, "#AAAAAA") for s in df_plot["status_material"]]
    bordas_pt = [COR_BORDA.get(s, "#888888") for s in df_plot["status_material"]]

    fig.add_trace(go_plt.Scatter(
        x=df_plot["sc/ha"],
        y=df_plot["Cultivar"],
        mode="markers+text",
        name="",
        showlegend=False,
        marker=dict(color=cores_pt, size=16, line=dict(color=bordas_pt, width=1.5)),
        text=[f"  {v:.1f} ({n})" for v, n in zip(df_plot["sc/ha"], df_plot["n_obs"])],
        textposition="middle right",
        textfont=dict(size=13, color="#000000"),
        hovertemplate="<b>%{y}</b><br>sc/ha: %{x:.1f}<extra></extra>",
    ))

    for status, cor in COR_STATUS_PLOT.items():
        if status in df_plot["status_material"].values:
            fig.add_trace(go_plt.Scatter(
                x=[None], y=[None], mode="markers", name=status,
                marker=dict(color=cor, size=12, line=dict(color=COR_BORDA.get(status, "#888"), width=1.5)),
            ))

    col_chk1, col_chk2 = st.columns(2)
    mostrar_lsd_lol  = col_chk1.checkbox("Mostrar linhas de corte LSD", value=True, key="chk_lsd_lol")
    mostrar_media_lol = col_chk2.checkbox("Mostrar linha de média", value=True, key="chk_media_lol")

    if mostrar_media_lol:
        fig.add_vline(
            x=media_plot,
            line=dict(color="#888888", width=1.5, dash="dash"),
            annotation_text=f"Média: {media_plot:.1f}",
            annotation_position="top",
            annotation_font=dict(size=13, color="#333333", weight="bold"),
        )

    # Linhas de corte LSD
    sc_vals_asc  = df_plot["sc/ha"].tolist()
    cultivares_asc = df_plot["Cultivar"].tolist()
    if lsd_sc and mostrar_lsd_lol:
        lider_p2   = df_apres["sc/ha"].max()
        sc_desc2   = df_apres.sort_values("sc/ha", ascending=False)["sc/ha"].tolist()
        cult_desc2 = df_apres.sort_values("sc/ha", ascending=False)["Cultivar"].tolist()
        for i in range(1, len(sc_desc2)):
            v = sc_desc2[i]
            if v is not None and (lider_p2 - v) > lsd_sc:
                y_corte = cult_desc2[i-1]
                idx_y   = cultivares_asc.index(y_corte) if y_corte in cultivares_asc else None
                if idx_y is not None and idx_y > 0:
                    fig.add_hline(
                        y=idx_y - 0.5,
                        line=dict(color="#FF0000", width=2.5, dash="dot"),
                    )
                    fig.add_annotation(
                        x=0.02, xref="paper",
                        y=idx_y - 0.5, yref="y",
                        text=f"LSD: {lsd_sc:.1f}",
                        showarrow=False,
                        xanchor="left",
                        yanchor="bottom",
                        font=dict(size=12, color="#FF0000", weight="bold"),
                    )
                lider_p2 = v

    x_max = df_plot["sc/ha"].max()
    x_range_max = round(x_max * 1.18, 1)  # 18% de folga para o texto não cortar

    altura_fig = max(400, len(df_plot) * 28 + 80)
    fig.update_layout(
        height=altura_fig,
        margin=dict(l=180, r=40, t=40, b=60),
        plot_bgcolor="#FFFFFF", paper_bgcolor="#FFFFFF",
        font=dict(family="Helvetica Neue, sans-serif", size=15, color="#000000"),
        xaxis=dict(
            title=dict(text="sc/ha", font=dict(size=16, color="#000000", weight="bold")),
            showgrid=False, zeroline=False, showline=False,
            tickfont=dict(size=14, color="#000000", weight="bold"),
            range=[0, x_range_max],
        ),
        yaxis=dict(
            showgrid=True, gridcolor="#EEEEEE", gridwidth=1,
            zeroline=False, showline=False,
            tickfont=dict(size=14, color="#000000", weight="bold"),
            categoryorder="array",
            categoryarray=df_plot["Cultivar"].tolist()[::-1],
            showticklabels=False,
        ),
        legend=dict(
            title=dict(text="Status", font=dict(size=14, color="#000000", weight="bold")),
            orientation="h", yanchor="bottom", y=1.01, xanchor="left", x=0,
            font=dict(size=14, color="#000000", weight="bold"),
        ),
        showlegend=True,
    )
    fig.update_traces(textfont=dict(size=14, color="#000000", weight="bold"), selector=dict(mode="markers+text"))

    # Annotations coloridas por status no eixo Y
    COR_STATUS_TEXTO = {
        "CHECK":    "#C46A3A",
        "STINE":    "#2976B6",
        "LINHAGEM": "#009900",
        "DP2":      "#5A8A4A",
    }
    for cultivar in df_plot["Cultivar"].tolist():
        status = df_plot[df_plot["Cultivar"] == cultivar]["status_material"].iloc[0]
        cor    = COR_STATUS_TEXTO.get(status, "#333333")
        fig.add_annotation(
            x=0, xref="paper",
            y=cultivar, yref="y",
            text=f"<b>{cultivar}</b>",
            showarrow=False,
            xanchor="right",
            yanchor="middle",
            font=dict(size=13, color=cor, weight="bold"),
        )

    st.plotly_chart(fig, use_container_width=True)
    st.caption(
        "ℹ️ **LSD (5%)** — Se a diferença de produtividade entre dois cultivares for maior que este valor, "
        "ela é real e não fruto do acaso (95% de confiança). "
        "A linha vermelha pontilhada marca onde começa essa diferença significativa em relação ao melhor material do grupo. "
        "Tecnicamente: t(α/2, gl_resíduo) × √(2 × QMR / nº de locais)."
    )

# SEÇÃO 7 — TABELA DE APRESENTAÇÃO
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Apresentação",
    "Quais materiais estão no grupo de elite?",
    "Médias por cultivar com agrupamento estatístico pelo LSD. Use os filtros da sidebar para definir o conjunto de análise.",
)

with st.popover("ℹ️ Como interpretar · Legenda", use_container_width=False):
    st.markdown("""
**📌 Como interpretar esta tabela**

Cada linha representa a **média** de um cultivar considerando todos os locais nos filtros ativos.
Os materiais são ordenados por **sc/ha decrescente**.

A **linha vermelha** separa grupos estatisticamente distintos com base no **LSD (5%)**:
materiais **acima da linha** não diferem significativamente do melhor desempenho
e são os **candidatos ao avanço**. Materiais abaixo formam grupos inferiores.

---

**🎨 Legenda de cores**
""")
    col1, col2, col3 = st.columns(3)
    col1.markdown('<div style="background:#C4DFB4;padding:6px 10px;border-radius:4px;text-align:center;font-size:13px;"><b>LINHAGEM / DP2</b><br>materiais em avaliação</div>', unsafe_allow_html=True)
    col2.markdown('<div style="background:#F4B184;padding:6px 10px;border-radius:4px;text-align:center;font-size:13px;"><b>CHECK</b><br>testemunhas externas</div>', unsafe_allow_html=True)
    col3.markdown('<div style="background:#2976B6;color:#fff;padding:6px 10px;border-radius:4px;text-align:center;font-size:13px;"><b>STINE</b><br>materiais comerciais Stine</div>', unsafe_allow_html=True)
    st.markdown("""
---

**📊 Produção Relativa**

Calculada em relação à base selecionada no seletor abaixo:
- **Média geral do ensaio** → 100% = média de todos os materiais no conjunto filtrado
- **Maior produtividade** → 100% = média do cultivar com maior kg/ha
- **Testemunha** → 100% = média da testemunha selecionada
""")

# ── Seletor de Produção Relativa próprio da seção 4 ──────────────────────────
col_ref4, col_test4, _ = st.columns([2, 2, 3])

with col_ref4:
    base_rel4 = st.selectbox(
        "Base da Produção Relativa",
        options=["Média geral do ensaio", "Maior produtividade", "Testemunha"],
        index=0,
        key="base_rel4",
    )

with col_test4:
    if base_rel4 == "Testemunha":
        testemunhas4 = sorted(
            ta_filtrado[ta_filtrado["status_material"].isin(["CHECK", "STINE"])]["dePara"].dropna().unique().tolist()
        )
        if testemunhas4:
            testemunha4_sel = st.selectbox("Selecione a testemunha", options=testemunhas4, key="test4_sel")
        else:
            st.warning("Nenhuma testemunha disponível nos filtros atuais.")
            testemunha4_sel = None
    else:
        testemunha4_sel = None

if base_rel4 == "Média geral do ensaio":
    ref_valor4 = ta_filtrado["kg_ha"].mean()
elif base_rel4 == "Maior produtividade":
    ref_valor4 = ta_filtrado["kg_ha"].max()
elif base_rel4 == "Testemunha" and testemunha4_sel:
    ref_valor4 = ta_filtrado[ta_filtrado["dePara"] == testemunha4_sel]["kg_ha"].mean()
else:
    ref_valor4 = ta_filtrado["kg_ha"].mean()

# ── Calcular médias por cultivar (df_apres já calculado antes da Seção 6) ─────
if df_apres.empty:
    st.warning("⚠️ Nenhum dado para exibir.")
else:
    df_apres = df_apres.sort_values("sc/ha", ascending=False).reset_index(drop=True)

    # Calcular Prod. Rel. sobre médias da tabela
    kg_medias = df_apres["kg/ha"].dropna()
    if base_rel4 == "Média geral do ensaio":
        ref_valor4_calc = kg_medias.mean()
    elif base_rel4 == "Maior produtividade":
        ref_valor4_calc = kg_medias.max()
    elif base_rel4 == "Testemunha" and testemunha4_sel:
        t_val = df_apres.loc[df_apres["Cultivar"] == testemunha4_sel, "kg/ha"]
        ref_valor4_calc = t_val.values[0] if len(t_val) > 0 else kg_medias.mean()
    else:
        ref_valor4_calc = kg_medias.mean()

    df_apres.insert(
        df_apres.columns.get_loc("sc/ha") + 1,
        "Prod. Rel. (%)",
        (df_apres["kg/ha"] / ref_valor4_calc * 100).round(1) if ref_valor4_calc else None,
    )

    # ── LSD e CV da ANOVA ─────────────────────────────────────────────────────
    lsd_apres = _lsd_apres  # já calculado antes da Seção 6
    lsd_sc    = round(lsd_apres / 60, 2) if isinstance(lsd_apres, (int, float)) and not np.isnan(lsd_apres) else None

    try:
        d_anova = ta_filtrado[["kg_ha","dePara","cod_fazenda"]].dropna().copy()
        d_anova = d_anova[d_anova["kg_ha"] > 0].reset_index(drop=True)
        y_a     = d_anova["kg_ha"].values.astype(float)
        grand_mean_anova = y_a.mean()
        X_c = pd.get_dummies(d_anova["dePara"],       drop_first=True).values.astype(float)
        X_l = pd.get_dummies(d_anova["cod_fazenda"],  drop_first=True).values.astype(float)
        X_a = np.hstack([np.ones((len(y_a), 1)), X_c, X_l])
        beta_a, _, rank_a, _ = np.linalg.lstsq(X_a, y_a, rcond=None)
        ss_res_a = np.sum((y_a - X_a @ beta_a) ** 2)
        gl_res_a = len(y_a) - rank_a
        qmr_a    = ss_res_a / gl_res_a if gl_res_a > 0 else np.nan
        cv_anova = round(np.sqrt(qmr_a) / grand_mean_anova * 100, 1) if not np.isnan(qmr_a) else "—"
    except Exception:
        cv_anova = "—"

    # Linhas de corte automáticas entre grupos
    linhas_corte = set()
    if lsd_sc is not None:
        sc_vals = df_apres["sc/ha"].tolist()
        lider = sc_vals[0]
        for i in range(1, len(sc_vals)):
            v = sc_vals[i]
            if v is None:
                continue
            if (lider - v) > lsd_sc:
                linhas_corte.add(i - 1)
                lider = v

    COR_STATUS = {
        "CHECK":    "#F4B184",
        "STINE":    "#2976B6",
        "LINHAGEM": "#00FF01",
        "DP2":      "#C4DFB4",
    }
    COR_TEXTO = {
        "CHECK":    "#1A1A1A",
        "STINE":    "#FFFFFF",
        "LINHAGEM": "#1A1A1A",
        "DP2":      "#1A1A1A",
        "":         "#000000",
    }

    # Colunas visíveis (exclui status_material que é só para cor)
    cols_show = ["Cultivar"] + [c for c in df_apres.columns if c not in ("Cultivar", "status_material")]

    # Rodapé — média geral
    rodape = {"Cultivar": "Média Geral", "status_material": ""}
    for c in cols_show[1:]:
        vals = df_apres[c].dropna() if c in df_apres.columns else pd.Series([], dtype=float)
        if c == "Pop. Final":
            rodape[c] = int(round(vals.mean(), 0)) if len(vals) > 0 else "—"
        else:
            rodape[c] = round(vals.mean(), 1) if len(vals) > 0 else "—"

    # ── Renderizar tabela em HTML ──────────────────────────────────────────────
    html = """
<style>
.tb-apres { width:100%; border-collapse:collapse; font-size:15px; font-family:'Helvetica Neue',sans-serif; }
.tb-apres th {
    background:#F2F2F2; color:#000000 !important; padding:8px 10px;
    text-align:center; border:1px solid #ccc; white-space:nowrap; font-weight:700; font-size:15px;
}
.tb-apres th:first-child { text-align:left; }
.tb-apres td { padding:7px 10px; border:1px solid #ddd; text-align:center; white-space:nowrap; font-size:15px; }
.tb-apres td:first-child { text-align:left; font-weight:500; }
.tb-apres td[data-fg="white"], .tb-apres td[data-fg="white"] * { color: #FFFFFF !important; }
.tb-apres td[data-fg="dark"],  .tb-apres td[data-fg="dark"]  * { color: #1A1A1A !important; }
.tb-apres tr.corte td { border-bottom: 10px solid #FF0000 !important; }
.tb-apres tr.rodape td {
    background:#D9D9D9 !important; font-weight:700; border-top:2px solid #888;
    font-size:15px; color:#000000 !important;
}
.tb-apres tr.rodape-info td { font-size:15px !important; font-weight:700 !important; border-top:none; color:#000000 !important; }
.tb-apres tr.rodape-info td:first-child  { background:#D9D9D9 !important; border:1px solid #ddd; }
.tb-apres tr.rodape-info td:nth-child(2) { background:#D9D9D9 !important; border:1px solid #ddd; }
.tb-apres tr.rodape-info td:nth-child(n+3) { background:#FFFFFF !important; border:none; }
</style>
<table class="tb-apres">
<thead><tr>
"""
    for c in cols_show:
        html += f"<th>{c}</th>"
    html += "</tr></thead><tbody>"

    for i, row in df_apres.iterrows():
        status  = row.get("status_material", "")
        bg      = COR_STATUS.get(status, "#FFFFFF")
        fg      = COR_TEXTO.get(status, "#000000")
        data_fg = "white" if fg == "#FFFFFF" else "dark"
        classe  = "corte" if i in linhas_corte else ""
        html += f'<tr class="{classe}">'
        for c in cols_show:
            val = row.get(c, "")
            val = "—" if val is None or (isinstance(val, float) and np.isnan(val)) else val
            html += f'<td data-fg="{data_fg}" style="background:{bg};">{val}</td>'
        html += "</tr>"

    html += '<tr class="rodape">'
    for c in cols_show:
        val = rodape.get(c, "")
        html += f"<td>{val}</td>"
    html += "</tr>"

    lsd_fmt  = round(lsd_sc, 2) if lsd_sc else "—"
    cv_geral = cv_anova
    n_locais = ta_filtrado["cod_fazenda"].nunique()

    def _rodape_info_row(label, valor, n_cols):
        cells  = f'<td style="background:#D9D9D9;border:1px solid #ddd;font-size:15px;font-weight:700;color:#000000 !important;">{label}</td>'
        cells += f'<td style="background:#D9D9D9;border:1px solid #ddd;font-size:15px;font-weight:700;text-align:left;color:#000000 !important;">{valor}</td>'
        cells += f'<td colspan="{n_cols-2}" style="background:#FFFFFF;border:none;"></td>'
        return f'<tr class="rodape-info">{cells}</tr>'

    html += _rodape_info_row("CV (%)", f"{cv_geral}%", len(cols_show))
    html += _rodape_info_row("LSD sc/ha (5%)", f"{lsd_fmt}", len(cols_show))
    html += _rodape_info_row("Locais", n_locais, len(cols_show))
    html += "</tbody></table>"

    import streamlit.components.v1 as components
    iframe_height = 60 + (len(df_apres) + 5) * 38
    components.html(html, height=iframe_height, scrolling=False)
    st.caption(
        "ℹ️ **CV (%) desta tabela** = √QMR ÷ Média Geral × 100, onde QMR é o Quadrado Médio do Resíduo da ANOVA conjunta "
        "(modelo: y = μ + cultivar + local + erro). Desconta os efeitos de cultivar e local, "
        "restando apenas o erro experimental — é o indicador correto da precisão do experimento."
    )
    st.caption(
        "ℹ️ **LSD sc/ha (5%)** — Se a diferença de produtividade entre dois cultivares for maior que este valor, "
        "ela é real e não fruto do acaso (95% de confiança). "
        "A linha vermelha na tabela marca onde começa essa diferença significativa em relação ao melhor material. "
        "Tecnicamente: t(α/2, gl_resíduo) × √(2 × QMR / nº de locais)."
    )

    # ── Exportar para Excel com formatação completa ───────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    if st.button("⬇️ Exportar Excel com formatação", type="primary"):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tabela de Apresentação"

        thin   = Side(style="thin", color="CCCCCC")
        BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

        for ci, col in enumerate(cols_show, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.font      = Font(bold=True, color="1A1A1A", name="Arial", size=10)
            cell.fill      = PatternFill("solid", start_color="F2F2F2")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = BORDER
            ws.column_dimensions[get_column_letter(ci)].width = max(12, len(col)+2)
        ws.row_dimensions[1].height = 28

        for ri, row in df_apres.iterrows():
            status   = row.get("status_material","")
            bg_hex   = COR_STATUS.get(status,"FFFFFF").replace("#","")
            fg_hex   = COR_TEXTO.get(status,"1A1A1A").replace("#","")
            is_corte = ri in linhas_corte

            for ci, col in enumerate(cols_show, 1):
                val = row.get(col, None)
                if isinstance(val, float) and np.isnan(val): val = None
                cell = ws.cell(row=ri+2, column=ci, value=val)
                cell.font      = Font(name="Arial", size=10, color=fg_hex)
                cell.fill      = PatternFill("solid", start_color=bg_hex)
                cell.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")
                b_bottom = Side(style="thick", color="FF0000") if is_corte else thin
                cell.border  = Border(left=thin, right=thin, top=thin, bottom=b_bottom)

        n_data = len(df_apres)
        rodape_rows = [
            ("Média Geral", {c: rodape.get(c,"") for c in cols_show}),
            ("CV (%)",      {cols_show[0]: "CV (%)",         cols_show[1]: f"{cv_geral}%"}),
            ("LSD",         {cols_show[0]: "LSD sc/ha (5%)", cols_show[1]: f"{lsd_fmt}"}),
            ("Locais",      {cols_show[0]: "Locais",          cols_show[1]: n_locais}),
        ]
        for rj, (_, rdata) in enumerate(rodape_rows):
            for ci, col in enumerate(cols_show, 1):
                val = rdata.get(col, None)
                cell = ws.cell(row=n_data+2+rj, column=ci, value=val)
                cell.font      = Font(name="Arial", size=10, bold=(rj==0))
                cell.fill      = PatternFill("solid", start_color="F0F0F0")
                cell.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")
                cell.border    = BORDER

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        st.download_button(
            label="📥 Baixar Excel",
            data=buf,
            file_name="tabela_apresentacao.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

st.divider()

# ════════════════════════════════════════════════════════════════════════════════
# ── GRÁFICO 2 — BOXPLOT COM DISTRIBUIÇÃO POR LOCAL ────────────────────────────
secao_titulo(
    "Visualização",
    "Como é a distribuição de produtividade por cultivar?",
    contexto_str,
)

with st.popover("ℹ️ Como interpretar · Distribuição", use_container_width=False):
    st.markdown("""
**📌 Como ler este gráfico**

Cada barra representa a distribuição de produtividade (sc/ha) de um cultivar considerando todos os locais nos filtros ativos.

---

**🔲 Elementos da barra**
""")
    col1, col2, col3 = st.columns(3)
    col1.markdown('<div style="background:rgba(93,174,139,0.12);border:1px solid #ccc;padding:8px;border-radius:4px;text-align:center;font-size:13px;"><b>Barra clara</b><br>Range total<br>(mín → máx)</div>', unsafe_allow_html=True)
    col2.markdown('<div style="background:rgba(93,174,139,0.45);border:1px solid #ccc;padding:8px;border-radius:4px;text-align:center;font-size:13px;"><b>Barra escura</b><br>Onde a maioria dos locais se concentra<br><span style="font-size:11px;color:#555;">(intervalo interquartil Q1→Q3, 50% dos dados)</span></div>', unsafe_allow_html=True)
    col3.markdown('<div style="background:#fff;border:2px solid #333;padding:8px;border-radius:4px;text-align:center;font-size:13px;"><b>Linha preta</b><br>Média<br>com valor e nº de observações</div>', unsafe_allow_html=True)
    st.markdown("""
---

**⚫ Pontos individuais**

Cada ponto representa um local (fazenda). Passe o mouse sobre o ponto para ver o **código da fazenda** e o valor de sc/ha.
Pontos muito afastados da barra indicam locais com comportamento atípico — ambientes muito favoráveis ou adversos.

> **n** = número de observações de produção válidas (sc/ha > 0) do cultivar no conjunto filtrado.

---

**📐 Entendendo cada medida**

> ⚠️ **Média e Mediana são medidas diferentes e podem contar histórias distintas sobre o mesmo cultivar.** Quando estiverem próximas, o desempenho é equilibrado. Quando divergirem, vale investigar quais ambientes estão puxando a média para cima ou para baixo.

- **Média:** é a soma de todas as produtividades dividida pelo número de locais. É o valor mais conhecido, mas pode ser "puxada" para cima por uma área excepcional ou para baixo por uma área muito ruim — por isso nem sempre representa bem o comportamento típico do cultivar.

- **Mediana (Q2):** é o valor "do meio" quando todos os resultados estão em ordem. Em campo: é a produtividade que representa o "típico" do cultivar — metade das áreas produz menos que isso e metade produz mais. **Diferente da média, não sofre distorção por áreas muito ruins ou muito boas.**

- **Q1 (primeiro quartil):** é o ponto que separa os 25% piores resultados dos 75% restantes. Em campo: é o **teto inferior** — abaixo de Q1 estão as situações em que o cultivar costuma ir mal. Um Q1 alto significa que mesmo nos piores ambientes o cultivar ainda entrega uma produtividade razoável — sinal de segurança na recomendação. *(início da barra escura no gráfico)*

- **Q3 (terceiro quartil):** é o ponto que separa os 75% melhores resultados dos 25% mais altos. Em campo: acima de Q3 estão as áreas onde o cultivar performa entre os melhores resultados que ele consegue entregar. *(fim da barra escura no gráfico)*

---

**📊 Interpretação prática**

- **Barra curta + média alta** → cultivar consistente e produtivo — ideal para recomendação ampla
- **Barra longa** → alta variabilidade entre locais — pode ter adaptação específica
- **Média alta mas Q3 largo** → desempenho instável, risco maior em recomendação
""")

if df_apres.empty:
    st.warning("⚠️ Nenhum dado para exibir.")
else:
    df_box = ta_filtrado[["dePara", "sc_ha", "status_material", "cod_fazenda"]].copy()
    df_box = df_box.rename(columns={"dePara": "Cultivar", "sc_ha": "sc/ha"})
    df_box = df_box[df_box["sc/ha"] > 0].dropna(subset=["sc/ha"])

    ordem_box = (
        df_box.groupby("Cultivar")["sc/ha"]
        .mean()
        .sort_values(ascending=False)
        .index.tolist()
    )

    COR_STATUS_BOX = {
        "CHECK":    "rgba(244,177,132,0.5)",
        "STINE":    "rgba(41,118,182,0.45)",
        "LINHAGEM": "rgba(0,255,1,0.45)",
        "DP2":      "rgba(196,223,180,0.45)",
    }
    COR_STATUS_ESC = {
        "CHECK":    "#C46A3A",
        "STINE":    "#1A4F7A",
        "LINHAGEM": "#009900",
        "DP2":      "#7AAF6A",
    }
    COR_STATUS_CLA = {
        "CHECK":    "rgba(244,177,132,0.15)",
        "STINE":    "rgba(41,118,182,0.12)",
        "LINHAGEM": "rgba(0,255,1,0.12)",
        "DP2":      "rgba(196,223,180,0.12)",
    }

    fig2 = go_plt.Figure()

    for cultivar in ordem_box:
        grp = df_box[df_box["Cultivar"] == cultivar]["sc/ha"].dropna()
        if len(grp) < 2:
            continue
        status  = df_box[df_box["Cultivar"] == cultivar]["status_material"].mode()[0]
        cor_esc = COR_STATUS_BOX.get(status, "#888888")
        cor_cla = COR_STATUS_CLA.get(status, "rgba(150,150,150,0.2)")

        q1, med, q3 = grp.quantile(0.25), grp.median(), grp.quantile(0.75)
        vmin, vmax  = grp.min(), grp.max()
        media       = grp.mean()
        n = len(grp)

        fig2.add_trace(go_plt.Bar(
            x=[vmax - vmin], base=vmin, y=[cultivar], orientation="h",
            marker_color=cor_cla, marker_line_width=0, width=0.35,
            showlegend=False, hoverinfo="skip",
        ))
        fig2.add_trace(go_plt.Bar(
            x=[q3 - q1], base=q1, y=[cultivar], orientation="h",
            marker_color=cor_esc, marker_line_width=0, width=0.35,
            showlegend=False, hoverinfo="skip",
        ))
        n_cultivares = len(ordem_box)
        idx = n_cultivares - 1 - ordem_box.index(cultivar)
        fig2.add_shape(
            type="line",
            x0=media, x1=media,
            y0=idx - 0.18, y1=idx + 0.18,
            line=dict(color="#000000", width=2.5),
        )
        fig2.add_annotation(
            x=media,
            y=cultivar,
            text=f"<b>{media:.1f}</b> ({n})",
            showarrow=False,
            xanchor="center",
            yanchor="bottom",
            yshift=22,
            font=dict(size=13, color="#000000", weight="bold", family="Helvetica Neue, sans-serif"),
        )

    for cultivar in ordem_box:
        grp2 = df_box[df_box["Cultivar"] == cultivar].dropna(subset=["sc/ha"])
        if grp2.empty:
            continue
        status  = grp2["status_material"].mode()[0]
        cor_pt  = COR_STATUS_ESC.get(status, "#555555")

        fig2.add_trace(go_plt.Scatter(
            x=grp2["sc/ha"],
            y=[cultivar] * len(grp2),
            mode="markers",
            showlegend=False,
            legendgroup=status,
            marker=dict(color=cor_pt, size=8, opacity=0.85, line=dict(color="#FFFFFF", width=0.8)),
            customdata=grp2["cod_fazenda"].tolist(),
            hovertemplate="<b>%{customdata}</b><br>sc/ha: %{x:.1f}<extra></extra>",
        ))

    # ── Legenda de status (controla barras + pontos via legendgroup) ──────────
    COR_LEG = {"CHECK": "#F4B184", "STINE": "#2976B6", "LINHAGEM": "#00FF01", "DP2": "#C4DFB4"}
    for status, cor in COR_LEG.items():
        if status in df_box["status_material"].values:
            fig2.add_trace(go_plt.Scatter(
                x=[None], y=[None], mode="markers", name=status,
                legendgroup=status,
                marker=dict(color=cor, size=12),
                showlegend=True,
            ))

    # ── Linhas de corte LSD ───────────────────────────────────────────────────
    lsd_box    = calcular_lsd(ta_filtrado)
    lsd_box_sc = round(lsd_box / 60, 2) if isinstance(lsd_box, (int, float)) and not np.isnan(lsd_box) else None

    mostrar_lsd = st.checkbox("Mostrar linhas de corte LSD", value=True, key="chk_lsd_box")

    if lsd_box_sc is not None and mostrar_lsd:
        medias_box = df_box.groupby("Cultivar")["sc/ha"].mean()
        lider_box  = medias_box[ordem_box[0]]
        for i in range(1, len(ordem_box)):
            v = medias_box.get(ordem_box[i], None)
            if v is None:
                continue
            if (lider_box - v) > lsd_box_sc:
                idx_corte = len(ordem_box) - 1 - i + 0.5
                fig2.add_shape(
                    type="line",
                    x0=0, x1=1, xref="paper",
                    y0=idx_corte, y1=idx_corte,
                    line=dict(color="#FF0000", width=2, dash="dot"),
                )
                fig2.add_annotation(
                    x=0.02, xref="paper",
                    y=idx_corte, yref="y",
                    text=f"LSD: {lsd_box_sc:.1f}",
                    showarrow=False,
                    xanchor="left",
                    yanchor="bottom",
                    font=dict(size=11, color="#FF0000", weight="bold"),
                )
                lider_box = v

    altura_box = max(500, len(ordem_box) * 75 + 80)
    fig2.update_layout(
        barmode="overlay",
        height=altura_box,
        margin=dict(l=180, r=60, t=40, b=60),
        plot_bgcolor="#FFFFFF", paper_bgcolor="#FFFFFF",
        font=dict(family="Helvetica Neue, sans-serif", size=14, color="#000000"),
        xaxis=dict(
            title=dict(text="sc/ha", font=dict(size=16, color="#000000", weight="bold")),
            showgrid=False, zeroline=False, showline=False,
            tickfont=dict(size=14, color="#000000", weight="bold"),
            rangemode="tozero",
        ),
        yaxis=dict(
            showgrid=True, gridcolor="#EEEEEE", gridwidth=1,
            zeroline=False, showline=False,
            tickfont=dict(size=14, color="#000000", weight="bold"),
            categoryorder="array",
            categoryarray=ordem_box[::-1],
            showticklabels=False,
        ),
        legend=dict(
            title=dict(text="Status", font=dict(size=14, color="#000000", weight="bold")),
            orientation="h", yanchor="bottom", y=1.01, xanchor="left", x=0,
            font=dict(size=14, color="#000000", weight="bold"),
        ),
        showlegend=True,
    )
    fig2.update_traces(textfont=dict(size=14, color="#000000", weight="bold"), selector=dict(mode="markers+text"))

    # Annotations coloridas por status no eixo Y
    COR_STATUS_TEXTO = {
        "CHECK":    "#C46A3A",
        "STINE":    "#2976B6",
        "LINHAGEM": "#009900",
        "DP2":      "#5A8A4A",
    }
    status_box_map = df_box.drop_duplicates("Cultivar").set_index("Cultivar")["status_material"].to_dict()
    for cultivar in ordem_box:
        status = status_box_map.get(cultivar, "")
        cor    = COR_STATUS_TEXTO.get(status, "#333333")
        fig2.add_annotation(
            x=0, xref="paper",
            y=cultivar, yref="y",
            text=f"<b>{cultivar}</b>",
            showarrow=False,
            xanchor="right",
            yanchor="middle",
            font=dict(size=13, color=cor, weight="bold"),
        )

    st.plotly_chart(fig2, use_container_width=True)

    # Dicionário de locais
    df_dic = (
        ta_filtrado[["cod_fazenda", "nomeFazenda", "cidade_nome", "estado_sigla"]]
        .drop_duplicates()
        .sort_values("cod_fazenda")
        .rename(columns={
            "cod_fazenda":  "Código",
            "nomeFazenda":  "Local",
            "cidade_nome":  "Cidade",
            "estado_sigla": "Estado",
        })
        .reset_index(drop=True)
    )
    with st.popover(f"📍 Dicionário de locais ({len(df_dic)} locais)", use_container_width=False):
        st.markdown("Código gerado automaticamente · passe o mouse sobre os pontos do gráfico para identificar o local.")
        busca_dic = st.text_input("🔍 Buscar", placeholder="Código, local ou cidade...", key="busca_dic")
        df_dic_show = df_dic[
            df_dic.apply(lambda r: busca_dic.strip().lower() in " ".join(r.astype(str)).lower(), axis=1)
        ] if busca_dic.strip() else df_dic
        st.dataframe(df_dic_show, hide_index=True, use_container_width=True)

st.divider()

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 7 — ESTABILIDADE E ÍNDICE DE CONFIANÇA
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Estabilidade",
    "Qual o comportamento do material e sua previsibilidade entre locais e safras?",
    "Cultivares com alto índice de confiança mantêm desempenho consistente independente do ambiente — menor risco na recomendação.",
)

with st.popover("ℹ️ Como interpretar · Índice de Confiança", use_container_width=False):
    st.markdown("""
**📌 Como ler esta tabela**

Cada linha representa um cultivar com seu desempenho consolidado no conjunto filtrado (locais e safras ativos).
Como os ensaios são **sem repetição**, a variação observada reflete diferenças entre ambientes — não erro experimental.

---

**📋 Colunas**
""")
    col1, col2 = st.columns(2)
    col1.markdown("""
| Coluna | Significado |
|--------|-------------|
| **Cultivar** | Nome do material |
| **[Safra]** | Nº de observações válidas naquela safra (— = não avaliado) |
| **N Total** | Total de observações válidas no conjunto filtrado |
| **Média (sc/ha)** | Média de produtividade no conjunto |
| **Mín / Máx** | Menor e maior produtividade observada |
| **Amplitude** | Máx − Mín — variação total entre ambientes |
""")
    col2.markdown("""
| Coluna | Significado |
|--------|-------------|
| **Q1 / Mediana / Q3** | Quartis — faixa central de 50% dos locais |
| **IQR (sc/ha)** | Q3 − Q1 — dispersão dos 50% centrais |
| **CV (%)** | Desvio padrão ÷ Média × 100 |
| **IQR Rel. (%)** | IQR ÷ Mediana × 100 — base do índice de confiança |
| **Índice de Confiança** | Classificação de 1 a 5 estrelas |
""")
    st.markdown("""
---

**📐 Entendendo cada medida**

> ⚠️ **Média e Mediana são medidas diferentes e podem contar histórias distintas sobre o mesmo cultivar.** Quando estiverem próximas, o desempenho é equilibrado. Quando divergirem, vale investigar quais ambientes estão puxando a média para cima ou para baixo.

- **Média:** é a soma de todas as produtividades dividida pelo número de locais. É o valor mais conhecido, mas pode ser "puxada" para cima por uma área excepcional ou para baixo por uma área muito ruim — por isso nem sempre representa bem o comportamento típico do cultivar.

- **Mediana (Q2):** é o valor "do meio" quando todos os resultados estão em ordem. Em campo: é a produtividade que representa o "típico" do cultivar — metade das áreas produz menos que isso e metade produz mais. **Diferente da média, não sofre distorção por áreas muito ruins ou muito boas.**

- **Q1 (primeiro quartil):** é o ponto que separa os 25% piores resultados dos 75% restantes. Em campo: é o **teto inferior** — abaixo de Q1 estão as situações em que o cultivar costuma ir mal. Um Q1 alto significa que mesmo nos piores ambientes o cultivar ainda entrega uma produtividade razoável — sinal de segurança na recomendação.

- **Q3 (terceiro quartil):** é o ponto que separa os 75% melhores resultados dos 25% mais altos. Em campo: acima de Q3 estão as áreas onde o cultivar performa entre os melhores resultados que ele consegue entregar.
- **IQR** → a distância entre Q1 e Q3 — representa a faixa onde estão os **50% centrais** dos locais. Quanto menor, mais concentrado e previsível é o desempenho

---

**⭐ Como interpretar as estrelas**

Baseado no **IQR Relativo (%)** — mais robusto para ensaios sem repetição pois ignora valores extremos e representa os 50% centrais dos ambientes:

| Estrelas | IQR Rel. (%) | Interpretação |
|----------|-------------|---------------|
| ⭐⭐⭐⭐⭐ | < 15% | Altamente consistente — previsível em qualquer ambiente |
| ⭐⭐⭐⭐ | 15 – 25% | Consistente — pequena variação entre locais |
| ⭐⭐⭐ | 25 – 35% | Moderado — variação aceitável, atenção ao ambiente |
| ⭐⭐ | 35 – 45% | Instável — desempenho dependente do ambiente |
| ⭐ | > 45% | Imprevisível — alto risco de recomendação ampla |

> **⚠️** indica N Total < 5 — índice calculado com poucos pontos, interpretar com cautela.

---

**🔢 Como é calculado**

> **IQR Rel. (%)** = (Q3 − Q1) ÷ Mediana × 100

> **CV (%)** = Desvio Padrão ÷ Média × 100

Quando múltiplas safras estão ativas, a variação entre safras também entra no cálculo — o índice se torna mais exigente e representa estabilidade ampla entre locais **e** anos.
""")

# ── Cálculo ──────────────────────────────────────────────────────────────────
df_estab = ta_filtrado[ta_filtrado["sc_ha"] > 0][["dePara", "sc_ha", "safra"]].dropna().copy()

# N por safra para cada cultivar
safras_ativas = sorted(df_estab["safra"].unique().tolist())

def estrelas(iqr_rel, n):
    if pd.isna(iqr_rel): return "—"
    aviso = " ⚠️" if n < 5 else ""
    if iqr_rel < 15:     return f"⭐⭐⭐⭐⭐{aviso}"
    if iqr_rel < 25:     return f"⭐⭐⭐⭐{aviso}"
    if iqr_rel < 35:     return f"⭐⭐⭐{aviso}"
    if iqr_rel < 45:     return f"⭐⭐{aviso}"
    return f"⭐{aviso}"

rows_estab = []
for cultivar, grp in df_estab.groupby("dePara"):
    s       = grp["sc_ha"]
    n       = len(s)
    media   = s.mean()
    dp      = s.std()
    cv      = (dp / media * 100) if media > 0 else np.nan
    q1      = s.quantile(0.25)
    med     = s.median()
    q3      = s.quantile(0.75)
    iqr     = q3 - q1
    iqr_rel = (iqr / med * 100) if med > 0 else np.nan

    row = {"Cultivar": cultivar}
    for safra in safras_ativas:
        n_safra = len(grp[grp["safra"] == safra])
        row[str(safra)] = n_safra if n_safra > 0 else "—"
    row.update({
        "N Total":             n,
        "Média (sc/ha)":       round(media, 1),
        "Mín":                 round(s.min(), 1),
        "Máx":                 round(s.max(), 1),
        "Amplitude":           round(s.max() - s.min(), 1),
        "Q1":                  round(q1, 1),
        "Mediana":             round(med, 1),
        "Q3":                  round(q3, 1),
        "IQR (sc/ha)":         round(iqr, 1),
        "CV (%)":              round(cv, 1) if not np.isnan(cv) else None,
        "IQR Rel. (%)":        round(iqr_rel, 1) if not np.isnan(iqr_rel) else None,
        "Índice de Confiança": estrelas(iqr_rel, n),
    })
    rows_estab.append(row)

df_ic = (
    pd.DataFrame(rows_estab)
    .sort_values("IQR Rel. (%)", ascending=True)
    .reset_index(drop=True)
)

ag_table(df_ic, height=min(600, 80 + len(df_ic) * 36))
exportar_excel(df_ic, nome_arquivo="estabilidade.xlsx", label="⬇️ Exportar Estabilidade", key="exp_estab")
st.caption(
    "ℹ️ Ensaios sem repetição — variação reflete diferenças entre ambientes. "
    "**Índice de Confiança** baseado no IQR Relativo (%) — mais robusto a valores extremos que o CV. "
    "Com múltiplas safras ativas, a variação entre anos também entra no cálculo."
)

st.divider()

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 8 — ADAPTABILIDADE E ESTABILIDADE
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Adaptabilidade e Estabilidade",
    "Como cada cultivar reage aos diferentes ambientes?",
    "Cultivares responsivos se destacam em ambientes favoráveis. Cultivares estáveis mantêm desempenho mesmo em condições adversas. Os melhores combinam as duas características.",
)


# ── Cálculo ───────────────────────────────────────────────────────────────────
df_er = ta_filtrado[ta_filtrado["sc_ha"] > 0][["dePara", "cod_fazenda", "sc_ha", "status_material"]].dropna().copy()

# Índice ambiental: média do local − média geral
media_geral = df_er["sc_ha"].mean()
media_local = df_er.groupby("cod_fazenda")["sc_ha"].mean()
df_er["idx_amb"] = df_er["cod_fazenda"].map(media_local) - media_geral

# Mínimo de locais para análise
MIN_LOCAIS = 4
contagem = df_er.groupby("dePara")["cod_fazenda"].nunique()
cultivares_validos = contagem[contagem >= MIN_LOCAIS].index.tolist()
df_er = df_er[df_er["dePara"].isin(cultivares_validos)]

# Eberhart & Russell via OLS por cultivar
rows_er = []
for cultivar, grp in df_er.groupby("dePara"):
    y = grp["sc_ha"].values
    x = grp["idx_amb"].values
    n = len(y)
    status = grp["status_material"].iloc[0]

    # Regressão simples: y = a + b*x
    X = np.column_stack([np.ones(n), x])
    try:
        beta, _, rank, _ = np.linalg.lstsq(X, y, rcond=None)
        a, b = beta
        y_hat = X @ beta
        ss_res = np.sum((y - y_hat) ** 2)
        gl_res = n - 2
        s2 = (ss_res / gl_res - 1) if gl_res > 0 else np.nan  # desvio dos desvios (E&R)
        s2 = max(s2, 0)  # não pode ser negativo
    except Exception:
        b, s2 = np.nan, np.nan

    rows_er.append({
        "dePara":           cultivar,
        "status_material":  status,
        "b":                round(b, 3) if not np.isnan(b) else None,
        "s2":               round(s2, 3) if not np.isnan(s2) else None,
        "n_locais":         grp["cod_fazenda"].nunique(),
    })

if not rows_er:
    st.warning("⚠️ Nenhum cultivar com locais suficientes para calcular adaptabilidade e estabilidade. Reduza o filtro de locais mínimos ou amplie a seleção.")
else:
    df_er_res = pd.DataFrame(rows_er).dropna(subset=["b", "s2"])

    # Lin & Binns — Pi por cultivar
    maximo_local = df_er.groupby("cod_fazenda")["sc_ha"].max().rename("max_local")
    df_pi = df_er.join(maximo_local, on="cod_fazenda")
    df_pi["diff"] = df_pi["max_local"] - df_pi["sc_ha"]  # diferença simples por local

    pi_rows = []
    for cultivar, grp in df_pi.groupby("dePara"):
        n_loc   = grp["cod_fazenda"].nunique()
        pi      = grp["diff"].sum() / (2 * n_loc)         # Pi original para ranking/scatter
        delta   = grp["diff"].mean()                       # média simples das diferenças
        pi_rows.append({"dePara": cultivar, "Pi": round(pi, 2), "delta_lider": round(delta, 1)})

    df_pi_res = pd.DataFrame(pi_rows)
    df_final = df_er_res.merge(df_pi_res, on="dePara", how="left")

    # Classificação por quadrante
    b_medio  = 1.0
    s2_medio = df_final["s2"].median()

    def classificar_quadrante(b, s2, s2_med):
        if pd.isna(b) or pd.isna(s2): return "—"
        if b >= b_medio and s2 <= s2_med: return "🟢 Alta Performance"
        if b >= b_medio and s2 >  s2_med: return "🟡 Ambiente Favorável"
        if b <  b_medio and s2 <= s2_med: return "🔵 Ampla Adaptação"
        return "🔴 Atenção"

    df_final["Quadrante"] = df_final.apply(
        lambda r: classificar_quadrante(r["b"], r["s2"], s2_medio), axis=1
    )

    # Pi normalizado para tamanho do ponto (invertido — menor Pi = ponto menor)
    pi_max = df_final["Pi"].max()
    pi_min = df_final["Pi"].min()
    df_final["pi_size"] = 8 + 24 * (1 - (df_final["Pi"] - pi_min) / (pi_max - pi_min + 1e-9))

    # ── Gráfico de regressão E&R ──────────────────────────────────────────────────
    st.markdown("#### Índice Ambiental (Regressão de Eberhart & Russell)")
    st.caption("Cada reta representa um cultivar. Inclinação (b) indica adaptabilidade — retas mais inclinadas respondem mais a ambientes favoráveis.")

    with st.popover("ℹ️ Como interpretar · Regressão E&R", use_container_width=False):
        st.markdown("""
    **📌 O que este gráfico mostra**

    Cada reta representa a relação entre a produtividade de um cultivar e a qualidade do ambiente — quanto melhor o ambiente (eixo X), quanto mais o cultivar produz (eixo Y).

    ---

    **📐 Como ler os elementos**

    - **Eixo X** → produtividade média do ambiente (sc/ha) — ambientes à direita são mais favoráveis
    - **Eixo Y** → produtividade do cultivar (sc/ha)
    - **Cada reta** → comportamento de um cultivar ao longo dos ambientes
    - **Pontos** → observações reais de cada local
    - **Linha tracejada cinza** → referência b = 1,0 — cultivar que acompanha exatamente a média dos ambientes
    - **Linha vertical pontilhada** → média geral do conjunto

    ---

    **📐 Como interpretar a inclinação (b)**

    - **Reta mais inclinada que a referência (b > 1)** → cultivar **responsivo** — ganha mais que a média em ambientes favoráveis, mas perde mais em ambientes ruins
    - **Reta paralela à referência (b ≈ 1)** → **adaptação ampla** — acompanha a média em qualquer ambiente
    - **Reta menos inclinada que a referência (b < 1)** → cultivar **pouco responsivo** — estável, mas não aproveita ambientes de alto potencial

    ---

    **📐 O que significa estar acima ou abaixo da reta de referência**

    A posição vertical da reta em relação à tracejada indica o **nível geral de produtividade** do cultivar:

    - **Reta acima da tracejada** → cultivar produz **acima da média geral** em todos os ambientes
    - **Reta abaixo da tracejada** → cultivar produz **abaixo da média geral** em todos os ambientes

    Combinando posição e inclinação:

    | Posição | Inclinação (b) | Interpretação |
    |---------|---------------|---------------|
    | Acima + b > 1 | Mais íngreme que a ref. | Alto potencial — lidera e ainda cresce mais em bons ambientes |
    | Acima + b < 1 | Mais plana que a ref. | Produtivo e estável — bom piso mesmo em ambientes adversos |
    | Abaixo + b > 1 | Mais íngreme que a ref. | Começa atrás mas se aproxima dos líderes em ambientes favoráveis |
    | Abaixo + b < 1 | Mais plana que a ref. | Abaixo da média e pouco responsivo — perfil de risco |

    ---

    **📐 O que observar na dispersão dos pontos**

    - **Pontos próximos da reta** → cultivar previsível (s² baixo) — o que a reta promete é o que o cultivar entrega
    - **Pontos espalhados** → comportamento imprevisível (s² alto) — mesmo em ambientes similares, o cultivar reage de forma diferente
    """)


    # Seletor de cultivares — default: top 3 por média (evita sobreposição de labels)
    top3_default = (
        df_er.groupby("dePara")["sc_ha"].mean()
        .sort_values(ascending=False)
        .head(3)
        .index.tolist()
    )
    cultivares_er = sorted(df_final["dePara"].tolist())
    sel_cultivares = st.multiselect(
        "Selecione os cultivares para exibir:",
        options=cultivares_er,
        default=[c for c in top3_default if c in cultivares_er],
        key="sel_er_reg",
    )

    if not sel_cultivares:
        st.info("Selecione ao menos um cultivar para exibir o gráfico.")
    else:
        fig_reg = go_plt.Figure()

        # Range do índice ambiental para traçar as retas (cálculo interno)
        idx_min = df_er["idx_amb"].min()
        idx_max = df_er["idx_amb"].max()
        x_idx   = np.linspace(idx_min, idx_max, 100)
        # Eixo X em sc/ha real = índice + média geral
        x_scha  = x_idx + media_geral

        # Paleta distinta por cultivar — Plotly cicla automaticamente
        PALETA_ER = [
            "#2976B6","#E67E22","#27AE60","#9B59B6","#E74C3C",
            "#1ABC9C","#F39C12","#2ECC71","#8E44AD","#D35400",
        ]

        for idx_c, cultivar in enumerate(sel_cultivares):
            grp = df_er[df_er["dePara"] == cultivar]
            cor = PALETA_ER[idx_c % len(PALETA_ER)]

            y      = grp["sc_ha"].values
            x      = grp["idx_amb"].values
            X      = np.column_stack([np.ones(len(y)), x])
            beta, _, _, _ = np.linalg.lstsq(X, y, rcond=None)
            a, b   = beta

            row_res = df_final[df_final["dePara"] == cultivar]
            s2_val  = row_res["s2"].iloc[0] if not row_res.empty else np.nan
            pi_val  = row_res["Pi"].iloc[0] if not row_res.empty else np.nan

            y_line = a + b * x_idx

            # Reta de regressão
            fig_reg.add_trace(go_plt.Scatter(
                x=x_scha, y=y_line,
                mode="lines",
                name=cultivar,
                line=dict(color=cor, width=2),
                legendgroup=cultivar,
                hovertemplate=(
                    f"<b>{cultivar}</b><br>"
                    f"b: {b:.3f}<br>"
                    f"s²: {s2_val:.3f}<br>"
                    f"Pi: {pi_val:.1f}<extra></extra>"
                ),
            ))

            # Pontos observados
            x_obs = grp["idx_amb"].values + media_geral
            fig_reg.add_trace(go_plt.Scatter(
                x=x_obs,
                y=grp["sc_ha"].values,
                mode="markers",
                name=cultivar,
                marker=dict(color=cor, size=7, opacity=0.6,
                            line=dict(color=cor, width=1)),
                legendgroup=cultivar,
                showlegend=False,
                hovertemplate=(
                    f"<b>{cultivar}</b><br>"
                    "Média do local: %{x:.1f} sc/ha<br>"
                    "Produtividade: %{y:.1f} sc/ha<extra></extra>"
                ),
            ))

        # Linha de referência b=1
        y_ref = media_geral + 1.0 * x_idx
        fig_reg.add_trace(go_plt.Scatter(
            x=x_scha, y=y_ref,
            mode="lines",
            name="Referência (b=1)",
            line=dict(color="#555555", width=1.8, dash="dash"),
            hoverinfo="skip",
        ))

        # Linha vertical x = média geral (índice = 0)
        fig_reg.add_vline(x=media_geral, line=dict(color="#444444", width=1.5, dash="dot"),
                          annotation_text=f"Média geral: {media_geral:.1f}",
                          annotation_position="top",
                          annotation_font=dict(size=12, color="#222222", weight="bold"))

        fig_reg.update_layout(
            xaxis_title="Produtividade média do ambiente (sc/ha)",
            yaxis_title="Produtividade do cultivar (sc/ha)",
            height=500,
            legend=dict(
                orientation="v", x=1.01, y=1, xanchor="left",
                font=dict(size=13, color="#111111"),
                bgcolor="rgba(255,255,255,0.85)",
                bordercolor="#DDDDDD", borderwidth=1,
            ),
            margin=dict(t=40, b=60, l=60, r=160),
            plot_bgcolor="#FFFFFF",
            paper_bgcolor="#FFFFFF",
            font=dict(family="Helvetica Neue, sans-serif", size=14, color="#111111"),
            xaxis=dict(
                title=dict(font=dict(size=15, color="#111111", weight="bold")),
                tickfont=dict(size=13, color="#111111"),
                gridcolor="#E5E5E5",
            ),
            yaxis=dict(
                title=dict(font=dict(size=15, color="#111111", weight="bold")),
                tickfont=dict(size=13, color="#111111"),
                gridcolor="#E5E5E5",
            ),
        )

        st.plotly_chart(fig_reg, use_container_width=True)
        st.caption(
            "ℹ️ Reta mais inclinada que a referência (b > 1) = responsivo · "
            "Reta mais plana (b < 1) = estável · "
            "Pontos espalhados = comportamento imprevisível (s² alto)."
        )

    st.divider()
    st.markdown("#### Dispersão por Quadrantes — Adaptabilidade × Estabilidade")
    st.caption("Posição = b (adaptabilidade) × s² (estabilidade) · Tamanho do ponto = superioridade Lin & Binns (maior = melhor).")

    with st.popover("ℹ️ Como interpretar · Quadrantes", use_container_width=False):
        st.markdown("""
    **📌 O que este gráfico mostra**

    Este gráfico responde duas perguntas ao mesmo tempo sobre cada cultivar:
    1. **Ele aproveita bem ambientes favoráveis?** (adaptabilidade)
    2. **Ele é previsível entre os locais?** (estabilidade)

    Cada ponto é um cultivar. A posição no gráfico, o tamanho do ponto e a tabela abaixo contam juntos a história completa.

    ---

    **📐 Como ler os eixos**

    - **Eixo X (horizontal) → Adaptabilidade (b)**
      Mede o quanto o cultivar reage quando o ambiente melhora ou piora.
      - Pontos à **direita de 1.0** → o cultivar cresce mais que a média quando o ambiente é bom, mas também cai mais quando o ambiente é ruim. É **responsivo**.
      - Pontos à **esquerda de 1.0** → o cultivar varia menos — não sobe tanto nos bons ambientes, mas também não cai tanto nos ruins. É **pouco responsivo**.
      - Linha tracejada vertical em 1.0 = referência de adaptação ampla.

    - **Eixo Y (vertical) → Estabilidade (s²)**
      Mede o quanto o cultivar "sai da linha" — o quanto ele surpreende para cima ou para baixo além do que a adaptabilidade já explica.
      - Pontos **abaixo** da linha tracejada horizontal → comportamento **previsível** — o que a reta de regressão promete é o que o cultivar entrega
      - Pontos **acima** da linha → comportamento **imprevisível** — mesmo em ambientes parecidos, o cultivar age diferente

    ---

    **🟩 Os 4 perfis**

    | Perfil | Posição no gráfico | O que significa na prática |
    |--------|-------------------|---------------------------|
    | 🟢 **Alta Performance** | Direita + baixo | Responsivo e previsível — o melhor perfil para recomendação ampla |
    | 🟡 **Ambiente Favorável** | Direita + alto | Bom potencial mas imprevisível — apostar só em áreas de alto potencial |
    | 🔵 **Ampla Adaptação** | Esquerda + baixo | Consistente e seguro — boa escolha para ambientes adversos ou produtores avessos a risco |
    | 🔴 **Atenção** | Esquerda + alto | Não aproveita bons ambientes e ainda é imprevisível — recomendação com cautela |

    ---

    **🔵 O tamanho do ponto — Ranking de Superioridade**

    O tamanho do ponto representa **quão próximo o cultivar ficou do melhor resultado em cada local**:

    - **Ponto maior** → cultivar esteve próximo do líder na maioria dos locais — alta superioridade
    - **Ponto menor** → cultivar ficou consistentemente atrás dos melhores em cada local

    > ⚠️ **Atenção:** um cultivar pode estar no perfil 🟡 Ambiente Favorável com ponto **grande** — isso significa que ele foi responsivo, imprevisível, mas ainda assim competitivo. A combinação dos dois conta a história completa.

    ---

    **📊 Como ler o Ranking e o Δ Líder na tabela**

    - **Ranking** → posição do cultivar pela proximidade ao melhor de cada local. O 1º é o que, em média, mais vezes esteve perto do vencedor local — não necessariamente o de maior média geral.

    - **Δ Líder (sc/ha)** → a diferença média em sacas por hectare entre o cultivar e o melhor resultado de cada local:
      - **0.0** → foi o melhor em todos os locais
      - **-4.6** → perdeu em média 4.6 sc/ha para o líder de cada local
      - Quanto mais próximo de zero, mais competitivo

    > **Exemplo prático:** um cultivar com média geral alta mas Δ Líder de -12 significa que ele é bom no geral, mas em cada local havia sempre alguém claramente melhor. Já um Δ de -3 com média menor pode indicar um material mais consistentemente competitivo em todos os ambientes.
    """)


    # ── Gráfico de dispersão ──────────────────────────────────────────────────────
    COR_STATUS_PASTEL = {
        "CHECK":    "#F4B184",
        "STINE":    "#2976B6",
        "LINHAGEM": "#00FF01",
        "DP2":      "#C4DFB4",
    }
    COR_BORDA_PASTEL = {
        "CHECK":    "#C46A3A",
        "STINE":    "#1A4F7A",
        "LINHAGEM": "#009900",
        "DP2":      "#7AAF6A",
    }
    fig_er = go_plt.Figure()

    for status, cor in COR_STATUS_PASTEL.items():
        df_s = df_final[df_final["status_material"] == status]
        if df_s.empty: continue
        fig_er.add_trace(go_plt.Scatter(
            x=df_s["b"],
            y=df_s["s2"],
            mode="markers+text",
            name=status,
            text=df_s["dePara"],
            textposition="top center",
            textfont=dict(size=13, color="#111111", weight="bold"),
            marker=dict(
                color=cor,
                size=df_s["pi_size"],
                line=dict(color=COR_BORDA_PASTEL.get(status, "#888"), width=1.5),
                opacity=0.90,
            ),
            customdata=df_s[["dePara", "b", "s2", "Pi", "n_locais", "Quadrante"]].values,
            hovertemplate=(
                "<b>%{customdata[0]}</b><br>"
                "b: %{customdata[1]:.3f}<br>"
                "s²: %{customdata[2]:.3f}<br>"
                "Pi: %{customdata[3]:.1f}<br>"
                "Locais: %{customdata[4]}<br>"
                "Perfil: %{customdata[5]}<extra></extra>"
            ),
        ))

    # Linhas de referência
    fig_er.add_vline(x=1.0,      line=dict(color="#888", width=1.2, dash="dash"))
    fig_er.add_hline(y=s2_medio, line=dict(color="#888", width=1.2, dash="dash"))

    # Anotações dos quadrantes
    x_min_plot = df_final["b"].min() - 0.1
    x_max_plot = df_final["b"].max() + 0.1
    s2_max_plot = df_final["s2"].max()

    fig_er.add_annotation(x=x_max_plot, y=0,           xanchor="right", yanchor="bottom", text="🟢 Alta Performance",      showarrow=False, font=dict(size=13, color="#27AE60", weight="bold"), bgcolor="rgba(255,255,255,0.8)")
    fig_er.add_annotation(x=x_max_plot, y=s2_max_plot,  xanchor="right", yanchor="top",    text="🟡 Ambiente Favorável", showarrow=False, font=dict(size=13, color="#F39C12", weight="bold"), bgcolor="rgba(255,255,255,0.8)")
    fig_er.add_annotation(x=x_min_plot, y=0,           xanchor="left",  yanchor="bottom", text="🔵 Ampla Adaptação",     showarrow=False, font=dict(size=13, color="#2980B9", weight="bold"), bgcolor="rgba(255,255,255,0.8)")
    fig_er.add_annotation(x=x_min_plot, y=s2_max_plot,  xanchor="left",  yanchor="top",    text="🔴 Atenção",     showarrow=False, font=dict(size=13, color="#E74C3C", weight="bold"), bgcolor="rgba(255,255,255,0.8)")

    fig_er.update_layout(
        xaxis_title="b — Adaptabilidade (b = 1: adaptação ampla)",
        yaxis_title="s² — Estabilidade (menor = mais estável)",
        height=550,
        legend=dict(title="Status", orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0,
                    font=dict(size=13, color="#111111"),
                    visible=False),
        margin=dict(t=60, b=60, l=60, r=40),
        plot_bgcolor="#FFFFFF",
        paper_bgcolor="#FFFFFF",
        font=dict(family="Helvetica Neue, sans-serif", size=14, color="#111111"),
        xaxis=dict(
            title=dict(font=dict(size=15, color="#111111", weight="bold")),
            tickfont=dict(size=13, color="#111111", weight="bold"),
            gridcolor="#E5E5E5",
        ),
        yaxis=dict(
            title=dict(font=dict(size=15, color="#111111", weight="bold")),
            tickfont=dict(size=13, color="#111111", weight="bold"),
            gridcolor="#E5E5E5",
        ),
    )

    st.plotly_chart(fig_er, use_container_width=True)
    st.caption(
        f"ℹ️ **Eberhart & Russell:** b = adaptabilidade (referência = 1,0) · s² = desvio da regressão · s² médio do conjunto = {s2_medio:.3f}. "
        f"**Lin & Binns:** tamanho do ponto proporcional ao Pi — pontos **maiores** = maior superioridade (mais próximo do líder). "
        f"Cultivares com menos de {MIN_LOCAIS} locais excluídos da análise."
    )

    # ── Tabela resumo ─────────────────────────────────────────────────────────────

    def label_responsividade(b):
        if pd.isna(b): return "—"
        if b > 1.15:   return f"Alta ({b:.2f})"
        if b >= 0.85:  return f"Ampla ({b:.2f})"
        return f"Baixa ({b:.2f})"

    def label_estabilidade(s2, s2_med):
        if pd.isna(s2): return "—"
        if s2 <= s2_med * 0.5:  return f"Alta ({s2:.2f})"
        if s2 <= s2_med:        return f"Média ({s2:.2f})"
        return f"Baixa ({s2:.2f})"

    # Ranking por Pi (menor = melhor)
    df_final_tab = df_final.sort_values("Pi", ascending=True).reset_index(drop=True)
    df_final_tab["Ranking"] = df_final_tab.index + 1

    medalhas = {1: "🥇", 2: "🥈", 3: "🥉"}
    df_final_tab["Ranking"] = df_final_tab["Ranking"].apply(
        lambda r: f"{medalhas.get(r, '')} {r}º".strip()
    )

    # Δ Líder: média simples das diferenças por local (negativo = abaixo do melhor)
    df_final_tab["Δ Líder (sc/ha)"] = df_final_tab["delta_lider"].apply(
        lambda d: 0.0 if d < 0.01 else round(-d, 1)
    )

    df_tabela_er = pd.DataFrame({
        "Cultivar":        df_final_tab["dePara"],
        "Locais":          df_final_tab["n_locais"],
        "Responsividade":  df_final_tab["b"].apply(label_responsividade),
        "Estabilidade":    df_final_tab.apply(lambda r: label_estabilidade(r["s2"], s2_medio), axis=1),
        "Ranking":         df_final_tab["Ranking"],
        "Δ Líder (sc/ha)": df_final_tab["Δ Líder (sc/ha)"],
        "Perfil":          df_final_tab["Quadrante"],
    })

    ag_table(df_tabela_er, height=min(600, 80 + len(df_tabela_er) * 36))
    exportar_excel(df_tabela_er, nome_arquivo="adaptabilidade_estabilidade.xlsx", label="⬇️ Exportar Adaptabilidade e Estabilidade", key="exp_er")
    st.caption(
        "ℹ️ **Δ Líder (sc/ha)** = média simples de quanto o cultivar ficou abaixo do melhor resultado em cada local — "
        "0.0 significa que foi o melhor em todos os locais. "
        "**Ranking** baseado no índice Pi de Lin & Binns. "
        "**Responsividade** e **Estabilidade** pelos coeficientes b e s² de Eberhart & Russell."
    )

st.divider()

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 9 — HEATMAP DE DESEMPENHO POR LOCAL
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Heatmap",
    "Como cada cultivar se saiu em cada local?",
    "Visualize padrões de desempenho — onde cada cultivar foi forte, onde ficou atrás e quais locais são mais exigentes.",
)

with st.popover("ℹ️ Como interpretar · Heatmap", use_container_width=False):
    st.markdown("""
**📌 O que este gráfico mostra**

Cada célula representa o desempenho de um cultivar em um local específico. A cor indica se o cultivar foi bem ou mal **em relação ao melhor resultado daquele local** — não em termos absolutos.

---

**🎨 Como ler as cores**

- **Verde escuro** → desempenho excelente — próximo do líder do local (≥ 95%)
- **Verde claro** → desempenho competitivo (90–95%)
- **Amarelo** → zona de transição — acima da média, ainda não no topo (87–90%)
- **Creme** → borderline — próximo do limiar competitivo (~80–87%)
- **Salmão / Vermelho** → abaixo do esperado — ficou para trás naquele ambiente
- **Cinza** → cultivar não foi avaliado naquele local

> A cor é sempre **relativa ao local** — um cultivar pode ser verde em um local difícil e salmão em um local fácil. O que importa é a posição relativa entre os cultivares **dentro de cada coluna**.

---

**📊 Os dois modos de visualização**

- **Produção Relativa (%)** → mostra quanto o cultivar produziu em relação ao melhor do local (100% = foi o melhor). Exemplo: 87% significa que produziu 87% do que o líder daquele local produziu.

- **Ranking por local** → mostra a posição do cultivar naquele local (1º, 2º, 3º...). Mais intuitivo para comparar posições, mas não mostra a distância entre os colocados.

---

**📐 Como ler as linhas e colunas**

- **Linhas (cultivares)** → agrupados por status (CHECK, STINE, DP2), ordem alfabética dentro de cada grupo
- **Colunas (locais)** → ordenados por estado → cidade → código do local
- **Padrão de verde numa linha** → cultivar consistente — bom em muitos locais
- **Verde isolado numa coluna** → local muito seletivo — poucos cultivares foram bem ali
- **Linha toda salmão/vermelha** → cultivar abaixo do esperado em todos os ambientes

---

**💡 Dica de leitura**

Compare as colunas verticalmente — se uma coluna tem muitos verdes, aquele local favoreceu quase todos (ambiente fácil). Se tem muitos vermelhos, foi um ambiente exigente onde só os melhores se destacaram.
""")

# ── Dados ─────────────────────────────────────────────────────────────────────
# Escopo: cultivares e locais definidos pelo filtro
df_hm_scope = ta_filtrado[ta_filtrado["sc_ha"] > 0][
    ["dePara", "status_material", "cod_fazenda", "estado_sigla", "cidade_nome"]
].dropna().drop_duplicates()

cultivares_no_filtro = df_hm_scope["dePara"].unique().tolist()
locais_no_filtro     = df_hm_scope["cod_fazenda"].unique().tolist()

# Valores: calculados sobre ta_raw restrito aos mesmos locais (contexto completo)
df_hm_full = ta_raw[
    (ta_raw["sc_ha"] > 0) &
    (ta_raw["cod_fazenda"].isin(locais_no_filtro))
][["dePara", "cod_fazenda", "sc_ha"]].dropna().copy()

# Máximo real por local (todos os cultivares do local, não só os filtrados)
maximo_por_local = df_hm_full.groupby("cod_fazenda")["sc_ha"].max().rename("max_local")
df_hm_full = df_hm_full.join(maximo_por_local, on="cod_fazenda")
df_hm_full["prod_rel"] = (df_hm_full["sc_ha"] / df_hm_full["max_local"] * 100).round(1)

# Ranking real por local (posição entre todos os cultivares do local)
df_hm_full["ranking_local"] = (
    df_hm_full.groupby("cod_fazenda")["sc_ha"]
    .rank(ascending=False, method="min")
    .astype(int)
)

# Total de cultivares por local (para mostrar "Xº de N" no hover)
total_por_local = df_hm_full.groupby("cod_fazenda")["dePara"].nunique().rename("total_local")
df_hm_full = df_hm_full.join(total_por_local, on="cod_fazenda")

# Diferença absoluta para o líder do local (sc/ha)
df_hm_full["diff_sc"] = ((df_hm_full["sc_ha"] - df_hm_full["max_local"]) / 60).round(1)

# Filtrar apenas os cultivares do escopo para montar o pivot
df_hm_plot = df_hm_full[df_hm_full["dePara"].isin(cultivares_no_filtro)].copy()

# Pivot de total por local (para hover)
pivot_total  = df_hm_plot.pivot_table(index="dePara", columns="cod_fazenda", values="total_local", aggfunc="first")
pivot_diff   = df_hm_plot.pivot_table(index="dePara", columns="cod_fazenda", values="diff_sc",     aggfunc="mean")

# Ordenação dos locais: estado → cidade → cod_fazenda
locais_meta = (
    df_hm_scope[["cod_fazenda", "estado_sigla", "cidade_nome"]]
    .drop_duplicates()
    .sort_values("cod_fazenda")
)
locais_ordem = locais_meta["cod_fazenda"].tolist()

# Ordenação dos cultivares: status order → alfabético
STATUS_ORDER = ["CHECK", "STINE", "DP2", "LINHAGEM"]
cult_status = (
    df_hm_scope[["dePara", "status_material"]]
    .drop_duplicates()
    .assign(_ord=lambda d: d["status_material"].apply(
        lambda s: STATUS_ORDER.index(s) if s in STATUS_ORDER else 99))
    .sort_values(["_ord", "dePara"])
)
cultivares_ordem = cult_status["dePara"].tolist()

# Pivot tables
pivot_rel   = df_hm_plot.pivot_table(index="dePara", columns="cod_fazenda", values="prod_rel",      aggfunc="mean")
pivot_rank  = df_hm_plot.pivot_table(index="dePara", columns="cod_fazenda", values="ranking_local", aggfunc="min")
pivot_total = pivot_total if "pivot_total" in dir() else pivot_rank * 0

pivot_rel   = pivot_rel.reindex(index=cultivares_ordem,  columns=locais_ordem)
pivot_rank  = pivot_rank.reindex(index=cultivares_ordem, columns=locais_ordem)
pivot_total = pivot_total.reindex(index=cultivares_ordem, columns=locais_ordem)
pivot_diff  = pivot_diff.reindex(index=cultivares_ordem,  columns=locais_ordem)

# ── Abas dos heatmaps ────────────────────────────────────────────────────────
# ── Helper: rótulos brancos nas zonas escuras do heatmap ─────────────────────
# A paleta RdYlGn tem duas zonas escuras: vermelho-escuro (início) e verde-escuro (fim).
# Em ambas o texto preto perde contraste — adicionamos annotations brancas.
# A zona "clara" (laranja/amarelo/verde-claro) fica com o textfont preto padrão.
_HM_ZMIN = 0    # zmin fixo — 0 para não clipar nenhum valor

def _add_white_labels(fig, pivot_val, pivot_diff, rows, cols, mostrar=True, zmin=0, zmax=100):
    """Annotations brancas nas zonas vermelho-escuro e verde-escuro da paleta RdYlGn."""
    _rng = zmax - zmin
    _lo  = zmin + 0.15 * _rng   # ≤15% — vermelho escuro → texto branco
    _hi  = zmin + 0.95 * _rng   # ≥95% — verde escuro → texto branco
    for i, row_id in enumerate(rows):
        for j, col_id in enumerate(cols):
            try:
                v = pivot_val.iloc[i, j]
                d = pivot_diff.iloc[i, j]
            except (IndexError, KeyError):
                continue
            if np.isnan(v) or (_lo < v < _hi):
                continue   # zona clara — texto preto do heatmap já está ok
            d_ok = not np.isnan(d)
            if v >= 99.5:
                txt = f"{v:.0f}%<br>líder" if mostrar else f"{v:.0f}%"
            else:
                txt = (f"{v:.0f}%<br>{d:+.1f} sc" if mostrar and d_ok else f"{v:.0f}%")
            fig.add_annotation(
                x=col_id, xref="x",
                y=row_id, yref="y",
                text=f"<b>{txt}</b>",
                showarrow=False,
                font=dict(size=13, color="#FFFFFF", weight="bold"),
                align="center",
            )

def _add_nan_mask(fig, z_vals, x, y, xgap=2, ygap=2):
    """Sobrepõe um trace cinza em células NaN para distinguir 'não avaliado' do creme neutro."""
    z_mask = [[1 if (v is None or (isinstance(v, float) and np.isnan(v))) else float("nan")
               for v in row] for row in z_vals]
    fig.add_trace(go_plt.Heatmap(
        z=z_mask,
        x=x, y=y,
        colorscale=[[0, "#DCDCDC"], [1, "#DCDCDC"]],
        zmin=0, zmax=1,
        zauto=False,
        showscale=False,
        xgap=xgap, ygap=ygap,
        hoverinfo="skip",
    ))

tab_hm_local, tab_hm_reg, tab_hm_filtro = st.tabs(["📍 Por Local", "🗺️ Por Região", "🔍 Relativo ao Filtro"])

with tab_hm_local:

    col_modo_hm, col_rotulo_hm = st.columns([3, 1])
    with col_modo_hm:
        modo_hm = st.radio(
            "Visualizar por:",
            options=["Produção Relativa (%)", "Ranking por local"],
            horizontal=True,
            key="radio_heatmap",
        )
    with col_rotulo_hm:
        mostrar_rotulos_hm = st.checkbox("Mostrar rótulos", value=True, key="chk_rotulos_hm")

    # ── Gráfico ───────────────────────────────────────────────────────────────────
    _lo_hm = _HM_ZMIN + 0.15 * (100 - _HM_ZMIN)   # ≤15% — zona vermelha escura
    _hi_hm = _HM_ZMIN + 0.95 * (100 - _HM_ZMIN)   # ≥95% — zona verde escura
    if modo_hm == "Produção Relativa (%)":
        pivot_plot  = pivot_rel
        colorscale  = COLORSCALE_PERF
        zmin, zmax  = _HM_ZMIN, 100
        colorbar_title = "Prod. Rel. (%)"
        # Porcentagem fixa — diferença em sc controlada pelo checkbox
        text_mat = []
        hover_mat = []
        for i, cultivar in enumerate(cultivares_ordem):
            row_text  = []
            row_hover = []
            for j, local in enumerate(locais_ordem):
                v = pivot_rel.iloc[i, j]  if i < len(pivot_rel)  and j < len(pivot_rel.columns)  else float("nan")
                d = pivot_diff.iloc[i, j] if i < len(pivot_diff) and j < len(pivot_diff.columns) else float("nan")
                if np.isnan(v):
                    row_text.append("")
                    row_hover.append("—")
                elif not (_lo_hm < v < _hi_hm):
                    row_text.append("")   # zona escura — annotation branca adicionada depois
                    row_hover.append(f"{v:.0f}% · líder do local" if v >= 99.5 else f"{v:.0f}% · {d:+.1f} sc/ha vs líder")
                elif v >= 99.5:
                    row_text.append(f"{v:.0f}%<br>líder" if mostrar_rotulos_hm else f"{v:.0f}%")
                    row_hover.append(f"{v:.0f}% · líder do local")
                else:
                    row_text.append(f"{v:.0f}%<br>{d:+.1f} sc" if mostrar_rotulos_hm else f"{v:.0f}%")
                    row_hover.append(f"{v:.0f}% · {d:+.1f} sc/ha vs líder")
            text_mat.append(row_text)
            hover_mat.append(row_hover)
    else:
        pivot_plot  = pivot_rank
        colorscale  = [[0, "#1a9850"], [0.5, "#fee08b"], [1, "#d73027"]]
        zmin        = 1
        zmax        = int(pivot_rank.max().max()) if not pivot_rank.empty else 10
        colorbar_title = "Ranking"
        text_mat  = []
        hover_mat = []
        for i, cultivar in enumerate(cultivares_ordem):
            row_text  = []
            row_hover = []
            for j, local in enumerate(locais_ordem):
                r = pivot_rank.iloc[i, j]  if i < len(pivot_rank)  and j < len(pivot_rank.columns)  else float("nan")
                t = pivot_total.iloc[i, j] if i < len(pivot_total) and j < len(pivot_total.columns) else float("nan")
                if np.isnan(r):
                    row_text.append("")
                    row_hover.append("—")
                else:
                    total_str = f" de {int(t)}" if not np.isnan(t) else ""
                    row_text.append(f"{int(r)}º")
                    row_hover.append(f"{int(r)}º{total_str}")
            text_mat.append(row_text)
            hover_mat.append(row_hover)

    z_vals = pivot_plot.values.tolist()

    n_cult    = len(cultivares_ordem)
    n_loc     = len(locais_ordem)
    row_h     = 52 if modo_hm == "Produção Relativa (%)" else 38
    altura_hm = max(350, n_cult * row_h + 100)

    fig_hm = go_plt.Figure(go_plt.Heatmap(
        z=z_vals,
        x=locais_ordem,
        y=cultivares_ordem,
        text=text_mat,
        customdata=hover_mat,
        texttemplate="%{text}",
        textfont=dict(size=13, color="#111111", weight="bold"),
        colorscale=colorscale,
        zmin=zmin, zmax=zmax,
        zauto=False,
        xgap=2, ygap=2,
        colorbar=dict(
            title=dict(text=colorbar_title, font=dict(size=12)),
            tickfont=dict(size=11),
            thickness=14,
        ),
        hovertemplate="<b>%{y}</b> · %{x}<br>" + colorbar_title + ": %{customdata}<extra></extra>",
    ))

    _add_nan_mask(fig_hm, z_vals, locais_ordem, cultivares_ordem)

    # Separadores visuais entre grupos de status
    cult_status_map = cult_status.set_index("dePara")["status_material"].to_dict()
    for i, cultivar in enumerate(cultivares_ordem[:-1]):
        status_atual = cult_status_map.get(cultivar, "")
        status_prox  = cult_status_map.get(cultivares_ordem[i+1], "")
        if status_atual != status_prox:
            fig_hm.add_shape(
                type="line",
                x0=0, x1=1, xref="paper",
                y0=i + 0.5, y1=i + 0.5, yref="y",
                line=dict(color="#333333", width=2),
            )

    fig_hm.update_layout(
        height=altura_hm,
        xaxis=dict(
            side="bottom",
            tickfont=dict(size=12, color="#111111", weight="bold"),
            title=dict(text="<b>Local (cod_fazenda)</b>", font=dict(size=15, color="#111111")),
        ),
        yaxis=dict(
            tickfont=dict(size=13, color="#111111", weight="bold"),
            autorange="reversed",
            showticklabels=False,  # esconde ticks padrão — usamos annotations coloridas
        ),
        margin=dict(t=30, b=80, l=180, r=60),
        plot_bgcolor="#FFFFFF",
        paper_bgcolor="#FFFFFF",
        font=dict(family="Helvetica Neue, sans-serif", size=12, color="#111111"),
    )

    # Annotations coloridas por status no eixo Y
    COR_STATUS_TEXTO_HM = {
        "CHECK":    "#C46A3A",
        "STINE":    "#2976B6",
        "LINHAGEM": "#009900",
        "DP2":      "#5A8A4A",
    }
    for i, cultivar in enumerate(cultivares_ordem):
        status = cult_status_map.get(cultivar, "")
        cor    = COR_STATUS_TEXTO_HM.get(status, "#333333")
        fig_hm.add_annotation(
            x=-0.01, xref="paper",
            y=i,     yref="y",
            text=f"<b>{cultivar}</b>",
            showarrow=False,
            xanchor="right",
            yanchor="middle",
            font=dict(size=13, color=cor, weight="bold"),
        )

    if modo_hm == "Produção Relativa (%)":
        _add_white_labels(fig_hm, pivot_rel, pivot_diff, cultivares_ordem, locais_ordem, mostrar_rotulos_hm, zmin=_HM_ZMIN, zmax=100)

    st.plotly_chart(fig_hm, use_container_width=True)
    st.caption(
        "ℹ️ Escala: vermelho → salmão → creme (~80%) → amarelo (~87%) → verde (≥90%) → verde escuro (100%) · "
        "Cinza = cultivar não avaliado naquele local · "
        "Linha preta = divisão entre grupos de status."
    )

    # Dicionário de locais
    df_dic_hm = (
        ta_filtrado[["cod_fazenda", "nomeFazenda", "cidade_nome", "estado_sigla"]]
        .drop_duplicates()
        .sort_values(["estado_sigla", "cidade_nome", "cod_fazenda"])
        .rename(columns={
            "cod_fazenda":  "Código",
            "nomeFazenda":  "Local",
            "cidade_nome":  "Cidade",
            "estado_sigla": "Estado",
        })
        .reset_index(drop=True)
    )
    with st.popover(f"📍 Dicionário de locais ({len(df_dic_hm)} locais)", use_container_width=False):
        st.markdown("Referência dos códigos exibidos nas colunas do heatmap.")
        busca_dic_hm = st.text_input("🔍 Buscar", placeholder="Código, local ou cidade...", key="busca_dic_hm")
        df_dic_hm_show = df_dic_hm[
            df_dic_hm.apply(lambda r: busca_dic_hm.strip().lower() in " ".join(r.astype(str)).lower(), axis=1)
        ] if busca_dic_hm.strip() else df_dic_hm
        st.dataframe(df_dic_hm_show, hide_index=True, use_container_width=True)




with tab_hm_reg:
    col_reg_nivel, col_reg_rotulo = st.columns([3, 1])
    with col_reg_nivel:
        nivel_reg = st.radio(
            "Agrupar por:",
            options=["Região Macro", "Região Micro", "Estado"],
            horizontal=True,
            key="radio_hm_reg",
        )
    with col_reg_rotulo:
        mostrar_rotulos_reg = st.checkbox("Mostrar rótulos", value=True, key="chk_rotulos_reg")

    _col_reg_map = {
        "Região Macro": "regiao_macro",
        "Região Micro": "regiao_micro",
        "Estado":       "estado_sigla",
    }
    _col_reg = _col_reg_map[nivel_reg]

    if _col_reg not in ta_filtrado.columns:
        st.info(f"Coluna '{_col_reg}' não disponível nos dados filtrados.")
    else:
        _df_reg_hm = (
            ta_filtrado[ta_filtrado["sc_ha"] > 0][["dePara", _col_reg, "sc_ha"]]
            .dropna()
            .groupby(["dePara", _col_reg])["sc_ha"]
            .mean()
            .reset_index()
        )
        _lider_reg = _df_reg_hm.groupby(_col_reg)["sc_ha"].max().rename("max_reg")
        _df_reg_hm = _df_reg_hm.join(_lider_reg, on=_col_reg)
        _df_reg_hm["prod_rel_reg"] = (_df_reg_hm["sc_ha"] / _df_reg_hm["max_reg"] * 100).round(1)
        _df_reg_hm["diff_sc_reg"]  = ((_df_reg_hm["sc_ha"] - _df_reg_hm["max_reg"]) / 60).round(1)

        _pivot_reg = _df_reg_hm.pivot_table(
            index="dePara", columns=_col_reg,
            values="prod_rel_reg", aggfunc="mean"
        )
        _pivot_diff_reg = _df_reg_hm.pivot_table(
            index="dePara", columns=_col_reg,
            values="diff_sc_reg", aggfunc="mean"
        )
        _regioes_ordem = sorted(_pivot_reg.columns.tolist())
        _pivot_reg      = _pivot_reg.reindex(index=cultivares_ordem, columns=_regioes_ordem)
        _pivot_diff_reg = _pivot_diff_reg.reindex(index=cultivares_ordem, columns=_regioes_ordem)

        _text_mat_reg  = []
        _hover_mat_reg = []
        _lo_reg = _HM_ZMIN + 0.15 * (100 - _HM_ZMIN)
        _hi_reg = _HM_ZMIN + 0.95 * (100 - _HM_ZMIN)
        for i, cultivar in enumerate(cultivares_ordem):
            _row_t, _row_h = [], []
            for j, reg in enumerate(_regioes_ordem):
                v = _pivot_reg.iloc[i, j]      if i < len(_pivot_reg)      and j < len(_pivot_reg.columns)      else float("nan")
                d = _pivot_diff_reg.iloc[i, j] if i < len(_pivot_diff_reg) and j < len(_pivot_diff_reg.columns) else float("nan")
                if np.isnan(v):
                    _row_t.append(""); _row_h.append("—")
                elif not (_lo_reg < v < _hi_reg):
                    _row_t.append("")   # zona escura — annotation branca adicionada depois
                    _row_h.append(f"{v:.0f}% · líder da região" if v >= 99.5 else f"{v:.0f}% · {d:+.1f} sc/ha vs líder")
                elif v >= 99.5:
                    _row_t.append(f"{v:.0f}%<br>líder" if mostrar_rotulos_reg else f"{v:.0f}%")
                    _row_h.append(f"{v:.0f}% · líder da região")
                else:
                    _row_t.append(f"{v:.0f}%<br>{d:+.1f} sc" if mostrar_rotulos_reg else f"{v:.0f}%")
                    _row_h.append(f"{v:.0f}% · {d:+.1f} sc/ha vs líder")
            _text_mat_reg.append(_row_t)
            _hover_mat_reg.append(_row_h)

        _colorscale_reg = COLORSCALE_PERF
        _altura_r = max(350, len(cultivares_ordem) * 52 + 100)

        fig_hm_reg = go_plt.Figure(go_plt.Heatmap(
            z=_pivot_reg.values.tolist(),
            x=_regioes_ordem,
            y=cultivares_ordem,
            text=_text_mat_reg,
            customdata=_hover_mat_reg,
            texttemplate="%{text}",
            textfont=dict(size=13, color="#111111", weight="bold"),
            colorscale=_colorscale_reg,
            zmin=_HM_ZMIN, zmax=100,
            zauto=False,
            xgap=2, ygap=2,
            colorbar=dict(
                title=dict(text="Prod. Rel. (%)", font=dict(size=12)),
                tickfont=dict(size=11),
                thickness=14,
            ),
            hovertemplate="<b>%{y}</b> · %{x}<br>Prod. Rel.: %{customdata}<extra></extra>",
        ))

        _add_nan_mask(fig_hm_reg, _pivot_reg.values.tolist(), _regioes_ordem, cultivares_ordem)

        for i, cultivar in enumerate(cultivares_ordem[:-1]):
            if cult_status_map.get(cultivar, "") != cult_status_map.get(cultivares_ordem[i+1], ""):
                fig_hm_reg.add_shape(
                    type="line", x0=0, x1=1, xref="paper",
                    y0=i + 0.5, y1=i + 0.5, yref="y",
                    line=dict(color="#333333", width=2),
                )

        fig_hm_reg.update_layout(
            height=_altura_r,
            xaxis=dict(
                side="bottom",
                tickfont=dict(size=12, color="#111111", weight="bold"),
                title=dict(text=f"<b>{nivel_reg}</b>", font=dict(size=14, color="#111111")),
            ),
            yaxis=dict(
                tickfont=dict(size=13, color="#111111", weight="bold"),
                autorange="reversed",
                showticklabels=False,
            ),
            margin=dict(t=30, b=80, l=180, r=60),
            plot_bgcolor="#FFFFFF", paper_bgcolor="#FFFFFF",
            font=dict(family="Helvetica Neue, sans-serif", size=12, color="#111111"),
        )

        for i, cultivar in enumerate(cultivares_ordem):
            status = cult_status_map.get(cultivar, "")
            cor    = COR_STATUS_TEXTO_HM.get(status, "#333333")
            fig_hm_reg.add_annotation(
                x=-0.01, xref="paper", y=i, yref="y",
                text=f"<b>{cultivar}</b>",
                showarrow=False, xanchor="right", yanchor="middle",
                font=dict(size=13, color=cor, weight="bold"),
            )

        _add_white_labels(fig_hm_reg, _pivot_reg, _pivot_diff_reg, cultivares_ordem, _regioes_ordem, mostrar_rotulos_reg, zmin=_HM_ZMIN, zmax=100)

        st.plotly_chart(fig_hm_reg, use_container_width=True)
        st.caption(
            f"ℹ️ Produção relativa média por {nivel_reg.lower()} · % em relação ao líder de cada região. "
            "Escala: vermelho → salmão → creme (~80%) → amarelo (~87%) → verde (≥90%) → verde escuro (100%)."
        )

# ── Tab 3 — Heatmap Relativo ao Filtro ───────────────────────────────────────
with tab_hm_filtro:
    st.caption(
        "🔍 Neste heatmap o 100% é o **melhor cultivar dentro do filtro atual** em cada local — "
        "não o melhor absoluto. Use para comparar materiais dentro de um nicho de GM ou status. "
        "O cabeçalho de cada coluna mostra o gap entre o líder do filtro e o líder absoluto daquele local."
    )

    col_rotulo_ft = st.columns([3, 1])[1]
    with col_rotulo_ft:
        mostrar_rotulos_ft = st.checkbox("Mostrar rótulos", value=True, key="chk_rotulos_ft")

    # ── Cálculo: max recalculado apenas sobre cultivares do filtro ────────────
    df_ft_base = ta_filtrado[ta_filtrado["sc_ha"] > 0][
        ["dePara", "cod_fazenda", "sc_ha"]
    ].dropna().copy()

    # Máximo do filtro por local
    max_filtro_local = df_ft_base.groupby("cod_fazenda")["sc_ha"].max().rename("max_filtro")
    df_ft_base = df_ft_base.join(max_filtro_local, on="cod_fazenda")
    df_ft_base["prod_rel_ft"] = (df_ft_base["sc_ha"] / df_ft_base["max_filtro"] * 100).round(1)
    df_ft_base["diff_sc_ft"]  = ((df_ft_base["sc_ha"] - df_ft_base["max_filtro"]) / 60).round(1)

    # Máximo absoluto por local (de ta_raw, mesmos locais) para calcular gap
    locais_ft = df_ft_base["cod_fazenda"].unique().tolist()
    max_abs_local = (
        ta_raw[(ta_raw["sc_ha"] > 0) & (ta_raw["cod_fazenda"].isin(locais_ft))]
        .groupby("cod_fazenda")["sc_ha"].max()
        .rename("max_abs")
    )
    # Gap = (max_filtro - max_abs) / max_abs * 100  — negativo = filtro abaixo do absoluto
    gap_series = ((max_filtro_local / max_abs_local - 1) * 100).round(1)

    pivot_rel_ft  = df_ft_base.pivot_table(
        index="dePara", columns="cod_fazenda", values="prod_rel_ft", aggfunc="mean"
    )
    pivot_diff_ft = df_ft_base.pivot_table(
        index="dePara", columns="cod_fazenda", values="diff_sc_ft", aggfunc="mean"
    )

    pivot_rel_ft  = pivot_rel_ft.reindex(index=cultivares_ordem,  columns=locais_ordem)
    pivot_diff_ft = pivot_diff_ft.reindex(index=cultivares_ordem, columns=locais_ordem)

    # zmin dinâmico: 3pp abaixo do mínimo real do filtro, com piso em 60%
    _vals_ft = pivot_rel_ft.stack().dropna()
    _zmin_ft = max(60, int(_vals_ft.min()) - 3) if not _vals_ft.empty else 60
    _lo_ft   = _zmin_ft + 0.15 * (100 - _zmin_ft)
    _hi_ft   = _zmin_ft + 0.95 * (100 - _zmin_ft)

    # ── Textos e hover ────────────────────────────────────────────────────────
    text_mat_ft  = []
    hover_mat_ft = []
    for i, cultivar in enumerate(cultivares_ordem):
        row_t, row_h = [], []
        for j, local in enumerate(locais_ordem):
            v = pivot_rel_ft.iloc[i, j]  if i < len(pivot_rel_ft)  and j < len(pivot_rel_ft.columns)  else float("nan")
            d = pivot_diff_ft.iloc[i, j] if i < len(pivot_diff_ft) and j < len(pivot_diff_ft.columns) else float("nan")
            if np.isnan(v):
                row_t.append(""); row_h.append("—")
            elif not (_lo_ft < v < _hi_ft):
                row_t.append("")   # zona escura — annotation branca adicionada depois
                row_h.append(f"{v:.0f}% · líder do filtro neste local" if v >= 99.5 else f"{v:.0f}% · {d:+.1f} sc/ha vs líder do filtro")
            elif v >= 99.5:
                row_t.append(f"{v:.0f}%<br>líder" if mostrar_rotulos_ft else f"{v:.0f}%")
                row_h.append(f"{v:.0f}% · líder do filtro neste local")
            else:
                row_t.append(f"{v:.0f}%<br>{d:+.1f} sc" if mostrar_rotulos_ft else f"{v:.0f}%")
                row_h.append(f"{v:.0f}% · {d:+.1f} sc/ha vs líder do filtro")
        text_mat_ft.append(row_t)
        hover_mat_ft.append(row_h)

    n_cult_ft = len(cultivares_ordem)
    altura_ft = max(350, n_cult_ft * 52 + 120)

    fig_ft = go_plt.Figure(go_plt.Heatmap(
        z=pivot_rel_ft.values.tolist(),
        x=locais_ordem,
        y=cultivares_ordem,
        text=text_mat_ft,
        customdata=hover_mat_ft,
        texttemplate="%{text}",
        textfont=dict(size=13, color="#111111", weight="bold"),
        colorscale=COLORSCALE_PERF,
        zmin=_zmin_ft, zmax=100,
        zauto=False,
        xgap=2, ygap=2,
        colorbar=dict(
            title=dict(text=f"Prod. Rel. Filtro (%) · mín={_zmin_ft}%", font=dict(size=11)),
            tickfont=dict(size=11),
            thickness=14,
        ),
        hovertemplate="<b>%{y}</b> · %{x}<br>%{customdata}<extra></extra>",
    ))

    _add_nan_mask(fig_ft, pivot_rel_ft.values.tolist(), locais_ordem, cultivares_ordem)

    # Separadores entre grupos de status
    for i, cultivar in enumerate(cultivares_ordem[:-1]):
        if cult_status_map.get(cultivar, "") != cult_status_map.get(cultivares_ordem[i+1], ""):
            fig_ft.add_shape(
                type="line", x0=0, x1=1, xref="paper",
                y0=i + 0.5, y1=i + 0.5, yref="y",
                line=dict(color="#333333", width=2),
            )

    # Annotations de gap no header de cada coluna
    for local in locais_ordem:
        gap = gap_series.get(local, float("nan"))
        if np.isnan(gap):
            continue
        if gap >= -0.5:
            gap_txt = "✓"
            gap_cor = "#1a9850"
        else:
            gap_txt = f"▼{abs(gap):.0f}%"
            gap_cor = "#d73027" if gap < -5 else "#f46d43"

        fig_ft.add_annotation(
            x=local, xref="x",
            y=1.01,  yref="paper",
            text=f"<b>{gap_txt}</b>",
            showarrow=False,
            yanchor="bottom",
            font=dict(size=13, color=gap_cor, weight="bold"),
        )

    fig_ft.update_layout(
        height=altura_ft,
        xaxis=dict(
            side="bottom",
            tickfont=dict(size=12, color="#111111", weight="bold"),
            title=dict(text="<b>Local (cod_fazenda)</b>", font=dict(size=15, color="#111111")),
        ),
        yaxis=dict(
            tickfont=dict(size=13, color="#111111", weight="bold"),
            autorange="reversed",
            showticklabels=False,
        ),
        margin=dict(t=50, b=80, l=180, r=60),
        plot_bgcolor="#FFFFFF",
        paper_bgcolor="#FFFFFF",
        font=dict(family="Helvetica Neue, sans-serif", size=12, color="#111111"),
    )

    # Labels coloridas por status no eixo Y
    for i, cultivar in enumerate(cultivares_ordem):
        status = cult_status_map.get(cultivar, "")
        cor    = COR_STATUS_TEXTO_HM.get(status, "#333333")
        fig_ft.add_annotation(
            x=-0.01, xref="paper",
            y=i,     yref="y",
            text=f"<b>{cultivar}</b>",
            showarrow=False,
            xanchor="right",
            yanchor="middle",
            font=dict(size=13, color=cor, weight="bold"),
        )

    _add_white_labels(fig_ft, pivot_rel_ft, pivot_diff_ft, cultivares_ordem, locais_ordem, mostrar_rotulos_ft, zmin=_zmin_ft, zmax=100)

    st.plotly_chart(fig_ft, use_container_width=True)
    st.caption(
        "ℹ️ 100% = melhor cultivar do filtro naquele local (não o absoluto). "
        "Cabeçalho: ✓ = líder do filtro = líder absoluto · ▼X% = líder do filtro está X% abaixo do melhor absoluto. "
        "Cinza = cultivar não avaliado · Escala: vermelho → creme (~80%) → amarelo (~87%) → verde (≥90%)."
    )

    # Dicionário de locais
    df_dic_ft = (
        ta_filtrado[["cod_fazenda", "nomeFazenda", "cidade_nome", "estado_sigla"]]
        .drop_duplicates()
        .sort_values(["estado_sigla", "cidade_nome", "cod_fazenda"])
        .rename(columns={
            "cod_fazenda":  "Código",
            "nomeFazenda":  "Local",
            "cidade_nome":  "Cidade",
            "estado_sigla": "Estado",
        })
        .reset_index(drop=True)
    )
    with st.popover(f"📍 Dicionário de locais ({len(df_dic_ft)} locais)", use_container_width=False):
        st.markdown("Referência dos códigos exibidos nas colunas do heatmap.")
        busca_dic_ft = st.text_input("🔍 Buscar", placeholder="Código, local ou cidade...", key="busca_dic_ft")
        df_dic_ft_show = df_dic_ft[
            df_dic_ft.apply(lambda r: busca_dic_ft.strip().lower() in " ".join(r.astype(str)).lower(), axis=1)
        ] if busca_dic_ft.strip() else df_dic_ft
        st.dataframe(df_dic_ft_show, hide_index=True, use_container_width=True)

st.markdown(
    '<p style="font-size:13px;color:#374151;text-align:center;">Painel JAUM DTC · Stine Seed · '
    'Desenvolvido por <a href="https://www.linkedin.com/in/eng-agro-andre-ferreira/" '
    'target="_blank" style="color:#27AE60;text-decoration:none;">Andre Ferreira</a></p>',
    unsafe_allow_html=True,
)
