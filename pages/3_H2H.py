"""
pages/3_H2H.py — Head-to-Head · JAUM DTC

Dois modos em abas:
  · Tab 1 — Tabela de Classificação: Produto 1 vs todos os adversários disponíveis
  · Tab 2 — Análise por Local: par específico (Prod1 × Prod2), cards + donut + barras
"""

import io
import numpy as np
import pandas as pd
import streamlit as st
import plotly.graph_objects as go_plt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from utils.theme import aplicar_tema, page_header, secao_titulo
from utils.loader import carregar_2023, carregar_2024, carregar_2025
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, JsCode, ColumnsAutoSizeMode

st.set_page_config(
    page_title="H2H · JAUM DTC",
    page_icon="⚔️",
    layout="wide",
    initial_sidebar_state="expanded",
)

aplicar_tema()

st.markdown("""
<style>
.jaum-header img { height: 200px !important; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# Constantes
# ─────────────────────────────────────────────────────────────────────────────
STATUS_P1      = ["STINE", "DP2", "LINHAGEM"]
EMPATE_MARGEM  = 1.0   # ± sc/ha

COR_VITORIA = "#27AE60"
COR_EMPATE  = "#FFFF00"
COR_DERROTA = "#FF0000"


def classificar_h2h(pct: float) -> tuple[str, str]:
    """Retorna (label, cor_fundo_hex) pelo % de vitórias."""
    if pd.isna(pct):
        return "—", "#F3F4F6"
    if pct <= 45:
        return "Restrito",        "#FF0000"
    elif pct <= 55:
        return "Competitivo",     "#FFFF00"
    elif pct <= 75:
        return "Superior",        "#87CEFF"
    else:
        return "Alta Performance","#90EE90"

# Texto por classe (Restrito = branco, demais = preto)
COR_TEXTO_CLASSE = {
    "Alta Performance": "#1A1A1A",
    "Superior":         "#1A1A1A",
    "Competitivo":      "#1A1A1A",
    "Restrito":         "#FFFFFF",
    "—":                "#6B7280",
}


# ─────────────────────────────────────────────────────────────────────────────
# Helper Excel
# ─────────────────────────────────────────────────────────────────────────────
def to_excel(df: pd.DataFrame) -> bytes:
    buf  = io.BytesIO()
    wb   = openpyxl.Workbook()
    ws   = wb.active
    df   = df.reset_index(drop=True)
    df   = df.loc[:, ~df.columns.str.startswith("_") & ~df.columns.str.startswith("::") & (df.columns != "_selectedRowNodeInfo")].copy()
    thin = Side(style="thin", color="CCCCCC")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(row=1, column=ci, value=str(col))
        c.font      = Font(bold=True, name="Arial", size=10)
        c.fill      = PatternFill("solid", start_color="F2F2F2")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = brd
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(12, len(str(col)) + 2)
    ws.row_dimensions[1].height = 28

    for ri, row_data in enumerate(df.itertuples(index=False), start=2):
        for ci, val in enumerate(row_data, 1):
            try:
                val = None if (val is None or (isinstance(val, float) and np.isnan(val))) else val
            except (TypeError, ValueError):
                pass
            c = ws.cell(row=ri, column=ci, value=val)
            c.font      = Font(name="Arial", size=10)
            c.alignment = Alignment(horizontal="left" if ci == 1 else "center", vertical="center")
            c.border    = brd

    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# Helper AgGrid com coluna Classe colorida
# ─────────────────────────────────────────────────────────────────────────────
def ag_table_h2h(df: pd.DataFrame, height: int = 480):
    """AgGrid padrão JAUM com cellStyle colorido na coluna Classe."""

    classe_style = JsCode("""
    function(params) {
        const v = params.value;
        if (v === 'Alta Performance') return {
            'backgroundColor': '#90EE90', 'color': '#1A1A1A',
            'fontWeight': '700', 'textAlign': 'center'
        };
        if (v === 'Superior') return {
            'backgroundColor': '#87CEFF', 'color': '#1A1A1A',
            'fontWeight': '700', 'textAlign': 'center'
        };
        if (v === 'Competitivo') return {
            'backgroundColor': '#FFFF00', 'color': '#1A1A1A',
            'fontWeight': '700', 'textAlign': 'center'
        };
        if (v === 'Restrito') return {
            'backgroundColor': '#FF0000', 'color': '#FFFFFF',
            'fontWeight': '700', 'textAlign': 'center'
        };
        return {};
    }
    """)

    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_default_column(
        resizable=True, sortable=True, filter=True,
        suppressMenu=False,
        menuTabs=["generalMenuTab", "filterMenuTab", "columnsMenuTab"],
        cellStyle={"fontSize": "13px", "color": "#000000",
                   "fontFamily": "Helvetica Neue, sans-serif"},
    )
    if "Classe" in df.columns:
        gb.configure_column("Classe", cellStyle=classe_style, minWidth=140)
    gb.configure_grid_options(
        headerHeight=36, rowHeight=32, domLayout="normal",
        suppressMenuHide=True, suppressColumnVirtualisation=True,
        suppressContextMenu=False, enableRangeSelection=True,
    )
    go = gb.build()
    go["defaultColDef"]["headerClass"] = "ag-header-black"
    go["onFirstDataRendered"] = JsCode(
        "function(params) { params.api.sizeColumnsToFit(); }"
    )
    AgGrid(
        df, gridOptions=go, height=height,
        update_mode=GridUpdateMode.NO_UPDATE,
        fit_columns_on_grid_load=False,
        columns_auto_size_mode=2,
        allow_unsafe_jscode=True,
        enable_enterprise_modules=True,
        custom_css={
            ".ag-header":                       {"background-color": "#4A4A4A !important"},
            ".ag-header-row":                   {"background-color": "#4A4A4A !important"},
            ".ag-header-cell":                  {"background-color": "#4A4A4A !important"},
            ".ag-header-cell-label":            {"color": "#FFFFFF !important", "font-weight": "700"},
            ".ag-header-cell-text":             {"color": "#FFFFFF !important", "font-size": "13px !important", "font-weight": "700 !important"},
            ".ag-icon":                         {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-header-icon":                  {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-header-cell-menu-button":      {"opacity": "1 !important", "visibility": "visible !important"},
            ".ag-header-cell-menu-button span": {"color": "#FFFFFF !important"},
            ".ag-icon-menu":                    {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-icon-filter":                  {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-cell":                         {"font-size": "13px !important"},
            ".ag-row":                          {"font-size": "13px !important"},
        },
        theme="streamlit",
        use_container_width=True,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Carregamento de dados (cacheado)
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def carregar_concat() -> pd.DataFrame:
    frames = []
    for loader in [carregar_2023, carregar_2024, carregar_2025]:
        d = loader()
        if d.get("ok") and d.get("ta_faixa") is not None:
            frames.append(d["ta_faixa"])
    if not frames:
        return pd.DataFrame()
    df = pd.concat(frames, ignore_index=True)
    for col in ["sc_ha", "kg_ha"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


with st.spinner("Carregando dados..."):
    ta_raw = carregar_concat()

if ta_raw.empty:
    st.error("❌ Nenhum dado disponível. Verifique a página de Diagnóstico.")
    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
# Header
# ─────────────────────────────────────────────────────────────────────────────
page_header(
    "Head-to-Head",
    "Compare cultivares diretamente nos locais em que ambos foram avaliados simultaneamente.",
    imagem="Data analysis-pana.png",
)

# ─────────────────────────────────────────────────────────────────────────────
# Sidebar — filtros encadeados (padrão JAUM · expanders + checkboxes)
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(
        '<p style="font-size:11px;font-weight:600;color:#6B7280;text-transform:uppercase;'
        'letter-spacing:0.05em;padding: 0.5rem;">Filtros</p>',
        unsafe_allow_html=True,
    )

    if st.button("🔄 Limpar filtros", use_container_width=True):
        for key in list(st.session_state.keys()):
            if any(key.startswith(p) for p in [
                "h2h_safra_", "h2h_macro_", "h2h_micro_", "h2h_estado_",
                "h2h_cidade_", "h2h_fazenda_", "h2h_resp_", "h2h_p2status_",
            ]):
                del st.session_state[key]
        st.rerun()

    def checkboxes_h2h(label, opcoes, default_all=True, defaults=None, prefix=""):
        sel = []
        for o in opcoes:
            checked = (o in defaults) if defaults is not None else default_all
            if st.checkbox(str(o), value=checked, key=f"{prefix}_{o}"):
                sel.append(o)
        return sel

    # ── 1. Safra — padrão só 2025/26 ──────────────────────────────────────────
    with st.expander("📅 Safra", expanded=True):
        safras_all     = sorted(ta_raw["safra"].dropna().unique().tolist())
        safra_default  = [s for s in safras_all if "2025" in str(s)] or safras_all[-1:]
        safras_sel     = checkboxes_h2h("Safra", safras_all, defaults=safra_default, prefix="h2h_safra")

    ta_f1 = ta_raw[ta_raw["safra"].isin(safras_sel)] if safras_sel else ta_raw.iloc[0:0]

    # ── 2. Região Macro ────────────────────────────────────────────────────────
    with st.expander("🗺️ Região Macro", expanded=False):
        macros_all = sorted(ta_f1["regiao_macro"].dropna().unique().tolist())
        macros_sel = checkboxes_h2h("Macro", macros_all, prefix="h2h_macro")

    ta_f2 = ta_f1[ta_f1["regiao_macro"].isin(macros_sel)] if macros_sel else ta_f1.iloc[0:0]

    # ── 3. Região Micro ────────────────────────────────────────────────────────
    with st.expander("📍 Região Micro", expanded=False):
        micros_all = sorted(ta_f2["regiao_micro"].dropna().unique().tolist())
        micros_sel = checkboxes_h2h("Micro", micros_all, prefix="h2h_micro")

    ta_f3 = ta_f2[ta_f2["regiao_micro"].isin(micros_sel)] if micros_sel else ta_f2.iloc[0:0]

    # ── 4. Estado ──────────────────────────────────────────────────────────────
    with st.expander("🏛️ Estado", expanded=False):
        estados_all = sorted(ta_f3["estado_sigla"].dropna().unique().tolist())
        estados_sel = checkboxes_h2h("Estado", estados_all, prefix="h2h_estado")

    ta_f4 = ta_f3[ta_f3["estado_sigla"].isin(estados_sel)] if estados_sel else ta_f3.iloc[0:0]

    # ── 5. Cidade ──────────────────────────────────────────────────────────────
    with st.expander("🏙️ Cidade", expanded=False):
        cidades_all = sorted(ta_f4["cidade_nome"].dropna().unique().tolist())
        cidades_sel = checkboxes_h2h("Cidade", cidades_all, prefix="h2h_cidade")

    ta_f5 = ta_f4[ta_f4["cidade_nome"].isin(cidades_sel)] if cidades_sel else ta_f4.iloc[0:0]

    # ── 6. Fazenda ─────────────────────────────────────────────────────────────
    with st.expander("🏡 Fazenda", expanded=False):
        fazendas_all = sorted(ta_f5["nomeFazenda"].dropna().unique().tolist())
        fazendas_sel = checkboxes_h2h("Fazenda", fazendas_all, prefix="h2h_fazenda")

    ta_f6 = ta_f5[ta_f5["nomeFazenda"].isin(fazendas_sel)] if fazendas_sel else ta_f5.iloc[0:0]

    # ── 7. Responsável ─────────────────────────────────────────────────────────
    with st.expander("👤 Responsável", expanded=False):
        resps_all = sorted(ta_f6["nomeResponsavel"].dropna().unique().tolist())
        resps_sel = checkboxes_h2h("Resp", resps_all, prefix="h2h_resp")

    ta_filtrado = ta_f6[ta_f6["nomeResponsavel"].isin(resps_sel)] if resps_sel else ta_f6.iloc[0:0]

    # ── 8. Status do adversário (Produto 2) ────────────────────────────────────
    with st.expander("🏷️ Status do Adversário (Prod. 2)", expanded=True):
        status_p2_all     = sorted(ta_filtrado["status_material"].dropna().unique().tolist())
        status_p2_default = [s for s in ["CHECK"] if s in status_p2_all]
        status_p2_sel     = checkboxes_h2h(
            "p2status", status_p2_all,
            defaults=status_p2_default,
            prefix="h2h_p2status",
        )

# Aviso se sem dados
if ta_filtrado.empty:
    st.warning("⚠️ Nenhum dado para os filtros selecionados.")
    st.stop()

# Pools separados por papel
df_p1 = ta_filtrado[ta_filtrado["status_material"].isin(STATUS_P1)].copy()
df_p2 = ta_filtrado[ta_filtrado["status_material"].isin(status_p2_sel)].copy() if status_p2_sel else pd.DataFrame()


# ─────────────────────────────────────────────────────────────────────────────
# Função de cruzamento (cacheada — o "join pesado")
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def cruzar_por_local(df1: pd.DataFrame, df2: pd.DataFrame) -> pd.DataFrame:
    """
    Cross-join de df1 × df2 por cod_fazenda.
    Retorna uma linha por (dePara_1, dePara_2, cod_fazenda) com sc_ha e kg_ha de cada um.
    """
    cols_base = ["dePara", "status_material", "cod_fazenda", "sc_ha", "kg_ha"]
    d1 = df1[[c for c in cols_base if c in df1.columns]].dropna(subset=["sc_ha"]).copy()
    d2 = df2[[c for c in cols_base if c in df2.columns]].dropna(subset=["sc_ha"]).copy()
    merged = d1.merge(d2, on="cod_fazenda", suffixes=("_1", "_2"))
    return merged[merged["dePara_1"] != merged["dePara_2"]].reset_index(drop=True)


# ─────────────────────────────────────────────────────────────────────────────
# Tabs
# ─────────────────────────────────────────────────────────────────────────────
tab1, tab2 = st.tabs(["📋  Tabela de Classificação", "📊  Análise por Local"])


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 1 — Tabela de Classificação
# ═══════════════════════════════════════════════════════════════════════════════
with tab1:
    secao_titulo(
        "HEAD-TO-HEAD · TABELA",
        "Classificação vs Adversários",
        "Escolha o Produto 1 (STINE / DP2 / LINHAGEM) e veja como ele se comporta "
        "contra cada adversário nos locais em que ambos foram avaliados.",
    )

    if df_p1.empty:
        st.warning("⚠️ Nenhum cultivar STINE/DP2/LINHAGEM encontrado com os filtros atuais.")
        st.stop()

    if df_p2.empty:
        st.warning("⚠️ Nenhum adversário disponível. Verifique o filtro 'Status (Produto 2)' no sidebar.")
        st.stop()

    cultivares_p1_t1 = sorted(df_p1["dePara"].dropna().unique())
    col_sel, col_btn = st.columns([4, 1])
    with col_sel:
        p1_t1 = st.selectbox(
            "Produto 1 (STINE / DP2 / LINHAGEM)",
            cultivares_p1_t1,
            key="p1_t1",
        )
    with col_btn:
        st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
        btn_t1 = st.button("▶ Rodar Análise", type="primary", key="btn_t1", use_container_width=True)

    key_t1 = f"res_t1__{p1_t1}"

    if btn_t1:
        with st.spinner("Calculando confrontos..."):
            df_cross = cruzar_por_local(
                df_p1[df_p1["dePara"] == p1_t1],
                df_p2,
            )
            rows = []
            for prod2, grp in df_cross.groupby("dePara_2"):
                n    = len(grp)
                diff = grp["sc_ha_1"] - grp["sc_ha_2"]
                vit  = int((diff > EMPATE_MARGEM).sum())
                emp  = int((diff.abs() <= EMPATE_MARGEM).sum())
                pct  = round(vit / n * 100, 1) if n > 0 else np.nan
                sc1  = grp["sc_ha_1"].mean()
                sc2  = grp["sc_ha_2"].mean()
                kg1  = grp["kg_ha_1"].mean() if "kg_ha_1" in grp.columns else np.nan
                kg2  = grp["kg_ha_2"].mean() if "kg_ha_2" in grp.columns else np.nan
                dif_kg  = (kg1 - kg2) if not (np.isnan(kg1) or np.isnan(kg2)) else np.nan
                dif_pct = ((sc1 / sc2) - 1) * 100 if (sc2 != 0 and not np.isnan(sc2)) else np.nan
                classe, cor = classificar_h2h(pct)
                dif_sc  = (sc1 - sc2) if not (np.isnan(sc1) or np.isnan(sc2)) else np.nan
                rows.append({
                    "Produto 1":       p1_t1,
                    "Kg/ha Prod 1":    round(kg1, 0) if not np.isnan(kg1) else None,
                    "SCs/ha Prod 1":   round(sc1, 1),
                    "Produto 2":       prod2,
                    "Kg/ha Prod 2":    round(kg2, 0) if not np.isnan(kg2) else None,
                    "SCs/ha Prod 2":   round(sc2, 1),
                    "Qtd. Vitórias":   vit,
                    "N° Comparações":  n,
                    "Dif. %":          round(dif_pct, 1) if not np.isnan(dif_pct) else None,
                    "Dif. (KG)":       round(dif_kg,  0) if not np.isnan(dif_kg)  else None,
                    "Dif. (SC)":       round(dif_sc,  1) if not np.isnan(dif_sc)  else None,
                    "% Vitórias":      pct,
                    "Classe":          classe,
                    "_cor":            cor,
                })

            df_res = pd.DataFrame(rows)
            if not df_res.empty:
                df_res = df_res.sort_values("% Vitórias", ascending=False)
            df_res = df_res.reset_index(drop=True)
            st.session_state[key_t1] = df_res

    if key_t1 in st.session_state:
        df_res = st.session_state[key_t1]

        if df_res.empty:
            st.info("Nenhum confronto encontrado — o cultivar não compartilha locais com os adversários selecionados.")
        else:
            # ── Título dinâmico + subtítulo de filtros ──────────────────────
            _all_safras = sorted(ta_raw["safra"].dropna().unique().tolist())
            _all_macros = sorted(ta_raw["regiao_macro"].dropna().unique().tolist())
            _all_micros = sorted(ta_raw["regiao_micro"].dropna().unique().tolist())

            filtros_ativos = []
            if safras_sel and set(safras_sel) != set(_all_safras):
                filtros_ativos.append(" / ".join(str(s) for s in safras_sel))
            if macros_sel and set(macros_sel) != set(_all_macros):
                filtros_ativos.append("Macro: " + ", ".join(macros_sel))
            if micros_sel and set(micros_sel) != set(_all_micros):
                filtros_ativos.append("Micro: " + ", ".join(micros_sel))
            if estados_sel and set(estados_sel) != set(sorted(ta_raw["estado_sigla"].dropna().unique().tolist())):
                filtros_ativos.append(", ".join(estados_sel))

            n_locais_ctx  = ta_filtrado["cod_fazenda"].nunique()
            n_cidades_ctx = ta_filtrado["cidade_nome"].nunique()
            filtros_ativos.append(f"{n_cidades_ctx} cidades · {n_locais_ctx} locais")
            contexto_str = "  ·  ".join(filtros_ativos) if filtros_ativos else "Todos os dados"

            COR_STATUS_TITULO = {
                "CHECK":    "#F4B184",
                "STINE":    "#2976B6",
                "DP2":      "#C4DFB4",
                "LINHAGEM": "#00FF01",
            }

            status_p1_label = df_p1[df_p1["dePara"] == p1_t1]["status_material"].iloc[0] \
                if not df_p1[df_p1["dePara"] == p1_t1].empty else ""
            cor_p1 = COR_STATUS_TITULO.get(status_p1_label, "#27AE60")
            n_adv = len(df_res)

            st.markdown(
                f'<div style="margin:0.5rem 0 0.2rem;">'
                f'<p style="font-size:13px;font-weight:600;color:#6B7280;text-transform:uppercase;'
                f'letter-spacing:0.05em;margin:0 0 4px;">Análise H2H · Produto 1</p>'
                f'<h2 style="font-size:1.9rem;font-weight:700;color:#1A1A1A;margin:0;line-height:1.2;">'
                f'<span style="color:#27AE60;">{p1_t1}</span>'
                f'<span style="font-size:1rem;font-weight:500;color:#6B7280;margin-left:10px;">'
                f'{status_p1_label} · {n_adv} adversários</span>'
                f'</h2>'
                f'<p style="font-size:14px;color:#6B7280;margin:4px 0 0;">{contexto_str}</p>'
                f'</div>',
                unsafe_allow_html=True,
            )

            st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)

            # ── Popover — como interpretar ──────────────────────────────────
            with st.popover("ℹ️ Como interpretar esta tabela", use_container_width=False):
                st.markdown("""
**📌 O que esta tabela mostra**

Cada linha é um confronto direto entre o **Produto 1** selecionado e um adversário (Produto 2), calculado exclusivamente nos locais onde **ambos foram avaliados simultaneamente**.

---

**📐 Como ler as colunas**

- **Kg/ha · Scs/ha Prod 1 / Prod 2** → médias de produtividade *apenas nos locais compartilhados* (não é a média geral do cultivar)
- **Qtd. Vitórias** → número de locais em que Prod 1 superou Prod 2 por mais de 1 sc/ha
- **N° Comparações** → total de locais com ambos avaliados
- **% Vitórias** → Vitórias ÷ Comparações × 100 — base da classificação
- **Dif. %** → média de (sc1 / sc2 − 1) × 100 por local — quanto Prod 1 produz a mais ou a menos em termos relativos
- **Dif. (Kg)** → média da diferença absoluta em kg/ha por local

> ⚠️ **Empate**: diferença de até **±1 sc/ha** não é contabilizada como vitória nem derrota.

---

**🎨 Legenda das cores — % de Vitórias**

""")
                col_a, col_b, col_c, col_d = st.columns(4)
                col_a.markdown(
                    '<div style="background:#90EE90;border-radius:6px;padding:8px;text-align:center;">'
                    '<b style="color:#1A1A1A;">Alta Performance</b><br>'
                    '<span style="font-size:12px;color:#1A1A1A;">&gt; 75% de vitórias</span></div>',
                    unsafe_allow_html=True,
                )
                col_b.markdown(
                    '<div style="background:#87CEFF;border-radius:6px;padding:8px;text-align:center;">'
                    '<b style="color:#1A1A1A;">Superior</b><br>'
                    '<span style="font-size:12px;color:#1A1A1A;">56 – 75% de vitórias</span></div>',
                    unsafe_allow_html=True,
                )
                col_c.markdown(
                    '<div style="background:#FFFF00;border-radius:6px;padding:8px;text-align:center;">'
                    '<b style="color:#1A1A1A;">Competitivo</b><br>'
                    '<span style="font-size:12px;color:#1A1A1A;">46 – 55% de vitórias</span></div>',
                    unsafe_allow_html=True,
                )
                col_d.markdown(
                    '<div style="background:#FF0000;border-radius:6px;padding:8px;text-align:center;">'
                    '<b style="color:#FFFFFF;">Restrito</b><br>'
                    '<span style="font-size:12px;color:#FFFFFF;">≤ 45% de vitórias</span></div>',
                    unsafe_allow_html=True,
                )
                st.markdown("""
---

**💡 Como interpretar**

- **Alta Performance** → Prod 1 vence em mais de 3/4 dos locais — material consistentemente superior ao adversário neste ambiente
- **Superior** → vence na maioria dos locais — boa performance geral
- **Competitivo** → resultado equilibrado — nenhum material se destaca claramente
- **Restrito** → Prod 1 perde na maioria dos locais frente a este adversário — atenção ao posicionamento
""")

            st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
            # ── Mini-resumo por classe ──────────────────────────────────────
            contagem  = df_res["Classe"].value_counts()
            total_cls = len(df_res)
            c1, c2, c3, c4 = st.columns(4)
            for col_ui, label, cor_txt in zip(
                [c1, c2, c3, c4],
                ["Alta Performance", "Superior", "Competitivo", "Restrito"],
                ["#27AE60",           "#1E40AF",  "#F2C811",    "#FF0000"],
            ):
                n_cls  = int(contagem.get(label, 0))
                pct_cl = f"{n_cls / total_cls * 100:.0f}%" if total_cls > 0 else "—"
                col_ui.markdown(
                    f'<div style="border:1px solid #E5E7EB;border-radius:10px;'
                    f'padding:10px 14px;background:#FFFFFF;text-align:center;'
                    f'box-shadow:0 1px 4px rgba(0,0,0,0.07);">'
                    f'<p style="margin:0;font-size:14px;font-weight:600;color:#374151;">{label}</p>'
                    f'<p style="margin:4px 0 0;font-size:2.2rem;font-weight:700;color:{cor_txt};">'
                    f'{n_cls} <span style="font-size:1.2rem;font-weight:500;">({pct_cl})</span></p>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
            st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)

            # ── Tabela AgGrid ───────────────────────────────────────────────
            df_show = df_res.drop(columns=["_cor"])
            ag_table_h2h(df_show, height=min(680, int((36 + 32 * len(df_show) + 20) * 1.3)))

            st.download_button(
                "⬇️ Exportar Excel",
                data=to_excel(df_show),
                file_name=f"h2h_{p1_t1}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_t1",
            )

    else:
        st.info("👆 Selecione o Produto 1 e clique em **▶ Rodar Análise** para calcular.")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 2 — Análise por Local (gráfico)
# ═══════════════════════════════════════════════════════════════════════════════
with tab2:
    secao_titulo(
        "HEAD-TO-HEAD · GRÁFICO",
        "Diferença de Produtividade por Local",
        "Selecione um par específico e veja a diferença de sc/ha em cada local compartilhado.",
    )

    if df_p1.empty or df_p2.empty:
        st.warning("⚠️ Dados insuficientes com os filtros atuais.")
        st.stop()

    cults_p1_t2 = sorted(df_p1["dePara"].dropna().unique())
    col_p1, col_p2, col_b2 = st.columns([2, 2, 1])

    with col_p1:
        p1_t2 = st.selectbox("Produto 1 (STINE / DP2 / LINHAGEM)", cults_p1_t2, key="p1_t2")

    # Restringe Produto 2 aos adversários que têm ao menos 1 local em comum com p1_t2
    locais_p1_t2 = set(df_p1[df_p1["dePara"] == p1_t2]["cod_fazenda"].dropna().unique())
    adv_disp = sorted(
        df_p2[df_p2["cod_fazenda"].isin(locais_p1_t2)]["dePara"].dropna().unique()
    )

    with col_p2:
        if adv_disp:
            p2_t2 = st.selectbox("Produto 2 (adversário)", adv_disp, key="p2_t2")
        else:
            st.warning("Nenhum adversário com locais em comum para este cultivar.")
            p2_t2 = None

    with col_b2:
        st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
        btn_t2 = st.button("▶ Rodar Análise", type="primary", key="btn_t2", use_container_width=True)

    key_t2 = f"res_t2__{p1_t2}__{p2_t2}"

    if btn_t2 and p2_t2:
        with st.spinner("Calculando..."):
            d1_loc = (
                df_p1[df_p1["dePara"] == p1_t2][["cod_fazenda", "sc_ha", "kg_ha"]]
                .dropna(subset=["sc_ha"])
            )
            d2_loc = (
                df_p2[df_p2["dePara"] == p2_t2][["cod_fazenda", "sc_ha", "kg_ha"]]
                .dropna(subset=["sc_ha"])
            )
            df_loc = d1_loc.merge(d2_loc, on="cod_fazenda", suffixes=("_1", "_2")).copy()
            df_loc["diff_sc"] = df_loc["sc_ha_1"] - df_loc["sc_ha_2"]
            df_loc["resultado"] = df_loc["diff_sc"].apply(
                lambda x: "Vitória" if x > EMPATE_MARGEM
                else ("Empate" if abs(x) <= EMPATE_MARGEM else "Derrota")
            )
            df_loc = df_loc.sort_values("diff_sc").reset_index(drop=True)
            st.session_state[key_t2] = df_loc

    if key_t2 in st.session_state and p2_t2:
        df_loc = st.session_state[key_t2]

        if df_loc.empty:
            st.info("Nenhum local compartilhado encontrado para este par.")
        else:
            n_loc   = len(df_loc)
            n_vit   = int((df_loc["resultado"] == "Vitória").sum())
            n_emp   = int((df_loc["resultado"] == "Empate").sum())
            n_der   = int((df_loc["resultado"] == "Derrota").sum())
            vit_sc  = df_loc.loc[df_loc["resultado"] == "Vitória", "diff_sc"]
            der_sc  = df_loc.loc[df_loc["resultado"] == "Derrota", "diff_sc"]
            max_vit = float(vit_sc.max()) if len(vit_sc) > 0 else np.nan
            med_vit = float(vit_sc.mean()) if len(vit_sc) > 0 else np.nan
            min_der = float(der_sc.min()) if len(der_sc) > 0 else np.nan
            med_der = float(der_sc.mean()) if len(der_sc) > 0 else np.nan

            # ── Título dinâmico + subtítulo de filtros ──────────────────────
            _all_safras = sorted(ta_raw["safra"].dropna().unique().tolist())
            _all_macros = sorted(ta_raw["regiao_macro"].dropna().unique().tolist())
            _all_micros = sorted(ta_raw["regiao_micro"].dropna().unique().tolist())

            filtros_ativos_t2 = []
            if safras_sel and set(safras_sel) != set(_all_safras):
                filtros_ativos_t2.append(" / ".join(str(s) for s in safras_sel))
            if macros_sel and set(macros_sel) != set(_all_macros):
                filtros_ativos_t2.append("Macro: " + ", ".join(macros_sel))
            if micros_sel and set(micros_sel) != set(_all_micros):
                filtros_ativos_t2.append("Micro: " + ", ".join(micros_sel))
            if estados_sel and set(estados_sel) != set(sorted(ta_raw["estado_sigla"].dropna().unique().tolist())):
                filtros_ativos_t2.append(", ".join(estados_sel))

            filtros_ativos_t2.append(f"{n_loc} locais compartilhados")
            contexto_str_t2 = "  ·  ".join(filtros_ativos_t2) if filtros_ativos_t2 else "Todos os dados"

            status_p2_label = df_p2[df_p2["dePara"] == p2_t2]["status_material"].iloc[0] \
                if not df_p2[df_p2["dePara"] == p2_t2].empty else ""
            status_p1_t2_label = df_p1[df_p1["dePara"] == p1_t2]["status_material"].iloc[0] \
                if not df_p1[df_p1["dePara"] == p1_t2].empty else ""

            st.markdown(
                f'<div style="margin:0.5rem 0 0.2rem;">'
                f'<p style="font-size:13px;font-weight:600;color:#6B7280;text-transform:uppercase;'
                f'letter-spacing:0.05em;margin:0 0 4px;">Análise H2H · Confronto Direto</p>'
                f'<h2 style="font-size:1.9rem;font-weight:700;color:#1A1A1A;margin:0;line-height:1.2;">'
                f'<span style="color:#27AE60;">{p1_t2}</span>'
                f'<span style="font-size:0.85rem;font-weight:500;color:#6B7280;margin-left:6px;">{status_p1_t2_label}</span>'
                f'<span style="font-size:1.1rem;font-weight:500;color:#6B7280;margin:0 12px;">vs</span>'
                f'<span style="color:#1A1A1A;">{p2_t2}</span>'
                f'<span style="font-size:0.85rem;font-weight:500;color:#6B7280;margin-left:6px;">{status_p2_label}</span>'
                f'</h2>'
                f'<p style="font-size:14px;color:#6B7280;margin:4px 0 0;">{contexto_str_t2}</p>'
                f'</div>',
                unsafe_allow_html=True,
            )

            st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

            # ── Popovers — interpretação + dicionário de locais ─────────────
            col_pop1, col_pop2, _ = st.columns([2, 2, 4])

            with col_pop1:
                with st.popover("ℹ️ Como interpretar", use_container_width=True):
                    st.markdown("""
**📌 O que este painel mostra**

Confronto direto entre **Produto 1** e **Produto 2** nos locais onde **ambos foram avaliados simultaneamente** na mesma safra.

---

**📐 Como ler os cards**

- **Locais avaliados** → total de locais com ambos os cultivares presentes
- **Vitórias** → locais em que Prod 1 superou Prod 2 por mais de **+1 sc/ha**
  - *Max*: maior diferença positiva em sc/ha
  - *Média*: média das diferenças nos locais de vitória
- **Empates** → diferença entre **−1 e +1 sc/ha** — resultado estatisticamente indiferente
- **Derrotas** → locais em que Prod 2 superou Prod 1 por mais de 1 sc/ha
  - *Min*: maior diferença negativa (pior local)
  - *Média*: média das diferenças nos locais de derrota

---

**🎨 Cores do gráfico de barras**
""")
                    col_a, col_b, col_c = st.columns(3)
                    col_a.markdown(
                        f'<div style="background:{COR_VITORIA};border-radius:6px;padding:8px;text-align:center;">'
                        f'<b style="color:#FFFFFF;">Vitória</b><br>'
                        f'<span style="font-size:11px;color:#FFFFFF;">&gt; +1 sc/ha</span></div>',
                        unsafe_allow_html=True,
                    )
                    col_b.markdown(
                        f'<div style="background:{COR_EMPATE};border-radius:6px;padding:8px;text-align:center;">'
                        f'<b style="color:#1A1A1A;">Empate</b><br>'
                        f'<span style="font-size:11px;color:#1A1A1A;">±1 sc/ha</span></div>',
                        unsafe_allow_html=True,
                    )
                    col_c.markdown(
                        f'<div style="background:{COR_DERROTA};border-radius:6px;padding:8px;text-align:center;">'
                        f'<b style="color:#FFFFFF;">Derrota</b><br>'
                        f'<span style="font-size:11px;color:#FFFFFF;">&lt; −1 sc/ha</span></div>',
                        unsafe_allow_html=True,
                    )
                    st.markdown("""
---

**💡 Como interpretar o gráfico**

- Barras para a **direita** (verde) → Prod 1 venceu naquele local
- Barras para a **esquerda** (vermelho) → Prod 2 venceu naquele local
- Barra **amarela** → resultado empatado (dentro da margem de ±1 sc/ha)
- O valor em cada barra mostra a **diferença em sc/ha** naquele local
- Locais ordenados do **pior para o melhor** resultado do Prod 1 (de baixo para cima)
""")

            with col_pop2:
                with st.popover("📍 Dicionário de Locais", use_container_width=True):
                    df_dic = (
                        ta_filtrado[
                            ta_filtrado["cod_fazenda"].isin(df_loc["cod_fazenda"])
                        ][["cod_fazenda", "nomeFazenda", "cidade_nome", "estado_sigla"]]
                        .drop_duplicates()
                        .sort_values(["estado_sigla", "cidade_nome", "cod_fazenda"])
                        .rename(columns={
                            "cod_fazenda":  "Código",
                            "nomeFazenda":  "Local",
                            "cidade_nome":  "Cidade",
                            "estado_sigla": "Estado",
                        })
                        .reset_index(drop=True)
                    )
                    st.markdown(
                        f"Referência dos **{len(df_dic)} locais** exibidos no gráfico de barras."
                    )
                    st.dataframe(df_dic, hide_index=True, use_container_width=True)

            st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

            # ── Cards compactos ─────────────────────────────────────────────
            pct_vit = f"{n_vit / n_loc * 100:.0f}%" if n_loc > 0 else "—"
            pct_emp = f"{n_emp / n_loc * 100:.0f}%" if n_loc > 0 else "—"
            pct_der = f"{n_der / n_loc * 100:.0f}%" if n_loc > 0 else "—"

            COR_EMPATE_CARD = "#D4A800"  # amarelo escuro — legível no fundo branco

            card = (
                "border:1px solid #E5E7EB;border-radius:10px;"
                "padding:12px 16px;background:#FFFFFF;text-align:center;"
                "box-shadow:0 1px 4px rgba(0,0,0,0.07);"
            )
            c1, c2, c3, c4 = st.columns(4)

            with c1:
                st.markdown(
                    f'<div style="{card}">'
                    f'<p style="margin:0;font-size:12px;color:#6B7280;">📍 Locais avaliados</p>'
                    f'<p style="margin:6px 0 0;font-size:1.9rem;font-weight:700;color:#1A1A1A;">{n_loc}</p>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

            with c2:
                sub = (
                    f'<p style="margin:2px 0;font-size:14px;font-weight:600;color:{COR_VITORIA};">Max: {max_vit:+.1f} sc/ha</p>'
                    f'<p style="margin:0;font-size:14px;font-weight:600;color:{COR_VITORIA};">Média: {med_vit:+.1f} sc/ha</p>'
                ) if not np.isnan(max_vit) else '<p style="margin:2px 0;font-size:14px;">&nbsp;</p><p style="margin:0;font-size:14px;">&nbsp;</p>'
                st.markdown(
                    f'<div style="{card}border-top:3px solid {COR_VITORIA};">'
                    f'<p style="margin:0;font-size:15px;font-weight:700;color:#1A1A1A;">✅ Vitórias</p>'
                    f'{sub}'
                    f'<p style="margin:6px 0;font-size:1.9rem;font-weight:700;color:{COR_VITORIA};">'
                    f'{n_vit} <span style="font-size:1rem;font-weight:600;">({pct_vit})</span></p>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

            with c3:
                st.markdown(
                    f'<div style="{card}border-top:3px solid {COR_EMPATE_CARD};">'
                    f'<p style="margin:0;font-size:15px;font-weight:700;color:#1A1A1A;">— Empates</p>'
                    f'<p style="margin:2px 0;font-size:14px;font-weight:600;color:{COR_EMPATE_CARD};">Entre ±1 sc/ha</p>'
                    f'<p style="margin:0;font-size:14px;">&nbsp;</p>'
                    f'<p style="margin:6px 0;font-size:1.9rem;font-weight:700;color:{COR_EMPATE_CARD};">'
                    f'{n_emp} <span style="font-size:1rem;font-weight:600;">({pct_emp})</span></p>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

            with c4:
                sub = (
                    f'<p style="margin:2px 0;font-size:14px;font-weight:600;color:{COR_DERROTA};">Min: {min_der:+.1f} sc/ha</p>'
                    f'<p style="margin:0;font-size:14px;font-weight:600;color:{COR_DERROTA};">Média: {med_der:+.1f} sc/ha</p>'
                ) if not np.isnan(min_der) else '<p style="margin:2px 0;font-size:14px;">&nbsp;</p><p style="margin:0;font-size:14px;">&nbsp;</p>'
                st.markdown(
                    f'<div style="{card}border-top:3px solid {COR_DERROTA};">'
                    f'<p style="margin:0;font-size:15px;font-weight:700;color:#1A1A1A;">✕ Derrotas</p>'
                    f'{sub}'
                    f'<p style="margin:6px 0;font-size:1.9rem;font-weight:700;color:{COR_DERROTA};">'
                    f'{n_der} <span style="font-size:1rem;font-weight:600;">({pct_der})</span></p>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

            st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)

            # ── Donut + Mapa ────────────────────────────────────────────────
            col_donut, col_mapa = st.columns([1, 2])

            with col_donut:
                fig_d = go_plt.Figure(go_plt.Pie(
                    labels=["Vitórias", "Empates", "Derrotas"],
                    values=[n_vit, n_emp, n_der],
                    hole=0.55,
                    marker_colors=[COR_VITORIA, COR_EMPATE, COR_DERROTA],
                    textinfo="label+percent",
                    textposition="outside",
                    textfont=dict(size=12, family="Helvetica Neue, sans-serif", color="#111111"),
                    hovertemplate="%{label}: %{value} local(is) (%{percent})<extra></extra>",
                    sort=False,
                    pull=[0.03, 0.03, 0.03],
                    domain=dict(x=[0.15, 0.85], y=[0.05, 0.90]),
                ))
                fig_d.update_layout(
                    title=dict(
                        text="Resultado Geral do Head",
                        font=dict(size=13, color="#111111"),
                        x=0.5,
                        xanchor="center",
                        y=0.99,
                        yanchor="top",
                    ),
                    showlegend=False,
                    height=420,
                    margin=dict(t=80, b=20, l=60, r=60),
                    paper_bgcolor="#FFFFFF",
                    font=dict(family="Helvetica Neue, sans-serif"),
                )
                st.plotly_chart(fig_d, use_container_width=True)

            with col_mapa:
                # Join com lat/lon da base
                df_coords = (
                    ta_filtrado[["cod_fazenda", "latitude", "longitude", "nomeFazenda", "cidade_nome", "estado_sigla"]]
                    .drop_duplicates("cod_fazenda")
                )
                df_map = df_loc.merge(df_coords, on="cod_fazenda", how="left").dropna(subset=["latitude", "longitude"])

                if not df_map.empty:
                    import folium
                    from streamlit_folium import st_folium

                    COR_MAP = {
                        "Vitória": "#27AE60",
                        "Empate":  "#FFDD00",
                        "Derrota": "#E74C3C",
                    }

                    lat_c = df_map["latitude"].mean()
                    lon_c = df_map["longitude"].mean()

                    m = folium.Map(
                        location=[lat_c, lon_c],
                        zoom_start=5,
                        tiles="OpenStreetMap",
                    )

                    for _, row in df_map.iterrows():
                        cor    = COR_MAP.get(row["resultado"], "#888888")
                        popup  = folium.Popup(
                            f"<b>{row['nomeFazenda']}</b><br>"
                            f"{row['cidade_nome']} — {row['estado_sigla']}<br>"
                            f"<b>{p1_t2}:</b> {row['sc_ha_1']:.1f} sc/ha<br>"
                            f"<b>{p2_t2}:</b> {row['sc_ha_2']:.1f} sc/ha<br>"
                            f"<b>Diferença:</b> {row['diff_sc']:+.1f} sc/ha<br>"
                            f"<b>Resultado:</b> {row['resultado']}",
                            max_width=260,
                        )
                        folium.CircleMarker(
                            location=[row["latitude"], row["longitude"]],
                            radius=9,
                            color="#FFFFFF",
                            weight=1.5,
                            fill=True,
                            fill_color=cor,
                            fill_opacity=0.9,
                            popup=popup,
                            tooltip=f"{row['nomeFazenda']} | {row['cidade_nome']} — {row['estado_sigla']} · {row['resultado']} ({row['diff_sc']:+.1f} sc/ha)",
                        ).add_to(m)

                    st.markdown(
                        '<p style="font-size:13px;font-weight:600;color:#4A4A4A;margin:0 0 6px;">📍 Locais por Resultado</p>',
                        unsafe_allow_html=True,
                    )
                    st_folium(m, use_container_width=True, height=420, returned_objects=[])
                else:
                    st.info("Coordenadas não disponíveis para os locais deste confronto.")

            st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

            # ── Barras (largura total) ──────────────────────────────────────
            cores_bar = df_loc["resultado"].map({
                "Vitória": COR_VITORIA,
                "Empate":  COR_EMPATE,
                "Derrota": COR_DERROTA,
            }).tolist()

            fig_b = go_plt.Figure(go_plt.Bar(
                x=df_loc["diff_sc"].round(1),
                y=df_loc["cod_fazenda"],
                orientation="h",
                marker_color=cores_bar,
                text=df_loc["diff_sc"].round(1),
                textposition="outside",
                textfont=dict(size=11, color="#111111"),
                hovertemplate="<b>%{y}</b><br>Diferença: %{x:+.1f} sc/ha<extra></extra>",
            ))
            fig_b.add_vline(x=0, line_color="#333333", line_width=2)
            fig_b.update_layout(
                title=dict(
                    text=f"Diferença de Produtividade por Local — {p1_t2} × {p2_t2}",
                    font=dict(size=13, color="#111111"),
                ),
                xaxis=dict(
                    title="Diferença (sc/ha)",
                    tickfont=dict(size=11, color="#111111"),
                    zerolinecolor="#CCCCCC",
                ),
                yaxis=dict(
                    title="Local",
                    tickfont=dict(size=11, color="#111111"),
                ),
                height=max(380, n_loc * 28 + 100),
                margin=dict(t=50, b=50, l=130, r=100),
                plot_bgcolor="#FFFFFF",
                paper_bgcolor="#FFFFFF",
                font=dict(family="Helvetica Neue, sans-serif", size=12, color="#111111"),
            )
            st.plotly_chart(fig_b, use_container_width=True)

            # ── Tabela local a local (expansível) ───────────────────────────
            with st.expander("📋 Ver tabela de dados por local"):
                df_exp = df_loc[["cod_fazenda", "sc_ha_1", "sc_ha_2", "diff_sc", "resultado"]].copy()
                df_exp.columns = [
                    "Local",
                    f"Scs/ha — {p1_t2}",
                    f"Scs/ha — {p2_t2}",
                    "Diferença (sc/ha)",
                    "Resultado",
                ]
                st.dataframe(df_exp, hide_index=True, use_container_width=True)
                st.download_button(
                    "⬇️ Exportar Excel",
                    data=to_excel(df_exp),
                    file_name=f"h2h_local_{p1_t2}_vs_{p2_t2}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_t2",
                )

    else:
        st.info("👆 Selecione os dois cultivares e clique em **▶ Rodar Análise** para calcular.")


# ─────────────────────────────────────────────────────────────────────────────
# Footer
# ─────────────────────────────────────────────────────────────────────────────
st.markdown(
    '<p style="font-size:13px;color:#374151;text-align:center;margin-top:2rem;">'
    "Painel JAUM DTC · Stine Seed · "
    'Desenvolvido por <a href="https://www.linkedin.com/in/eng-agro-andre-ferreira/" '
    'target="_blank" style="color:#27AE60;text-decoration:none;">Andre Ferreira</a></p>',
    unsafe_allow_html=True,
)
