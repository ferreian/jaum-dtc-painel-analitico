"""
pages/1_Diagnostico.py — Diagnóstico de Dados
"""
import numpy as np
import pandas as pd
import streamlit as st

from utils.theme import aplicar_tema, page_header, secao_titulo
from utils.loader import carregar_2023, carregar_2024, carregar_2025

st.set_page_config(
    page_title="Diagnóstico · JAUM DTC",
    page_icon="🔄",
    layout="wide",
    initial_sidebar_state="expanded",
)

aplicar_tema()
page_header("Diagnóstico de Dados", "Status de carregamento e checagem de inconsistências.", imagem="App development-bro.png")

st.markdown("""
<style>
.secao-label {
    font-size: 11px; font-weight: 600; color: #6B7280;
    text-transform: uppercase; letter-spacing: 0.05em;
    margin: 1.5rem 0 0.5rem;
}
.resumo-card { background:#fff; border:1px solid #E5E7EB; border-radius:10px; padding:1rem 1.2rem; }
.resumo-card-ok   { border-left: 4px solid #27AE60; }
.resumo-card-warn { border-left: 4px solid #F59E0B; }
</style>
""", unsafe_allow_html=True)

# ── Botão de atualização ──────────────────────────────────────────────────────
col_btn, col_info = st.columns([1, 3])
with col_btn:
    if st.button("▶ Atualizar dados", type="primary", use_container_width=True):
        st.cache_data.clear()
        st.rerun()
with col_info:
    st.info("Os dados de 2025 são atualizados a cada 1 hora. Clique para forçar atualização.", icon="ℹ️")

st.divider()

# ── Carregamento ──────────────────────────────────────────────────────────────
with st.spinner("Carregando dados das três safras..."):
    dados_2023 = carregar_2023()
    dados_2024 = carregar_2024()
    dados_2025 = carregar_2025()

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 1 — STATUS
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo("Dados", "Status de Carregamento", "Verifique se as três safras foram carregadas corretamente.")

col1, col2, col3 = st.columns(3)

with col1:
    if dados_2023["ok"]:
        df23 = dados_2023["ta_faixa"]
        st.success("✅ **Safra 2023/24** — carregado")
    else:
        st.error(f"❌ **Safra 2023/24** — {dados_2023.get('erro', '')}")
        df23 = None

with col2:
    if dados_2024["ok"]:
        ta_24 = dados_2024.get("ta_faixa")
        erros = dados_2024.get("erros", [])
        if erros:
            st.warning("⚠️ **Safra 2024/25** — parcialmente carregado")
            with st.expander(f"{len(erros)} erro(s) no carregamento"):
                for e in erros:
                    st.text(e)
        else:
            st.success("✅ **Safra 2024/25** — carregado")
    else:
        st.error(f"❌ **Safra 2024/25** — {', '.join(dados_2024.get('erros', []))}")
        ta_24 = None

with col3:
    if dados_2025["ok"]:
        ta_25 = dados_2025.get("ta_faixa")
        st.success("✅ **Safra 2025/26** — carregado via Supabase")
    else:
        st.error(f"❌ **Safra 2025/26** — {dados_2025.get('erro', '')}")
        ta_25 = None

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 2 — VISÃO CONSOLIDADA
# ════════════════════════════════════════════════════════════════════════════════
frames = [d.get("ta_faixa") for d in [dados_2023, dados_2024, dados_2025]
          if d.get("ok") and d.get("ta_faixa") is not None]

if not frames:
    st.warning("⚠️ Nenhum dado disponível para consolidar.")
    st.stop()

ta_concat = pd.concat(frames, ignore_index=True)

# ── Filtro por responsável ───────────────────────────────────────────────────
with st.sidebar:
    st.markdown(
        '<p style="font-size:11px;font-weight:600;color:#6B7280;text-transform:uppercase;'
        'letter-spacing:0.05em;padding:0.5rem 0 0.2rem;">Filtrar por Responsável</p>',
        unsafe_allow_html=True,
    )
    # Lista vem SOMENTE dos tabelões de 2025
    _tabeloes_raw_sidebar = dados_2025.get("tabeloes", {}) if dados_2025.get("ok") else {}
    _resps_tabeloes = set()
    for _df_tb in _tabeloes_raw_sidebar.values():
        if "nomeResponsavel" in _df_tb.columns:
            _resps_tabeloes.update(_df_tb["nomeResponsavel"].dropna().unique().tolist())

    if _resps_tabeloes:
        _resps_all = ["Todos"] + sorted(_resps_tabeloes)
        _resp_sel  = st.selectbox("Responsável", options=_resps_all, key="diag_resp")
    else:
        st.caption("Responsáveis disponíveis após carregar dados de 2025.")
        _resp_sel = "Todos"

st.divider()
secao_titulo("Consolidação", "Como estão os dados hoje", "Retrato atual dos ensaios avaliados, fazendas e cultivares nas três safras.")

col_m1, col_m2, col_m3, col_m4 = st.columns(4)
col_m1.metric("Total de ensaios",     f"{len(ta_concat):,}")
col_m2.metric("Safras",               ta_concat["safra"].nunique())
col_m3.metric("Fazendas",             ta_concat["cod_fazenda"].nunique())
col_m4.metric("Cultivares avaliados", ta_concat["nome"].nunique())

diag = (
    ta_concat.groupby("safra", dropna=False)
    .agg(Ensaios=("safra","count"), Fazendas=("cod_fazenda","nunique"),
         Cultivares=("nome","nunique"), Regioes=("regiao_macro","nunique"))
    .reset_index()
    .rename(columns={"safra":"Safra","Regioes":"Regiões Macro"})
)
st.dataframe(diag, use_container_width=True, hide_index=True)

# ════════════════════════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 3 — CHECAGEM DE INCONSISTÊNCIAS POR AVALIAÇÃO (tabelões enriquecidos 2025)
# ════════════════════════════════════════════════════════════════════════════════
if not dados_2025["ok"]:
    st.info("ℹ️ Checagem disponível apenas quando os dados de 2025 estão carregados.")
    st.stop()

tabeloes_raw = dados_2025.get("tabeloes", {})

# Filtro por responsável — nomeResponsavel já vem nos tabelões enriquecidos
if _resp_sel != "Todos":
    tabeloes = {
        k: df[df["nomeResponsavel"] == _resp_sel].copy()
        if "nomeResponsavel" in df.columns else df.copy()
        for k, df in tabeloes_raw.items()
    }
else:
    tabeloes = tabeloes_raw

st.divider()
secao_titulo("Qualidade", "Checagem de Inconsistências por Avaliação",
             "Cada avaliação (Av1–Av7) é verificada de forma independente — nulos, intervalos, duplicatas e consistência.")

st.divider()

def _safe(nome):
    return tabeloes.get(nome, pd.DataFrame())

# Colunas de contexto — já vêm nos tabelões enriquecidos
COL_CTX = [
    "nomeResponsavel", "safra", "cod_fazenda", "nomeFazenda",
    "cidade_nome", "estado_sigla", "regiao_macro",
    "nome", "dePara", "status_material", "tipoTeste", "indexTratamento",
]

def _ctx(df):
    return [c for c in COL_CTX if c in df.columns]

def _show(df):
    """Exibe dataframe garantindo nomeResponsavel como 1ª coluna."""
    if df is None or df.empty:
        return df
    _prioridade = ["nomeResponsavel", "safra", "cod_fazenda", "nomeFazenda",
                   "cidade_nome", "estado_sigla", "nome", "dePara",
                   "status_material", "tipoTeste", "indexTratamento"]
    _ctx_p = [c for c in _prioridade if c in df.columns]
    _resto  = [c for c in df.columns if c not in _ctx_p]
    return df[_ctx_p + _resto]

def _chk_nulos(df, cols):
    """Retorna (rows, dfs, n_alertas) de nulos críticos."""
    rows, dfs, n = [], {}, 0
    for col in cols:
        if col not in df.columns:
            rows.append({"Variável": col, "Nulos": "—", "% Nulos": "—", "Status": "⚪ Coluna ausente"})
            continue
        qtd = int(df[col].isnull().sum())
        pct = round(qtd / max(len(df), 1) * 100, 1)
        status = "🔴 Crítico (>20%)" if pct > 20 else ("⚠️ Atenção" if qtd > 0 else "✅ OK")
        if pct > 20:
            n += 1
            dfs[col] = df.loc[df[col].isnull(), _ctx(df) + [col]]
        rows.append({"Variável": col, "Nulos": qtd, "% Nulos": pct, "Status": status})
    return rows, dfs, n

def _chk_range(df, checks):
    """checks = [(col, lo, hi, unidade), ...]. Retorna (rows, dfs, n_alertas)."""
    rows, dfs, n = [], {}, 0
    for col, lo, hi, unidade in checks:
        if col not in df.columns:
            continue
        s = pd.to_numeric(df[col], errors="coerce")
        mask = s.notna() & ~s.between(lo, hi)
        qtd = int(mask.sum())
        if qtd > 0:
            n += 1
            dfs[col] = df.loc[mask, _ctx(df) + [col]]
        rows.append({"Variável": col, "Intervalo": f"[{lo}, {hi}] {unidade}",
                     "Ocorrências": qtd, "Status": "⚠️" if qtd > 0 else "✅ OK"})
    return rows, dfs, n

def _chk_dups(df, chave):
    """Retorna (n, df_dups)."""
    cols = [c for c in chave if c in df.columns]
    if len(cols) < 2:
        return 0, pd.DataFrame()
    mask = df.duplicated(subset=cols, keep=False)
    n = int(mask.sum())
    return n, df.loc[mask, _ctx(df) + cols].sort_values(cols) if n > 0 else pd.DataFrame()

def _chk_orfaos(df):
    """Registros sem fazenda ou sem cultivar."""
    rows, dfs, n = [], {}, 0
    for campo, desc in [("cod_fazenda", "Sem fazenda"), ("dePara", "Sem cultivar (dePara)")]:
        if campo not in df.columns:
            continue
        mask = df[campo].isnull()
        qtd = int(mask.sum())
        if qtd > 0:
            n += 1
            dfs[desc] = df.loc[mask, _ctx(df)]
        rows.append({"Verificação": desc, "Ocorrências": qtd,
                     "Status": "⚠️" if qtd > 0 else "✅ OK"})
    return rows, dfs, n

def _chk_outlier(df, col, lo, hi, label, dp_col=None, dp_thresh=None):
    """Outlier ±3DP + CV. Retorna (row_dict, df_out, n_alert)."""
    if df.empty or col not in df.columns:
        return None, pd.DataFrame(), 0
    s = df[col].dropna()
    s = s[(s >= lo) & (s <= hi)]
    if len(s) < 3:
        return None, pd.DataFrame(), 0
    media, dp = s.mean(), s.std()
    cv = dp / media * 100 if media > 0 else 0
    mask_out = (df[col].notna() & df[col].between(lo, hi) &
                ((df[col] < media - 3*dp) | (df[col] > media + 3*dp)))
    n_out = int(mask_out.sum())
    n_alert = 1 if (cv > 30 or n_out > 0) else 0
    df_out = df.loc[mask_out, _ctx(df) + [col]] if n_out > 0 else pd.DataFrame()
    row = {"Variável": label, "N": len(s), "Média": round(media, 2), "CV (%)": round(cv, 1),
           "Outliers (±3DP)": n_out,
           "Status": "🔴 CV>30%" if cv > 30 else ("⚠️ CV>15%" if cv > 15 else "✅ OK")}
    return row, df_out, n_alert

_COR_AV = {"Alta": "#E74C3C", "Média": "#F39C12", "Baixa": "#3498DB", "OK": "#27AE60"}
_BG_AV  = {"Alta": "#FDF0EF", "Média": "#FFFBF0", "Baixa": "#EBF5FB", "OK": "#E9F7EF"}

def _render_av(titulo, icone, alertas_total, sev, conteudo_fn):
    """Renderiza expander de uma avaliação."""
    _cor = _COR_AV[sev if alertas_total > 0 else "OK"]
    _bg  = _BG_AV [sev if alertas_total > 0 else "OK"]
    _icon_exp = "🔴" if (alertas_total > 0 and sev == "Alta") else \
                "⚠️" if (alertas_total > 0 and sev == "Média") else \
                "🔵" if (alertas_total > 0 and sev == "Baixa") else "✅"
    with st.expander(
        f"{_icon_exp} **{titulo}** — "
        f"{alertas_total} alerta(s)" if alertas_total > 0 else f"✅ **{titulo}** — sem alertas",
        expanded=(alertas_total > 0 and sev == "Alta"),
    ):
        st.markdown(f"""
<div style="background:{_bg};border-left:4px solid {_cor};border-radius:0 8px 8px 0;
            padding:8px 14px;margin-bottom:12px;">
  <span style="font-size:13px;font-weight:700;color:{_cor};">{_icon_exp} {alertas_total} alerta(s)</span>
  <span style="font-size:11px;color:{_cor};background:white;border:1px solid {_cor};
               border-radius:20px;padding:1px 8px;margin-left:8px;">Severidade {sev if alertas_total > 0 else "OK"}</span>
</div>
""", unsafe_allow_html=True)
        conteudo_fn()

# ────────────────────────────────────────────────────────────────────────────
# Av1 — Qualidade Inicial
# ────────────────────────────────────────────────────────────────────────────
_df1 = _safe("tb_av1")
_r1_nulos, _d1_nulos, _n1_nulos = _chk_nulos(_df1, ["notaMedia","notaUniformidade","notaDensidade","notaVigor"]) if not _df1.empty else ([], {}, 0)
_r1_range, _d1_range, _n1_range = _chk_range(_df1, [
    ("notaUniformidade", 0, 5, "escala 0–5"),
    ("notaDensidade",    0, 5, "escala 0–5"),
    ("notaVigor",        0, 5, "escala 0–5"),
    ("notaDaninhas",     0, 5, "escala 0–5"),
    ("notaPragas",       0, 5, "escala 0–5"),
    ("notaDoencas",      0, 5, "escala 0–5"),
]) if not _df1.empty else ([], {}, 0)
_n1_dups, _d1_dups  = _chk_dups(_df1, ["fazendaRef","idBaseRef"])
_r1_orf, _d1_orf, _n1_orf = _chk_orfaos(_df1) if not _df1.empty else ([], {}, 0)
_n1_total = _n1_nulos + _n1_range + (1 if _n1_dups > 0 else 0) + _n1_orf

def _av1_content():
    if _df1.empty:
        st.info("Tabelão Av1 não disponível.")
        return
    st.caption(f"**{len(_df1)} registros** no tabelão Av1")
    if _r1_nulos:
        st.markdown("**📋 Nulos Críticos**")
        st.dataframe(pd.DataFrame(_r1_nulos), use_container_width=True, hide_index=True)
        for k, d in _d1_nulos.items():
            st.markdown(f"→ `{k}`: {len(d)} registro(s)")
            st.dataframe(_show(d), use_container_width=True, hide_index=True)
    if _r1_range:
        st.markdown("**📏 Valores Fora de Intervalo**")
        st.dataframe(pd.DataFrame(_r1_range), use_container_width=True, hide_index=True)
        for k, d in _d1_range.items():
            st.markdown(f"→ `{k}`: {len(d)} registro(s)")
            st.dataframe(_show(d), use_container_width=True, hide_index=True)
    if _n1_dups > 0:
        st.markdown(f"**♻️ Duplicatas** — {_n1_dups} registro(s)")
        st.dataframe(_show(_d1_dups), use_container_width=True, hide_index=True)
    if _r1_orf:
        st.markdown("**🔗 Registros Órfãos**")
        st.dataframe(pd.DataFrame(_r1_orf), use_container_width=True, hide_index=True)
        for k, d in _d1_orf.items():
            st.dataframe(_show(d), use_container_width=True, hide_index=True)

_render_av("Av1 — Qualidade Inicial", "📋", _n1_total, "Média", _av1_content)

# ────────────────────────────────────────────────────────────────────────────
# Av2 — Nível de Doenças
# ────────────────────────────────────────────────────────────────────────────
_df2 = _safe("tb_av2")
_r2_nulos, _d2_nulos, _n2_nulos = _chk_nulos(_df2, ["notaMedia"]) if not _df2.empty else ([], {}, 0)
_r2_range, _d2_range, _n2_range = _chk_range(_df2, [
    ("notaPhytophthora",  0, 9, "escala 0–9"),
    ("notaAnomalia",      0, 9, "escala 0–9"),
    ("notaOidio",         0, 9, "escala 0–9"),
    ("notaManchaParda",   0, 9, "escala 0–9"),
    ("notaManchaAlvo",    0, 9, "escala 0–9"),
    ("notaManchaOlhoRa",  0, 9, "escala 0–9"),
    ("notaCercospora",    0, 9, "escala 0–9"),
    ("notaAntracnose",    0, 9, "escala 0–9"),
    ("notaDFC",           0, 9, "escala 0–9"),
]) if not _df2.empty else ([], {}, 0)
_n2_dups, _d2_dups = _chk_dups(_df2, ["fazendaRef","idBaseRef"])
_r2_orf, _d2_orf, _n2_orf = _chk_orfaos(_df2) if not _df2.empty else ([], {}, 0)
_n2_total = _n2_nulos + _n2_range + (1 if _n2_dups > 0 else 0) + _n2_orf

def _av2_content():
    if _df2.empty:
        st.info("Tabelão Av2 não disponível.")
        return
    st.caption(f"**{len(_df2)} registros** no tabelão Av2")
    if _r2_nulos:
        st.markdown("**📋 Nulos Críticos**")
        st.dataframe(pd.DataFrame(_r2_nulos), use_container_width=True, hide_index=True)
    if _r2_range:
        st.markdown("**📏 Valores Fora de Intervalo**")
        st.dataframe(pd.DataFrame(_r2_range), use_container_width=True, hide_index=True)
        for k, d in _d2_range.items():
            st.markdown(f"→ `{k}`: {len(d)} registro(s)")
            st.dataframe(_show(d), use_container_width=True, hide_index=True)
    if _n2_dups > 0:
        st.markdown(f"**♻️ Duplicatas** — {_n2_dups} registro(s)")
        st.dataframe(_show(_d2_dups), use_container_width=True, hide_index=True)
    if _r2_orf:
        st.markdown("**🔗 Registros Órfãos**")
        st.dataframe(pd.DataFrame(_r2_orf), use_container_width=True, hide_index=True)

_render_av("Av2 — Nível de Doenças", "🦠", _n2_total, "Média", _av2_content)

# ────────────────────────────────────────────────────────────────────────────
# Av3 — Floração
# ────────────────────────────────────────────────────────────────────────────
_df3 = _safe("tb_av3")
_r3_nulos, _d3_nulos, _n3_nulos = _chk_nulos(_df3, ["dataInicioFloracao","corFlor","habitoCrescimento"]) if not _df3.empty else ([], {}, 0)
_n3_datas = 0; _d3_datas = pd.DataFrame()
if not _df3.empty and {"dataInicioFloracao","dataFimFloracao"}.issubset(_df3.columns):
    _ini = pd.to_datetime(_df3["dataInicioFloracao"].astype(str), errors="coerce")
    _fim = pd.to_datetime(_df3["dataFimFloracao"].astype(str),    errors="coerce")
    _mask_dt = _fim.notna() & _ini.notna() & (_fim < _ini)
    if int(_mask_dt.sum()) > 0:
        _n3_datas = 1
        _d3_datas = _df3.loc[_mask_dt, _ctx(_df3) + ["dataInicioFloracao","dataFimFloracao"]]
_n3_dups, _d3_dups = _chk_dups(_df3, ["fazendaRef","idBaseRef"])
_r3_orf, _d3_orf, _n3_orf = _chk_orfaos(_df3) if not _df3.empty else ([], {}, 0)
_n3_total = _n3_nulos + _n3_datas + (1 if _n3_dups > 0 else 0) + _n3_orf

def _av3_content():
    if _df3.empty:
        st.info("Tabelão Av3 não disponível.")
        return
    st.caption(f"**{len(_df3)} registros** no tabelão Av3")
    if _r3_nulos:
        st.markdown("**📋 Nulos Críticos**")
        st.dataframe(pd.DataFrame(_r3_nulos), use_container_width=True, hide_index=True)
    if _n3_datas > 0:
        st.markdown(f"**📅 Floração invertida** — fim anterior ao início: {len(_d3_datas)} registro(s)")
        st.dataframe(_show(_d3_datas), use_container_width=True, hide_index=True)
    else:
        st.success("✅ Datas de floração consistentes.")
    if _n3_dups > 0:
        st.markdown(f"**♻️ Duplicatas** — {_n3_dups} registro(s)")
        st.dataframe(_show(_d3_dups), use_container_width=True, hide_index=True)
    if _r3_orf:
        st.markdown("**🔗 Registros Órfãos**")
        st.dataframe(pd.DataFrame(_r3_orf), use_container_width=True, hide_index=True)

_render_av("Av3 — Floração", "🌸", _n3_total, "Média", _av3_content)

# ────────────────────────────────────────────────────────────────────────────
# Av4 — Morfologia
# ────────────────────────────────────────────────────────────────────────────
_df4 = _safe("tb_av4")
_r4_nulos, _d4_nulos, _n4_nulos = _chk_nulos(_df4, ["media_ALT","media_AIV"]) if not _df4.empty else ([], {}, 0)
_r4_range, _d4_range, _n4_range = _chk_range(_df4, [
    ("media_ALT", 30, 120, "cm"),
    ("media_AIV",  6,  35, "cm"),
]) if not _df4.empty else ([], {}, 0)
_n4_dups, _d4_dups = _chk_dups(_df4, ["fazendaRef","idBaseRef"])
_r4_orf, _d4_orf, _n4_orf = _chk_orfaos(_df4) if not _df4.empty else ([], {}, 0)
_r4_out_alt, _d4_out_alt, _n4_out_alt = _chk_outlier(_df4, "media_ALT", 30, 120, "Altura de Planta (cm)", "std_ALT", 12.0)
_r4_out_aiv, _d4_out_aiv, _n4_out_aiv = _chk_outlier(_df4, "media_AIV",  6,  35, "Inserção 1ª Vagem (cm)", "std_AIV", 4.0)
_n4_total = _n4_nulos + _n4_range + (1 if _n4_dups > 0 else 0) + _n4_orf + _n4_out_alt + _n4_out_aiv

def _av4_content():
    if _df4.empty:
        st.info("Tabelão Av4 não disponível.")
        return
    st.caption(f"**{len(_df4)} registros** no tabelão Av4")
    if _r4_nulos:
        st.markdown("**📋 Nulos Críticos**")
        st.dataframe(pd.DataFrame(_r4_nulos), use_container_width=True, hide_index=True)
    if _r4_range:
        st.markdown("**📏 Valores Fora de Intervalo**")
        st.dataframe(pd.DataFrame(_r4_range), use_container_width=True, hide_index=True)
        for k, d in _d4_range.items():
            st.markdown(f"→ `{k}`: {len(d)} registro(s)")
            st.dataframe(_show(d), use_container_width=True, hide_index=True)
    rows_out4 = [r for r in [_r4_out_alt, _r4_out_aiv] if r]
    if rows_out4:
        st.markdown("**📊 Outliers / CV Alto**")
        st.dataframe(pd.DataFrame(rows_out4), use_container_width=True, hide_index=True)
        for d in [_d4_out_alt, _d4_out_aiv]:
            if not d.empty:
                st.dataframe(_show(d), use_container_width=True, hide_index=True)
    if _n4_dups > 0:
        st.markdown(f"**♻️ Duplicatas** — {_n4_dups} registro(s)")
        st.dataframe(_show(_d4_dups), use_container_width=True, hide_index=True)
    if _r4_orf:
        st.markdown("**🔗 Registros Órfãos**")
        st.dataframe(pd.DataFrame(_r4_orf), use_container_width=True, hide_index=True)

_render_av("Av4 — Morfologia", "🌿", _n4_total, "Alta", _av4_content)

# ────────────────────────────────────────────────────────────────────────────
# Av5 — Vagens
# ────────────────────────────────────────────────────────────────────────────
_df5 = _safe("tb_av5")
_r5_nulos, _d5_nulos, _n5_nulos = _chk_nulos(_df5, ["media_totalVagens"]) if not _df5.empty else ([], {}, 0)
_r5_range, _d5_range, _n5_range = _chk_range(_df5, [
    ("media_VTS",         0,  80, "n"),
    ("media_VTM",         0, 100, "n"),
    ("media_VTI",         0,  60, "n"),
    ("media_totalVagens", 0, 150, "n"),
]) if not _df5.empty else ([], {}, 0)
_n5_dups, _d5_dups = _chk_dups(_df5, ["fazendaRef","idBaseRef"])
_r5_orf, _d5_orf, _n5_orf = _chk_orfaos(_df5) if not _df5.empty else ([], {}, 0)
# Soma dos terços vs total
_n5_tercos = 0; _d5_tercos = pd.DataFrame()
if not _df5.empty and {"media_VTS","media_VTM","media_VTI","media_totalVagens"}.issubset(_df5.columns):
    _d = _df5.copy()
    _d["soma_tercos"] = _d[["media_VTS","media_VTM","media_VTI"]].sum(axis=1)
    _d = _d[_d["media_totalVagens"].notna() & (_d["media_totalVagens"] > 0)].copy()
    _d["diff_pct"] = ((_d["soma_tercos"] - _d["media_totalVagens"]).abs() / _d["media_totalVagens"] * 100).round(1)
    _ms5 = _d["diff_pct"] > 5
    if int(_ms5.sum()) > 0:
        _n5_tercos = 1
        _c5 = [c for c in _ctx(_df5) + ["media_VTS","media_VTM","media_VTI","soma_tercos","media_totalVagens","diff_pct"] if c in _d.columns]
        _d5_tercos = _d.loc[_ms5, _c5]
_n5_total = _n5_nulos + _n5_range + _n5_tercos + (1 if _n5_dups > 0 else 0) + _n5_orf

def _av5_content():
    if _df5.empty:
        st.info("Tabelão Av5 não disponível.")
        return
    st.caption(f"**{len(_df5)} registros** no tabelão Av5")
    if _r5_nulos:
        st.markdown("**📋 Nulos Críticos**")
        st.dataframe(pd.DataFrame(_r5_nulos), use_container_width=True, hide_index=True)
    if _r5_range:
        st.markdown("**📏 Valores Fora de Intervalo**")
        st.dataframe(pd.DataFrame(_r5_range), use_container_width=True, hide_index=True)
        for k, d in _d5_range.items():
            st.markdown(f"→ `{k}`: {len(d)} registro(s)")
            st.dataframe(_show(d), use_container_width=True, hide_index=True)
    if _n5_tercos > 0:
        st.markdown(f"**🔢 Σ Terços ≠ Total de Vagens (>5%)** — {len(_d5_tercos)} registro(s)")
        st.dataframe(_show(_d5_tercos), use_container_width=True, hide_index=True)
    if _n5_dups > 0:
        st.markdown(f"**♻️ Duplicatas** — {_n5_dups} registro(s)")
        st.dataframe(_show(_d5_dups), use_container_width=True, hide_index=True)
    if _r5_orf:
        st.markdown("**🔗 Registros Órfãos**")
        st.dataframe(pd.DataFrame(_r5_orf), use_container_width=True, hide_index=True)

_render_av("Av5 — Vagens", "🫘", _n5_total, "Média", _av5_content)

# ────────────────────────────────────────────────────────────────────────────
# Av6 — Maturidade
# ────────────────────────────────────────────────────────────────────────────
_df6 = _safe("tb_av6")
_r6_nulos, _d6_nulos, _n6_nulos = _chk_nulos(_df6, ["DMF","notaAC","notaAV","notaQF"]) if not _df6.empty else ([], {}, 0)
# Ciclo fora de intervalo
_n6_ciclo = 0; _d6_ciclo = pd.DataFrame()
if not _df6.empty and "dias_ate_DMF" in _df6.columns:
    _mask6 = _df6["dias_ate_DMF"].notna() & ~_df6["dias_ate_DMF"].between(80, 160)
    if int(_mask6.sum()) > 0:
        _n6_ciclo = 1
        _d6_ciclo = _df6.loc[_mask6, _ctx(_df6) + ["dataPlantioSoja","DMF","dias_ate_DMF"]]
# GM vs GM_visual
_n6_gm = 0; _d6_gm = pd.DataFrame()
if not _df6.empty and {"GM_visual","gm"}.issubset(_df6.columns):
    _df6c = _df6.copy()
    _df6c["gm_num"]        = pd.to_numeric(_df6c["gm"],        errors="coerce")
    _df6c["GM_visual_num"] = pd.to_numeric(_df6c["GM_visual"], errors="coerce")
    _df6v = _df6c[_df6c["gm_num"].notna() & _df6c["GM_visual_num"].notna() & (_df6c["GM_visual_num"] > 0)].copy()
    _df6v["diff_gm"] = (_df6v["gm_num"] - _df6v["GM_visual_num"]).abs()
    _d6_gm = _df6v[_df6v["diff_gm"] > 30].copy()
    _n6_gm = len(_d6_gm)
_n6_dups, _d6_dups = _chk_dups(_df6, ["fazendaRef","idBaseRef"])
_r6_orf, _d6_orf, _n6_orf = _chk_orfaos(_df6) if not _df6.empty else ([], {}, 0)
_n6_total = _n6_nulos + _n6_ciclo + (1 if _n6_gm > 0 else 0) + (1 if _n6_dups > 0 else 0) + _n6_orf

def _av6_content():
    if _df6.empty:
        st.info("Tabelão Av6 não disponível.")
        return
    st.caption(f"**{len(_df6)} registros** no tabelão Av6")
    if _r6_nulos:
        st.markdown("**📋 Nulos Críticos**")
        st.dataframe(pd.DataFrame(_r6_nulos), use_container_width=True, hide_index=True)
        for k, d in _d6_nulos.items():
            st.markdown(f"→ `{k}`: {len(d)} registro(s)")
            st.dataframe(_show(d), use_container_width=True, hide_index=True)
    if _n6_ciclo > 0:
        st.markdown(f"**📅 Ciclo fora de [80–160] dias** — {len(_d6_ciclo)} registro(s)")
        st.dataframe(_show(_d6_ciclo), use_container_width=True, hide_index=True)
    else:
        st.success("✅ Ciclo dentro do intervalo esperado.")
    if _n6_gm > 0:
        st.markdown(f"**🎯 GM vs GM Visual — divergência > 3.0** — {_n6_gm} registro(s)")
        _cols_gm = [c for c in _ctx(_df6) + ["gm_num","GM_visual_num","diff_gm"] if c in _d6_gm.columns]
        st.dataframe(_show(_d6_gm[_cols_gm].sort_values("diff_gm", ascending=False).reset_index(drop=True)), use_container_width=True, hide_index=True)
    else:
        st.success("✅ GM vs GM visual consistentes.")
    if _n6_dups > 0:
        st.markdown(f"**♻️ Duplicatas** — {_n6_dups} registro(s)")
        st.dataframe(_show(_d6_dups), use_container_width=True, hide_index=True)
    if _r6_orf:
        st.markdown("**🔗 Registros Órfãos**")
        st.dataframe(pd.DataFrame(_r6_orf), use_container_width=True, hide_index=True)

_render_av("Av6 — Maturidade", "🌾", _n6_total, "Alta", _av6_content)

# ────────────────────────────────────────────────────────────────────────────
# Av7 — Produtividade
# ────────────────────────────────────────────────────────────────────────────
_df7 = _safe("tb_av7")
_r7_nulos, _d7_nulos, _n7_nulos = _chk_nulos(_df7, ["kg_ha","sc_ha","pesoMilGraos_corrigido","umidadeParcela"]) if not _df7.empty else ([], {}, 0)
_r7_range, _d7_range, _n7_range = _chk_range(_df7, [
    ("umidadeParcela",             0, 100, "%"),
    ("umidadeAmostraPesoMilGraos", 0, 100, "%"),
    ("gm",                        50,  99, "GM inteiro"),
    ("kg_ha",                    900, 6000, "kg/ha"),
]) if not _df7.empty else ([], {}, 0)
_r7_out_sc,  _d7_out_sc,  _n7_out_sc  = _chk_outlier(_df7, "sc_ha",  15, 100, "Produtividade (sc/ha)")
_r7_out_pmg, _d7_out_pmg, _n7_out_pmg = _chk_outlier(_df7, "pesoMilGraos_corrigido", 80, 200, "PMG (g)")
# Consistência agronômica
_n7_agro = 0; _dfs7_agro = {}; _rows7_agro = []
if not _df7.empty:
    if {"populacao","pop_plantasFinal_ha"}.issubset(_df7.columns):
        _ds = _df7[_df7["populacao"].notna() & _df7["pop_plantasFinal_ha"].notna() & (_df7["populacao"] > 0)].copy()
        _ds["perda_stand_pct"] = ((_ds["populacao"] - _ds["pop_plantasFinal_ha"]) / _ds["populacao"] * 100).round(1)
        _ms7 = _ds["perda_stand_pct"] > 30
        _n7s = int(_ms7.sum())
        if _n7s > 0:
            _n7_agro += 1
            _dfs7_agro["Perda de stand > 30%"] = _ds.loc[_ms7, _ctx(_df7) + ["populacao","pop_plantasFinal_ha","perda_stand_pct"]]
        _rows7_agro.append({"Verificação": "Perda de stand > 30%", "Ocorrências": _n7s, "Status": "⚠️" if _n7s > 0 else "✅ OK"})
    if "pop_plantasFinal_ha" in _df7.columns:
        _mp = _df7["pop_plantasFinal_ha"].notna() & (_df7["pop_plantasFinal_ha"] > 350000)
        _np = int(_mp.sum())
        if _np > 0:
            _n7_agro += 1
            _dfs7_agro["Pop. Final > 350k"] = _df7.loc[_mp, _ctx(_df7) + ["populacao","pop_plantasFinal_ha"]]
        _rows7_agro.append({"Verificação": "Pop. Final > 350.000 pl/ha", "Ocorrências": _np, "Status": "⚠️" if _np > 0 else "✅ OK"})
# Qualidade inicial vs final (Av1 × Av6 dentro do contexto Av7)
_n7_qual = 0; _d7_qual = pd.DataFrame()
_df1_q = _safe("tb_av1"); _df6_q = _safe("tb_av6")
if not _df1_q.empty and not _df6_q.empty:
    _chq = [c for c in ["fazendaRef","idBaseRef"] if c in _df1_q.columns and c in _df6_q.columns]
    if _chq and "notaMedia" in _df1_q.columns and "notaQF" in _df6_q.columns:
        _q1 = _df1_q[_chq + _ctx(_df1_q) + ["notaMedia"]].rename(columns={"notaMedia":"notaQI"})
        _q6 = _df6_q[_chq + ["notaQF"]].dropna(subset=["notaQF"])
        _dq = _q1.merge(_q6, on=_chq, how="inner")
        _dq = _dq[_dq["notaQI"].notna() & _dq["notaQF"].notna()].copy()
        _dq["diff"] = (_dq["notaQI"] - _dq["notaQF"]).abs()
        _d7_qual = _dq[_dq["diff"] > 2].sort_values("diff", ascending=False)
        _n7_qual = len(_d7_qual)

_n7_dups, _d7_dups = _chk_dups(_df7, ["fazendaRef","idBaseRef"])
_r7_orf, _d7_orf, _n7_orf = _chk_orfaos(_df7) if not _df7.empty else ([], {}, 0)
_n7_total = (_n7_nulos + _n7_range + _n7_out_sc + _n7_out_pmg + _n7_agro +
             (1 if _n7_qual > 0 else 0) + (1 if _n7_dups > 0 else 0) + _n7_orf)

def _av7_content():
    if _df7.empty:
        st.info("Tabelão Av7 não disponível.")
        return
    st.caption(f"**{len(_df7)} registros** no tabelão Av7")
    if _r7_nulos:
        st.markdown("**📋 Nulos Críticos**")
        st.dataframe(pd.DataFrame(_r7_nulos), use_container_width=True, hide_index=True)
        for k, d in _d7_nulos.items():
            st.markdown(f"→ `{k}`: {len(d)} registro(s)")
            st.dataframe(_show(d), use_container_width=True, hide_index=True)
    if _r7_range:
        st.markdown("**📏 Valores Fora de Intervalo**")
        st.dataframe(pd.DataFrame(_r7_range), use_container_width=True, hide_index=True)
        for k, d in _d7_range.items():
            st.markdown(f"→ `{k}`: {len(d)} registro(s)")
            st.dataframe(_show(d), use_container_width=True, hide_index=True)
    rows_out7 = [r for r in [_r7_out_sc, _r7_out_pmg] if r]
    if rows_out7:
        st.markdown("**📊 Outliers / CV Alto**")
        st.dataframe(pd.DataFrame(rows_out7), use_container_width=True, hide_index=True)
        for d in [_d7_out_sc, _d7_out_pmg]:
            if not d.empty:
                st.dataframe(_show(d), use_container_width=True, hide_index=True)
    if _rows7_agro:
        st.markdown("**🌱 Consistência Agronômica**")
        st.dataframe(pd.DataFrame(_rows7_agro), use_container_width=True, hide_index=True)
        for k, d in _dfs7_agro.items():
            st.markdown(f"→ {k}: {len(d)} registro(s)")
            st.dataframe(_show(d), use_container_width=True, hide_index=True)
    if _n7_qual > 0:
        st.markdown(f"**⭐ Qualidade Inicial vs Final (Av1×Av6) — |diff| > 2** — {_n7_qual} registro(s)")
        _cols_q7 = [c for c in _ctx(_df1_q) + ["notaQI","notaQF","diff"] if c in _d7_qual.columns]
        st.dataframe(_show(_d7_qual[_cols_q7].reset_index(drop=True)), use_container_width=True, hide_index=True)
    else:
        st.success("✅ Qualidade inicial vs final consistentes.")
    if _n7_dups > 0:
        st.markdown(f"**♻️ Duplicatas** — {_n7_dups} registro(s)")
        st.dataframe(_show(_d7_dups), use_container_width=True, hide_index=True)
    if _r7_orf:
        st.markdown("**🔗 Registros Órfãos**")
        st.dataframe(pd.DataFrame(_r7_orf), use_container_width=True, hide_index=True)

_render_av("Av7 — Produtividade", "🏆", _n7_total, "Alta", _av7_content)


# ── Exportação Excel ──────────────────────────────────────────────────────────
secao_titulo("Exportação", "Baixar Relatório de Inconsistências", "Planilha Excel organizada por avaliação.")

if "excel_diag_bytes" not in st.session_state:
    st.session_state["excel_diag_bytes"] = None

_gerar = False
if st.session_state["excel_diag_bytes"] is None:
    _col_btn, _ = st.columns([1, 3])
    with _col_btn:
        _gerar = st.button("⬇️ Gerar relatório Excel", type="primary", use_container_width=True)
else:
    _col_btn, _ = st.columns([1, 3])
    with _col_btn:
        st.download_button(
            label="📥 Baixar Excel",
            data=st.session_state["excel_diag_bytes"],
            file_name="diagnostico_jaum_dtc.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_excel_diag",
            use_container_width=True,
        )
    if st.button("↩ Gerar novamente"):
        st.session_state["excel_diag_bytes"] = None
        st.rerun()

if _gerar:
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    _HF  = PatternFill("solid", start_color="1F4E79")
    _HFt = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    _AF  = PatternFill("solid", start_color="FCE4D6")
    _WF  = PatternFill("solid", start_color="FFF2CC")
    _OF  = PatternFill("solid", start_color="E2EFDA")
    _NF  = Font(name="Arial", size=10)
    _tn  = Side(style="thin", color="BFBFBF")
    _BR  = Border(left=_tn, right=_tn, top=_tn, bottom=_tn)

    def _hdr(ws, cols, widths):
        ws.append(cols)
        for ci, (cell, w) in enumerate(zip(ws[1], widths), 1):
            cell.font = _HFt; cell.fill = _HF
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _BR
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 28

    def _row(ws, idx, fill=None):
        for cell in ws[idx]:
            cell.font = _NF
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = _BR
            if fill: cell.fill = fill

    def _garantir_resp(df):
        """Garante nomeResponsavel como 1ª coluna, depois as demais de COL_CTX."""
        if df is None or df.empty:
            return df
        _prioridade = ["nomeResponsavel", "safra", "cod_fazenda", "nomeFazenda",
                       "cidade_nome", "estado_sigla", "nome", "dePara",
                       "status_material", "tipoTeste", "indexTratamento"]
        _ctx_presentes = [c for c in _prioridade if c in df.columns]
        _resto = [c for c in df.columns if c not in _ctx_presentes]
        return df[_ctx_presentes + _resto]

    def _df2ws(ws, df, fill=None, obs=None):
        df = _garantir_resp(df).reset_index(drop=True)
        if obs:
            ws["A1"] = obs
            ws["A1"].font = Font(bold=True, name="Arial", size=10, color="9C0006")
            ws.row_dimensions[1].height = 18
        _hdr(ws, list(df.columns), [max(14, len(c) + 2) for c in df.columns])
        if df.empty:
            ws.append(["Nenhum registro encontrado"]); _row(ws, ws.max_row, _OF); return
        for _, r in df.iterrows():
            ws.append([None if pd.isna(v) else (str(v) if hasattr(v, "date") else v) for v in r])
            _row(ws, ws.max_row, fill or _AF)
        ws.freeze_panes = "B2"  # congela nomeResponsavel

    def _safe_name(s):
        return s.replace("/","_").replace("\\","_").replace("*","_")\
                .replace("?","_").replace("[","_").replace("]","_").replace(":","_")[:31]

    wb = openpyxl.Workbook(); wb.remove(wb.active)

    # ── Aba Resumo por avaliação ──────────────────────────────────────────────
    ws_r = wb.create_sheet("Resumo")
    _hdr(ws_r, ["Avaliação", "Alertas", "Status"], [28, 10, 10])
    for _av_nome, _av_n in [
        ("Av1 — Qualidade Inicial",  _n1_total),
        ("Av2 — Nível de Doenças",   _n2_total),
        ("Av3 — Floração",           _n3_total),
        ("Av4 — Morfologia",         _n4_total),
        ("Av5 — Vagens",             _n5_total),
        ("Av6 — Maturidade",         _n6_total),
        ("Av7 — Produtividade",      _n7_total),
    ]:
        ws_r.append([_av_nome, _av_n, "⚠️" if _av_n > 0 else "✅"])
        _row(ws_r, ws_r.max_row, _AF if _av_n > 0 else _OF)
    ws_r.freeze_panes = "A2"

    # ── Abas por avaliação ────────────────────────────────────────────────────
    _abas = []

    # Av1
    for k, d in _d1_nulos.items():   _abas.append((_safe_name(f"Av1 Nulo {k}"),   d, _AF, f"Av1 — Nulo: {k}"))
    for k, d in _d1_range.items():   _abas.append((_safe_name(f"Av1 Range {k}"),  d, _WF, f"Av1 — Fora intervalo: {k}"))
    if _n1_dups > 0:                  _abas.append((_safe_name("Av1 Duplicatas"),  _d1_dups, _WF, "Av1 — Duplicatas"))
    for k, d in _d1_orf.items():     _abas.append((_safe_name(f"Av1 Orfao {k}"),  d, _AF, f"Av1 — Órfão: {k}"))

    # Av2
    for k, d in _d2_range.items():   _abas.append((_safe_name(f"Av2 Range {k}"),  d, _WF, f"Av2 — Fora intervalo: {k}"))
    if _n2_dups > 0:                  _abas.append((_safe_name("Av2 Duplicatas"),  _d2_dups, _WF, "Av2 — Duplicatas"))
    for k, d in _d2_orf.items():     _abas.append((_safe_name(f"Av2 Orfao {k}"),  d, _AF, f"Av2 — Órfão: {k}"))

    # Av3
    for k, d in _d3_nulos.items():   _abas.append((_safe_name(f"Av3 Nulo {k}"),   d, _AF, f"Av3 — Nulo: {k}"))
    if _n3_datas > 0:                 _abas.append((_safe_name("Av3 Floracao inv"),_d3_datas, _WF, "Av3 — Floração invertida"))
    if _n3_dups > 0:                  _abas.append((_safe_name("Av3 Duplicatas"),  _d3_dups, _WF, "Av3 — Duplicatas"))

    # Av4
    for k, d in _d4_nulos.items():   _abas.append((_safe_name(f"Av4 Nulo {k}"),   d, _AF, f"Av4 — Nulo: {k}"))
    for k, d in _d4_range.items():   _abas.append((_safe_name(f"Av4 Range {k}"),  d, _WF, f"Av4 — Fora intervalo: {k}"))
    if not _d4_out_alt.empty:         _abas.append((_safe_name("Av4 Outlier ALT"), _d4_out_alt, _WF, "Av4 — Outlier Altura"))
    if not _d4_out_aiv.empty:         _abas.append((_safe_name("Av4 Outlier AIV"), _d4_out_aiv, _WF, "Av4 — Outlier AIV"))
    if _n4_dups > 0:                  _abas.append((_safe_name("Av4 Duplicatas"),  _d4_dups, _WF, "Av4 — Duplicatas"))

    # Av5
    for k, d in _d5_range.items():   _abas.append((_safe_name(f"Av5 Range {k}"),  d, _WF, f"Av5 — Fora intervalo: {k}"))
    if _n5_tercos > 0:                _abas.append((_safe_name("Av5 Tercos"),      _d5_tercos, _WF, "Av5 — Σterços ≠ totalVagens"))
    if _n5_dups > 0:                  _abas.append((_safe_name("Av5 Duplicatas"),  _d5_dups, _WF, "Av5 — Duplicatas"))

    # Av6
    for k, d in _d6_nulos.items():   _abas.append((_safe_name(f"Av6 Nulo {k}"),   d, _AF, f"Av6 — Nulo: {k}"))
    if _n6_ciclo > 0:                 _abas.append((_safe_name("Av6 Ciclo"),        _d6_ciclo, _WF, "Av6 — Ciclo fora [80-160] dias"))
    if _n6_gm > 0:                    _abas.append((_safe_name("Av6 GM visual"),    _d6_gm, _WF, "Av6 — GM vs GM visual divergente"))
    if _n6_dups > 0:                  _abas.append((_safe_name("Av6 Duplicatas"),   _d6_dups, _WF, "Av6 — Duplicatas"))

    # Av7
    for k, d in _d7_nulos.items():   _abas.append((_safe_name(f"Av7 Nulo {k}"),   d, _AF, f"Av7 — Nulo: {k}"))
    for k, d in _d7_range.items():   _abas.append((_safe_name(f"Av7 Range {k}"),  d, _WF, f"Av7 — Fora intervalo: {k}"))
    if not _d7_out_sc.empty:          _abas.append((_safe_name("Av7 Outlier sc"),  _d7_out_sc,  _WF, "Av7 — Outlier sc/ha"))
    if not _d7_out_pmg.empty:         _abas.append((_safe_name("Av7 Outlier PMG"), _d7_out_pmg, _WF, "Av7 — Outlier PMG"))
    for k, d in _dfs7_agro.items():  _abas.append((_safe_name(f"Av7 Agro {k}"),   d, _AF, f"Av7 — {k}"))
    if _n7_qual > 0:                  _abas.append((_safe_name("Av7 Qualidade"),    _d7_qual, _WF, "Av7 — Qualidade inicial vs final"))
    if _n7_dups > 0:                  _abas.append((_safe_name("Av7 Duplicatas"),   _d7_dups, _WF, "Av7 — Duplicatas"))

    for nome_aba, df_ab, fill_ab, obs_ab in _abas:
        ws_ab = wb.create_sheet(nome_aba)
        ws_ab.sheet_view.showGridLines = False
        _df2ws(ws_ab, df_ab, fill=fill_ab, obs=f"{len(df_ab)} registro(s) | {obs_ab}")

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    st.session_state["excel_diag_bytes"] = buf.getvalue()
    st.rerun()

st.divider()
st.markdown(
    '<p style="font-size:13px;color:#374151;text-align:center;">Painel JAUM DTC · Stine Seed · '
    'Desenvolvido por <a href="https://www.linkedin.com/in/eng-agro-andre-ferreira/" '
    'target="_blank" style="color:#27AE60;text-decoration:none;">Andre Ferreira</a></p>',
    unsafe_allow_html=True,
)
