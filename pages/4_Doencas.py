"""
pages/4_Doencas.py — Avaliação de Doenças
"""
import numpy as np
import pandas as pd
import streamlit as st
import plotly.graph_objects as go_plt

# ── Cores globais por status ──────────────────────────────────────────────────
COR_STATUS_PLOT = {
    "CHECK":    "#F4B184",
    "STINE":    "#2976B6",
    "LINHAGEM": "#00FF01",
    "DP2":      "#C4DFB4",
}
COR_BORDA = {
    "CHECK":    "#C46A3A",
    "STINE":    "#1A4F7A",
    "LINHAGEM": "#009900",
    "DP2":      "#7AAF6A",
}
COR_STATUS_TEXTO = {
    "CHECK":    "#C46A3A",
    "STINE":    "#2976B6",
    "LINHAGEM": "#009900",
    "DP2":      "#5A8A4A",
}

# ── Classes de reação a doenças ───────────────────────────────────────────────
ORDEM_CLASS = ["AS", "S", "MT", "T", "R"]
COR_CLASS = {
    "AS": "#8B0000",   # vermelho escuro — Altamente Susceptível
    "S":  "#E63946",   # vermelho        — Susceptível
    "MT": "#FFD600",   # amarelo         — Moderadamente Tolerante
    "T":  "#70C96E",   # verde claro     — Tolerante
    "R":  "#1E7A34",   # verde escuro    — Resistente
}
COR_TEXTO_CLASS = {
    "AS": "#FFFFFF",
    "S":  "#FFFFFF",
    "MT": "#1A1A1A",
    "T":  "#1A1A1A",
    "R":  "#FFFFFF",
}
LABEL_CLASS = {
    "AS": "AS — Altamente Susceptível (1–2)",
    "S":  "S — Susceptível",
    "MT": "MT — Moderadamente Tolerante",
    "T":  "T — Tolerante",
    "R":  "R — Resistente",
}

STATUS_ORDER_HM = ["CHECK", "STINE", "DP2", "LINHAGEM"]

COR_STATUS_TEXTO_HM = {
    "CHECK":    "#C46A3A",
    "STINE":    "#2976B6",
    "LINHAGEM": "#009900",
    "DP2":      "#5A8A4A",
}

COR_STATUS_TEXTO_HM = {
    "CHECK":    "#C46A3A",
    "STINE":    "#2976B6",
    "LINHAGEM": "#009900",
    "DP2":      "#5A8A4A",
}

def nota_para_classe(moda):
    """Converte moda da nota (1–9) para sigla de classe."""
    if moda is None or (isinstance(moda, float) and np.isnan(moda)):
        return None
    m = float(moda)
    if m <= 2:   return "AS"
    elif m <= 4: return "S"
    elif m <= 6: return "MT"
    elif m <= 8: return "T"
    else:        return "R"

# ── Doenças disponíveis ───────────────────────────────────────────────────────
DOENCAS = {
    "Phytophthora":    {"nota": "notaPhytophthora",    "inc": "inc_notaPhytophthora",    "class": "class_notaPhytophthora"},
    "Anomalia":        {"nota": "notaAnomalia",        "inc": "inc_notaAnomalia",        "class": "class_notaAnomalia"},
    "Oídio":           {"nota": "notaOidio",           "inc": "inc_notaOidio",           "class": "class_notaOidio"},
    "Mancha Parda":    {"nota": "notaManchaParda",     "inc": "inc_notaManchaParda",     "class": "class_notaManchaParda"},
    "Mancha Alvo":     {"nota": "notaManchaAlvo",      "inc": "inc_notaManchaAlvo",      "class": "class_notaManchaAlvo"},
    "Mancha Olho-Rã":  {"nota": "notaManchaOlhoRa",   "inc": "inc_notaManchaOlhoRa",   "class": "class_notaManchaOlhoRa"},
    "Cercospora":      {"nota": "notaCercospora",      "inc": "inc_notaCercospora",      "class": "class_notaCercospora"},
    "Antracnose":      {"nota": "notaAntracnose",      "inc": "inc_notaAntracnose",      "class": "class_notaAntracnose"},
    "DFC":             {"nota": "notaDFC",             "inc": "inc_notaDFC",             "class": "class_notaDFC"},
}

from utils.theme import aplicar_tema, page_header, secao_titulo
from utils.loader import carregar_2023, carregar_2024, carregar_2025
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, JsCode, ColumnsAutoSizeMode

st.set_page_config(
    page_title="Doenças · JAUM DTC",
    page_icon="🌿",
    layout="wide",
    initial_sidebar_state="expanded",
)

aplicar_tema()

st.markdown("""
<style>
.jaum-header img { height: 200px !important; }
[data-testid="stDataFrame"] td,
[data-testid="stDataFrame"] th,
[data-testid="stDataFrame"] [role="columnheader"],
[data-testid="stDataFrame"] [role="columnheader"] span,
[data-testid="stDataFrame"] [role="columnheader"] div {
    font-size: 14px !important;
    font-weight: 700 !important;
    color: #000000 !important;
    opacity: 1 !important;
}
[data-testid="stCaptionContainer"] p,
[data-testid="stCaptionContainer"] {
    color: #374151 !important;
    opacity: 1 !important;
}
</style>
""", unsafe_allow_html=True)

# ── Helper AgGrid ─────────────────────────────────────────────────────────────
def ag_table(df, height=400):
    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_default_column(
        resizable=True,
        sortable=True,
        filter=True,
        suppressMenu=False,
        menuTabs=["generalMenuTab", "filterMenuTab", "columnsMenuTab"],
        cellStyle={"fontSize": "14px", "color": "#000000", "fontFamily": "Helvetica Neue, sans-serif"},
    )
    gb.configure_grid_options(
        headerHeight=36,
        rowHeight=32,
        domLayout="normal",
        suppressMenuHide=True,
        suppressColumnVirtualisation=True,
        suppressContextMenu=False,
        enableRangeSelection=True,
    )
    go = gb.build()
    go["defaultColDef"]["headerClass"] = "ag-header-black"
    go["onFirstDataRendered"] = JsCode("function(params) { params.api.sizeColumnsToFit(); }")
    AgGrid(
        df,
        gridOptions=go,
        height=height,
        update_mode=GridUpdateMode.NO_UPDATE,
        fit_columns_on_grid_load=False,
        columns_auto_size_mode=2,
        allow_unsafe_jscode=True,
        enable_enterprise_modules=True,
        custom_css={
            ".ag-header":                      {"background-color": "#4A4A4A !important"},
            ".ag-header-row":                  {"background-color": "#4A4A4A !important"},
            ".ag-header-cell":                 {"background-color": "#4A4A4A !important"},
            ".ag-header-cell-label":           {"color": "#FFFFFF !important", "font-weight": "700"},
            ".ag-header-cell-text":            {"color": "#FFFFFF !important", "font-size": "13px !important", "font-weight": "700 !important"},
            ".ag-icon":                        {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-header-icon":                 {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-header-cell-menu-button":     {"opacity": "1 !important", "visibility": "visible !important"},
            ".ag-header-cell-menu-button span":{"color": "#FFFFFF !important"},
            ".ag-icon-menu":                   {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-icon-filter":                 {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-cell":                        {"font-size": "13px !important", "color": "#000000 !important"},
            ".ag-row":                         {"font-size": "13px !important"},
        },
        theme="streamlit",
        use_container_width=True,
    )


# ── Helper exportar Excel ─────────────────────────────────────────────────────
def exportar_excel(df, nome_arquivo="tabela.xlsx", label="⬇️ Exportar Excel", key=None):
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    buf = io.BytesIO()
    wb  = openpyxl.Workbook()
    ws  = wb.active

    df = df.reset_index(drop=True)
    df = df.loc[:, ~df.columns.str.startswith("::") & ~df.columns.str.startswith("_")].copy()

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=str(col))
        cell.font      = Font(bold=True, name="Arial", size=10, color="1A1A1A")
        cell.fill      = PatternFill("solid", start_color="F2F2F2")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(12, len(str(col)) + 2)
    ws.row_dimensions[1].height = 28

    for ri, row_data in enumerate(df.itertuples(index=False), start=2):
        for ci, val in enumerate(row_data, 1):
            try:
                if val is None or (isinstance(val, float) and np.isnan(val)):
                    val = None
                elif hasattr(val, "__class__") and type(val).__name__ in ("NAType", "NaTType"):
                    val = None
                else:
                    if pd.isna(val):
                        val = None
            except (TypeError, ValueError):
                pass
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="left" if ci == 1 else "center", vertical="center")
            cell.border    = border

    wb.save(buf)
    buf.seek(0)
    st.download_button(
        label=label,
        data=buf,
        file_name=nome_arquivo,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=key,
    )


# ── Header ────────────────────────────────────────────────────────────────────
page_header(
    "Avaliação de Doenças",
    "Analise o comportamento sanitário dos materiais — nota, incidência e classificação por doença. "
    "Identifique cultivares com melhor perfil de sanidade e compare com os checks.",
    imagem="Researchers-pana.png",
)

# ── Carregamento ──────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def carregar_concat():
    frames = []
    for loader in [carregar_2023, carregar_2024, carregar_2025]:
        d = loader()
        if d.get("ok") and d.get("ta_faixa") is not None:
            frames.append(d["ta_faixa"])
    if not frames:
        return pd.DataFrame()
    df = pd.concat(frames, ignore_index=True)
    df["gm"] = pd.to_numeric(df["gm"], errors="coerce")
    df["gm_cat"] = (df["gm"] / 10).round(1)
    return df

with st.spinner("Carregando dados..."):
    ta_raw = carregar_concat()

if ta_raw.empty:
    st.error("❌ Nenhum dado disponível. Verifique a página de Diagnóstico.")
    st.stop()

# Filtra apenas registros que têm ao menos uma coluna de doença preenchida
_cols_doenca = [v["nota"] for v in DOENCAS.values() if v["nota"] in ta_raw.columns]
ta_raw = ta_raw[ta_raw[_cols_doenca].notna().any(axis=1)].copy() if _cols_doenca else ta_raw

if "GM_visual" in ta_raw.columns:
    med_gm = ta_raw["GM_visual"].dropna()
    med_gm = med_gm[med_gm > 0]
    if len(med_gm) > 0 and med_gm.median() > 10:
        ta_raw["GM_visual"] = (ta_raw["GM_visual"] / 10).round(1)

# ── Sidebar — Filtros encadeados ──────────────────────────────────────────────
with st.sidebar:
    st.markdown(
        '<p style="font-size:11px;font-weight:600;color:#6B7280;text-transform:uppercase;'
        'letter-spacing:0.05em;padding:0.5rem;">Filtros</p>',
        unsafe_allow_html=True,
    )

    if st.button("🔄 Limpar filtros", use_container_width=True, key="btn_limpar_dc"):
        for k in list(st.session_state.keys()):
            if any(k.startswith(p) for p in ["dc_safra_", "dc_macro_", "dc_micro_",
                                               "dc_estado_", "dc_cidade_", "dc_fazenda_",
                                               "dc_resp_", "dc_status_", "dc_cult_"]):
                del st.session_state[k]
        st.rerun()

    def checkboxes_dc(label, opcoes, default_all=True, defaults=None, prefix=""):
        sel = []
        for o in opcoes:
            checked = (o in defaults) if defaults is not None else default_all
            if st.checkbox(str(o), value=checked, key=f"{prefix}_{o}"):
                sel.append(o)
        return sel

    with st.expander("📅 Safra", expanded=True):
        safras_all   = sorted(ta_raw["safra"].dropna().unique().tolist())
        safra_default = [s for s in safras_all if "2025" in str(s)] or safras_all[-1:]
        safras_sel   = checkboxes_dc("Safra", safras_all, defaults=safra_default, prefix="dc_safra")

    ta_f1 = ta_raw[ta_raw["safra"].isin(safras_sel)] if safras_sel else ta_raw.iloc[0:0]

    with st.expander("🗺️ Região Macro", expanded=False):
        macros_all = sorted(ta_f1["regiao_macro"].dropna().unique().tolist())
        macros_sel = checkboxes_dc("Macro", macros_all, prefix="dc_macro")

    ta_f2 = ta_f1[ta_f1["regiao_macro"].isin(macros_sel)] if macros_sel else ta_f1.iloc[0:0]

    with st.expander("📍 Região Micro", expanded=False):
        micros_all = sorted(ta_f2["regiao_micro"].dropna().unique().tolist())
        micros_sel = checkboxes_dc("Micro", micros_all, prefix="dc_micro")

    ta_f3 = ta_f2[ta_f2["regiao_micro"].isin(micros_sel)] if micros_sel else ta_f2.iloc[0:0]

    with st.expander("🏛️ Estado", expanded=False):
        estados_all = sorted(ta_f3["estado_sigla"].dropna().unique().tolist())
        estados_sel = checkboxes_dc("Estado", estados_all, prefix="dc_estado")

    ta_f4 = ta_f3[ta_f3["estado_sigla"].isin(estados_sel)] if estados_sel else ta_f3.iloc[0:0]

    with st.expander("🏙️ Cidade", expanded=False):
        cidades_all = sorted(ta_f4["cidade_nome"].dropna().unique().tolist())
        cidades_sel = checkboxes_dc("Cidade", cidades_all, prefix="dc_cidade")

    ta_f5 = ta_f4[ta_f4["cidade_nome"].isin(cidades_sel)] if cidades_sel else ta_f4.iloc[0:0]

    with st.expander("🏡 Fazenda", expanded=False):
        fazendas_all = sorted(ta_f5["nomeFazenda"].dropna().unique().tolist())
        fazendas_sel = checkboxes_dc("Fazenda", fazendas_all, prefix="dc_fazenda")

    ta_f6 = ta_f5[ta_f5["nomeFazenda"].isin(fazendas_sel)] if fazendas_sel else ta_f5.iloc[0:0]

    with st.expander("👤 Responsável", expanded=False):
        resps_all = sorted(ta_f6["nomeResponsavel"].dropna().unique().tolist())
        resps_sel = checkboxes_dc("Resp", resps_all, prefix="dc_resp")

    ta_f7 = ta_f6[ta_f6["nomeResponsavel"].isin(resps_sel)] if resps_sel else ta_f6.iloc[0:0]

    with st.expander("🏷️ Status do Material", expanded=False):
        status_all = sorted(ta_f7["status_material"].dropna().unique().tolist())
        status_sel = checkboxes_dc("Status", status_all, prefix="dc_status")

    ta_f8 = ta_f7[ta_f7["status_material"].isin(status_sel)] if status_sel else ta_f7.iloc[0:0]

    with st.expander("🌱 Cultivar", expanded=False):
        cultivares_all = sorted(ta_f8["dePara"].dropna().unique().tolist())
        cultivares_sel = checkboxes_dc("Cult", cultivares_all, prefix="dc_cult")

    ta_f9 = ta_f8[ta_f8["dePara"].isin(cultivares_sel)] if cultivares_sel else ta_f8.iloc[0:0]

    with st.expander("🎯 Grupo de Maturidade", expanded=False):
        gm_min = float(round(ta_f9["gm_cat"].min(), 1)) if not ta_f9.empty else 5.0
        gm_max = float(round(ta_f9["gm_cat"].max(), 1)) if not ta_f9.empty else 9.0
        if gm_min >= gm_max:
            gm_max = round(gm_min + 0.1, 1)
        gm_range = st.slider("GM", min_value=gm_min, max_value=gm_max,
                             value=(gm_min, gm_max), step=0.1, format="%.1f", key="dc_gm")

    ta_filtrado = ta_f9[ta_f9["gm_cat"].between(gm_range[0], gm_range[1])]

# ── Aviso se sem dados ────────────────────────────────────────────────────────
if ta_filtrado.empty:
    st.warning("⚠️ Nenhum dado para os filtros selecionados.")
    st.stop()

# ── Contexto de filtros ───────────────────────────────────────────────────────
_all_safras  = sorted(ta_raw["safra"].dropna().unique().tolist())
_all_macros  = sorted(ta_raw["regiao_macro"].dropna().unique().tolist())
_all_micros  = sorted(ta_raw["regiao_micro"].dropna().unique().tolist())

filtros_ativos = []
if safras_sel and set(safras_sel) != set(_all_safras):
    filtros_ativos.append(" / ".join(str(s) for s in safras_sel))
if macros_sel and set(macros_sel) != set(_all_macros):
    filtros_ativos.append("Macro: " + ", ".join(macros_sel))
if micros_sel and set(micros_sel) != set(_all_micros):
    filtros_ativos.append("Micro: " + ", ".join(micros_sel))
if estados_sel and set(estados_sel) != set(sorted(ta_raw["estado_sigla"].dropna().unique().tolist())):
    filtros_ativos.append(", ".join(estados_sel))

n_locais_ctx  = ta_filtrado["cod_fazenda"].nunique()
n_cidades_ctx = ta_filtrado["cidade_nome"].nunique()
filtros_ativos.append(f"{n_cidades_ctx} cidades · {n_locais_ctx} locais")
contexto_str = "  ·  ".join(filtros_ativos) if filtros_ativos else "Todos os dados"


# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 1 — AUDITORIA
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Auditoria",
    "Quais são os dados por ensaio?",
    "Visão individual de cada observação com dados de produção e sanidade. Use para conferência antes da análise.",
)

col_map_audit = {
    "safra":               "Safra",
    "cod_fazenda":         "Cód. Local",
    "nomeFazenda":         "Fazenda",
    "cidade_nome":         "Cidade",
    "estado_sigla":        "Estado",
    "regiao_macro":        "Região Macro",
    "regiao_micro":        "Região Micro",
    "nomeResponsavel":     "Responsável",
    "dePara":              "Cultivar",
    "status_material":     "Status",
    "gm_cat":              "GM",
    "kg_ha":               "kg/ha",
    "sc_ha":               "sc/ha",
    "umidadeParcela":      "Umidade (%)",
    "pop_plantasFinal_ha": "Pop. Final (pl/ha)",
    # ── Doenças ──────────────────────────────────────────────────────────────
    "notaPhytophthora":    "Nota Phytophthora",
    "notaAnomalia":        "Nota Anomalia",
    "notaOidio":           "Nota Oídio",
    "notaManchaParda":     "Nota M. Parda",
    "notaManchaAlvo":      "Nota M. Alvo",
    "notaManchaOlhoRa":    "Nota M. Olho-Rã",
    "notaCercospora":      "Nota Cercospora",
    "notaAntracnose":      "Nota Antracnose",
    "notaDFC":             "Nota DFC",
    "notaMedia_av2":       "Nota Média Av2",
    "class_notaPhytophthora":  "Class. Phytophthora",
    "class_notaAnomalia":      "Class. Anomalia",
    "class_notaOidio":         "Class. Oídio",
    "class_notaManchaParda":   "Class. M. Parda",
    "class_notaManchaAlvo":    "Class. M. Alvo",
    "class_notaManchaOlhoRa":  "Class. M. Olho-Rã",
    "class_notaCercospora":    "Class. Cercospora",
    "class_notaAntracnose":    "Class. Antracnose",
    "class_notaDFC":           "Class. DFC",
}

cols_disp = [c for c in col_map_audit.keys() if c in ta_filtrado.columns]
df_audit  = ta_filtrado[cols_disp].copy()

# Arredondar notas para 1 casa decimal
for col in df_audit.columns:
    if col.startswith("nota") or col == "notaMedia_av2":
        df_audit[col] = pd.to_numeric(df_audit[col], errors="coerce")
        df_audit[col] = df_audit[col].where(df_audit[col] > 0).round(1)

# Recalcular incidência: nota 1–5 = presente, nota 6–9 = ausente, 0/NaN = excluído
for doenca, cols_d in DOENCAS.items():
    col_nota = cols_d["nota"]
    col_inc  = f"inc_{col_nota}"  # nome da coluna recalculada
    if col_nota in df_audit.columns:
        nota_num = pd.to_numeric(df_audit[col_nota], errors="coerce")
        df_audit[col_inc] = nota_num.apply(
            lambda n: 1 if (pd.notna(n) and 1 <= n <= 5) else (0 if (pd.notna(n) and n >= 6) else None)
        )

# Renomear para exibição
rename_inc = {f"inc_{cols_d['nota']}": f"Inc. {doenca}" for doenca, cols_d in DOENCAS.items()
              if f"inc_{cols_d['nota']}" in df_audit.columns}
df_audit = df_audit.rename(columns={**col_map_audit, **rename_inc})

ag_table(df_audit, height=min(400, 36 + 32 * len(df_audit) + 20))
exportar_excel(df_audit, nome_arquivo="auditoria_doencas.xlsx",
               label="⬇️ Exportar Auditoria", key="exp_audit_dc")

st.divider()


# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 2 — DESCRITIVA POR DOENÇA
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Resumo por Doença",
    "Como cada cultivar se comporta sanitariamente?",
    contexto_str,
)

with st.popover("ℹ️ Como interpretar", use_container_width=False):
    st.markdown("""
Nota mais frequente (moda), incidência e classificação de cada doença por cultivar — sobre todos os locais nos filtros ativos, excluindo notas 0.

**Colunas**
- **Nota** → nota mais frequente nos locais (escala 1–9, sendo 9 o melhor)
- **Inc.** → % de locais onde a doença foi detectada (notas 1–5)
- **Classe** → derivada da nota mais frequente
""")
    cols_leg = st.columns(len(COR_CLASS))
    for i, (cls, cor) in enumerate(COR_CLASS.items()):
        cols_leg[i].markdown(
            f'<div style="background:{cor};border-radius:6px;padding:8px;text-align:center;'
            f'font-size:12px;font-weight:600;">{LABEL_CLASS[cls]}</div>',
            unsafe_allow_html=True,
        )
    st.markdown("""
Compare os lançamentos (STINE / LINHAGEM) com os checks para avaliar o diferencial sanitário.
""")

# ── Montar tabela resumo por cultivar × doença ────────────────────────────────
resumo_rows = []
for cultivar, grp in ta_filtrado.groupby("dePara", dropna=True):
    row = {
        "Cultivar":       cultivar,
        "Status":         grp["status_material"].mode()[0] if not grp["status_material"].mode().empty else "",
        "GM":             grp["gm_cat"].mode()[0] if not grp["gm_cat"].mode().empty else "",
        "Locais":         grp["cod_fazenda"].nunique(),
        "kg/ha":          round(grp["kg_ha"].dropna().mean(), 1) if "kg_ha" in grp.columns else None,
        "sc/ha":          round(grp["sc_ha"].dropna().mean(), 1) if "sc_ha" in grp.columns else None,
    }
    for doenca, cols in DOENCAS.items():
        col_nota  = cols["nota"]

        if col_nota in grp.columns:
            s = pd.to_numeric(grp[col_nota], errors="coerce")
            s_val = s[s > 0]
            if len(s_val) > 0:
                moda = s_val.mode()[0]
                inc  = round((s_val.between(1, 5)).sum() / len(s_val) * 100, 1)
                row[f"Nota {doenca}"]   = round(moda, 1)
                row[f"Inc. {doenca} (%)"] = inc
                row[f"Classe {doenca}"] = nota_para_classe(moda)
            else:
                row[f"Nota {doenca}"]     = None
                row[f"Inc. {doenca} (%)"] = None
                row[f"Classe {doenca}"]   = None
        else:
            row[f"Nota {doenca}"]     = None
            row[f"Inc. {doenca} (%)"] = None
            row[f"Classe {doenca}"]   = None

    resumo_rows.append(row)

df_resumo = (
    pd.DataFrame(resumo_rows)
    .sort_values(["Status", "sc/ha"], ascending=[True, False])
    .reset_index(drop=True)
)

# ── AgGrid com cores nas colunas de Classe ────────────────────────────────────
def ag_table_doencas(df, height=400):
    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_default_column(
        resizable=True,
        sortable=True,
        filter=True,
        suppressMenu=False,
        menuTabs=["generalMenuTab", "filterMenuTab", "columnsMenuTab"],
        cellStyle={"fontSize": "13px", "color": "#000000", "fontFamily": "Helvetica Neue, sans-serif"},
    )
    gb.configure_grid_options(
        headerHeight=36,
        rowHeight=32,
        domLayout="normal",
        suppressMenuHide=True,
        suppressColumnVirtualisation=True,
        enableRangeSelection=True,
    )

    # Colorir colunas de Classe
    js_class_color = JsCode("""
    function(params) {
        var v = params.value;
        if (v === 'AS') return { background: '#8B0000', color: '#FFFFFF', fontWeight: '700' };
        if (v === 'S')  return { background: '#E63946', color: '#FFFFFF', fontWeight: '700' };
        if (v === 'MT') return { background: '#FFD600', color: '#1A1A1A', fontWeight: '700' };
        if (v === 'T')  return { background: '#70C96E', color: '#1A1A1A', fontWeight: '700' };
        if (v === 'R')  return { background: '#1E7A34', color: '#FFFFFF', fontWeight: '700' };
        return {};
    }
    """)

    _center_style = {"fontSize": "13px", "color": "#000000", "fontFamily": "Helvetica Neue, sans-serif", "textAlign": "center"}

    for col in df.columns:
        if col.startswith("Classe "):
            gb.configure_column(col, cellStyle=js_class_color, width=90, headerClass="ag-header-center")
        elif col.startswith("Nota "):
            gb.configure_column(col, width=80, cellStyle=_center_style, headerClass="ag-header-center")
        elif col.startswith("Inc. "):
            gb.configure_column(col, width=85, cellStyle=_center_style, headerClass="ag-header-center")

    gb.configure_column("Cultivar", pinned="left", width=170)
    gb.configure_column("Status",   width=90)
    gb.configure_column("GM",       width=60)
    gb.configure_column("Locais",   width=70)
    gb.configure_column("kg/ha",    width=80)
    gb.configure_column("sc/ha",    width=80)

    go = gb.build()
    go["defaultColDef"]["headerClass"] = "ag-header-black"
    go["onFirstDataRendered"] = JsCode("function(params) { params.api.sizeColumnsToFit(); }")

    AgGrid(
        df,
        gridOptions=go,
        height=height,
        update_mode=GridUpdateMode.NO_UPDATE,
        fit_columns_on_grid_load=False,
        allow_unsafe_jscode=True,
        enable_enterprise_modules=True,
        custom_css={
            ".ag-header":                      {"background-color": "#4A4A4A !important"},
            ".ag-header-row":                  {"background-color": "#4A4A4A !important"},
            ".ag-header-cell":                 {"background-color": "#4A4A4A !important"},
            ".ag-header-cell-label":           {"color": "#FFFFFF !important", "font-weight": "700"},
            ".ag-header-cell-text":            {"color": "#FFFFFF !important", "font-size": "13px !important", "font-weight": "700 !important"},
            ".ag-icon":                        {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-header-icon":                 {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-header-cell-menu-button":     {"opacity": "1 !important", "visibility": "visible !important"},
            ".ag-header-cell-menu-button span":{"color": "#FFFFFF !important"},
            ".ag-icon-menu":                   {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-icon-filter":                 {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-row":                         {"font-size": "13px !important"},
            ".ag-header-center .ag-header-cell-label": {"justify-content": "center !important"},
        },
        theme="streamlit",
        use_container_width=True,
    )

ag_table_doencas(df_resumo, height=min(680, 36 + 32 * len(df_resumo) + 20))
exportar_excel(df_resumo, nome_arquivo="resumo_doencas.xlsx",
               label="⬇️ Exportar Resumo de Doenças", key="exp_resumo_dc")

st.caption(
    "ℹ️ **Nota** = nota mais frequente (moda) nos locais avaliados, escala 1–9 (9 = melhor sanidade). "
    "**Inc.** = % de locais com doença detectada (notas 1–5). "
    "**Classe** = derivada da Nota · AS (pior) → S → MT → T → R (melhor)."
)

st.divider()


# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 3 — APRESENTAÇÃO
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Apresentação",
    "Visão consolidada de sanidade por cultivar",
    contexto_str,
)

SIGLAS = {
    "Phytophthora":   "PHY",
    "Anomalia":       "ANO",
    "Oídio":          "OID",
    "Mancha Parda":   "MPA",
    "Mancha Alvo":    "MAL",
    "Mancha Olho-Rã": "OLR",
    "Cercospora":     "CER",
    "Antracnose":     "ANT",
    "DFC":            "DFC",
}

def moda_para_classe(moda):
    if moda is None or (isinstance(moda, float) and np.isnan(moda)):
        return "—"
    if moda <= 2: return "AS"
    if moda <= 4: return "S"
    if moda <= 6: return "MT"
    if moda <= 8: return "T"
    return "R"

doencas_disp = [
    d for d, cols in DOENCAS.items()
    if cols["nota"] in ta_filtrado.columns
    and pd.to_numeric(ta_filtrado[cols["nota"]], errors="coerce").gt(0).any()
]

with st.popover("ℹ️ Como interpretar esta tabela", use_container_width=False):
    st.markdown("""
Para cada doença selecionada, três indicadores por cultivar — calculados sobre todos os locais nos filtros ativos, excluindo notas 0.

**Colunas por doença**
- **Nota** → nota mais frequente (moda), escala 1–9 (9 = melhor sanidade)
- **%** → % de locais onde a doença foi detectada (notas 1–5)
- **Classe** → derivada da nota: AS · S · MT · T · R

**Cores das linhas** → status do material (CHECK, STINE, LINHAGEM, DP2)

**Alertas automáticos** → sinalizados cultivares com classe MT, S ou AS.
""")

with st.expander("🌿 Selecionar doenças", expanded=True):
    _cols = st.columns(3)
    doencas_sel = []
    for i, d in enumerate(doencas_disp):
        default_check = i < 5
        if _cols[i % 3].checkbox(f"{SIGLAS.get(d, d)} — {d}", value=default_check, key=f"dc_chk_{d}"):
            doencas_sel.append(d)

if not doencas_sel:
    st.info("Selecione ao menos uma doença para gerar a tabela.")
else:
    apres_rows = []
    for cultivar, grp in ta_filtrado.groupby("dePara", dropna=True):
        row = {
            "Cultivar":        cultivar,
            "status_material": grp["status_material"].mode()[0] if not grp["status_material"].mode().empty else "",
            "GM":              grp["gm_cat"].mode()[0] if not grp["gm_cat"].mode().empty else "",
            "Locais":          grp["cod_fazenda"].nunique(),
            "sc/ha":           round(grp["sc_ha"].dropna().mean(), 1) if "sc_ha" in grp.columns else None,
            "kg/ha":           round(grp["kg_ha"].dropna().mean(), 1) if "kg_ha" in grp.columns else None,
        }
        for doenca in doencas_sel:
            col_nota = DOENCAS[doenca]["nota"]
            sig      = SIGLAS.get(doenca, doenca)
            if col_nota in grp.columns:
                s     = pd.to_numeric(grp[col_nota], errors="coerce")
                s_val = s[s > 0]
                if len(s_val) > 0:
                    moda_val = float(s_val.mode()[0])
                    inc_pct  = round((s_val.between(1, 5)).sum() / len(s_val) * 100, 1)
                    row[f"{sig}_tipica"]   = round(moda_val, 1)
                    row[f"{sig}_pct"]    = inc_pct
                    row[f"{sig}_classe"] = moda_para_classe(moda_val)
                else:
                    row[f"{sig}_tipica"]   = None
                    row[f"{sig}_pct"]    = None
                    row[f"{sig}_classe"] = None
            else:
                row[f"{sig}_tipica"]   = None
                row[f"{sig}_pct"]    = None
                row[f"{sig}_classe"] = None
        apres_rows.append(row)

    df_apres_dc = (
        pd.DataFrame(apres_rows)
        .sort_values("sc/ha", ascending=False)
        .reset_index(drop=True)
    )

    COR_STATUS_AP = {
        "CHECK":    "#F4B184",
        "STINE":    "#2976B6",
        "LINHAGEM": "#00FF01",
        "DP2":      "#C4DFB4",
    }
    COR_TEXTO_AP = {
        "CHECK":    "#1A1A1A",
        "STINE":    "#FFFFFF",
        "LINHAGEM": "#1A1A1A",
        "DP2":      "#1A1A1A",
        "":         "#000000",
    }

    thead_row1 = (
        '<th rowspan="2" style="text-align:left;">Cultivar</th>'
        '<th rowspan="2">GM</th>'
        '<th rowspan="2">Locais</th>'
        '<th rowspan="2">sc/ha</th>'
        '<th rowspan="2">kg/ha</th>'
    )
    thead_row2 = ""
    for doenca in doencas_sel:
        thead_row1 += f'<th colspan="3" style="text-align:center;border-left:2px solid #AAAAAA;">{doenca}</th>'
        thead_row2 += '<th style="border-left:2px solid #AAAAAA;text-align:center;">Nota</th><th>%</th><th>Classe</th>'

    html = """
<style>
.tb-dc2 { width:100%; border-collapse:collapse; font-size:14px; font-family:'Helvetica Neue',sans-serif; }
.tb-dc2 th {
    background:#E8E8E8; color:#1A1A1A !important; padding:7px 8px;
    text-align:center; border:1px solid #BBBBBB; white-space:nowrap; font-weight:700;
}
.tb-dc2 th:first-child { text-align:left; }
.tb-dc2 td { padding:6px 8px; border:1px solid #ddd; text-align:center; white-space:nowrap; }
.tb-dc2 td:first-child { text-align:left; font-weight:600; }
.tb-dc2 td[data-fg="white"] { color: #FFFFFF !important; }
.tb-dc2 td[data-fg="dark"]  { color: #1A1A1A !important; }
.tb-dc2 td.sep { border-left:2px solid #AAAAAA !important; }
.tb-dc2 tr.rodape td {
    background:#D9D9D9 !important; font-weight:700; border-top:2px solid #888;
    color:#000000 !important;
}
</style>
""" + f"""
<table class="tb-dc2">
<thead>
  <tr>{thead_row1}</tr>
  <tr>{thead_row2}</tr>
</thead>
<tbody>
"""

    def fmt(v):
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return "—"
        return str(v)

    alertas = []

    for _, row in df_apres_dc.iterrows():
        status   = row.get("status_material", "")
        bg       = COR_STATUS_AP.get(status, "#FFFFFF")
        fg       = COR_TEXTO_AP.get(status, "#000000")
        data_fg  = "white" if fg == "#FFFFFF" else "dark"
        cultivar = row["Cultivar"]

        html += "<tr>"
        html += f'<td data-fg="{data_fg}" style="background:{bg};">{cultivar}</td>'
        html += f'<td data-fg="{data_fg}" style="background:{bg};">{fmt(row.get("GM"))}</td>'
        html += f'<td data-fg="{data_fg}" style="background:{bg};">{fmt(row.get("Locais"))}</td>'
        html += f'<td data-fg="{data_fg}" style="background:{bg};">{fmt(row.get("sc/ha"))}</td>'
        html += f'<td data-fg="{data_fg}" style="background:{bg};">{fmt(row.get("kg/ha"))}</td>'

        for doenca in doencas_sel:
            sig    = SIGLAS.get(doenca, doenca)
            moda   = row.get(f"{sig}_tipica")
            pct    = row.get(f"{sig}_pct")
            classe = row.get(f"{sig}_classe") or "—"

            if classe in ("AS", "S", "MT"):
                pct_str = f"{pct:.0f}%" if pct is not None and not (isinstance(pct, float) and np.isnan(pct)) else "—"
                alertas.append((cultivar, doenca, pct_str, classe))

            html += f'<td class="sep" data-fg="{data_fg}" style="background:{bg};">{fmt(moda)}</td>'
            html += f'<td data-fg="{data_fg}" style="background:{bg};">{fmt(pct)}</td>'
            html += f'<td data-fg="{data_fg}" style="background:{bg};">{classe}</td>'
        html += "</tr>"

    html += "</tbody></table>"

    st.markdown(html, unsafe_allow_html=True)
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

    # Alertas automáticos
    if alertas:
        seen = {}
        for cultivar, doenca, pct_str, classe in alertas:
            chave = (cultivar, doenca)
            seen[chave] = (pct_str, classe)
        linhas = [f"**{cv}** → {d}: incidência {p} · classe {cl}" for (cv, d), (p, cl) in seen.items()]
        st.warning("⚠️ **Atenção — cultivares com Moderadamente Tolerante (MT), Susceptível (S) ou Altamente Susceptível (AS):**\n\n" + "  \n".join(linhas))

    # ── Rodapé de locais por doença ───────────────────────────────────────────
    n_locais_total_ap = ta_filtrado["cod_fazenda"].nunique()
    _linhas_rodape = []
    for doenca in doencas_sel:
        col_nota_ap = DOENCAS[doenca]["nota"]
        if col_nota_ap not in ta_filtrado.columns:
            continue
        s_ap = pd.to_numeric(ta_filtrado[col_nota_ap], errors="coerce")
        n_loc_ap = ta_filtrado[s_ap > 0]["cod_fazenda"].nunique()
        if n_loc_ap < n_locais_total_ap:
            _linhas_rodape.append(
                f"**{SIGLAS.get(doenca, doenca)}**: {n_loc_ap} de {n_locais_total_ap} locais com nota registrada"
            )
    if _linhas_rodape:
        st.caption("ℹ️ Locais avaliados por doença (demais locais sem nota registrada): " + " · ".join(_linhas_rodape))

    # ── Export colorido ───────────────────────────────────────────────────────
    if st.button("⬇️ Exportar Excel com formatação", type="primary", key="exp_apres_dc"):
        import io, openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        COR_STATUS_XL = {
            "CHECK":    "F4B184",
            "STINE":    "2976B6",
            "LINHAGEM": "00FF01",
            "DP2":      "C4DFB4",
        }
        COR_TEXTO_XL = {
            "CHECK":    "1A1A1A",
            "STINE":    "FFFFFF",
            "LINHAGEM": "1A1A1A",
            "DP2":      "1A1A1A",
            "":         "000000",
        }

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tabela de Apresentação"

        thin   = Side(style="thin",   color="CCCCCC")
        medium = Side(style="medium", color="AAAAAA")
        BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ── Linha 1 — cabeçalho grupos ────────────────────────────────────────
        base_cols  = ["Cultivar", "GM", "Locais", "sc/ha", "kg/ha"]
        n_base     = len(base_cols)
        doenca_cols = []  # lista de (doenca, [col_keys])
        for doenca in doencas_sel:
            sig = SIGLAS.get(doenca, doenca)
            doenca_cols.append((doenca, [f"{sig}_tipica", f"{sig}_pct", f"{sig}_classe"]))

        # Merge células de base_cols (rowspan 2 = merge rows 1-2)
        for ci, col in enumerate(base_cols, 1):
            ws.merge_cells(start_row=1, start_column=ci, end_row=2, end_column=ci)
            cell = ws.cell(row=1, column=ci, value=col)
            cell.font      = Font(bold=True, name="Arial", size=10, color="1A1A1A")
            cell.fill      = PatternFill("solid", start_color="E8E8E8")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = BORDER
            ws.column_dimensions[get_column_letter(ci)].width = max(14, len(col) + 2)

        # Cabeçalho de grupos de doença (row 1, colspan 3)
        ci_start = n_base + 1
        for doenca, sub_cols in doenca_cols:
            ci_end = ci_start + len(sub_cols) - 1
            ws.merge_cells(start_row=1, start_column=ci_start, end_row=1, end_column=ci_end)
            cell = ws.cell(row=1, column=ci_start, value=doenca)
            cell.font      = Font(bold=True, name="Arial", size=10, color="1A1A1A")
            cell.fill      = PatternFill("solid", start_color="E8E8E8")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = Border(left=medium, right=medium, top=thin, bottom=thin)
            ci_start = ci_end + 1

        # ── Linha 2 — sub-cabeçalhos por doença ──────────────────────────────
        SUB_LABELS = {"_tipica": "Nota", "_pct": "Inc. (%)", "_classe": "Classe"}
        ci_start = n_base + 1
        for doenca, sub_cols in doenca_cols:
            for j, sub_key in enumerate(sub_cols):
                label = SUB_LABELS.get(sub_key.split("_", 1)[1] if "_" in sub_key else sub_key,
                                       sub_key.split("_")[-1].capitalize())
                # Pega apenas o sufixo (_tipica, _pct, _classe)
                for suffix, lbl in [("_tipica", "Nota"), ("_pct", "Inc. (%)"), ("_classe", "Classe")]:
                    if sub_key.endswith(suffix):
                        label = lbl
                        break
                cell = ws.cell(row=2, column=ci_start, value=label)
                cell.font      = Font(bold=True, name="Arial", size=10, color="1A1A1A")
                cell.fill      = PatternFill("solid", start_color="E8E8E8")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                left_b = medium if j == 0 else thin
                cell.border = Border(left=left_b, right=thin, top=thin, bottom=thin)
                ws.column_dimensions[get_column_letter(ci_start)].width = 14
                ci_start += 1

        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 22

        # ── Dados ─────────────────────────────────────────────────────────────
        all_col_keys = base_cols + [sk for _, subs in doenca_cols for sk in subs]

        for ri, row in df_apres_dc.iterrows():
            status = row.get("status_material", "")
            bg_hex = COR_STATUS_XL.get(status, "FFFFFF")
            fg_hex = COR_TEXTO_XL.get(status, "000000")
            xl_row = ri + 3  # dados começam na linha 3

            # base cols
            for ci, col in enumerate(base_cols, 1):
                val = row.get(col)
                if isinstance(val, float) and np.isnan(val): val = None
                cell = ws.cell(row=xl_row, column=ci, value=val)
                cell.font      = Font(name="Arial", size=10, color=fg_hex)
                cell.fill      = PatternFill("solid", start_color=bg_hex)
                cell.alignment = Alignment(horizontal="left" if ci == 1 else "center", vertical="center")
                cell.border    = BORDER

            # colunas de doença
            ci_start = n_base + 1
            COR_CLASS_XL = {
                "AS": ("8B0000", "FFFFFF"),
                "S":  ("E63946", "FFFFFF"),
                "MT": ("FFD600", "1A1A1A"),
                "T":  ("70C96E", "1A1A1A"),
                "R":  ("1E7A34", "FFFFFF"),
            }
            for doenca, sub_cols in doenca_cols:
                for j, sub_key in enumerate(sub_cols):
                    val = row.get(sub_key)
                    if val is None or (isinstance(val, float) and np.isnan(val)): val = None
                    is_classe = sub_key.endswith("_classe")
                    if is_classe and val and val in COR_CLASS_XL:
                        cell_bg, cell_fg = COR_CLASS_XL[val]
                    else:
                        cell_bg, cell_fg = bg_hex, fg_hex
                    cell = ws.cell(row=xl_row, column=ci_start, value=val)
                    cell.font      = Font(name="Arial", size=10, color=cell_fg, bold=is_classe)
                    cell.fill      = PatternFill("solid", start_color=cell_bg)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    left_b = medium if j == 0 else thin
                    cell.border = Border(left=left_b, right=thin, top=thin, bottom=thin)
                    ci_start += 1

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        st.download_button(
            label="📥 Baixar Excel",
            data=buf,
            file_name="apresentacao_doencas.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="exp_apres_dc_dl",
        )

st.divider()

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 3B — RANKING POR QUARTIL SANITÁRIO
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Ranking por Quartil Sanitário",
    "Quais cultivares se mantêm no topo da sanidade nos melhores locais?",
    contexto_str,
)

with st.popover("ℹ️ Como interpretar", use_container_width=False):
    st.markdown("""
**Top 25% do grupo (%)**
Em cada local, os cultivares são comparados entre si. Os que ficam entre os melhores 25% daquele ambiente "entram no top". Esta coluna mostra em quantos % dos locais o cultivar conseguiu isso.

Um cultivar com 80% significa que em 8 de cada 10 locais ele foi melhor que a maioria — mesmo nos locais onde a doença foi severa para todos.

*(Tecnicamente: cultivar com nota ≥ 3º quartil — Q3 — do grupo naquele local)*

---

**Nota**
A nota do meio — se você ordenar todas as notas do cultivar do menor para o maior, é o valor central. Representa o comportamento habitual do cultivar, sem ser puxada por um local excepcionalmente bom ou ruim.

*(Tecnicamente: mediana das notas médias por local)*

---

**Como ler as duas colunas juntas**

| Top 25% | Nota típica | O que significa |
|---|---|---|
| Alto | Alta | Consistentemente bom — destaque real |
| Alto | Baixa | Irregular — brilha em alguns locais, afunda em outros |
| Baixo | Alta | Mediano em todo lugar — nunca decepciona, nunca surpreende |
| Baixo | Baixa | Perfil sanitário fraco na maioria dos ambientes |

---

**sc/ha médio** → produtividade média nos locais avaliados — para cruzar sanidade com produção.
""")

if not doencas_sel:
    st.info("Selecione ao menos uma doença na Seção 3 para exibir o ranking.")
elif "sc_ha" not in ta_filtrado.columns:
    st.info("Coluna sc_ha não encontrada na base — ranking indisponível.")
else:
    doenca_rank = st.selectbox(
        "Doença para o ranking",
        options=doencas_sel,
        format_func=lambda d: f"{SIGLAS.get(d, d)} — {d}",
        key="dc_doenca_rank",
    )

    col_nota_rank = DOENCAS[doenca_rank]["nota"]

    if col_nota_rank not in ta_filtrado.columns:
        st.warning("Coluna de nota não encontrada para esta doença.")
    else:
        df_rank = ta_filtrado[["dePara", "status_material", "cod_fazenda", col_nota_rank, "sc_ha"]].copy()
        df_rank[col_nota_rank] = pd.to_numeric(df_rank[col_nota_rank], errors="coerce")
        df_rank["sc_ha"]       = pd.to_numeric(df_rank["sc_ha"],       errors="coerce")
        df_rank = df_rank[df_rank[col_nota_rank] > 0].dropna(subset=[col_nota_rank])

        # Q3 por local
        q3_por_local = (
            df_rank.groupby("cod_fazenda")[col_nota_rank]
            .quantile(0.75)
            .rename("_q3")
        )
        df_rank = df_rank.merge(q3_por_local, on="cod_fazenda", how="left")
        df_rank["_top"] = (df_rank[col_nota_rank] >= df_rank["_q3"]).astype(int)

        # Média por cultivar × local antes de agregar
        df_rank_agg = (
            df_rank.groupby(["dePara", "status_material", "cod_fazenda"])
            .agg(
                nota_media   = (col_nota_rank, "mean"),
                _top_local   = ("_top", "max"),   # 1 se em alguma rep ficou no top
                sc_ha_media  = ("sc_ha", "mean"),
            )
            .reset_index()
        )

        # Ranking por cultivar
        rank_rows = []
        for cultivar, grp in df_rank_agg.groupby("dePara"):
            status   = grp["status_material"].mode()[0] if not grp["status_material"].mode().empty else ""
            n_locais = len(grp)
            n_top    = int(grp["_top_local"].sum())
            pct_top  = round(n_top / n_locais * 100, 1) if n_locais > 0 else 0
            nota_med = round(grp["nota_media"].mean(), 2)
            sc_med   = round(grp["sc_ha_media"].mean(), 1) if grp["sc_ha_media"].notna().any() else None
            rank_rows.append({
                "Cultivar":             cultivar,
                "Status":               status,
                "Locais":               n_locais,
                "Top 25% do grupo (%)": pct_top,
                "Nota":                 round(grp["nota_media"].median(), 2),
                "sc/ha médio":          sc_med,
            })

        df_ranking = (
            pd.DataFrame(rank_rows)
            .sort_values("Top 25% do grupo (%)", ascending=False)
            .reset_index(drop=True)
        )

        gb_rk = GridOptionsBuilder.from_dataframe(df_ranking)
        gb_rk.configure_default_column(
            resizable=True, sortable=True, filter=True,
            cellStyle={"fontSize": "13px", "fontFamily": "Helvetica Neue, sans-serif"},
        )
        gb_rk.configure_grid_options(headerHeight=36, rowHeight=32, domLayout="normal")
        gb_rk.configure_column("Cultivar",        pinned="left", width=170)
        gb_rk.configure_column("Status",          width=90)
        gb_rk.configure_column("Locais",          width=80)
        gb_rk.configure_column("Top 25% do grupo (%)", width=155,
            cellStyle=JsCode("""function(p){
                var v = p.value;
                if(v >= 75) return {background:'#1E7A34', color:'#FFFFFF', fontWeight:'700'};
                if(v >= 50) return {background:'#70C96E', color:'#1A1A1A', fontWeight:'700'};
                if(v >= 25) return {background:'#FFD600', color:'#1A1A1A'};
                return {background:'#E63946', color:'#FFFFFF', fontWeight:'700'};
            }"""))
        gb_rk.configure_column("Nota", width=90,
            cellStyle=JsCode("""function(p){
                var v = p.value;
                if(v >= 7) return {background:'#70C96E', color:'#1A1A1A'};
                if(v >= 5) return {background:'#FFD600', color:'#1A1A1A'};
                return {background:'#E63946', color:'#FFFFFF'};
            }"""))
        gb_rk.configure_column("sc/ha médio", width=110)

        go_rk = gb_rk.build()
        go_rk["defaultColDef"]["headerClass"] = "ag-header-black"

        AgGrid(
            df_ranking,
            gridOptions=go_rk,
            height=min(500, 36 + 32 * len(df_ranking) + 20),
            update_mode=GridUpdateMode.NO_UPDATE,
            fit_columns_on_grid_load=False,
            allow_unsafe_jscode=True,
            enable_enterprise_modules=True,
            custom_css={
                ".ag-header":                  {"background-color": "#4A4A4A !important"},
                ".ag-header-row":              {"background-color": "#4A4A4A !important"},
                ".ag-header-cell":             {"background-color": "#4A4A4A !important"},
                ".ag-header-cell-label":       {"color": "#FFFFFF !important", "font-weight": "700"},
                ".ag-header-cell-text":        {"color": "#FFFFFF !important", "font-size": "13px !important", "font-weight": "700 !important"},
                ".ag-icon":                    {"color": "#FFFFFF !important", "opacity": "1 !important"},
                ".ag-header-icon":             {"color": "#FFFFFF !important", "opacity": "1 !important"},
                ".ag-header-cell-menu-button": {"opacity": "1 !important", "visibility": "visible !important"},
                ".ag-icon-menu":               {"color": "#FFFFFF !important", "opacity": "1 !important"},
                ".ag-icon-filter":             {"color": "#FFFFFF !important", "opacity": "1 !important"},
                ".ag-row":                     {"font-size": "13px !important"},
            },
            theme="streamlit",
            use_container_width=True,
        )

        n_locais_total_rk = ta_filtrado["cod_fazenda"].nunique()
        _locais_rk = df_rank_agg["cod_fazenda"].nunique()
        _caption_rk = (
            f"ℹ️ **Top 25% do grupo (%)** = % de locais em que o cultivar ficou entre os 25% melhores do grupo naquele local (nota ≥ Q3). "
            f"Ordenado do maior para o menor. {len(df_ranking)} cultivares avaliados para {doenca_rank}"
        )
        if _locais_rk < n_locais_total_rk:
            _caption_rk += f" ({_locais_rk} de {n_locais_total_rk} locais ativos com nota registrada)."
        else:
            _caption_rk += "."
        st.caption(_caption_rk)

        exportar_excel(df_ranking, nome_arquivo="ranking_quartil_sanitario.xlsx",
                       label="⬇️ Exportar Ranking por Quartil", key="exp_rank_dc")

st.divider()

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 4 — GRÁFICO DE LINHAS POR LOCAL
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Desempenho por Local",
    "Como cada cultivar se comportou em cada local?",
    contexto_str,
)

col_interp, col_dic = st.columns([1, 9])
with col_interp:
    with st.popover("ℹ️ Como interpretar", use_container_width=True):
        st.markdown("""
Cada linha representa um cultivar. O eixo Y mostra a nota da doença **do pior (topo) ao melhor (base)** — quanto mais baixa a linha, melhor a sanidade do cultivar naquele local.

**🎨 Elementos visuais**

- **Linha** → trajetória do cultivar entre os locais avaliados, cor única por cultivar
- **Hover** → ao passar o mouse sobre a linha: cultivar · local · status · nota · incidência % · classe
- **Coluna sombreada em vermelho** → local com maior incidência da doença — o mesmo critério do card de alerta acima do gráfico

**📐 Referência no gráfico**

- **Linha vermelha tracejada (nota 6)** → limite MT/T — acima desta linha o cultivar está em zona de atenção

**📐 Escala da nota**

| Nota | Classe | Significado |
|---|---|---|
| 9 | R | Resistente |
| 7–8 | T | Tolerante |
| 5–6 | MT | Moderadamente Tolerante |
| 3–4 | S | Susceptível |
| 1–2 | AS | Altamente Susceptível |

**💡 Dica de leitura**

Compare a altura das linhas entre cultivares no mesmo local. Cultivares com linha consistentemente próxima à base têm melhor perfil sanitário naquela doença.

**⚠️ Card de alerta — como o percentual é calculado**

O card aparece quando um local tem **incidência média > 50%**. O cálculo é:

```
% incidência do local = registros com nota 1–5 (todos os cultivares)
                        ─────────────────────────────────────────────  × 100
                        total de registros com nota > 0 naquele local
```

**Importante:** é a incidência **do ambiente** (local), não de um cultivar específico. Todos os cultivares avaliados naquele local entram no denominador — o número reflete a pressão da doença no ambiente independentemente do material.

Exemplo: se num local foram feitas 15 avaliações (de cultivares diferentes) e 10 delas tiveram nota entre 1 e 5 → incidência do local = **67%**.
""")

with col_dic:
    _df_dic = (
        ta_filtrado[["cod_fazenda", "nomeFazenda", "cidade_nome", "estado_sigla"]]
        .drop_duplicates()
        .sort_values(["estado_sigla", "cidade_nome", "cod_fazenda"])
        .rename(columns={
            "cod_fazenda":  "Código",
            "nomeFazenda":  "Local",
            "cidade_nome":  "Cidade",
            "estado_sigla": "Estado",
        })
        .reset_index(drop=True)
    )
    with st.popover(f"📍 Dicionário de locais ({len(_df_dic)})", use_container_width=False):
        st.markdown("Referência dos códigos exibidos no eixo X do gráfico.")
        st.dataframe(_df_dic, hide_index=True, use_container_width=True)

doenca_graf = st.selectbox(
    "Selecione a doença",
    options=doencas_disp,
    format_func=lambda d: f"{SIGLAS.get(d, d)} — {d}",
    key="dc_doenca_graf",
)

col_nota_g = DOENCAS[doenca_graf]["nota"]

# ── Preparar dados ────────────────────────────────────────────────────────────
if col_nota_g not in ta_filtrado.columns:
    st.warning("Coluna de nota não encontrada para esta doença.")
else:
    df_g = ta_filtrado[["dePara", "status_material", "cod_fazenda", "cidade_nome",
                         "estado_sigla", col_nota_g]].copy()
    df_g[col_nota_g] = pd.to_numeric(df_g[col_nota_g], errors="coerce")
    df_g = df_g[df_g[col_nota_g] > 0].dropna(subset=[col_nota_g])

    if df_g.empty:
        st.info("Nenhuma avaliação disponível para esta doença nos filtros ativos.")
    else:
        # Nota por cultivar × local (sem repetição em faixa — mean = valor único)
        df_g_agg = (
            df_g.groupby(["dePara", "status_material", "cod_fazenda", "cidade_nome", "estado_sigla"])
            [col_nota_g].mean().round(1).reset_index()
        )

        # Incidência por cultivar × local (nota 1–5)
        df_inc = df_g.copy()
        df_inc["_inc"] = df_inc[col_nota_g].between(1, 5).astype(int)
        df_inc_agg = (
            df_inc.groupby(["dePara", "cod_fazenda"])
            .agg(_n=("_inc", "count"), _inc=("_inc", "sum"))
            .reset_index()
        )
        df_inc_agg["inc_pct"] = (df_inc_agg["_inc"] / df_inc_agg["_n"] * 100).round(1)
        df_g_agg = df_g_agg.merge(df_inc_agg[["dePara", "cod_fazenda", "inc_pct"]],
                                   on=["dePara", "cod_fazenda"], how="left")

        # Classe por ponto
        df_g_agg["classe"] = df_g_agg[col_nota_g].apply(moda_para_classe)

        # Ordenar locais: estado → cidade → cod_fazenda
        locais_ord = (
            df_g_agg[["cod_fazenda", "cidade_nome", "estado_sigla"]]
            .drop_duplicates()
            .sort_values(["estado_sigla", "cidade_nome", "cod_fazenda"])
            ["cod_fazenda"].tolist()
        )
        df_g_agg["_x"] = df_g_agg["cod_fazenda"].apply(lambda x: locais_ord.index(x))

        # Paleta de cores — uma cor por cultivar
        cultivares_g = sorted(df_g_agg["dePara"].unique().tolist())
        palette = [
            "#1F77B4","#FF7F0E","#2CA02C","#D62728","#9467BD",
            "#8C564B","#E377C2","#7F7F7F","#BCBD22","#17BECF",
            "#AEC7E8","#FFBB78","#98DF8A","#FF9896","#C5B0D5",
            "#C49C94","#F7B6D2","#C7C7C7","#DBDB8D","#9EDAE5",
            "#393B79","#637939","#8C6D31","#843C39","#7B4173",
        ]
        cor_cultivar = {c: palette[i % len(palette)] for i, c in enumerate(cultivares_g)}

        # ── Card de alerta — locais com alta incidência ───────────────────────
        inc_por_local = (
            df_inc_agg.groupby("cod_fazenda")
            .apply(lambda g: round(g["_inc"].sum() / g["_n"].sum() * 100, 1))
            .reset_index()
            .rename(columns={0: "inc_media"})
        )
        locais_alerta = inc_por_local[inc_por_local["inc_media"] > 50]["cod_fazenda"].tolist()
        if locais_alerta:
            # Enriquecer com nome da cidade
            dic_local = (
                ta_filtrado[["cod_fazenda", "cidade_nome", "estado_sigla"]]
                .drop_duplicates()
                .set_index("cod_fazenda")
            )
            detalhes = []
            for local in locais_alerta:
                inc_val = inc_por_local.loc[inc_por_local["cod_fazenda"] == local, "inc_media"].values[0]
                cidade  = dic_local.loc[local, "cidade_nome"] if local in dic_local.index else ""
                estado  = dic_local.loc[local, "estado_sigla"] if local in dic_local.index else ""
                detalhes.append(f"**{local}** ({cidade} — {estado}): {inc_val:.0f}%")
            st.warning(
                f"⚠️ **{len(locais_alerta)} local(is) com incidência média > 50% em {doenca_graf}:**  \n"
                + "  \n".join(detalhes)
            )

        # ── Montar figura ─────────────────────────────────────────────────────
        fig = go_plt.Figure()

        for cultivar in cultivares_g:
            df_c = df_g_agg[df_g_agg["dePara"] == cultivar].sort_values("_x")
            if df_c.empty:
                continue
            status = df_c["status_material"].mode()[0] if not df_c["status_material"].mode().empty else ""
            cor    = cor_cultivar[cultivar]

            # Linha suavizada com hover
            fig.add_trace(go_plt.Scatter(
                x=df_c["cod_fazenda"],
                y=df_c[col_nota_g],
                mode="lines",
                name=cultivar,
                line=dict(color=cor, width=2, shape="spline", smoothing=0.8),
                legendgroup=cultivar,
                showlegend=True,
                customdata=df_c[["status_material", "inc_pct", "classe"]].values,
                hovertemplate=(
                    f"<b>{cultivar}</b> · %{{x}}<br>"
                    f"Status: %{{customdata[0]}}<br>"
                    f"Nota: %{{y}}<br>"
                    f"Incidência: %{{customdata[1]:.0f}}%<br>"
                    f"Classe: %{{customdata[2]}}"
                    "<extra></extra>"
                ),
            ))

        # Linha de referência nota 6
        fig.add_hline(
            y=6, line_dash="dot", line_color="#E74C3C", line_width=1.5,
            annotation_text="Limite MT/T (nota 6)",
            annotation_position="top right",
            annotation_font=dict(size=11, color="#E74C3C"),
        )



        # ── Destaque no local com maior incidência (mesmo critério do card) ────
        if not inc_por_local.empty:
            local_critico = inc_por_local.loc[inc_por_local["inc_media"].idxmax(), "cod_fazenda"]
            inc_critica   = inc_por_local["inc_media"].max()
            if inc_critica > 0 and local_critico in locais_ord:
                dic_local_g = (
                    ta_filtrado[["cod_fazenda", "cidade_nome"]]
                    .drop_duplicates().set_index("cod_fazenda")
                )
                cidade_critica = dic_local_g.loc[local_critico, "cidade_nome"] if local_critico in dic_local_g.index else ""
                fig.add_shape(
                    type="rect",
                    xref="x", yref="paper",
                    x0=locais_ord.index(local_critico) - 0.5,
                    x1=locais_ord.index(local_critico) + 0.5,
                    y0=0, y1=1,
                    fillcolor="rgba(231,76,60,0.07)",
                    line=dict(width=0),
                )
                fig.add_annotation(
                    x=local_critico,
                    y=0.7,
                    yref="y",
                    text=f"⚠️ maior pressão<br>{cidade_critica} · {inc_critica:.0f}% incidência",
                    showarrow=False,
                    xanchor="center",
                    yanchor="top",
                    font=dict(size=10, color="#E74C3C"),
                    bgcolor="rgba(255,255,255,0.85)",
                    bordercolor="#E74C3C",
                    borderwidth=1,
                    borderpad=3,
                )

        n_locais_g  = len(locais_ord)
        altura_graf = max(450, min(700, 350 + n_locais_g * 8))

        fig.update_layout(
            height=altura_graf,
            yaxis=dict(
                title=dict(text="<b>Nota (1 = pior · 9 = melhor)</b>", font=dict(size=14, color="#111111", weight="bold")),
                range=[9.3, 0.7],
                tickvals=list(range(1, 10)),
                tickfont=dict(size=12, color="#111111", weight="bold"),
                gridcolor="#EEEEEE",
                zeroline=False,
            ),
            xaxis=dict(
                title=dict(text="<b>Local (cod_fazenda)</b>", font=dict(size=14, color="#111111", weight="bold")),
                tickangle=-45,
                tickfont=dict(size=11, color="#111111", weight="bold"),
                categoryorder="array",
                categoryarray=locais_ord,
                gridcolor="#EEEEEE",
            ),
            legend=dict(
                title=dict(text="<b>Cultivar</b>", font=dict(size=12, color="#111111")),
                font=dict(size=12, color="#111111", weight="bold"),
                itemsizing="constant",
                bgcolor="rgba(255,255,255,0.85)",
                bordercolor="#DDDDDD",
                borderwidth=1,
            ),
            plot_bgcolor="#FAFAFA",
            paper_bgcolor="#FFFFFF",
            margin=dict(t=40, b=100, l=60, r=20),
            font=dict(family="Helvetica Neue, sans-serif", size=12, color="#111111"),
            hovermode="closest",
        )

        n_locais_total = ta_filtrado["cod_fazenda"].nunique()
        st.plotly_chart(fig, use_container_width=True)
        _caption_loc = (
            f"ℹ️ Eixo Y invertido — pior nota no topo, melhor na base. "
            f"Linha tracejada vermelha = limite nota 6 (zona de atenção acima). "
            f"{len(cultivares_g)} cultivares · {n_locais_g} locais com avaliação de {doenca_graf}"
        )
        if n_locais_g < n_locais_total:
            _caption_loc += f" (de {n_locais_total} locais ativos — os demais não têm nota registrada para esta doença)."
        else:
            _caption_loc += "."
        st.caption(_caption_loc)

st.divider()

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 5 — DELTA VS REFERÊNCIA
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Delta vs Referência",
    "Como cada cultivar se compara ao check ou stine de referência?",
    contexto_str,
)

_col_interp_delta, _col_dic_delta = st.columns([1, 9])

with _col_interp_delta:
    with st.popover("ℹ️ Como interpretar", use_container_width=False):
        st.markdown("""
Para a doença e referência selecionadas, cada barra mostra a diferença de nota entre o cultivar e a referência **no mesmo local**.

**Como ler**
- **Barra verde (positivo)** → cultivar ficou acima da referência naquele local (melhor sanidade)
- **Barra vermelha (negativo)** → cultivar ficou abaixo da referência naquele local (pior sanidade)
- **Apenas locais onde ambos foram avaliados** entram no cálculo

**Delta**
```
delta = nota_cultivar_no_local − nota_referência_no_local
```

Um cultivar consistentemente positivo em vários locais tem vantagem sanitária real sobre a referência.
""")

with _col_dic_delta:
    with st.popover(f"📍 Dicionário de locais ({ta_filtrado['cod_fazenda'].nunique()})", use_container_width=False):
        _df_dic_delta = (
            ta_filtrado[["cod_fazenda", "nomeFazenda", "cidade_nome", "estado_sigla"]]
            .drop_duplicates()
            .sort_values(["estado_sigla", "cidade_nome", "cod_fazenda"])
            .rename(columns={
                "cod_fazenda":  "Código",
                "nomeFazenda":  "Local",
                "cidade_nome":  "Cidade",
                "estado_sigla": "Estado",
            })
            .reset_index(drop=True)
        )
        st.markdown("Referência dos códigos exibidos no eixo X do gráfico.")
        st.dataframe(_df_dic_delta, hide_index=True, use_container_width=True)

# ── Seletores ─────────────────────────────────────────────────────────────────
_col_d_delta, _col_ref_delta = st.columns(2)

with _col_d_delta:
    doenca_delta = st.selectbox(
        "Doença",
        options=doencas_disp,
        format_func=lambda d: f"{SIGLAS.get(d, d)} — {d}",
        key="dc_doenca_delta",
    )

# Cultivares elegíveis como referência: CHECK ou STINE no filtro ativo
_refs_disponiveis = sorted(
    ta_filtrado[ta_filtrado["status_material"].isin(["CHECK", "STINE"])]["dePara"]
    .dropna().unique().tolist()
)

if not _refs_disponiveis:
    st.info("Nenhum cultivar CHECK ou STINE disponível nos filtros ativos para usar como referência.")
else:
    with _col_ref_delta:
        ref_delta = st.selectbox(
            "Referência (CHECK / STINE)",
            options=_refs_disponiveis,
            key="dc_ref_delta",
        )

    col_nota_delta = DOENCAS[doenca_delta]["nota"]

    if col_nota_delta not in ta_filtrado.columns:
        st.warning("Coluna de nota não encontrada para esta doença.")
    else:
        # ── Preparar dados ────────────────────────────────────────────────────
        df_delta = ta_filtrado[["dePara", "status_material", "cod_fazenda",
                                 "cidade_nome", "estado_sigla", col_nota_delta]].copy()
        df_delta[col_nota_delta] = pd.to_numeric(df_delta[col_nota_delta], errors="coerce")
        df_delta = df_delta[df_delta[col_nota_delta] > 0].dropna(subset=[col_nota_delta])

        # Média por cultivar × local
        df_delta_agg = (
            df_delta.groupby(["dePara", "status_material", "cod_fazenda",
                               "cidade_nome", "estado_sigla"])
            [col_nota_delta].mean().round(1).reset_index()
        )

        # Nota da referência por local
        df_ref = df_delta_agg[df_delta_agg["dePara"] == ref_delta][
            ["cod_fazenda", col_nota_delta]
        ].rename(columns={col_nota_delta: "_nota_ref"})

        if df_ref.empty:
            st.info(f"**{ref_delta}** não tem avaliações para {doenca_delta} nos filtros ativos.")
        else:
            # Join — apenas locais onde ambos foram avaliados
            df_delta_agg = df_delta_agg.merge(df_ref, on="cod_fazenda", how="inner")
            df_delta_agg["delta"] = (df_delta_agg[col_nota_delta] - df_delta_agg["_nota_ref"]).round(1)

            # Remover a própria referência das barras
            df_plot = df_delta_agg[df_delta_agg["dePara"] != ref_delta].copy()

            if df_plot.empty:
                st.info("Nenhum outro cultivar foi avaliado nos mesmos locais que a referência.")
            else:
                # Ordenar locais
                locais_delta = (
                    df_plot[["cod_fazenda", "cidade_nome", "estado_sigla"]]
                    .drop_duplicates()
                    .sort_values(["estado_sigla", "cidade_nome", "cod_fazenda"])
                    ["cod_fazenda"].tolist()
                )

                # Ordenar cultivares por delta médio (melhor para pior)
                ordem_cultivares = (
                    df_plot.groupby("dePara")["delta"].mean()
                    .sort_values(ascending=False)
                    .index.tolist()
                )

                # ── Figura ────────────────────────────────────────────────────
                fig_delta = go_plt.Figure()

                for cultivar in ordem_cultivares:
                    df_c = df_plot[df_plot["dePara"] == cultivar].set_index("cod_fazenda")
                    status = df_c["status_material"].mode()[0] if not df_c["status_material"].mode().empty else ""

                    deltas_x = []
                    deltas_y = []
                    cores    = []
                    hover    = []

                    for local in locais_delta:
                        if local in df_c.index:
                            d_val = df_c.loc[local, "delta"]
                            nota_cult = df_c.loc[local, col_nota_delta]
                            nota_ref  = df_c.loc[local, "_nota_ref"]
                            deltas_x.append(local)
                            deltas_y.append(d_val)
                            cores.append("#1E7A34" if d_val >= 0 else "#E63946")
                            hover.append(
                                f"<b>{cultivar}</b> · {local}<br>"
                                f"Nota: {nota_cult} · Ref ({ref_delta}): {nota_ref}<br>"
                                f"Delta: {d_val:+.1f}"
                            )

                    if not deltas_x:
                        continue

                    fig_delta.add_trace(go_plt.Bar(
                        name=cultivar,
                        x=deltas_x,
                        y=deltas_y,
                        marker_color=cores,
                        text=[f"{v:+.1f}" for v in deltas_y],
                        textposition="outside",
                        textfont=dict(size=10, color="#333333"),
                        hovertext=hover,
                        hoverinfo="text",
                        legendgroup=cultivar,
                        showlegend=True,
                        offsetgroup=cultivar,
                    ))

                # Linha zero
                fig_delta.add_hline(
                    y=0,
                    line_color="#333333",
                    line_width=1.5,
                )

                # Faixas de fundo: zona positiva (verde) e negativa (vermelha)
                fig_delta.add_hrect(
                    y0=0, y1=8,
                    fillcolor="rgba(30,122,52,0.04)",
                    line_width=0,
                )
                fig_delta.add_hrect(
                    y0=-8, y1=0,
                    fillcolor="rgba(230,57,70,0.04)",
                    line_width=0,
                )

                n_cultivares_delta = len(ordem_cultivares)
                altura_delta = max(420, min(680, 350 + len(locais_delta) * 12))

                fig_delta.update_layout(
                    height=altura_delta,
                    title=dict(
                        text=f"Delta de nota vs <b>{ref_delta}</b> — {doenca_delta}",
                        font=dict(size=14, color="#111111"),
                        x=0,
                        xanchor="left",
                    ),
                    barmode="group",
                    bargap=0.15,
                    bargroupgap=0.05,
                    xaxis=dict(
                        title=dict(text="<b>Local (cod_fazenda)</b>", font=dict(size=14, color="#111111", weight="bold")),
                        tickangle=-45,
                        tickfont=dict(size=11, color="#111111", weight="bold"),
                        categoryorder="array",
                        categoryarray=locais_delta,
                        gridcolor="#EEEEEE",
                    ),
                    yaxis=dict(
                        title=dict(text="<b>Delta (nota)</b>", font=dict(size=14, color="#111111", weight="bold")),
                        tickfont=dict(size=12, color="#111111", weight="bold"),
                        gridcolor="#EEEEEE",
                        zeroline=False,
                    ),
                    legend=dict(
                        title=dict(text="<b>Cultivar</b>", font=dict(size=12, color="#111111", weight="bold")),
                        font=dict(size=11, color="#111111", weight="bold"),
                        bgcolor="rgba(255,255,255,0.85)",
                        bordercolor="#DDDDDD",
                        borderwidth=1,
                    ),
                    plot_bgcolor="#FAFAFA",
                    paper_bgcolor="#FFFFFF",
                    margin=dict(t=50, b=120, l=60, r=20),
                    font=dict(family="Helvetica Neue, sans-serif", size=12, color="#111111"),
                    hovermode="x unified",
                )

                n_locais_total_delta = ta_filtrado["cod_fazenda"].nunique()
                st.plotly_chart(fig_delta, use_container_width=True)
                _caption_delta = (
                    f"ℹ️ Delta = nota do cultivar − nota de {ref_delta} no mesmo local. "
                    f"Verde = acima da referência · Vermelho = abaixo. "
                    f"{n_cultivares_delta} cultivares · {len(locais_delta)} locais com avaliação conjunta"
                )
                if len(locais_delta) < n_locais_total_delta:
                    _caption_delta += f" (de {n_locais_total_delta} locais ativos — os demais não têm avaliação conjunta para esta doença)."
                else:
                    _caption_delta += "."
                st.caption(_caption_delta)

                # ── Tabela resumo delta ───────────────────────────────────────
                resumo_delta_rows = []
                for cultivar in ordem_cultivares:
                    df_c = df_plot[df_plot["dePara"] == cultivar]
                    status = df_c["status_material"].mode()[0] if not df_c["status_material"].mode().empty else ""
                    n_loc     = len(df_c)
                    n_pos     = int((df_c["delta"] > 0).sum())
                    n_neg     = int((df_c["delta"] < 0).sum())
                    n_emp     = int((df_c["delta"] == 0).sum())
                    delta_med = round(df_c["delta"].mean(), 2)
                    resumo_delta_rows.append({
                        "Cultivar":         cultivar,
                        "Status":           status,
                        "Locais":           n_loc,
                        "▲ Acima ref":      n_pos,
                        "▼ Abaixo ref":     n_neg,
                        "= Empate":         n_emp,
                        "Delta médio":      delta_med,
                    })

                df_resumo_delta = pd.DataFrame(resumo_delta_rows)

                gb_dt = GridOptionsBuilder.from_dataframe(df_resumo_delta)
                gb_dt.configure_default_column(
                    resizable=True, sortable=True, filter=True,
                    cellStyle={"fontSize": "13px", "fontFamily": "Helvetica Neue, sans-serif"},
                )
                gb_dt.configure_grid_options(headerHeight=36, rowHeight=32, domLayout="normal")
                gb_dt.configure_column("Cultivar",     pinned="left", width=170)
                gb_dt.configure_column("Status",       width=90)
                gb_dt.configure_column("Locais",       width=80)
                gb_dt.configure_column("▲ Acima ref",  width=110,
                    cellStyle=JsCode("""function(p){
                        if(p.value>0) return {background:'#D5F5D5',color:'#1A1A1A',fontWeight:'700'};
                        return {};
                    }"""))
                gb_dt.configure_column("▼ Abaixo ref", width=110,
                    cellStyle=JsCode("""function(p){
                        if(p.value>0) return {background:'#FDDCDE',color:'#1A1A1A',fontWeight:'700'};
                        return {};
                    }"""))
                gb_dt.configure_column("= Empate",     width=90)
                gb_dt.configure_column("Delta médio",  width=110,
                    cellStyle=JsCode("""function(p){
                        var v=p.value;
                        if(v>0)  return {background:'#1E7A34',color:'#FFFFFF',fontWeight:'700'};
                        if(v<0)  return {background:'#E63946',color:'#FFFFFF',fontWeight:'700'};
                        return {background:'#FFD600',color:'#1A1A1A'};
                    }"""))

                go_dt = gb_dt.build()
                go_dt["defaultColDef"]["headerClass"] = "ag-header-black"

                AgGrid(
                    df_resumo_delta,
                    gridOptions=go_dt,
                    height=min(500, 36 + 32 * len(df_resumo_delta) + 20),
                    update_mode=GridUpdateMode.NO_UPDATE,
                    fit_columns_on_grid_load=False,
                    allow_unsafe_jscode=True,
                    enable_enterprise_modules=True,
                    custom_css={
                        ".ag-header":                  {"background-color": "#4A4A4A !important"},
                        ".ag-header-row":              {"background-color": "#4A4A4A !important"},
                        ".ag-header-cell":             {"background-color": "#4A4A4A !important"},
                        ".ag-header-cell-label":       {"color": "#FFFFFF !important", "font-weight": "700"},
                        ".ag-header-cell-text":        {"color": "#FFFFFF !important", "font-size": "13px !important", "font-weight": "700 !important"},
                        ".ag-icon":                    {"color": "#FFFFFF !important", "opacity": "1 !important"},
                        ".ag-header-icon":             {"color": "#FFFFFF !important", "opacity": "1 !important"},
                        ".ag-header-cell-menu-button": {"opacity": "1 !important", "visibility": "visible !important"},
                        ".ag-icon-menu":               {"color": "#FFFFFF !important", "opacity": "1 !important"},
                        ".ag-icon-filter":             {"color": "#FFFFFF !important", "opacity": "1 !important"},
                        ".ag-row":                     {"font-size": "13px !important"},
                    },
                    theme="streamlit",
                    use_container_width=True,
                )
                st.caption(
                    f"ℹ️ **▲ Acima ref** = nº de locais em que o cultivar superou {ref_delta}. "
                    f"**Delta médio** = média dos deltas em todos os locais avaliados juntos."
                )
                exportar_excel(df_resumo_delta, nome_arquivo="delta_referencia.xlsx",
                               label="⬇️ Exportar Delta vs Referência", key="exp_delta_dc")

st.divider()

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 6 — EVOLUÇÃO TEMPORAL POR SAFRA
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Evolução por Safra",
    "O perfil sanitário dos cultivares melhorou ou piorou ao longo das safras?",
    contexto_str,
)

with st.popover("ℹ️ Como interpretar", use_container_width=False):
    st.markdown("""
Compara a nota média de cada cultivar entre as safras disponíveis na base, independente do filtro de safra ativo na sidebar — para garantir que todas as safras aparecem no gráfico.

Os demais filtros (macro, micro, estado, GM, status) continuam aplicados.

**Gráfico**
- **Linha subindo** → cultivar melhorando sanitariamente (nota maior = melhor)
- **Linha descendo** → cultivar com piora no perfil sanitário
- **Linha cinza tracejada** → média geral do grupo naquela safra (referência)
- Cultivares avaliados em apenas 1 safra são omitidos do gráfico

---

**Tabela de consistência — como cada coluna é calculada**

**Safras aval.** → quantas safras distintas o cultivar tem nota > 0 para a doença selecionada.

**Acima da média** → para cada safra, calcula a média geral do grupo. Se a nota do cultivar ≥ essa média, conta 1. A coluna soma quantas safras isso aconteceu.

**Consistência %** → `Acima da média ÷ Safras aval. × 100`. Colorida de vermelho (0%) a verde escuro (100%).

**Tendência** → regressão linear sobre as notas em ordem cronológica:
- **↑ melhora** → nota subindo entre safras
- **↓ piora** → nota caindo entre safras
- **→ estável** → variação pequena (slope entre −0.1 e +0.1)

---

**⚠️ Atenção ao interpretar juntos Consistência e Tendência**

Um cultivar pode ter **100% de consistência e ↓ piora** ao mesmo tempo — significa que ele está caindo, mas o grupo inteiro também caiu junto. A consistência mede posição relativa ao grupo, não evolução absoluta.

Para ver a variação absoluta, observe as colunas de nota por safra diretamente.
""")

_col_d_ev, _col_status_ev = st.columns(2)

with _col_d_ev:
    doenca_ev = st.selectbox(
        "Doença",
        options=doencas_disp,
        format_func=lambda d: f"{SIGLAS.get(d, d)} — {d}",
        key="dc_doenca_ev",
    )

with _col_status_ev:
    _status_ev_opts = sorted(ta_raw["status_material"].dropna().unique().tolist())
    status_ev_sel = st.multiselect(
        "Status",
        options=_status_ev_opts,
        default=_status_ev_opts,
        key="dc_status_ev",
    )

col_nota_ev = DOENCAS[doenca_ev]["nota"]

if col_nota_ev not in ta_raw.columns:
    st.warning("Coluna de nota não encontrada para esta doença.")
else:
    # Aplicar todos os filtros EXCETO safra — usar ta_raw como base
    _filtros_sem_safra = {
        "macro_nome":       ta_filtrado["macro_nome"].unique().tolist() if "macro_nome" in ta_filtrado.columns else None,
        "micro_nome":       ta_filtrado["micro_nome"].unique().tolist() if "micro_nome" in ta_filtrado.columns else None,
        "estado_sigla":     ta_filtrado["estado_sigla"].unique().tolist() if "estado_sigla" in ta_filtrado.columns else None,
        "gm_cat":           ta_filtrado["gm_cat"].unique().tolist() if "gm_cat" in ta_filtrado.columns else None,
    }

    df_ev = ta_raw.copy()
    for col, vals in _filtros_sem_safra.items():
        if vals is not None and col in df_ev.columns:
            df_ev = df_ev[df_ev[col].isin(vals)]

    if status_ev_sel:
        df_ev = df_ev[df_ev["status_material"].isin(status_ev_sel)]

    df_ev[col_nota_ev] = pd.to_numeric(df_ev[col_nota_ev], errors="coerce")
    df_ev = df_ev[df_ev[col_nota_ev] > 0].dropna(subset=[col_nota_ev, "safra", "dePara"])

    # Média por cultivar × safra
    df_ev_agg = (
        df_ev.groupby(["dePara", "status_material", "safra"])
        [col_nota_ev].mean().round(2).reset_index()
    )

    safras_ev = sorted(df_ev_agg["safra"].unique().tolist())

    if len(safras_ev) < 2:
        st.info("São necessárias ao menos 2 safras na base para exibir a evolução temporal. Verifique os filtros ativos.")
    else:
        # Manter apenas cultivares avaliados em 2+ safras
        contagem_safras = df_ev_agg.groupby("dePara")["safra"].nunique()
        cultivares_ev = contagem_safras[contagem_safras >= 2].index.tolist()

        n_single = (contagem_safras == 1).sum()

        if not cultivares_ev:
            st.info("Nenhum cultivar foi avaliado em mais de uma safra com os filtros ativos.")
        else:
            df_ev_plot = df_ev_agg[df_ev_agg["dePara"].isin(cultivares_ev)].copy()

            # Paleta por status
            palette_ev = [
                "#1F77B4","#FF7F0E","#2CA02C","#D62728","#9467BD",
                "#8C564B","#E377C2","#7F7F7F","#BCBD22","#17BECF",
                "#AEC7E8","#FFBB78","#98DF8A","#FF9896","#C5B0D5",
                "#C49C94","#F7B6D2","#C7C7C7","#DBDB8D","#9EDAE5",
            ]
            cultivares_ord_ev = sorted(cultivares_ev)
            cor_ev = {c: palette_ev[i % len(palette_ev)] for i, c in enumerate(cultivares_ord_ev)}

            fig_ev = go_plt.Figure()

            for cultivar in cultivares_ord_ev:
                df_c = df_ev_plot[df_ev_plot["dePara"] == cultivar].sort_values("safra")
                status = df_c["status_material"].mode()[0] if not df_c["status_material"].mode().empty else ""
                cor = cor_ev[cultivar]

                fig_ev.add_trace(go_plt.Scatter(
                    x=df_c["safra"].astype(str).tolist(),
                    y=df_c[col_nota_ev].tolist(),
                    mode="lines+markers",
                    name=cultivar,
                    line=dict(color=cor, width=2),
                    marker=dict(size=8, color=cor),
                    hovertemplate=(
                        f"<b>{cultivar}</b> ({status})<br>"
                        "Safra: %{x}<br>"
                        "Nota média: %{y:.2f}<extra></extra>"
                    ),
                ))

            # Média geral por safra (linha de referência)
            media_geral_ev = df_ev_plot.groupby("safra")[col_nota_ev].mean().round(2)
            fig_ev.add_trace(go_plt.Scatter(
                x=[str(s) for s in media_geral_ev.index],
                y=media_geral_ev.values.tolist(),
                mode="lines",
                name="Média geral",
                line=dict(color="#AAAAAA", width=1.5, dash="dot"),
                hovertemplate="<b>Média geral</b><br>Safra: %{x}<br>Nota: %{y:.2f}<extra></extra>",
            ))

            # Linha de referência nota 6
            fig_ev.add_hline(
                y=6, line_dash="dot", line_color="#E74C3C", line_width=1.5,
                annotation_text="Limite MT/T (nota 6)",
                annotation_position="top right",
                annotation_font=dict(size=11, color="#E74C3C"),
            )

            fig_ev.update_layout(
                height=480,
                xaxis=dict(
                    title=dict(text="<b>Safra</b>", font=dict(size=14, color="#111111", weight="bold")),
                    tickfont=dict(size=12, color="#111111", weight="bold"),
                    categoryorder="array",
                    categoryarray=[str(s) for s in safras_ev],
                    gridcolor="#EEEEEE",
                ),
                yaxis=dict(
                    title=dict(text="<b>Nota média (1 = pior · 9 = melhor)</b>", font=dict(size=14, color="#111111", weight="bold")),
                    range=[0.5, 9.5],
                    tickvals=list(range(1, 10)),
                    tickfont=dict(size=12, color="#111111", weight="bold"),
                    gridcolor="#EEEEEE",
                    zeroline=False,
                ),
                legend=dict(
                    title=dict(text="<b>Cultivar</b>", font=dict(size=12, color="#111111", weight="bold")),
                    font=dict(size=11, color="#111111", weight="bold"),
                    bgcolor="rgba(255,255,255,0.85)",
                    bordercolor="#DDDDDD",
                    borderwidth=1,
                ),
                plot_bgcolor="#FAFAFA",
                paper_bgcolor="#FFFFFF",
                margin=dict(t=40, b=60, l=60, r=20),
                font=dict(family="Helvetica Neue, sans-serif", size=12, color="#111111"),
                hovermode="x unified",
            )

            st.plotly_chart(fig_ev, use_container_width=True)

            _caption_ev = (
                f"ℹ️ Nota média por safra — apenas cultivares avaliados em 2+ safras ({len(cultivares_ev)} cultivares). "
                f"Linha cinza tracejada = média geral do grupo. "
            )
            if n_single > 0:
                _caption_ev += f"{int(n_single)} cultivar(es) com apenas 1 safra foram omitidos."
            st.caption(_caption_ev)

            # ── Tabela de consistência ────────────────────────────────────────
            st.markdown("##### Consistência por cultivar")
            st.caption("% das safras em que o cultivar ficou acima da média geral do grupo.")

            consist_rows = []
            for cultivar in cultivares_ord_ev:
                df_c = df_ev_plot[df_ev_plot["dePara"] == cultivar]
                status = df_c["status_material"].mode()[0] if not df_c["status_material"].mode().empty else ""
                n_safras_c = len(df_c)
                notas_por_safra = df_c.set_index("safra")[col_nota_ev]

                n_acima = 0
                tendencia_vals = []
                for safra in safras_ev:
                    if safra in notas_por_safra.index and safra in media_geral_ev.index:
                        if notas_por_safra[safra] >= media_geral_ev[safra]:
                            n_acima += 1
                        tendencia_vals.append(notas_por_safra[safra])

                # Tendência: coeficiente de regressão linear simples
                tendencia_txt = "—"
                if len(tendencia_vals) >= 2:
                    xs = list(range(len(tendencia_vals)))
                    n_ = len(xs)
                    mx, my = sum(xs)/n_, sum(tendencia_vals)/n_
                    num = sum((x-mx)*(y-my) for x,y in zip(xs,tendencia_vals))
                    den = sum((x-mx)**2 for x in xs) or 1
                    slope = num/den
                    if slope > 0.1:
                        tendencia_txt = "↑ melhora"
                    elif slope < -0.1:
                        tendencia_txt = "↓ piora"
                    else:
                        tendencia_txt = "→ estável"

                pct_consist = round(n_acima / n_safras_c * 100, 0) if n_safras_c > 0 else 0

                row_c = {
                    "Cultivar":       cultivar,
                    "Status":         status,
                    "Safras aval.":   n_safras_c,
                    "Acima da média": n_acima,
                    "Consistência %": pct_consist,
                    "Tendência":      tendencia_txt,
                }
                for safra in safras_ev:
                    if safra in notas_por_safra.index:
                        row_c[str(safra)] = notas_por_safra[safra]
                    else:
                        row_c[str(safra)] = None

                consist_rows.append(row_c)

            df_consist = pd.DataFrame(consist_rows).sort_values("Consistência %", ascending=False)

            gb_ev = GridOptionsBuilder.from_dataframe(df_consist)
            gb_ev.configure_default_column(
                resizable=True, sortable=True, filter=True,
                cellStyle={"fontSize": "13px", "fontFamily": "Helvetica Neue, sans-serif"},
            )
            gb_ev.configure_grid_options(headerHeight=36, rowHeight=32, domLayout="normal")
            gb_ev.configure_column("Cultivar",       pinned="left", width=170)
            gb_ev.configure_column("Status",         width=90)
            gb_ev.configure_column("Safras aval.",   width=100)
            gb_ev.configure_column("Acima da média", width=120)
            gb_ev.configure_column("Consistência %", width=120,
                cellStyle=JsCode("""function(p){
                    var v=p.value;
                    if(v>=75) return {background:'#1E7A34',color:'#FFFFFF',fontWeight:'700'};
                    if(v>=50) return {background:'#70C96E',color:'#1A1A1A',fontWeight:'700'};
                    if(v>=25) return {background:'#FFD600',color:'#1A1A1A'};
                    return {background:'#E63946',color:'#FFFFFF',fontWeight:'700'};
                }"""))
            gb_ev.configure_column("Tendência", width=110,
                cellStyle=JsCode("""function(p){
                    if(p.value && p.value.includes('melhora')) return {color:'#1E7A34',fontWeight:'700'};
                    if(p.value && p.value.includes('piora'))   return {color:'#E63946',fontWeight:'700'};
                    return {color:'#888888'};
                }"""))
            for safra in safras_ev:
                gb_ev.configure_column(str(safra), width=90,
                    cellStyle=JsCode(f"""function(p){{
                        if(p.value==null) return {{background:'#F0F0F0',color:'#999999'}};
                        if(p.value>=7) return {{background:'#70C96E',color:'#1A1A1A'}};
                        if(p.value>=5) return {{background:'#FFD600',color:'#1A1A1A'}};
                        return {{background:'#E63946',color:'#FFFFFF'}};
                    }}"""))

            go_ev = gb_ev.build()
            go_ev["defaultColDef"]["headerClass"] = "ag-header-black"

            AgGrid(
                df_consist,
                gridOptions=go_ev,
                height=min(500, 36 + 32 * len(df_consist) + 20),
                update_mode=GridUpdateMode.NO_UPDATE,
                fit_columns_on_grid_load=False,
                allow_unsafe_jscode=True,
                enable_enterprise_modules=True,
                custom_css={
                    ".ag-header":                  {"background-color": "#4A4A4A !important"},
                    ".ag-header-row":              {"background-color": "#4A4A4A !important"},
                    ".ag-header-cell":             {"background-color": "#4A4A4A !important"},
                    ".ag-header-cell-label":       {"color": "#FFFFFF !important", "font-weight": "700"},
                    ".ag-header-cell-text":        {"color": "#FFFFFF !important", "font-size": "13px !important", "font-weight": "700 !important"},
                    ".ag-icon":                    {"color": "#FFFFFF !important", "opacity": "1 !important"},
                    ".ag-header-icon":             {"color": "#FFFFFF !important", "opacity": "1 !important"},
                    ".ag-header-cell-menu-button": {"opacity": "1 !important", "visibility": "visible !important"},
                    ".ag-icon-menu":               {"color": "#FFFFFF !important", "opacity": "1 !important"},
                    ".ag-icon-filter":             {"color": "#FFFFFF !important", "opacity": "1 !important"},
                    ".ag-row":                     {"font-size": "13px !important"},
                },
                theme="streamlit",
                use_container_width=True,
            )
            exportar_excel(df_consist, nome_arquivo="evolucao_safras.xlsx",
                           label="⬇️ Exportar Evolução por Safra", key="exp_ev_dc")

            # ── Rodapé de locais por safra ────────────────────────────────────
            n_locais_total_ev = ta_filtrado["cod_fazenda"].nunique()
            _linhas_rodape_ev = []
            for safra in safras_ev:
                n_loc_ev = df_ev[df_ev["safra"] == safra]["cod_fazenda"].nunique()
                if n_loc_ev < n_locais_total_ev:
                    _linhas_rodape_ev.append(f"**{safra}**: {n_loc_ev} de {n_locais_total_ev} locais")
            if _linhas_rodape_ev:
                st.caption(
                    f"ℹ️ Locais com nota registrada para {doenca_ev} por safra: "
                    + " · ".join(_linhas_rodape_ev)
                    + " (demais locais sem nota registrada nessa safra)."
                )

st.divider()

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 7 — PERFIL MULTIDOENÇA
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Perfil Multidoença",
    "Qual é o perfil sanitário completo de cada cultivar?",
    contexto_str,
)

with st.popover("ℹ️ Como interpretar", use_container_width=False):
    st.markdown("""
Exibe o perfil sanitário de um cultivar em todas as doenças simultaneamente — dois visuais complementares lado a lado. A nota usada é a **menor nota observada** (pior caso) — revela vulnerabilidades que a média ou moda escondem. O hover mostra quantas observações com nota > 0 entraram no cálculo.

**Radar (esquerda)**
Cada eixo é uma doença. Quanto maior a área, melhor o perfil geral. Útil para ver equilíbrio vs pontos fracos — um cultivar com área irregular tem calcanhar de aquiles.

⚠️ A área do radar varia conforme a ordem das doenças nos eixos — use para percepção geral, não para comparação precisa de números.

**Barras horizontais (direita)**
Mesmas doenças, mesmos valores, sem distorção geométrica. Cada barra mostra a nota média — colorida pela classe. Use para comparar valores exatos entre cultivares.

**Como usar os dois juntos**
O radar dá a impressão visual rápida. As barras confirmam os números. Se o radar parecer grande mas as barras tiverem células amarelas, o cultivar é consistente mas não excelente.
""")

# ── Seletores ─────────────────────────────────────────────────────────────────
_cultivares_pm = sorted(ta_filtrado["dePara"].dropna().unique().tolist())

cultivares_pm_sel = st.multiselect(
    "Selecione cultivares para comparar (máx. 6 recomendado)",
    options=_cultivares_pm,
    default=_cultivares_pm[:min(4, len(_cultivares_pm))],
    key="dc_cultivares_pm",
)

if not cultivares_pm_sel:
    st.info("Selecione ao menos um cultivar para exibir o perfil.")
else:
    # ── Calcular menor nota + contagem por cultivar × doença ──────────────
    pm_rows = []
    for doenca in DOENCAS:
        col_nota_pm = DOENCAS[doenca]["nota"]
        if col_nota_pm not in ta_filtrado.columns:
            continue
        s_all = pd.to_numeric(ta_filtrado[col_nota_pm], errors="coerce")
        if s_all[s_all > 0].empty:
            continue
        for cultivar in cultivares_pm_sel:
            mask = (ta_filtrado["dePara"] == cultivar)
            s_c = s_all[mask]
            s_c = s_c[s_c > 0].dropna()
            if s_c.empty:
                continue
            nota_min  = float(s_c.min())
            n_obs     = int(len(s_c))
            n_total   = int(mask.sum())
            pm_rows.append({
                "cultivar": cultivar,
                "doenca":   SIGLAS.get(doenca, doenca),
                "nota":     nota_min,
                "classe":   nota_para_classe(nota_min),
                "n_obs":    n_obs,
                "n_total":  n_total,
            })

    if not pm_rows:
        st.info("Sem dados suficientes para os cultivares selecionados.")
    else:
        df_pm = pd.DataFrame(pm_rows)
        doencas_pm = sorted(df_pm["doenca"].unique().tolist())

        palette_pm = [
            "#1F77B4","#FF7F0E","#2CA02C","#D62728","#9467BD",
            "#8C564B","#E377C2","#7F7F7F","#BCBD22","#17BECF",
        ]
        cor_pm = {c: palette_pm[i % len(palette_pm)] for i, c in enumerate(cultivares_pm_sel)}

        col_radar, col_bar = st.columns(2)

        # ── RADAR ─────────────────────────────────────────────────────────
        with col_radar:
            fig_radar = go_plt.Figure()

            for cultivar in cultivares_pm_sel:
                df_c = df_pm[df_pm["cultivar"] == cultivar]
                notas_radar = []
                hover_radar = []
                for d in doencas_pm:
                    row = df_c[df_c["doenca"] == d]
                    if not row.empty:
                        notas_radar.append(float(row["nota"].values[0]))
                        hover_radar.append(f"{d}: {row['nota'].values[0]:.1f} ({int(row['n_obs'].values[0])}/{int(row['n_total'].values[0])} obs)")
                    else:
                        notas_radar.append(0)
                        hover_radar.append(f"{d}: sem dado")

                # Fechar o polígono
                theta = doencas_pm + [doencas_pm[0]]
                r     = notas_radar + [notas_radar[0]]

                # Converter hex para rgba com transparência
                hex_cor = cor_pm[cultivar].lstrip("#")
                r_c, g_c, b_c = int(hex_cor[0:2],16), int(hex_cor[2:4],16), int(hex_cor[4:6],16)
                fill_rgba = f"rgba({r_c},{g_c},{b_c},0.15)"

                fig_radar.add_trace(go_plt.Scatterpolar(
                    r=r,
                    theta=theta,
                    fill="toself",
                    name=cultivar,
                    line=dict(color=cor_pm[cultivar], width=2),
                    fillcolor=fill_rgba,
                    hovertemplate="<b>%{fullData.name}</b><br>%{theta}: %{r:.1f}<extra></extra>",
                ))

            fig_radar.update_layout(
                polar=dict(
                    radialaxis=dict(
                        visible=True,
                        range=[0, 9],
                        tickvals=[1, 3, 5, 7, 9],
                        tickfont=dict(size=10, color="#555555", weight="bold"),
                        gridcolor="#DDDDDD",
                    ),
                    angularaxis=dict(
                        tickfont=dict(size=11, color="#111111", weight="bold"),
                        gridcolor="#DDDDDD",
                    ),
                    bgcolor="#FAFAFA",
                ),
                legend=dict(
                    font=dict(size=11, color="#111111", weight="bold"),
                    bgcolor="rgba(255,255,255,0.85)",
                    bordercolor="#DDDDDD",
                    borderwidth=1,
                ),
                paper_bgcolor="#FFFFFF",
                margin=dict(t=40, b=40, l=40, r=40),
                height=420,
                font=dict(family="Helvetica Neue, sans-serif"),
            )
            st.plotly_chart(fig_radar, use_container_width=True)
            st.caption("⚠️ Área do radar varia com a ordem das doenças — use para impressão geral. Nota = menor valor observado.")

        # ── BARRAS HORIZONTAIS ─────────────────────────────────────────────
        with col_bar:
            fig_bar_pm = go_plt.Figure()

            # Ordenar doenças por nota média geral (pior → melhor) para leitura natural
            ordem_doencas_pm = (
                df_pm.groupby("doenca")["nota"].mean()
                .sort_values(ascending=True)
                .index.tolist()
            )

            for cultivar in cultivares_pm_sel:
                df_c = df_pm[df_pm["cultivar"] == cultivar].set_index("doenca")
                notas_bar = []
                cores_bar = []
                hover_bar = []
                for d in ordem_doencas_pm:
                    if d in df_c.index:
                        n = df_c.loc[d, "nota"]
                        cls = df_c.loc[d, "classe"]
                        n_obs = int(df_c.loc[d, "n_obs"])
                        n_tot = int(df_c.loc[d, "n_total"])
                        notas_bar.append(n)
                        cores_bar.append(COR_CLASS.get(cls, "#CCCCCC"))
                        hover_bar.append(f"<b>{cultivar}</b><br>{d}: {n:.1f} ({cls})<br>{n_obs}/{n_tot} obs com nota > 0")
                    else:
                        notas_bar.append(None)
                        cores_bar.append("#EEEEEE")
                        hover_bar.append(f"<b>{cultivar}</b><br>{d}: sem dado")

                fig_bar_pm.add_trace(go_plt.Bar(
                    name=cultivar,
                    y=ordem_doencas_pm,
                    x=notas_bar,
                    orientation="h",
                    marker_color=cor_pm[cultivar],
                    text=[f"{v:.1f} ({int(df_c.loc[d,'n_obs'])}/{int(df_c.loc[d,'n_total'])})" if v and d in df_c.index else "—"
                          for v, d in zip(notas_bar, ordem_doencas_pm)],
                    textposition="outside",
                    textfont=dict(size=11, weight="bold"),
                    hovertext=hover_bar,
                    hoverinfo="text",
                ))

            fig_bar_pm.add_vline(
                x=6, line_dash="dot", line_color="#E74C3C", line_width=1.5,
                annotation_text="nota 6",
                annotation_position="top",
                annotation_font=dict(size=13, color="#E74C3C", weight="bold"),
            )

            fig_bar_pm.update_layout(
                barmode="group",
                height=420,
                xaxis=dict(
                    title=dict(text="<b>Nota média (1 = pior · 9 = melhor)</b>", font=dict(size=14, color="#111111", weight="bold")),
                    range=[0, 10.5],
                    tickvals=list(range(0, 10)),
                    tickfont=dict(size=11, color="#111111", weight="bold"),
                    gridcolor="#EEEEEE",
                ),
                yaxis=dict(
                    tickfont=dict(size=11, color="#111111", weight="bold"),
                    gridcolor="#EEEEEE",
                    categoryorder="array",
                    categoryarray=ordem_doencas_pm,
                ),
                legend=dict(
                    font=dict(size=11, color="#111111", weight="bold"),
                    bgcolor="rgba(255,255,255,0.85)",
                    bordercolor="#DDDDDD",
                    borderwidth=1,
                ),
                plot_bgcolor="#FAFAFA",
                paper_bgcolor="#FFFFFF",
                margin=dict(t=40, b=40, l=10, r=60),
                font=dict(family="Helvetica Neue, sans-serif", size=12, color="#111111"),
                hovermode="y unified",
            )
            st.plotly_chart(fig_bar_pm, use_container_width=True)
            st.caption("Nota = menor valor observado (pior caso). Rótulo: nota (obs com nota > 0 / total de registros). Doenças ordenadas da menor para maior nota entre os cultivares selecionados.")

            # ── Rodapé de locais por doença ───────────────────────────────────
            n_locais_total_pm = ta_filtrado["cod_fazenda"].nunique()
            _linhas_rodape_pm = []
            for doenca in DOENCAS:
                col_nota_pm2 = DOENCAS[doenca]["nota"]
                if col_nota_pm2 not in ta_filtrado.columns:
                    continue
                s_pm2 = pd.to_numeric(ta_filtrado[col_nota_pm2], errors="coerce")
                n_loc_pm2 = ta_filtrado[s_pm2 > 0]["cod_fazenda"].nunique()
                if n_loc_pm2 < n_locais_total_pm:
                    _linhas_rodape_pm.append(f"**{SIGLAS.get(doenca, doenca)}**: {n_loc_pm2}/{n_locais_total_pm}")
            if _linhas_rodape_pm:
                st.caption(
                    "ℹ️ Locais com nota registrada (de " + str(n_locais_total_pm) + " ativos): "
                    + " · ".join(_linhas_rodape_pm)
                )

st.divider()

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 8 — HEATMAP CULTIVAR × LOCAL
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Heatmap por Local",
    "Como cada cultivar se saiu em cada local para uma doença?",
    contexto_str,
)

_col_interp_hm2, _col_dic_hm2 = st.columns([1, 9])

with _col_interp_hm2:
    with st.popover("ℹ️ Como interpretar", use_container_width=False):
        st.markdown("""
Para a doença selecionada, cada célula mostra a nota do cultivar naquele local. **•** indica que a doença foi detectada (incidência > 0%).

**Cores das células**
""")
        _cls_cols = st.columns(5)
        for i, (cls, cor) in enumerate(COR_CLASS.items()):
            fg = COR_TEXTO_CLASS[cls]
            _cls_cols[i].markdown(
                f'<div style="background:{cor};color:{fg};border-radius:6px;padding:6px;'
                f'text-align:center;font-size:12px;font-weight:700;">{cls}<br>'
                f'<span style="font-weight:400;font-size:11px;">{LABEL_CLASS[cls].split(" — ")[1]}</span></div>',
                unsafe_allow_html=True,
            )
        st.markdown("""
**Destaques automáticos**
- **Borda vermelha + ⚠️** → local com maior pressão da doença (menor nota mais frequente do grupo, abaixo de 6)
- **★ verde** → cultivar com melhor perfil geral no filtro ativo (maior nota mais frequente entre os locais)

Cultivar não avaliado em um local aparece como célula cinza (—).
""")

with _col_dic_hm2:
    with st.popover(f"📍 Dicionário de locais ({ta_filtrado['cod_fazenda'].nunique()})", use_container_width=False):
        _df_dic_hm2 = (
            ta_filtrado[["cod_fazenda", "nomeFazenda", "cidade_nome", "estado_sigla"]]
            .drop_duplicates()
            .sort_values(["estado_sigla", "cidade_nome", "cod_fazenda"])
            .rename(columns={
                "cod_fazenda":  "Código",
                "nomeFazenda":  "Local",
                "cidade_nome":  "Cidade",
                "estado_sigla": "Estado",
            })
            .reset_index(drop=True)
        )
        st.markdown("Referência dos códigos exibidos nas colunas do heatmap.")
        st.dataframe(_df_dic_hm2, hide_index=True, use_container_width=True)

doenca_hm2 = st.selectbox(
    "Selecione a doença",
    options=doencas_disp,
    format_func=lambda d: f"{SIGLAS.get(d, d)} — {d}",
    key="dc_doenca_hm2",
)

col_nota_hm2 = DOENCAS[doenca_hm2]["nota"]

if col_nota_hm2 not in ta_filtrado.columns:
    st.warning("Coluna de nota não encontrada para esta doença.")
else:
    df_hm2 = ta_filtrado[["dePara", "status_material", "cod_fazenda",
                            "cidade_nome", "estado_sigla", col_nota_hm2]].copy()
    df_hm2[col_nota_hm2] = pd.to_numeric(df_hm2[col_nota_hm2], errors="coerce")
    df_hm2 = df_hm2[df_hm2[col_nota_hm2] > 0].dropna(subset=[col_nota_hm2])

    if df_hm2.empty:
        st.info("Nenhuma avaliação disponível para esta doença nos filtros ativos.")
    else:
        # Nota por cultivar × local (sem repetição em faixa — mean = valor único)
        df_hm2_agg = (
            df_hm2.groupby(["dePara", "status_material", "cod_fazenda",
                             "cidade_nome", "estado_sigla"])
            [col_nota_hm2].mean().round(1).reset_index()
        )

        # Ordenação locais: estado → cidade → cod_fazenda
        locais_hm2 = (
            df_hm2_agg[["cod_fazenda", "cidade_nome", "estado_sigla"]]
            .drop_duplicates()
            .sort_values(["estado_sigla", "cidade_nome", "cod_fazenda"])
            ["cod_fazenda"].tolist()
        )

        # Ordenação cultivares: status → alfabético
        cult_status_hm2 = (
            df_hm2_agg[["dePara", "status_material"]]
            .drop_duplicates()
            .assign(_ord=lambda d: d["status_material"].apply(
                lambda s: STATUS_ORDER_HM.index(s) if s in STATUS_ORDER_HM else 99))
            .sort_values(["_ord", "dePara"])
        )
        cultivares_hm2  = cult_status_hm2["dePara"].tolist()
        status_map_hm2  = cult_status_hm2.set_index("dePara")["status_material"].to_dict()

        # Pivot
        pivot_hm2 = df_hm2_agg.pivot_table(
            index="dePara", columns="cod_fazenda",
            values=col_nota_hm2, aggfunc="mean"
        ).reindex(index=cultivares_hm2, columns=locais_hm2)

        fig_hm2 = go_plt.Figure()

        # Heatmap base invisível para estrutura
        fig_hm2.add_trace(go_plt.Heatmap(
            z=[[0] * len(locais_hm2)] * len(cultivares_hm2),
            x=locais_hm2,
            y=cultivares_hm2,
            colorscale=[[0, "#FFFFFF"], [1, "#FFFFFF"]],
            showscale=False,
            hoverinfo="skip",
        ))

        # Pivot incidência por cultivar × local
        df_hm2_inc = df_hm2.copy()
        df_hm2_inc["_inc"] = df_hm2_inc[col_nota_hm2].between(1, 5).astype(int)
        pivot_inc_hm2 = df_hm2_inc.groupby(["dePara", "cod_fazenda"]).apply(
            lambda g: round(g["_inc"].sum() / len(g) * 100, 1)
        ).unstack(fill_value=np.nan)

        # Células coloridas
        for i, cultivar in enumerate(cultivares_hm2):
            for j, local in enumerate(locais_hm2):
                v = pivot_hm2.loc[cultivar, local] if cultivar in pivot_hm2.index and local in pivot_hm2.columns else None
                if v is None or (isinstance(v, float) and np.isnan(v)):
                    cor = "#E0E0E0"
                    txt = "—"
                    fg  = "#888888"
                else:
                    cls = moda_para_classe(v)
                    cor = COR_CLASS.get(cls, "#CCCCCC")
                    fg  = COR_TEXTO_CLASS.get(cls, "#1A1A1A")
                    # asterisco se incidência > 0
                    inc_v = None
                    try:
                        inc_v = pivot_inc_hm2.loc[cultivar, local]
                    except Exception:
                        pass
                    asterisco = " •" if inc_v is not None and not (isinstance(inc_v, float) and np.isnan(inc_v)) and inc_v > 0 else ""
                    txt = f"{v:.1f}{asterisco}<br>{cls}"

                fig_hm2.add_shape(
                    type="rect",
                    x0=j - 0.5, x1=j + 0.5,
                    y0=i - 0.5, y1=i + 0.5,
                    fillcolor=cor,
                    line=dict(color="#FFFFFF", width=1),
                )
                fig_hm2.add_annotation(
                    x=j, y=i,
                    text=txt,
                    showarrow=False,
                    font=dict(size=10, color=fg, weight="bold"),
                    align="center",
                )

        # Separadores de status
        for i in range(len(cultivares_hm2) - 1):
            s1 = status_map_hm2.get(cultivares_hm2[i], "")
            s2 = status_map_hm2.get(cultivares_hm2[i + 1], "")
            if s1 != s2:
                fig_hm2.add_shape(
                    type="line",
                    x0=-0.5, x1=len(locais_hm2) - 0.5,
                    y0=i + 0.5, y1=i + 0.5,
                    line=dict(color="#333333", width=2),
                )

        # ── Destaque coluna local mais crítico (menor nota — moda) ──────────────
        moda_local_hm2 = (
            df_hm2_agg.groupby("cod_fazenda")[col_nota_hm2]
            .apply(lambda s: s.mode().iloc[0] if not s.mode().empty else np.nan)
        )
        if not moda_local_hm2.empty and len(locais_hm2) > 0:
            local_crit_hm2 = moda_local_hm2.idxmin()
            nota_crit_hm2  = round(moda_local_hm2.min(), 1)
            if local_crit_hm2 in locais_hm2 and nota_crit_hm2 < 6:
                j_crit = locais_hm2.index(local_crit_hm2)
                # Borda vermelha na coluna
                fig_hm2.add_shape(
                    type="rect",
                    x0=j_crit - 0.5, x1=j_crit + 0.5,
                    y0=-0.5, y1=len(cultivares_hm2) - 0.5,
                    fillcolor="rgba(0,0,0,0)",
                    line=dict(color="#E74C3C", width=2.5),
                )
                # Badge no topo da coluna
                fig_hm2.add_annotation(
                    x=j_crit, y=-0.5,
                    yref="y",
                    text=f"⚠️ {nota_crit_hm2}",
                    showarrow=False,
                    xanchor="center",
                    yanchor="bottom",
                    yshift=6,
                    font=dict(size=10, color="#E74C3C", family="Helvetica Neue, sans-serif"),
                    bgcolor="rgba(255,255,255,0.9)",
                    bordercolor="#E74C3C",
                    borderwidth=1,
                    borderpad=2,
                )

        # ── Cultivar com melhor perfil geral (maior moda) ────────────────────
        moda_cultivar_hm2 = (
            df_hm2_agg.groupby("dePara")[col_nota_hm2]
            .apply(lambda s: s.mode().iloc[0] if not s.mode().empty else np.nan)
        )
        if not moda_cultivar_hm2.empty and len(cultivares_hm2) > 0:
            melhor_cult = moda_cultivar_hm2.idxmax()
            nota_melhor = round(moda_cultivar_hm2.max(), 1)
            if melhor_cult in cultivares_hm2:
                i_melhor = cultivares_hm2.index(melhor_cult)
                fig_hm2.add_annotation(
                    x=len(locais_hm2) - 0.5,
                    y=i_melhor,
                    xref="x", yref="y",
                    text=f"★ melhor nota {nota_melhor}",
                    showarrow=False,
                    xanchor="left",
                    yanchor="middle",
                    xshift=8,
                    font=dict(size=11, color="#1E7A34", weight="bold", family="Helvetica Neue, sans-serif"),
                    bgcolor="rgba(255,255,255,0.85)",
                    bordercolor="#1E7A34",
                    borderwidth=1,
                    borderpad=2,
                )

        # Anotações Y coloridas por status
        for i, cultivar in enumerate(cultivares_hm2):
            status = status_map_hm2.get(cultivar, "")
            cor_s  = COR_STATUS_TEXTO_HM.get(status, "#333333")
            fig_hm2.add_annotation(
                x=-0.01, xref="paper",
                y=i,     yref="y",
                text=f"<b>{cultivar}</b>",
                showarrow=False,
                xanchor="right",
                yanchor="middle",
                font=dict(size=11, color=cor_s, weight="bold"),
            )

        altura_hm2 = max(400, len(cultivares_hm2) * 28 + 80)

        fig_hm2.update_layout(
            height=altura_hm2,
            xaxis=dict(
                side="bottom",
                tickfont=dict(size=11, color="#111111", weight="bold"),
                tickangle=-45,
                title=dict(text="<b>Local (cod_fazenda)</b>", font=dict(size=14, color="#111111", weight="bold")),
                categoryorder="array",
                categoryarray=locais_hm2,
            ),
            yaxis=dict(
                tickfont=dict(size=11, color="#111111", weight="bold"),
                autorange="reversed",
                showticklabels=False,
            ),
            margin=dict(t=30, b=100, l=200, r=140),
            plot_bgcolor="#FFFFFF",
            paper_bgcolor="#FFFFFF",
            font=dict(family="Helvetica Neue, sans-serif", size=12, color="#111111"),
        )

        st.plotly_chart(fig_hm2, use_container_width=True)
        n_locais_total_hm2 = ta_filtrado["cod_fazenda"].nunique()
        _caption_hm2 = (
            f"ℹ️ Cor da célula = classe derivada da nota do cultivar naquele local. "
            f"**•** = doença presente (incidência > 0%). "
            f"Células cinzas = cultivar não avaliado naquele local. "
            f"Linha separadora = divisão entre grupos de status. "
            f"**⚠️** = local com maior pressão (menor nota do grupo). "
            f"**★** = cultivar com melhor nota geral (maior nota mais frequente entre os locais). "
            f"{len(cultivares_hm2)} cultivares · {len(locais_hm2)} locais avaliados para {doenca_hm2}"
        )
        if len(locais_hm2) < n_locais_total_hm2:
            _caption_hm2 += f" (de {n_locais_total_hm2} locais ativos — os demais não têm nota registrada para esta doença)."
        else:
            _caption_hm2 += "."
        st.caption(_caption_hm2)

st.divider()

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 9 — CURVA DE SOBREVIVÊNCIA + MAPA DE COLAPSO
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Análise de Sobrevivência",
    "Qual cultivar mantém sanidade sob pressão crescente?",
    contexto_str,
)

with st.popover("ℹ️ Como interpretar", use_container_width=False):
    st.markdown("""
Cada curva mostra, para um dado nível de exigência sanitária (eixo X), **em quantos % dos locais o cultivar atingiu ao menos aquela nota**. Cultivar que se mantém alto mesmo com exigência alta é o mais resistente.

**Eixos**
- **X — Nota** → régua de exigência, de 1 (mínimo) a 9 (máximo)
- **Y — Sobrevivência** → % de locais em que o cultivar atingiu ao menos aquela nota

**Destaques automáticos**
- **▲ melhor / ▼ pior** → cultivares com maior e menor área sob a curva (ASC)
- **Faixa amarela** → zona de atenção (notas 4–6), onde a resistência começa a ceder
- **Linha vermelha tracejada** → fronteira MT/T (nota 6)

**Tabela ASC** → resume o desempenho geral em um único número. Quanto maior, melhor o perfil sanitário.

**Mapa de Colapso** → logo abaixo, mostra em quantos locais cada cultivar teve nota ≤ 4 (situação crítica).
""")

doenca_surv = st.selectbox(
    "Selecione a doença",
    options=doencas_disp,
    format_func=lambda d: f"{SIGLAS.get(d, d)} — {d}",
    key="dc_doenca_surv",
)

col_nota_sv = DOENCAS[doenca_surv]["nota"]

if col_nota_sv not in ta_filtrado.columns:
    st.warning("Coluna de nota não encontrada para esta doença.")
else:
    df_sv = ta_filtrado[["dePara", "status_material", "cod_fazenda", col_nota_sv]].copy()
    df_sv[col_nota_sv] = pd.to_numeric(df_sv[col_nota_sv], errors="coerce")
    df_sv = df_sv[df_sv[col_nota_sv] > 0].dropna(subset=[col_nota_sv])

    # Nota por cultivar × local (sem repetição em faixa — mean = valor único)
    df_sv_agg = (
        df_sv.groupby(["dePara", "status_material", "cod_fazenda"])
        [col_nota_sv].mean().round(1).reset_index()
    )

    cultivares_sv = sorted(df_sv_agg["dePara"].unique().tolist())

    if df_sv_agg.empty or len(cultivares_sv) == 0:
        st.info("Sem dados suficientes para esta doença nos filtros ativos.")
    else:
        thresholds = [t / 10 for t in range(10, 91)]  # 1.0 → 9.0 step 0.1

        # ── Calcular curvas e ASC ─────────────────────────────────────────────
        palette_sv = [
            "#1F77B4","#FF7F0E","#2CA02C","#D62728","#9467BD",
            "#8C564B","#E377C2","#7F7F7F","#BCBD22","#17BECF",
            "#AEC7E8","#FFBB78","#98DF8A","#FF9896","#C5B0D5",
            "#C49C94","#F7B6D2","#C7C7C7","#DBDB8D","#9EDAE5",
            "#393B79","#637939","#8C6D31","#843C39","#7B4173",
        ]
        cor_sv = {c: palette_sv[i % len(palette_sv)] for i, c in enumerate(sorted(cultivares_sv))}

        curvas   = {}
        asc_vals = {}

        for cultivar in cultivares_sv:
            notas = df_sv_agg[df_sv_agg["dePara"] == cultivar][col_nota_sv].tolist()
            n     = len(notas)
            sobrev = [sum(1 for v in notas if v >= t) / n * 100 for t in thresholds]
            curvas[cultivar]   = sobrev
            asc_vals[cultivar] = float(np.trapezoid(sobrev, thresholds))

        # Ordenar tabela ASC — maior para menor
        df_asc = (
            pd.DataFrame.from_dict(asc_vals, orient="index", columns=["Área sob a curva"])
            .reset_index().rename(columns={"index": "Cultivar"})
            .sort_values("Área sob a curva", ascending=False)
            .reset_index(drop=True)
        )
        df_asc["Área sob a curva"] = df_asc["Área sob a curva"].round(2)

        # ── Figura ────────────────────────────────────────────────────────────
        fig_sv = go_plt.Figure()

        for cultivar in cultivares_sv:
            status = df_sv_agg[df_sv_agg["dePara"] == cultivar]["status_material"].mode()
            status = status[0] if not status.empty else ""
            fig_sv.add_trace(go_plt.Scatter(
                x=thresholds,
                y=curvas[cultivar],
                mode="lines",
                name=cultivar,
                line=dict(color=cor_sv[cultivar], width=2, shape="spline", smoothing=0.5),
                customdata=[[status, round(asc_vals[cultivar], 2)]] * len(thresholds),
                hovertemplate=(
                    f"<b>{cultivar}</b><br>"
                    "Status: %{customdata[0]}<br>"
                    "Nota ≥ %{x:.1f}: %{y:.1f}% dos locais<br>"
                    "ASC: %{customdata[1]}"
                    "<extra></extra>"
                ),
            ))

        # Linha de referência nota 6
        fig_sv.add_vline(
            x=6, line_dash="dot", line_color="#E74C3C", line_width=1.5,
            annotation_text="Limite MT/T (nota 6)",
            annotation_position="top right",
            annotation_font=dict(size=13, color="#E74C3C", weight="bold"),
        )

        # ── Faixa sombreada zona de atenção (nota 4–6) ────────────────────────
        fig_sv.add_vrect(
            x0=4, x1=6,
            fillcolor="rgba(255, 214, 0, 0.10)",
            layer="below",
            line_width=0,
            annotation_text="zona de atenção",
            annotation_position="top left",
            annotation_font=dict(size=13, color="#B8860B", weight="bold"),
        )

        # ── Rótulo direto no final das curvas melhor e pior (por ASC) ─────────
        melhor_cult_sv = df_asc.iloc[0]["Cultivar"]
        pior_cult_sv   = df_asc.iloc[-1]["Cultivar"]
        for cultivar, label, cor_label, anchor in [
            (melhor_cult_sv, "▲ melhor", "#1E7A34", "bottom"),
            (pior_cult_sv,   "▼ pior",   "#8B0000", "top"),
        ]:
            if cultivar in curvas:
                y_final = curvas[cultivar][-1]  # valor na nota 9
                fig_sv.add_annotation(
                    x=9,
                    y=y_final,
                    text=f"<b>{cultivar}</b><br><span style='font-size:10px'>{label}</span>",
                    showarrow=False,
                    xanchor="left",
                    yanchor=anchor,
                    xshift=8,
                    font=dict(size=10, color=cor_label),
                    bgcolor="rgba(255,255,255,0.85)",
                    bordercolor=cor_label,
                    borderwidth=1,
                    borderpad=3,
                )

        # Anotação 100% no topo
        fig_sv.add_annotation(
            x=0, xref="x", y=100, yref="y",
            text="<b>100%</b>",
            showarrow=False,
            xanchor="right",
            yanchor="middle",
            font=dict(size=11, color="#555555"),
            xshift=-6,
        )

        # Tabela ASC dentro do gráfico (top 10)
        top10 = df_asc.head(10)
        tbl_header = ["<b>Cultivar</b>", "<b>Área sob a curva</b>"]
        tbl_cells_cultivar = top10["Cultivar"].tolist()
        tbl_cells_asc      = top10["Área sob a curva"].tolist()
        tbl_fill_colors    = ["#D5F5D5" if i < 3 else "#F9F9F9" for i in range(len(top10))]

        fig_sv.add_trace(go_plt.Table(
            domain=dict(x=[0.63, 0.88], y=[0.40, 1.0]),
            columnwidth=[100, 50],
            header=dict(
                values=tbl_header,
                fill_color="#4A4A4A",
                font=dict(color="white", size=11),
                align="center",
                height=26,
            ),
            cells=dict(
                values=[tbl_cells_cultivar, tbl_cells_asc],
                fill_color=[tbl_fill_colors, tbl_fill_colors],
                font=dict(color="#111111", size=11),
                align=["left", "center"],
                height=22,
            ),
        ))

        fig_sv.update_layout(
            height=520,
            xaxis=dict(
                title=dict(text="<b>Nota</b>", font=dict(size=14, color="#111111", weight="bold")),
                tickvals=list(range(1, 10)),
                tickfont=dict(size=12, color="#111111", weight="bold"),
                range=[0.9, 9.1],
                autorange=False,
                gridcolor="#CCCCCC",
                griddash="dot",
                gridwidth=1,
                domain=[0, 0.60],
            ),
            yaxis=dict(
                title=dict(text="<b>Sobrevivência (% de locais)</b>", font=dict(size=14, color="#111111", weight="bold")),
                tickformat=".0f",
                ticksuffix="%",
                range=[-5, 105],
                autorange=False,
                tickfont=dict(size=12, color="#111111", weight="bold"),
                gridcolor="#CCCCCC",
                griddash="dot",
                gridwidth=1,
                zeroline=False,
            ),
            legend=dict(
                title=dict(text="<b>Cultivar</b>", font=dict(size=12, color="#111111", weight="bold")),
                font=dict(size=11, color="#111111", weight="bold"),
                x=1.02, y=1,
                xanchor="left",
                bgcolor="rgba(255,255,255,0.85)",
                bordercolor="#DDDDDD",
                borderwidth=1,
            ),
            plot_bgcolor="#FAFAFA",
            paper_bgcolor="#FFFFFF",
            margin=dict(t=40, b=60, l=60, r=160),
            font=dict(family="Helvetica Neue, sans-serif", size=12, color="#111111"),
            hovermode="x unified",
        )

        n_locais_sv     = df_sv_agg["cod_fazenda"].nunique()
        n_locais_total_sv = ta_filtrado["cod_fazenda"].nunique()
        st.plotly_chart(fig_sv, use_container_width=True)
        _caption_sv = f"🌿 {doenca_surv} · {len(cultivares_sv)} cultivares · {n_locais_sv} locais avaliados"
        if n_locais_sv < n_locais_total_sv:
            _caption_sv += f" (de {n_locais_total_sv} locais ativos — os demais não têm nota registrada para esta doença)."
        else:
            _caption_sv += "."
        st.caption(_caption_sv)

        st.divider()

        # ── Mapa de colapso ───────────────────────────────────────────────────
        st.markdown("#### ⚠️ Mapa de Colapso — locais com S ou AS (nota ≤ 4)")

        colapso_rows = []
        for cultivar in cultivares_sv:
            df_c = df_sv_agg[df_sv_agg["dePara"] == cultivar]
            status     = df_c["status_material"].mode()[0] if not df_c["status_material"].mode().empty else ""
            n_total    = len(df_c)
            n_colapso  = int((df_c[col_nota_sv] <= 4).sum())
            n_mt       = int(((df_c[col_nota_sv] > 4) & (df_c[col_nota_sv] <= 6)).sum())
            n_ok       = int((df_c[col_nota_sv] > 6).sum())
            n_inc      = int((df_c[col_nota_sv].between(1, 5)).sum())
            locais_col = df_c[df_c[col_nota_sv] <= 4]["cod_fazenda"].tolist()
            locais_inc = df_c[df_c[col_nota_sv].between(1, 5)]["cod_fazenda"].tolist()
            colapso_rows.append({
                "Cultivar":              cultivar,
                "Status":                status,
                "Locais Aval.":          n_total,
                "S / AS (≤4)":           n_colapso,
                "MT (5–6)":              n_mt,
                "T / R (≥7)":            n_ok,
                "% Colapso":             round(n_colapso / n_total * 100, 1) if n_total > 0 else 0,
                "% Incidência":          round(n_inc    / n_total * 100, 1) if n_total > 0 else 0,
                "Locais em Colapso":     ", ".join(locais_col) if locais_col else "—",
                "Locais c/ Incidência":  ", ".join(locais_inc) if locais_inc else "—",
            })

        df_colapso = (
            pd.DataFrame(colapso_rows)
            .sort_values(["S / AS (≤4)", "% Colapso"], ascending=[False, False])
            .reset_index(drop=True)
        )

        # AgGrid colapso
        gb_col = GridOptionsBuilder.from_dataframe(df_colapso)
        gb_col.configure_default_column(
            resizable=True, sortable=True, filter=True,
            cellStyle={"fontSize": "13px", "fontFamily": "Helvetica Neue, sans-serif"},
        )
        gb_col.configure_grid_options(headerHeight=36, rowHeight=32, domLayout="normal")
        gb_col.configure_column("Cultivar",             pinned="left", width=170)
        gb_col.configure_column("Status",               width=90)
        gb_col.configure_column("Locais Aval.",         width=100)
        gb_col.configure_column("S / AS (≤4)",          width=100,
            cellStyle=JsCode("""function(p){
                var v=p.value;
                if(v>0) return {background:'#E63946',color:'#FFFFFF',fontWeight:'700'};
                return {background:'#D5F5D5',color:'#1A1A1A'};
            }"""))
        gb_col.configure_column("MT (5–6)",             width=90,
            cellStyle=JsCode("""function(p){
                var v=p.value;
                if(v>0) return {background:'#FFD600',color:'#1A1A1A'};
                return {};
            }"""))
        gb_col.configure_column("T / R (≥7)",           width=90,
            cellStyle=JsCode("""function(p){
                var v=p.value;
                if(v>0) return {background:'#70C96E',color:'#1A1A1A'};
                return {};
            }"""))
        gb_col.configure_column("% Colapso",            width=100,
            cellStyle=JsCode("""function(p){
                var v=p.value;
                if(v>=80) return {background:'#8B0000',color:'#FFFFFF',fontWeight:'700'};
                if(v>=50) return {background:'#E63946',color:'#FFFFFF',fontWeight:'700'};
                if(v>=20) return {background:'#FFD600',color:'#1A1A1A'};
                return {background:'#D5F5D5',color:'#1A1A1A'};
            }"""))
        gb_col.configure_column("% Incidência",         width=105,
            cellStyle=JsCode("""function(p){
                var v=p.value;
                if(v>=80) return {background:'#8B0000',color:'#FFFFFF',fontWeight:'700'};
                if(v>=50) return {background:'#E63946',color:'#FFFFFF',fontWeight:'700'};
                if(v>=20) return {background:'#FFD600',color:'#1A1A1A'};
                return {background:'#D5F5D5',color:'#1A1A1A'};
            }"""))
        gb_col.configure_column("Locais em Colapso",    width=280)
        gb_col.configure_column("Locais c/ Incidência", width=280)

        go_col = gb_col.build()
        go_col["defaultColDef"]["headerClass"] = "ag-header-black"

        AgGrid(
            df_colapso,
            gridOptions=go_col,
            height=min(620, 36 + 32 * len(df_colapso) + 20),
            update_mode=GridUpdateMode.NO_UPDATE,
            fit_columns_on_grid_load=False,
            allow_unsafe_jscode=True,
            enable_enterprise_modules=True,
            custom_css={
                ".ag-header":                      {"background-color": "#4A4A4A !important"},
                ".ag-header-row":                  {"background-color": "#4A4A4A !important"},
                ".ag-header-cell":                 {"background-color": "#4A4A4A !important"},
                ".ag-header-cell-label":           {"color": "#FFFFFF !important", "font-weight": "700"},
                ".ag-header-cell-text":            {"color": "#FFFFFF !important", "font-size": "13px !important", "font-weight": "700 !important"},
                ".ag-icon":                        {"color": "#FFFFFF !important", "opacity": "1 !important"},
                ".ag-header-icon":                 {"color": "#FFFFFF !important", "opacity": "1 !important"},
                ".ag-header-cell-menu-button":     {"opacity": "1 !important", "visibility": "visible !important"},
                ".ag-icon-menu":                   {"color": "#FFFFFF !important", "opacity": "1 !important"},
                ".ag-icon-filter":                 {"color": "#FFFFFF !important", "opacity": "1 !important"},
                ".ag-row":                         {"font-size": "13px !important"},
            },
            theme="streamlit",
            use_container_width=True,
        )
        st.caption(
            "ℹ️ **S / AS (≤4)** = nº de locais em colapso sanitário · "
            "**% Colapso** = % de locais com nota ≤4 · "
            "**% Incidência** = % de locais com doença detectada (notas 1–5) · "
            "**MT (5–6)** = zona de atenção · **T / R (≥7)** = sanidade adequada. "
            "Ordenado pelo maior nº de colapsos."
        )
        exportar_excel(df_colapso, nome_arquivo="colapso_doencas.xlsx",
                       label="⬇️ Exportar Mapa de Colapso", key="exp_colapso_dc")

st.markdown(
    '<p style="font-size:13px;color:#374151;text-align:center;margin-top:2rem;">Painel JAUM DTC · Stine Seed · '
    'Desenvolvido por <a href="https://www.linkedin.com/in/eng-agro-andre-ferreira/" '
    'target="_blank" style="color:#27AE60;text-decoration:none;">Andre Ferreira</a></p>',
    unsafe_allow_html=True,
)
