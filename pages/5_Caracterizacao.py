"""
pages/5_Caracterizacao.py — Caracterização Agronômica
"""
import numpy as np
import pandas as pd
import streamlit as st

from utils.theme import aplicar_tema, page_header, secao_titulo
from utils.loader import carregar_2023, carregar_2024, carregar_2025, carregar_av5_graos_faixa
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, JsCode

st.set_page_config(
    page_title="Caracterização Agronômica · JAUM DTC",
    page_icon="🌱",
    layout="wide",
    initial_sidebar_state="expanded",
)

aplicar_tema()

# ── CSS global ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
[data-testid="stCaptionContainer"] p,
[data-testid="stCaptionContainer"] {
    color: #374151 !important;
    opacity: 1 !important;
}
</style>
""", unsafe_allow_html=True)

# ── AgGrid helper ─────────────────────────────────────────────────────────────
AG_CSS = {
    ".ag-header":                      {"background-color": "#4A4A4A !important"},
    ".ag-header-row":                  {"background-color": "#4A4A4A !important"},
    ".ag-header-cell":                 {"background-color": "#4A4A4A !important"},
    ".ag-header-cell-label":           {"color": "#FFFFFF !important", "font-weight": "700"},
    ".ag-header-cell-text":            {"color": "#FFFFFF !important", "font-size": "13px !important", "font-weight": "700 !important"},
    ".ag-icon":                        {"color": "#FFFFFF !important", "opacity": "1 !important"},
    ".ag-header-icon":                 {"color": "#FFFFFF !important", "opacity": "1 !important"},
    ".ag-header-cell-menu-button":     {"opacity": "1 !important", "visibility": "visible !important"},
    ".ag-icon-menu":                   {"color": "#FFFFFF !important", "opacity": "1 !important"},
    ".ag-icon-filter":                 {"color": "#FFFFFF !important", "opacity": "1 !important"},
    ".ag-row":                         {"font-size": "13px !important"},
}

def ag_table(df, height=400, pinned_col="Cultivar"):
    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_default_column(
        resizable=True, sortable=True, filter=True,
        cellStyle={"fontSize": "13px", "fontFamily": "Helvetica Neue, sans-serif"},
    )
    gb.configure_grid_options(headerHeight=36, rowHeight=32, domLayout="normal")
    if pinned_col and pinned_col in df.columns:
        gb.configure_column(pinned_col, pinned="left", width=170)
    go = gb.build()
    go["defaultColDef"]["headerClass"] = "ag-header-black"
    AgGrid(
        df,
        gridOptions=go,
        height=height,
        update_mode=GridUpdateMode.NO_UPDATE,
        fit_columns_on_grid_load=False,
        allow_unsafe_jscode=True,
        enable_enterprise_modules=True,
        custom_css=AG_CSS,
        theme="streamlit",
        use_container_width=True,
    )

def exportar_excel(df, nome_arquivo="tabela.xlsx", label="⬇️ Exportar Excel", key=None):
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    buf = io.BytesIO()
    wb  = openpyxl.Workbook()
    ws  = wb.active
    df  = df.reset_index(drop=True)
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=str(col))
        cell.font      = Font(bold=True, name="Arial", size=10, color="1A1A1A")
        cell.fill      = PatternFill("solid", start_color="F2F2F2")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(12, len(str(col)) + 2)
    ws.row_dimensions[1].height = 28
    for ri, row_data in enumerate(df.itertuples(index=False), start=2):
        for ci, val in enumerate(row_data, 1):
            try:
                if val is None or (isinstance(val, float) and np.isnan(val)):
                    val = None
                elif pd.isna(val):
                    val = None
            except (TypeError, ValueError):
                pass
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="left" if ci == 1 else "center", vertical="center")
            cell.border    = border
    wb.save(buf)
    buf.seek(0)
    st.download_button(label=label, data=buf, file_name=nome_arquivo,
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=key)

# ── Carregamento ──────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def carregar_concat():
    frames = []
    for loader in [carregar_2023, carregar_2024, carregar_2025]:
        d = loader()
        if d.get("ok") and d.get("ta_faixa") is not None:
            frames.append(d["ta_faixa"])
    if not frames:
        return pd.DataFrame()
    df = pd.concat(frames, ignore_index=True)
    df["gm"]     = pd.to_numeric(df["gm"],     errors="coerce")
    df["gm_cat"] = (df["gm"] / 10).round(1)
    return df

with st.spinner("Carregando dados..."):
    ta_raw = carregar_concat()

if ta_raw.empty:
    st.error("❌ Nenhum dado disponível. Verifique a página de Diagnóstico.")
    st.stop()

if "GM_visual" in ta_raw.columns:
    med_gm = ta_raw["GM_visual"].dropna()
    med_gm = med_gm[med_gm > 0]
    if len(med_gm) > 0 and med_gm.median() > 10:
        ta_raw["GM_visual"] = (ta_raw["GM_visual"] / 10).round(1)

# ── Page header ───────────────────────────────────────────────────────────────
page_header(
    "Caracterização Agronômica",
    "Perfil morfológico e agronômico dos cultivares — floração, porte, hábito de crescimento e componentes de rendimento.",
    imagem="Creation process-bro.png",
)

# ── Sidebar — Filtros encadeados ──────────────────────────────────────────────
with st.sidebar:
    st.markdown(
        '<p style="font-size:11px;font-weight:600;color:#6B7280;text-transform:uppercase;'
        'letter-spacing:0.05em;padding:0.5rem;">Filtros</p>',
        unsafe_allow_html=True,
    )

    if st.button("🔄 Limpar filtros", use_container_width=True):
        for key in list(st.session_state.keys()):
            if any(key.startswith(p) for p in ["ca_safra_", "ca_macro_", "ca_micro_",
                                                "ca_estado_", "ca_cidade_", "ca_fazenda_",
                                                "ca_status_", "ca_cult_"]):
                del st.session_state[key]
        st.rerun()

    def checkboxes_ca(label, opcoes, default_all=True, defaults=None, prefix=""):
        sel = []
        for o in opcoes:
            checked = (o in defaults) if defaults is not None else default_all
            if st.checkbox(str(o), value=checked, key=f"{prefix}_{o}"):
                sel.append(o)
        return sel

    # 1. Safra
    with st.expander("📅 Safra", expanded=True):
        safras_all    = sorted(ta_raw["safra"].dropna().unique().tolist())
        safra_default = [s for s in safras_all if "2025" in str(s)] or safras_all[-1:]
        safras_sel    = checkboxes_ca("Safra", safras_all, defaults=safra_default, prefix="ca_safra")

    ta_f1 = ta_raw[ta_raw["safra"].isin(safras_sel)] if safras_sel else ta_raw.iloc[0:0]

    # 2. Região Macro
    with st.expander("🗺️ Região Macro", expanded=False):
        macros_all = sorted(ta_f1["regiao_macro"].dropna().unique().tolist())
        macros_sel = checkboxes_ca("Macro", macros_all, prefix="ca_macro")

    ta_f2 = ta_f1[ta_f1["regiao_macro"].isin(macros_sel)] if macros_sel else ta_f1.iloc[0:0]

    # 3. Região Micro
    with st.expander("📍 Região Micro", expanded=False):
        micros_all = sorted(ta_f2["regiao_micro"].dropna().unique().tolist())
        micros_sel = checkboxes_ca("Micro", micros_all, prefix="ca_micro")

    ta_f3 = ta_f2[ta_f2["regiao_micro"].isin(micros_sel)] if micros_sel else ta_f2.iloc[0:0]

    # 4. Estado
    with st.expander("🏛️ Estado", expanded=False):
        estados_all = sorted(ta_f3["estado_sigla"].dropna().unique().tolist())
        estados_sel = checkboxes_ca("Estado", estados_all, prefix="ca_estado")

    ta_f4 = ta_f3[ta_f3["estado_sigla"].isin(estados_sel)] if estados_sel else ta_f3.iloc[0:0]

    # 5. Cidade
    with st.expander("🏙️ Cidade", expanded=False):
        cidades_all = sorted(ta_f4["cidade_nome"].dropna().unique().tolist())
        cidades_sel = checkboxes_ca("Cidade", cidades_all, prefix="ca_cidade")

    ta_f5 = ta_f4[ta_f4["cidade_nome"].isin(cidades_sel)] if cidades_sel else ta_f4.iloc[0:0]

    # 6. Fazenda
    with st.expander("🏡 Fazenda", expanded=False):
        fazendas_all = sorted(ta_f5["nomeFazenda"].dropna().unique().tolist())
        fazendas_sel = checkboxes_ca("Fazenda", fazendas_all, prefix="ca_fazenda")

    ta_f6 = ta_f5[ta_f5["nomeFazenda"].isin(fazendas_sel)] if fazendas_sel else ta_f5.iloc[0:0]

    # 7. Status do material
    with st.expander("🏷️ Status do Material", expanded=False):
        status_all = sorted(ta_f6["status_material"].dropna().unique().tolist())
        status_sel = checkboxes_ca("Status", status_all, prefix="ca_status")

    ta_f7 = ta_f6[ta_f6["status_material"].isin(status_sel)] if status_sel else ta_f6.iloc[0:0]

    # 8. Cultivar
    with st.expander("🌱 Cultivar", expanded=False):
        cultivares_all = sorted(ta_f7["dePara"].dropna().unique().tolist())
        cultivares_sel = checkboxes_ca("Cult", cultivares_all, prefix="ca_cult")

    ta_f8 = ta_f7[ta_f7["dePara"].isin(cultivares_sel)] if cultivares_sel else ta_f7.iloc[0:0]

    # 9. GM — slider
    with st.expander("🎯 Grupo de Maturidade", expanded=False):
        gm_min = float(round(ta_f8["gm_cat"].min(), 1)) if not ta_f8.empty else 5.0
        gm_max = float(round(ta_f8["gm_cat"].max(), 1)) if not ta_f8.empty else 9.0
        if gm_min >= gm_max:
            gm_max = round(gm_min + 0.1, 1)
        gm_range = st.slider("GM", min_value=gm_min, max_value=gm_max,
                              value=(gm_min, gm_max), step=0.1, format="%.1f",
                              key="ca_gm_slider")

    ta_filtrado = ta_f8[ta_f8["gm_cat"].between(gm_range[0], gm_range[1])]

if ta_filtrado.empty:
    st.warning("⚠️ Nenhum dado para os filtros selecionados.")
    st.stop()

# ── Contexto ──────────────────────────────────────────────────────────────────
_all_safras  = sorted(ta_raw["safra"].dropna().unique().tolist())
_all_macros  = sorted(ta_raw["regiao_macro"].dropna().unique().tolist()) if "regiao_macro" in ta_raw.columns else []
_all_micros  = sorted(ta_raw["regiao_micro"].dropna().unique().tolist()) if "regiao_micro" in ta_raw.columns else []

filtros_ativos = []
if safras_sel and set(safras_sel) != set(_all_safras):
    filtros_ativos.append(" / ".join(str(s) for s in safras_sel))
if macros_sel and set(macros_sel) != set(_all_macros):
    filtros_ativos.append("Macro: " + ", ".join(macros_sel))
if micros_sel and set(micros_sel) != set(_all_micros):
    filtros_ativos.append("Micro: " + ", ".join(micros_sel))
if estados_sel:
    filtros_ativos.append(", ".join(estados_sel))

n_locais_ctx  = ta_filtrado["cod_fazenda"].nunique()
n_cidades_ctx = ta_filtrado["cidade_nome"].nunique()
contexto_str  = ("  ·  ".join(filtros_ativos) if filtros_ativos else "") + \
                f"  ·  {n_cidades_ctx} cidades · {n_locais_ctx} locais"
contexto_str  = contexto_str.lstrip("  ·  ")

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 1 — AUDITORIA
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Auditoria",
    "Quais são os dados por ensaio?",
    "Visão individual de cada observação. Use para conferência antes da análise.",
)

# Mapeamento de colunas brutas — ta_faixa
COL_MAP_AUDIT = {
    "safra":                "Safra",
    "cod_fazenda":          "Cód. Local",
    "nomeFazenda":          "Fazenda",
    "cidade_nome":          "Cidade",
    "estado_sigla":         "Estado",
    "regiao_macro":         "Região Macro",
    "regiao_micro":         "Região Micro",
    "dePara":               "Cultivar",
    "status_material":      "Status",
    "gm_cat":               "GM",
    "GM_visual":            "GM Visual",
    "sc_ha":                "sc/ha",
    "kg_ha":                "kg/ha",
    "pesoMilGraos_corrigido": "PMG (g)",
    "pop_plantas_ha":       "Pop. Inicial (pl/ha)",
    "pop_plantasFinal_ha":  "Pop. Final (pl/ha)",
    "dataInicioFloracao":   "Início Floração",
    "dataFimFloracao":      "Fim Floração",
    "corFlorNome":          "Cor Flor",
    "corFlorGrupo":         "Grupo Cor Flor",
    "habitoCrescimentoNome":"Hábito Crescimento",
    "habitoCrescimentoGrupo":"Grupo Hábito",
    "corPubNome":           "Cor Pubescência",
    "corPubGrupo":          "Grupo Cor Pub.",
    "media_AIV":            "AIV (cm)",
    "media_ALT":            "ALP (cm)",
    "media_ENG":            "ENG (nota)",
    "media_RV":             "Ramos Veg.",
    "media_RR":             "Ramos Rep.",
    "media_VTS":            "Vagens Sup.",
    "media_VTM":            "Vagens Med.",
    "media_VTI":            "Vagens Inf.",
    "media_totalVagens":    "Total Vagens",
    "DMF":                  "DMF",
    "dias_ate_DMF":         "Ciclo (dias)",
    "notaAC":               "Acamamento",
    "notaAV":               "Abert. Vagens",
    "notaQF":               "Qualidade Final",
}

# Mapeamento de colunas — tb_av5_graos
COL_MAP_GRAOS = {
    "safra":          "Safra",
    "cod_fazenda":    "Cód. Local",
    "nomeFazenda":    "Fazenda",
    "cidade_nome":    "Cidade",
    "estado_sigla":   "Estado",
    "regiao_macro":   "Região Macro",
    "regiao_micro":   "Região Micro",
    "dePara":         "Cultivar",
    "status_material":"Status",
    "planta":         "Planta",
    "totalVagensTS":  "Total Vagens Sup.",
    "mediaGraoTS":    "Grãos/Vagem Sup.",
    "totalVagensTM":  "Total Vagens Med.",
    "mediaGraoTM":    "Grãos/Vagem Med.",
    "totalVagensTI":  "Total Vagens Inf.",
    "mediaGraoTI":    "Grãos/Vagem Inf.",
}

tab_faixa, tab_graos = st.tabs(["🌱 Dados Agronômicos", "🫘 Grãos por Planta"])

with tab_faixa:
    cols_audit = [c for c in COL_MAP_AUDIT if c in ta_filtrado.columns]
    df_audit   = ta_filtrado[cols_audit].rename(columns=COL_MAP_AUDIT).copy()

    # Converter datas para string legível
    for col_data in ["Início Floração", "Fim Floração", "DMF"]:
        if col_data in df_audit.columns:
            df_audit[col_data] = pd.to_datetime(df_audit[col_data], errors="coerce").dt.strftime("%d/%m/%Y").fillna("")

    # Total Vagens com 1 casa decimal
    if "Total Vagens" in df_audit.columns:
        df_audit["Total Vagens"] = pd.to_numeric(df_audit["Total Vagens"], errors="coerce").round(1)

    ag_table(df_audit, height=min(400, 36 + 32 * len(df_audit) + 20))
    exportar_excel(df_audit, nome_arquivo="auditoria_caracterizacao.xlsx",
                   label="⬇️ Exportar Auditoria", key="exp_audit_ca")

with tab_graos:
    with st.spinner("Carregando dados de grãos..."):
        df_graos_raw = carregar_av5_graos_faixa()

    if df_graos_raw.empty:
        st.info("Dados de grãos por planta não disponíveis (disponível a partir de 2024).")
    else:
        # Aplicar os mesmos filtros da sidebar
        df_graos = df_graos_raw.copy()

        if safras_sel:
            df_graos = df_graos[df_graos["safra"].isin(safras_sel)]
        if "regiao_macro" in df_graos.columns and macros_sel:
            df_graos = df_graos[df_graos["regiao_macro"].isin(macros_sel)]
        if "regiao_micro" in df_graos.columns and micros_sel:
            df_graos = df_graos[df_graos["regiao_micro"].isin(micros_sel)]
        if "estado_sigla" in df_graos.columns and estados_sel:
            df_graos = df_graos[df_graos["estado_sigla"].isin(estados_sel)]
        if "cidade_nome" in df_graos.columns and cidades_sel:
            df_graos = df_graos[df_graos["cidade_nome"].isin(cidades_sel)]
        if "status_material" in df_graos.columns and status_sel:
            df_graos = df_graos[df_graos["status_material"].isin(status_sel)]
        if "dePara" in df_graos.columns and cultivares_sel:
            df_graos = df_graos[df_graos["dePara"].isin(cultivares_sel)]

        if df_graos.empty:
            st.info("Nenhum dado de grãos para os filtros selecionados.")
        else:
            cols_graos = [c for c in COL_MAP_GRAOS if c in df_graos.columns]
            df_graos_show = df_graos[cols_graos].rename(columns=COL_MAP_GRAOS)
            ag_table(df_graos_show, height=min(400, 36 + 32 * len(df_graos_show) + 20))
            st.caption(
                f"ℹ️ Cada linha = uma planta avaliada. "
                f"Sup. = Terço Superior · Med. = Terço Médio · Inf. = Terço Inferior. "
                f"{len(df_graos_show)} observações · 2023 não possui esta avaliação."
            )
            exportar_excel(df_graos_show, nome_arquivo="auditoria_graos.xlsx",
                           label="⬇️ Exportar Grãos", key="exp_audit_graos")

st.divider()

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 2 — TABELA RESUMO POR CULTIVAR
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Resumo por Cultivar",
    "Qual é o perfil agronômico de cada cultivar?",
    contexto_str,
)

with st.popover("ℹ️ Como interpretar", use_container_width=False):
    st.markdown("""
Cada linha representa um cultivar — médias calculadas sobre todos os locais nos filtros ativos, excluindo valores zero ou nulos.

**Colunas quantitativas** → média dos locais avaliados.

**Colunas qualitativas** (Cor da Flor, Hábito de Crescimento, Cor da Pubescência) → valor mais frequente (moda) entre os locais avaliados.

**Aba Dados Agronômicos** — produtividade, população, morfologia, floração, ciclo e notas de campo.

**Aba Grãos por Planta** — médias agregadas por cultivar das avaliações de vagens e grãos por terço da planta (disponível a partir de 2024).

**Siglas**
- **ENG** → Engalhamento · **AIV** → Altura de Inserção da 1ª Vagem · **ALP** → Altura de Planta
- **DP** → Desvio Padrão médio das plantas avaliadas no local — indica a variabilidade dentro da parcela
- **Ramos Veg.** → Ramos Vegetativos · **Ramos Rep.** → Ramos Reprodutivos
- **Vagens Sup. / Med. / Inf.** → Vagens no Terço Superior / Médio / Inferior da planta
- **Ciclo** → Dias entre plantio e DMF
- **Acamamento** → Nota de acamamento · **Abert. Vagens** → Nota de abertura de vagens · **QF** → Qualidade Final
""")

# ── Colunas numéricas — aba agronômica ────────────────────────────────────────
COLS_NUM_AGRO = {
    "sc_ha":                  "sc/ha",
    "kg_ha":                  "kg/ha",
    "pesoMilGraos_corrigido": "PMG (g)",
    "pop_plantas_ha":         "Pop. Inicial",
    "pop_plantasFinal_ha":    "Pop. Final",
    "GM_visual":              "GM Visual",
    "dias_ate_DMF":           "Ciclo (dias)",
    "media_AIV":              "AIV (cm)",
    "std_AIV":                "DP AIV",
    "media_ALT":              "ALP (cm)",
    "std_ALT":                "DP ALP",
    "media_ENG":              "ENG",
    "std_ENG":                "DP ENG",
    "media_RV":               "Ramos Veg.",
    "std_RV":                 "DP Ramos Veg.",
    "media_RR":               "Ramos Rep.",
    "std_RR":                 "DP Ramos Rep.",
    "media_VTS":              "Vagens Sup.",
    "std_VTS":                "DP Vagens Sup.",
    "media_VTM":              "Vagens Med.",
    "std_VTM":                "DP Vagens Med.",
    "media_VTI":              "Vagens Inf.",
    "std_VTI":                "DP Vagens Inf.",
    "media_totalVagens":      "Total Vagens",
    "notaAC":                 "Acamamento",
    "notaAV":                 "Abert. Vagens",
    "notaQF":                 "Qualidade Final",
}

# ── Colunas qualitativas ──────────────────────────────────────────────────────
COLS_QUAL = {
    "corFlorNome":           "Cor Flor",
    "habitoCrescimentoNome": "Hábito Crescimento",
    "corPubNome":            "Cor Pubescência",
}

# ── Colunas numéricas — aba grãos (base da agregação por cultivar) ────────────
# Ramos e vagens já estão na aba agronômica; aqui ficam só as colunas do tb_av5_graos
COLS_NUM_GRAOS: dict = {}  # preenchido abaixo após carregar av5

# ── Colunas numéricas — grãos de tb_av5_graos agregados por cultivar ─────────
COLS_AV5_RESUMO = {
    "totalVagensTS": "Total Vagens Sup.",
    "mediaGraoTS":   "Grãos/Vagem Sup.",
    "totalVagensTM": "Total Vagens Med.",
    "mediaGraoTM":   "Grãos/Vagem Med.",
    "totalVagensTI": "Total Vagens Inf.",
    "mediaGraoTI":   "Grãos/Vagem Inf.",
}

# ── Agregação por cultivar ────────────────────────────────────────────────────
resumo_rows_agro  = []
resumo_rows_graos = []

for cultivar, grp in ta_filtrado.groupby("dePara", dropna=True):
    status = grp["status_material"].mode()[0] if not grp["status_material"].mode().empty else ""
    gm     = round(grp["gm_cat"].dropna().median(), 1) if grp["gm_cat"].notna().any() else None
    base   = {"Cultivar": cultivar, "Status": status, "GM": gm, "Locais": grp["cod_fazenda"].nunique()}

    # Aba agronômica
    row_agro = base.copy()
    for col, label in COLS_NUM_AGRO.items():
        if col not in grp.columns:
            row_agro[label] = None
            continue
        s = pd.to_numeric(grp[col], errors="coerce").dropna()
        # Para médias, excluir zeros; para desvios padrão, manter zeros (válidos)
        if not col.startswith("std_") and col not in ("pop_plantas_ha", "pop_plantasFinal_ha"):
            s = s[s > 0]
        if s.empty:
            row_agro[label] = None
        else:
            row_agro[label] = int(round(s.mean(), 0)) if col in ("pop_plantas_ha", "pop_plantasFinal_ha") else round(s.mean(), 1)
    for col, label in COLS_QUAL.items():
        if col not in grp.columns:
            row_agro[label] = None
            continue
        s = grp[col].dropna().astype(str).str.strip()
        s = s[~s.isin(["", "nan", "None", "null", "NaN"])]
        row_agro[label] = s.mode().iloc[0] if not s.empty else None
    resumo_rows_agro.append(row_agro)

    # Aba grãos — linha base só com identificadores (dados vêm de tb_av5_graos)
    row_graos = base.copy()
    resumo_rows_graos.append(row_graos)

df_agro  = pd.DataFrame(resumo_rows_agro)
df_graos_resumo = pd.DataFrame(resumo_rows_graos)

# Agregar tb_av5_graos por cultivar e juntar na aba grãos
with st.spinner("Carregando dados de grãos..."):
    df_av5_raw = carregar_av5_graos_faixa()

if not df_av5_raw.empty:
    # Aplicar filtros da sidebar
    df_av5 = df_av5_raw.copy()
    if safras_sel and "safra" in df_av5.columns:
        df_av5 = df_av5[df_av5["safra"].isin(safras_sel)]
    if macros_sel and "regiao_macro" in df_av5.columns:
        df_av5 = df_av5[df_av5["regiao_macro"].isin(macros_sel)]
    if micros_sel and "regiao_micro" in df_av5.columns:
        df_av5 = df_av5[df_av5["regiao_micro"].isin(micros_sel)]
    if estados_sel and "estado_sigla" in df_av5.columns:
        df_av5 = df_av5[df_av5["estado_sigla"].isin(estados_sel)]
    if cidades_sel and "cidade_nome" in df_av5.columns:
        df_av5 = df_av5[df_av5["cidade_nome"].isin(cidades_sel)]
    if status_sel and "status_material" in df_av5.columns:
        df_av5 = df_av5[df_av5["status_material"].isin(status_sel)]
    if cultivares_sel and "dePara" in df_av5.columns:
        df_av5 = df_av5[df_av5["dePara"].isin(cultivares_sel)]

    if not df_av5.empty:
        cols_av5_disp = [c for c in COLS_AV5_RESUMO if c in df_av5.columns]
        if cols_av5_disp:
            for c in cols_av5_disp:
                df_av5[c] = pd.to_numeric(df_av5[c], errors="coerce")
                # Zeros em totalVagens indicam ausência de vagens no terço — excluir
                if c.startswith("totalVagens"):
                    df_av5[c] = df_av5[c].where(df_av5[c] > 0, other=np.nan)
            agg_av5 = (
                df_av5.groupby("dePara")[cols_av5_disp]
                .mean().round(1).reset_index()
                .rename(columns={"dePara": "Cultivar", **{c: COLS_AV5_RESUMO[c] for c in cols_av5_disp}})
            )
            df_graos_resumo = df_graos_resumo.merge(agg_av5, on="Cultivar", how="left")

# ── Função para montar e exibir AgGrid de resumo ─────────────────────────────
def _exibir_resumo(df_r, cols_num, cols_qual_labels, key_exp, filename):
    if df_r.empty:
        st.info("Nenhum dado disponível para os filtros ativos.")
        return

    if "sc/ha" in df_r.columns:
        df_r = df_r.sort_values("sc/ha", ascending=False).reset_index(drop=True)

    # Rodapé
    rodape = {"Cultivar": "Média Geral", "Status": "", "GM": "", "Locais": ""}
    for label in [v for v in cols_num.values() if v in df_r.columns]:
        vals = pd.to_numeric(df_r[label], errors="coerce").dropna()
        rodape[label] = round(vals.mean(), 1) if len(vals) > 0 else "—"
    for label in cols_qual_labels:
        rodape[label] = ""

    df_show = pd.concat([df_r, pd.DataFrame([rodape])], ignore_index=True)

    gb = GridOptionsBuilder.from_dataframe(df_show)
    gb.configure_default_column(
        resizable=True, sortable=True, filter=True,
        cellStyle={"fontSize": "13px", "fontFamily": "Helvetica Neue, sans-serif"},
    )
    gb.configure_grid_options(headerHeight=36, rowHeight=32, domLayout="normal")
    gb.configure_column("Cultivar", pinned="left", width=170)
    gb.configure_column("Status",   width=90)
    gb.configure_column("GM",       width=70)
    gb.configure_column("Locais",   width=80)

    widths_all = {
        "sc/ha": 80, "kg/ha": 90, "PMG (g)": 90, "Pop. Inicial": 120, "Pop. Final": 110,
        "GM Visual": 90, "Ciclo (dias)": 100,
        "ENG": 70, "DP ENG": 80, "AIV (cm)": 80, "DP AIV": 80, "ALP (cm)": 80, "DP ALP": 80,
        "Acamamento": 110, "Abert. Vagens": 120, "Qualidade Final": 120,
        "Cor Flor": 120, "Hábito Crescimento": 160, "Cor Pubescência": 140,
        "Ramos Veg.": 100, "DP Ramos Veg.": 110, "Ramos Rep.": 100, "DP Ramos Rep.": 110,
        "Vagens Sup.": 100, "DP Vagens Sup.": 110,
        "Vagens Med.": 100, "DP Vagens Med.": 110,
        "Vagens Inf.": 100, "DP Vagens Inf.": 110,
        "Total Vagens": 110,
        "Total Vagens Sup.": 140, "Grãos/Vagem Sup.": 130,
        "Total Vagens Med.": 140, "Grãos/Vagem Med.": 130,
        "Total Vagens Inf.": 140, "Grãos/Vagem Inf.": 130,
    }

    last_row = JsCode("""function(p) {
        if (p.rowIndex === p.api.getDisplayedRowCount() - 1)
            return {background: '#D9D9D9', fontWeight: '700', color: '#1A1A1A'};
    }""")

    for col in df_show.columns:
        if col in ("Cultivar", "Status", "GM", "Locais"):
            continue
        gb.configure_column(col, width=widths_all.get(col, 90), cellStyle=last_row)

    go = gb.build()
    go["defaultColDef"]["headerClass"] = "ag-header-black"

    AgGrid(
        df_show,
        gridOptions=go,
        height=min(600, 36 + 32 * len(df_show) + 20),
        update_mode=GridUpdateMode.NO_UPDATE,
        fit_columns_on_grid_load=False,
        allow_unsafe_jscode=True,
        enable_enterprise_modules=True,
        custom_css=AG_CSS,
        theme="streamlit",
        use_container_width=True,
    )

    n_loc_total = ta_filtrado["cod_fazenda"].nunique()
    st.caption(
        f"ℹ️ Valores numéricos = média dos locais com dado válido (> 0). "
        f"Colunas qualitativas = valor mais frequente (moda). "
        f"{len(df_r)} cultivares · {n_loc_total} locais nos filtros ativos."
    )
    exportar_excel(df_r, nome_arquivo=filename, label="⬇️ Exportar", key=key_exp)


# ── Exibir abas ───────────────────────────────────────────────────────────────
tab_r_agro, tab_r_graos = st.tabs(["🌱 Dados Agronômicos", "🫘 Grãos por Planta"])

with tab_r_agro:
    _exibir_resumo(df_agro, COLS_NUM_AGRO, list(COLS_QUAL.values()),
                   "exp_resumo_agro", "resumo_agronomico.xlsx")

with tab_r_graos:
    with st.popover("ℹ️ Como interpretar · Grãos por Planta", use_container_width=False):
        st.markdown("""
Esta aba mostra dados da avaliação de grãos por terço da planta — disponível a partir de 2024. Cada cultivar é avaliado em 5 plantas por parcela; os valores aqui são a média dessas 5 plantas, depois média entre os locais.

**Total Vagens Sup. / Med. / Inf.**
Quantas vagens foram contadas no terço Superior, Médio e Inferior da planta. Reflete a distribuição de vagens ao longo do caule.
- Cultivares com mais vagens no terço médio tendem a ter perfil mais estável
- Terço superior muito carregado pode indicar risco de acamamento

**Grãos/Vagem Sup. / Med. / Inf.**
Média de grãos por vagem em cada terço. Calculado como a soma ponderada de vagens com 1, 2, 3 e 4 grãos dividida pelo total de vagens.
- Valores típicos: 2.0 a 3.0 grãos/vagem
- Terço médio geralmente tem a maior média — é o terço mais produtivo
- Cultivares com mais grãos/vagem tendem a ter maior produtividade, especialmente quando combinado com alto total de vagens

**Como usar junto com a aba Dados Agronômicos**
Compare o **Total Vagens** (soma dos três terços, já na aba agronômica) com os **Grãos/Vagem** aqui. Um cultivar pode ter poucas vagens mas muitos grãos por vagem — ou muitas vagens com poucos grãos. O cruzamento com sc/ha revela qual estratégia converte melhor em produção.
""")
    _exibir_resumo(df_graos_resumo, COLS_AV5_RESUMO, [],
                   "exp_resumo_graos", "resumo_graos.xlsx")

st.divider()

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 3 — PERFIL VISUAL DA PLANTA
# ════════════════════════════════════════════════════════════════════════════════
import streamlit.components.v1 as components
import base64
from pathlib import Path as _Path

secao_titulo(
    "Perfil Visual da Planta",
    "Como as vagens se distribuem ao longo da planta?",
    contexto_str,
)

with st.popover("ℹ️ Como interpretar", use_container_width=False):
    st.markdown("""
A visualização mostra a distribuição percentual de vagens pelos três terços da planta para o cultivar selecionado, calculada sobre os locais nos filtros ativos.

**% Vagens por terço** → proporção de vagens em cada terço em relação ao total.
- **Terço Sup.** → vagens no terço superior da planta
- **Terço Med.** → vagens no terço médio — geralmente o mais carregado
- **Terço Inf.** → vagens no terço inferior

**Métricas laterais** → médias do cultivar nos locais avaliados:
- **ALP** → Altura de planta (cm)
- **PMG** → Peso de mil grãos corrigido (g)
- **AIV** → Altura de inserção da 1ª vagem (cm)

Cultivares com distribuição mais equilibrada entre os terços tendem a ter maior estabilidade produtiva.
""")

# Só mostrar se tiver dados de vagens
_cols_vagens = ["media_VTS", "media_VTM", "media_VTI"]
_tem_vagens  = all(c in ta_filtrado.columns for c in _cols_vagens)

if not _tem_vagens or df_agro.empty:
    st.info("Dados de distribuição de vagens não disponíveis para os filtros ativos.")
else:
    # Carregar imagem soja.png como base64 — uma vez só
    _img_path = _Path(__file__).parent.parent / "assets" / "soja.png"
    _img_b64  = ""
    if _img_path.exists():
        with open(_img_path, "rb") as _f:
            _img_b64 = base64.b64encode(_f.read()).decode()
    _img_src = f"data:image/png;base64,{_img_b64}" if _img_b64 else ""

    def _fmt(v, suffix=""):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return "—"
        return f"{v}{suffix}"

    def _card_html(row, titulo):
        vts = row.get("Vagens Sup.", None)
        vtm = row.get("Vagens Med.", None)
        vti = row.get("Vagens Inf.", None)
        # Excluir None, NaN e zeros do total (zero pode indicar ausência de dado)
        _vals = [v for v in [vts, vtm, vti] if v is not None and not pd.isna(v) and v > 0]
        total_v = sum(_vals)
        pct_sup = round(vts / total_v * 100, 1) if total_v > 0 and vts is not None and not pd.isna(vts) and vts > 0 else None
        pct_med = round(vtm / total_v * 100, 1) if total_v > 0 and vtm is not None and not pd.isna(vtm) and vtm > 0 else None
        pct_inf = round(vti / total_v * 100, 1) if total_v > 0 and vti is not None and not pd.isna(vti) and vti > 0 else None

        sem_vagens = total_v == 0

        alp    = row.get("ALP (cm)", None)
        pmg    = row.get("PMG (g)",  None)
        aiv    = row.get("AIV (cm)", None)
        flor   = row.get("Cor Flor", None)
        habito = row.get("Hábito Crescimento", None)
        pub    = row.get("Cor Pubescência", None)

        img_tag = f"<img src='{_img_src}' alt='Soja'/>" if _img_src else \
                  "<div style='width:280px;height:520px;background:#f3f4f6;border-radius:8px;display:flex;align-items:center;justify-content:center;color:#9CA3AF;font-size:12px;'>soja.png</div>"

        ciclo = row.get("Ciclo (dias)", None)

        return f"""
<style>
.pw{{display:flex;align-items:flex-start;justify-content:center;font-family:'Helvetica Neue',sans-serif;}}
.pe{{position:relative;width:150px;height:500px;flex-shrink:0;}}
.pd{{position:relative;width:170px;height:500px;flex-shrink:0;}}
.pct-bloco{{position:absolute;right:10px;text-align:right;}}
.tp{{font-size:30px;font-weight:800;color:#6B7280;line-height:1;}}
.tl{{font-size:11px;color:#9CA3AF;margin-top:3px;}}
.met-bloco{{position:absolute;left:14px;border-left:3px solid #E5E7EB;padding-left:10px;}}
.mt{{font-size:10px;color:#6B7280;font-weight:600;text-transform:uppercase;letter-spacing:.04em;}}
.mv{{font-size:16px;font-weight:700;color:#111827;}}
.mv-sm{{font-size:13px;font-weight:600;color:#374151;}}
.pi{{width:260px;flex-shrink:0;}}
.pi img{{width:100%;height:auto;display:block;}}
.tc{{text-align:center;font-size:15px;font-weight:700;color:#374151;letter-spacing:.06em;text-transform:uppercase;margin-bottom:12px;}}
</style>
<div class="tc">{titulo}</div>
<div class="pw">

  <!-- Esquerda: % posicionados proporcionalmente -->
  <div class="pe">
    {'<div class="pct-bloco" style="top:30px;"><div class="tp">' + _fmt(pct_sup,"%") + '</div><div class="tl">Terço Superior</div></div>' if not sem_vagens else '<div style="position:absolute;top:120px;right:10px;text-align:right;font-size:12px;color:#9CA3AF;line-height:1.4;">Avaliação de<br>vagens não<br>disponível</div>'}
    {'<div class="pct-bloco" style="top:200px;"><div class="tp">' + _fmt(pct_med,"%") + '</div><div class="tl">Terço Médio</div></div>' if not sem_vagens else ''}
    {'<div class="pct-bloco" style="top:370px;"><div class="tp">' + _fmt(pct_inf,"%") + '</div><div class="tl">Terço Inferior</div></div>' if not sem_vagens else ''}
  </div>

  <!-- Centro: imagem -->
  <div class="pi">{img_tag}</div>

  <!-- Direita: métricas posicionadas proporcionalmente -->
  <div class="pd">
    <div class="met-bloco" style="top:30px;">
      <div class="mt">Altura de Planta</div>
      <div class="mv">{_fmt(alp," cm")}</div>
      <div class="mt" style="margin-top:8px;">Ciclo</div>
      <div class="mv">{_fmt(ciclo," dias")}</div>
    </div>
    <div class="met-bloco" style="top:200px;">
      <div class="mt">Peso Mil Grãos</div>
      <div class="mv">{_fmt(pmg," g")}</div>
      <div class="mt" style="margin-top:8px;">Inserção 1ª Vagem</div>
      <div class="mv">{_fmt(aiv," cm")}</div>
    </div>
    <div class="met-bloco" style="top:360px;">
      <div class="mt">Cor da Flor</div>
      <div class="mv-sm">{_fmt(flor)}</div>
      <div class="mt" style="margin-top:5px;">Hábito</div>
      <div class="mv-sm">{_fmt(habito)}</div>
      <div class="mt" style="margin-top:5px;">Cor Pubescência</div>
      <div class="mv-sm">{_fmt(pub)}</div>
    </div>
  </div>

</div>"""

    # ── Wrapper com botão de download ────────────────────────────────────────
    def _card_com_download_ca(card_html_inner, cultivar_nome, card_id):
        _nome_arquivo = cultivar_nome.replace(" ", "_").replace("/", "-")
        return f"""
<script src="https://cdnjs.cloudflare.com/ajax/libs/html2canvas/1.4.1/html2canvas.min.js"></script>
<div id="{card_id}" style="background:#FFFFFF;padding:16px 12px 12px;
     border-radius:12px;border:1px solid #E5E7EB;">
{card_html_inner}
</div>
<div style="text-align:center;margin-top:10px;">
  <button onclick="
    html2canvas(document.getElementById(\'{card_id}\'),{{
      backgroundColor:\'#FFFFFF\',scale:2,useCORS:true,allowTaint:true
    }}).then(function(c){{
      var a=document.createElement(\'a\');
      a.download=\'{_nome_arquivo}_perfil.png\';
      a.href=c.toDataURL(\'image/png\');
      a.click();
    }});"
    style="background:#27AE60;color:#FFFFFF;border:none;border-radius:8px;
           padding:7px 18px;font-size:12px;font-weight:600;cursor:pointer;
           font-family:\'Helvetica Neue\',sans-serif;letter-spacing:.02em;">
    ⬇ Baixar imagem
  </button>
</div>"""

    # ── Separar cultivares por grupo ─────────────────────────────────────────
    STATUS_STINE = {"STINE", "DP2", "LINHAGEM"}
    STATUS_CHECK = {"CHECK"}

    df_stine = df_agro[df_agro["Status"].isin(STATUS_STINE)].copy()
    df_check = df_agro[df_agro["Status"].isin(STATUS_CHECK)].copy()

    col_l, col_r = st.columns(2)

    with col_l:
        cultivares_stine = df_stine["Cultivar"].dropna().tolist()
        if cultivares_stine:
            cult_s = st.selectbox("Stine / DP2 / Linhagem", options=cultivares_stine, key="ca_cult_stine")
            row_s  = df_stine[df_stine["Cultivar"] == cult_s].iloc[0]
            components.html(
                _card_com_download_ca(
                    _card_html(row_s, f"Caracterização — {cult_s}"),
                    cult_s, "ca-card-stine"
                ),
                height=700, scrolling=False,
            )
            n_s = ta_filtrado[ta_filtrado["dePara"] == cult_s]["cod_fazenda"].nunique()
            st.caption(f"ℹ️ {n_s} locais avaliados.")
        else:
            st.info("Nenhum material Stine/DP2/Linhagem nos filtros ativos.")

    with col_r:
        cultivares_check = df_check["Cultivar"].dropna().tolist()
        if cultivares_check:
            cult_c = st.selectbox("Check", options=cultivares_check, key="ca_cult_check")
            row_c  = df_check[df_check["Cultivar"] == cult_c].iloc[0]
            components.html(
                _card_com_download_ca(
                    _card_html(row_c, f"Caracterização — {cult_c}"),
                    cult_c, "ca-card-check"
                ),
                height=700, scrolling=False,
            )
            n_c = ta_filtrado[ta_filtrado["dePara"] == cult_c]["cod_fazenda"].nunique()
            st.caption(f"ℹ️ {n_c} locais avaliados.")
        else:
            st.info("Nenhum Check nos filtros ativos.")

st.divider()

st.markdown(
    '<p style="font-size:13px;color:#374151;text-align:center;">Painel JAUM DTC · Stine Seed · '
    'Desenvolvido por <a href="https://www.linkedin.com/in/eng-agro-andre-ferreira/" '
    'target="_blank" style="color:#27AE60;text-decoration:none;">Andre Ferreira</a></p>',
    unsafe_allow_html=True,
)
