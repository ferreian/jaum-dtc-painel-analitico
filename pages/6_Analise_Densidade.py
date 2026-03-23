"""
pages/7_Analise_Densidade.py — Análise Conjunta de Densidade
"""
import numpy as np
import pandas as pd
import streamlit as st

from utils.theme import aplicar_tema, page_header, secao_titulo
from utils.loader import carregar_2024, carregar_2025
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, JsCode, ColumnsAutoSizeMode

COR_STATUS_PLOT = {
    "CHECK":    "#F4B184",
    "STINE":    "#2976B6",
    "LINHAGEM": "#00FF01",
    "DP2":      "#C4DFB4",
}
COR_BORDA = {
    "CHECK":    "#C46A3A",
    "STINE":    "#1A4F7A",
    "LINHAGEM": "#009900",
    "DP2":      "#7AAF6A",
}

st.set_page_config(
    page_title="Análise Densidade · JAUM DTC",
    page_icon="🌿",
    layout="wide",
    initial_sidebar_state="expanded",
)

aplicar_tema()

# ── AgGrid helper ─────────────────────────────────────────────────────────────
def ag_table(df, height=400):
    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_default_column(
        resizable=True, sortable=True, filter=True,
        suppressMenu=False,
        menuTabs=["generalMenuTab", "filterMenuTab", "columnsMenuTab"],
        cellStyle={"fontSize": "14px", "color": "#000000", "fontFamily": "Helvetica Neue, sans-serif"},
    )
    gb.configure_grid_options(
        headerHeight=36, rowHeight=32, domLayout="normal",
        suppressMenuHide=True, suppressColumnVirtualisation=True,
        suppressContextMenu=False, enableRangeSelection=True,
    )
    go = gb.build()
    go["defaultColDef"]["headerClass"] = "ag-header-black"
    go["onFirstDataRendered"] = JsCode("function(params) { params.api.sizeColumnsToFit(); }")
    AgGrid(
        df, gridOptions=go, height=height,
        update_mode=GridUpdateMode.NO_UPDATE,
        fit_columns_on_grid_load=False, columns_auto_size_mode=2,
        allow_unsafe_jscode=True, enable_enterprise_modules=True,
        custom_css={
            ".ag-header":                      {"background-color": "#4A4A4A !important"},
            ".ag-header-row":                  {"background-color": "#4A4A4A !important"},
            ".ag-header-cell":                 {"background-color": "#4A4A4A !important"},
            ".ag-header-cell-label":           {"color": "#FFFFFF !important", "font-weight": "700"},
            ".ag-header-cell-text":            {"color": "#FFFFFF !important", "font-size": "13px !important", "font-weight": "700 !important"},
            ".ag-icon":                        {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-header-icon":                 {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-header-cell-menu-button":     {"opacity": "1 !important", "visibility": "visible !important"},
            ".ag-header-cell-menu-button span":{"color": "#FFFFFF !important"},
            ".ag-icon-menu":                   {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-icon-filter":                 {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-cell":                        {"font-size": "13px !important", "color": "#000000 !important"},
            ".ag-row":                         {"font-size": "13px !important"},
        },
        theme="streamlit", use_container_width=True,
    )


def exportar_excel(df, nome_arquivo="tabela.xlsx", label="⬇️ Exportar Excel", key=None):
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    buf = io.BytesIO()
    wb  = openpyxl.Workbook()
    ws  = wb.active
    df  = df.reset_index(drop=True)
    df  = df.loc[:, ~df.columns.str.startswith("::")].copy()
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=str(col))
        cell.font      = Font(bold=True, name="Arial", size=10, color="1A1A1A")
        cell.fill      = PatternFill("solid", start_color="F2F2F2")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(12, len(str(col)) + 2)
    ws.row_dimensions[1].height = 28
    for ri, row_data in enumerate(df.itertuples(index=False), start=2):
        for ci, val in enumerate(row_data, 1):
            try:
                if val is None or (isinstance(val, float) and np.isnan(val)):
                    val = None
                elif pd.isna(val):
                    val = None
            except (TypeError, ValueError):
                pass
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="left" if ci == 1 else "center", vertical="center")
            cell.border    = border
    wb.save(buf)
    buf.seek(0)
    st.download_button(label=label, data=buf, file_name=nome_arquivo,
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=key)


st.markdown("""
<style>
[data-testid="stDataFrame"] td,
[data-testid="stDataFrame"] th,
[data-testid="stDataFrame"] [role="columnheader"],
[data-testid="stDataFrame"] [role="columnheader"] span,
[data-testid="stDataFrame"] [role="columnheader"] div {
    font-size: 14px !important; font-weight: 700 !important;
    color: #000000 !important; opacity: 1 !important;
}
[data-testid="stCaptionContainer"] p,
[data-testid="stCaptionContainer"] { color: #374151 !important; opacity: 1 !important; }
</style>
""", unsafe_allow_html=True)

page_header(
    "Análise Conjunta · Densidade",
    "Avalie o efeito da população de plantas na produtividade. Compare cultivares em diferentes densidades de plantio e identifique os materiais mais responsivos e estáveis.",
    imagem="Business mission-amico.png",
)

# ── Carregamento — ta_densidade (2024 + 2025, sem 2023) ───────────────────────
@st.cache_data(show_spinner=False)
def carregar_concat_dens():
    frames = []
    for loader in [carregar_2024, carregar_2025]:
        d = loader()
        if d.get("ok") and d.get("ta_densidade") is not None:
            df = d["ta_densidade"]
            if not df.empty:
                frames.append(df)
    if not frames:
        return pd.DataFrame()
    df = pd.concat(frames, ignore_index=True)
    df["gm"]     = pd.to_numeric(df["gm"],     errors="coerce")
    df["gm_cat"] = (df["gm"] / 10).round(1)

    # ── Corrigir populações digitadas em mil plantas/ha (ex: 300 → 300.000) ──
    if "populacao" in df.columns:
        df["populacao"] = pd.to_numeric(df["populacao"], errors="coerce")
        mask = df["populacao"] < 1000
        df.loc[mask, "populacao"] = df.loc[mask, "populacao"] * 1000
        # Remover valores absurdos (> 1.000.000)
        df.loc[df["populacao"] > 1_000_000, "populacao"] = pd.NA

    # ── Criar pop_grupo automaticamente via K-Means ───────────────────────────
    if "populacao" in df.columns:
        try:
            from sklearn.cluster import KMeans
            pop_valid = df["populacao"].dropna()
            pop_valid = pop_valid[pop_valid > 0]
            if len(pop_valid) >= 4:
                # Detectar número ideal de clusters (max 6)
                valores = pop_valid.unique()
                n_clusters = min(len(valores), 6)
                km = KMeans(n_clusters=n_clusters, random_state=42, n_init=10)
                km.fit(pop_valid.values.reshape(-1, 1))
                centros = sorted(km.cluster_centers_.flatten())
                # Nomear grupos pelo centro arredondado para milhares
                def _nome_grupo(centro):
                    return f"{round(centro / 1000):.0f}k"
                nomes = [_nome_grupo(c) for c in centros]
                # Atribuir grupo a cada linha
                def _atribuir_grupo(p):
                    if pd.isna(p) or p <= 0:
                        return None
                    idx = int(np.argmin([abs(p - c) for c in centros]))
                    return nomes[idx]
                df["pop_grupo"] = df["populacao"].apply(_atribuir_grupo)
            else:
                df["pop_grupo"] = df["populacao"].apply(
                    lambda p: f"{round(p/1000):.0f}k" if pd.notna(p) and p > 0 else None
                )
        except Exception:
            # Fallback: arredondar para múltiplo de 50k mais próximo
            df["pop_grupo"] = df["populacao"].apply(
                lambda p: f"{round(p/50000)*50:.0f}k" if pd.notna(p) and p > 0 else None
            )

    return df

with st.spinner("Carregando dados..."):
    ta_raw = carregar_concat_dens()

if ta_raw.empty:
    st.error("❌ Nenhum dado de densidade disponível. Verifique a página de Diagnóstico.")
    st.stop()

if "GM_visual" in ta_raw.columns:
    med_gm = ta_raw["GM_visual"].dropna()
    med_gm = med_gm[med_gm > 0]
    if len(med_gm) > 0 and med_gm.median() > 10:
        ta_raw["GM_visual"] = (ta_raw["GM_visual"] / 10).round(1)

# ── Sidebar — Filtros encadeados (prefixo den_) ───────────────────────────────
with st.sidebar:
    st.markdown(
        '<p style="font-size:11px;font-weight:600;color:#6B7280;text-transform:uppercase;'
        'letter-spacing:0.05em;padding:0.5rem;">Filtros</p>',
        unsafe_allow_html=True,
    )

    if st.button("🔄 Limpar filtros", use_container_width=True):
        for key in list(st.session_state.keys()):
            if any(key.startswith(p) for p in ["den_safra_","den_macro_","den_micro_","den_estado_",
                                                "den_cidade_","den_fazenda_","den_resp_","den_status_",
                                                "den_cult_","den_pop_","den_gm_"]):
                del st.session_state[key]
        st.rerun()

    def checkboxes_den(label, opcoes, default_all=True, defaults=None, prefix=""):
        sel = []
        for o in opcoes:
            checked = (o in defaults) if defaults is not None else default_all
            if st.checkbox(str(o), value=checked, key=f"{prefix}_{o}"):
                sel.append(o)
        return sel

    # 1. Safra — padrão só 2025/26
    with st.expander("📅 Safra", expanded=True):
        safras_all    = sorted(ta_raw["safra"].dropna().unique().tolist())
        safra_default = [s for s in safras_all if "2025" in str(s)] or safras_all[-1:]
        safras_sel    = checkboxes_den("Safra", safras_all, defaults=safra_default, prefix="den_safra")

    ta_f1 = ta_raw[ta_raw["safra"].isin(safras_sel)] if safras_sel else ta_raw.iloc[0:0]

    # 2. Região Macro
    with st.expander("🗺️ Região Macro", expanded=False):
        macros_all = sorted(ta_f1["regiao_macro"].dropna().unique().tolist())
        macros_sel = checkboxes_den("Macro", macros_all, prefix="den_macro")

    ta_f2 = ta_f1[ta_f1["regiao_macro"].isin(macros_sel)] if macros_sel else ta_f1.iloc[0:0]

    # 3. Região Micro
    with st.expander("📍 Região Micro", expanded=False):
        micros_all = sorted(ta_f2["regiao_micro"].dropna().unique().tolist())
        micros_sel = checkboxes_den("Micro", micros_all, prefix="den_micro")

    ta_f3 = ta_f2[ta_f2["regiao_micro"].isin(micros_sel)] if micros_sel else ta_f2.iloc[0:0]

    # 4. Estado
    with st.expander("🏛️ Estado", expanded=False):
        estados_all = sorted(ta_f3["estado_sigla"].dropna().unique().tolist())
        estados_sel = checkboxes_den("Estado", estados_all, prefix="den_estado")

    ta_f4 = ta_f3[ta_f3["estado_sigla"].isin(estados_sel)] if estados_sel else ta_f3.iloc[0:0]

    # 5. Cidade
    with st.expander("🏙️ Cidade", expanded=False):
        cidades_all = sorted(ta_f4["cidade_nome"].dropna().unique().tolist())
        cidades_sel = checkboxes_den("Cidade", cidades_all, prefix="den_cidade")

    ta_f5 = ta_f4[ta_f4["cidade_nome"].isin(cidades_sel)] if cidades_sel else ta_f4.iloc[0:0]

    # 6. Fazenda
    with st.expander("🏡 Fazenda", expanded=False):
        fazendas_all = sorted(ta_f5["nomeFazenda"].dropna().unique().tolist())
        fazendas_sel = checkboxes_den("Fazenda", fazendas_all, prefix="den_fazenda")

    ta_f6 = ta_f5[ta_f5["nomeFazenda"].isin(fazendas_sel)] if fazendas_sel else ta_f5.iloc[0:0]

    # 7. Responsável
    with st.expander("👤 Responsável", expanded=False):
        resps_all = sorted(ta_f6["nomeResponsavel"].dropna().unique().tolist())
        resps_sel = checkboxes_den("Resp", resps_all, prefix="den_resp")

    ta_f7 = ta_f6[ta_f6["nomeResponsavel"].isin(resps_sel)] if resps_sel else ta_f6.iloc[0:0]

    # 8. Status do material
    with st.expander("🏷️ Status do Material", expanded=False):
        status_all = sorted(ta_f7["status_material"].dropna().unique().tolist())
        status_sel = checkboxes_den("Status", status_all, prefix="den_status")

    ta_f8 = ta_f7[ta_f7["status_material"].isin(status_sel)] if status_sel else ta_f7.iloc[0:0]

    # 9. Cultivar
    with st.expander("🌱 Cultivar", expanded=False):
        cultivares_all = sorted(ta_f8["dePara"].dropna().unique().tolist())
        cultivares_sel = checkboxes_den("Cult", cultivares_all, prefix="den_cult")

    ta_f9 = ta_f8[ta_f8["dePara"].isin(cultivares_sel)] if cultivares_sel else ta_f8.iloc[0:0]

    # 10. Grupo de Densidade
    with st.expander("👥 Grupo de Densidade", expanded=True):
        if "pop_grupo" in ta_f9.columns:
            grupos_all = sorted(
                ta_f9["pop_grupo"].dropna().unique().tolist(),
                key=lambda x: int(x.replace("k",""))
            )
        else:
            grupos_all = []
        grupos_sel = checkboxes_den("Pop", grupos_all, default_all=True, prefix="den_pop")

    ta_f10 = ta_f9[ta_f9["pop_grupo"].isin(grupos_sel)] if grupos_sel and "pop_grupo" in ta_f9.columns else ta_f9.iloc[0:0]

    # 11. GM — slider
    with st.expander("🎯 Grupo de Maturidade", expanded=False):
        gm_min = float(ta_f10["gm_cat"].min()) if not ta_f10.empty else 5.5
        gm_max = float(ta_f10["gm_cat"].max()) if not ta_f10.empty else 8.5
        gm_min = max(5.5, round(gm_min, 1))
        gm_max = min(8.5, round(gm_max, 1))
        if gm_min >= gm_max:
            gm_max = min(gm_min + 0.1, 8.5)
        gm_range = st.slider("GM", min_value=5.5, max_value=8.5,
                              value=(gm_min, gm_max), step=0.1, format="%.1f",
                              key="den_gm_slider")

    ta_filtrado = ta_f10[ta_f10["gm_cat"].between(gm_range[0], gm_range[1])]

if ta_filtrado.empty:
    st.warning("⚠️ Nenhum dado para os filtros selecionados.")
    st.stop()

# ── Contexto para subtítulos ──────────────────────────────────────────────────
_all_safras = sorted(ta_raw["safra"].dropna().unique().tolist())
filtros_ativos = []
if safras_sel and set(safras_sel) != set(_all_safras):
    filtros_ativos.append(" / ".join(str(s) for s in safras_sel))
if macros_sel:
    filtros_ativos.append("Macro: " + ", ".join(macros_sel))
if estados_sel:
    filtros_ativos.append(", ".join(estados_sel))
n_fazendas_ctx = ta_filtrado["cod_fazenda"].nunique()
n_cidades_ctx  = ta_filtrado["cidade_nome"].nunique()
contexto_str   = ("  ·  ".join(filtros_ativos) if filtros_ativos else "Todos os dados") + \
                 f"  ·  {n_cidades_ctx} cidades · {n_fazendas_ctx} locais"

# ── Seletor Produção Relativa (seção auditoria) ───────────────────────────────
col_ref, col_test, _ = st.columns([2, 2, 3])
with col_ref:
    base_rel = st.selectbox(
        "Base da Produção Relativa",
        options=["Média geral do ensaio", "Maior produtividade", "Testemunha"],
        index=0, key="den_base_rel",
    )
with col_test:
    if base_rel == "Testemunha":
        testemunhas = sorted(
            ta_filtrado[ta_filtrado["status_material"].isin(["CHECK","STINE"])]["dePara"].dropna().unique().tolist()
        )
        if testemunhas:
            testemunha_sel = st.selectbox("Selecione a testemunha", options=testemunhas, key="den_test_sel")
        else:
            st.warning("Nenhuma testemunha disponível nos filtros atuais.")
            testemunha_sel = None
    else:
        testemunha_sel = None

df_tabela = ta_filtrado.copy()
if base_rel == "Média geral do ensaio":
    ref_valor = df_tabela["kg_ha"].mean()
elif base_rel == "Maior produtividade":
    ref_valor = df_tabela["kg_ha"].max()
elif base_rel == "Testemunha" and testemunha_sel:
    ref_valor = df_tabela[df_tabela["dePara"] == testemunha_sel]["kg_ha"].mean()
else:
    ref_valor = df_tabela["kg_ha"].mean()

df_tabela["prod_relativa_pct"] = ((df_tabela["kg_ha"] / ref_valor) * 100).round(1) if ref_valor else np.nan

# ── Classificação de doenças por nota ────────────────────────────────────────
def classificar_doenca(nota):
    if pd.isna(nota):   return np.nan
    elif nota == 0:     return "Ausente"
    elif nota <= 2:     return "AS"
    elif nota <= 4:     return "S"
    elif nota <= 6:     return "MT"
    elif nota <= 8:     return "T"
    else:               return "AT"

_COLUNAS_DOENCA = [
    "notaPhytophthora", "notaAnomalia",    "notaOidio",
    "notaManchaParda",  "notaManchaAlvo",  "notaManchaOlhoRa",
    "notaCercospora",   "notaAntracnose",  "notaDFC",
]
for _col in _COLUNAS_DOENCA:
    if _col in df_tabela.columns:
        _col_class = _col.replace("nota", "class_nota")
        df_tabela[_col_class] = pd.to_numeric(df_tabela[_col], errors="coerce").apply(classificar_doenca)

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 1 — AUDITORIA
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Auditoria",
    "Quais são os dados por ensaio?",
    "Visão individual de cada observação. Use para auditoria e conferência dos dados antes da análise.",
)

col_map = {
    # ── Identificação ────────────────────────────────────────────────────────
    "safra":               "Safra",
    "cod_fazenda":         "Cód. Local",
    "nomeFazenda":         "Fazenda",
    "cidade_nome":         "Cidade",
    "estado_sigla":        "Estado",
    "regiao_macro":        "Região Macro",
    "regiao_micro":        "Região Micro",
    "nomeResponsavel":     "Responsável",
    "dePara":              "Cultivar",
    "status_material":     "Status",
    "gm_cat":              "GM",
    "GM_visual":           "GM Visual",
    # ── Densidade / Produtividade ─────────────────────────────────────────────
    "populacao":           "População (pl/ha)",
    "pop_grupo":           "Grupo Densidade",
    "kg_ha":               "kg/ha",
    "sc_ha":               "sc/ha",
    "prod_relativa_pct":   "Prod. Relativa (%)",
    "umidadeParcela":      "Umidade (%)",
    "pop_plantasFinal_ha": "Pop. Final (pl/ha)",
    # ── Agronômicas ───────────────────────────────────────────────────────────
    "media_AIV":           "AIV (cm)",
    "media_ALT":           "ALP (cm)",
    "notaAC":              "Acamamento",
    "dias_ate_DMF":        "Ciclo (dias)",
    "pesoMilGraos_corrigido": "PMG (g)",
    # ── Doenças — nota e classificação ───────────────────────────────────────
    "notaPhytophthora":        "Phytophthora (nota)",
    "class_notaPhytophthora":  "Phytophthora (class)",
    "notaAnomalia":            "Anomalia (nota)",
    "class_notaAnomalia":      "Anomalia (class)",
    "notaOidio":               "Oídio (nota)",
    "class_notaOidio":         "Oídio (class)",
    "notaManchaParda":         "M. Parda (nota)",
    "class_notaManchaParda":   "M. Parda (class)",
    "notaManchaAlvo":          "M. Alvo (nota)",
    "class_notaManchaAlvo":    "M. Alvo (class)",
    "notaManchaOlhoRa":        "M. Olho de Rã (nota)",
    "class_notaManchaOlhoRa":  "M. Olho de Rã (class)",
    "notaCercospora":          "Cercospora (nota)",
    "class_notaCercospora":    "Cercospora (class)",
    "notaAntracnose":          "Antracnose (nota)",
    "class_notaAntracnose":    "Antracnose (class)",
    "notaDFC":                 "DFC (nota)",
    "class_notaDFC":           "DFC (class)",
    # ── Morfológicas ──────────────────────────────────────────────────────────
    "corFlorNome":             "Cor da Flor",
    "habitoCrescimentoNome":   "Hábito de Crescimento",
    "corPubNome":              "Cor Pubescência",
}

cols_disp = [c for c in col_map if c in df_tabela.columns]
df_show   = df_tabela[cols_disp].rename(columns=col_map)

ag_table(df_show, height=min(400, 36 + 32 * len(df_show) + 20))
exportar_excel(df_show, nome_arquivo="auditoria_densidade.xlsx",
               label="⬇️ Exportar Auditoria", key="den_exp_audit")

st.divider()

# ── Verificação de dados suficientes ─────────────────────────────────────────
_tem_grupos = "pop_grupo" in ta_filtrado.columns and ta_filtrado["pop_grupo"].notna().any()
_ordem_grupos = sorted(
    ta_filtrado["pop_grupo"].dropna().unique().tolist(),
    key=lambda x: int(x.replace("k",""))
) if _tem_grupos else []


# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 1b — GRUPOS DE DENSIDADE (Schwabish)
# ════════════════════════════════════════════════════════════════════════════════
import plotly.graph_objects as _go_cl

# ── Constantes de estilo ──────────────────────────────────────────────────────
_CL_VERDE    = "#27AE60"   # destaque principal
_CL_CINZA    = "#CBD5E1"   # grupos secundários
_CL_TEXTO    = "#1A1A1A"
_CL_SUB      = "#6B7280"
_CL_GRID     = "#F0F0F0"
_CL_BG       = "#FFFFFF"
_CL_FONTE    = "Helvetica Neue, sans-serif"

def _cl_layout(fig, height=360, margin=None):
    """Aplica o layout limpo padrão Schwabish."""
    m = margin or dict(t=70, b=50, l=20, r=20)
    fig.update_layout(
        height=height,
        plot_bgcolor=_CL_BG, paper_bgcolor=_CL_BG,
        font=dict(family=_CL_FONTE, color=_CL_TEXTO),
        showlegend=False,
        margin=m,
        xaxis=dict(showgrid=False, zeroline=False, showline=False,
                   tickfont=dict(size=12, color="#1A1A1A", weight="bold")),
        yaxis=dict(showgrid=True, gridcolor=_CL_GRID, zeroline=False,
                   showline=False, tickfont=dict(size=12, color="#1A1A1A", weight="bold")),
    )
    return fig

if not _tem_grupos or "populacao" not in ta_filtrado.columns:
    st.info("Dados de clusterização não disponíveis.")
else:
    df_cl = ta_filtrado[
        ta_filtrado["populacao"].notna() &
        (ta_filtrado["populacao"] > 0) &
        ta_filtrado["pop_grupo"].notna()
    ].copy()

    # ── Pré-calcular resumo de cada cluster ──────────────────────────────────
    resumo_cl = []
    for grupo in _ordem_grupos:
        v = df_cl[df_cl["pop_grupo"] == grupo]["populacao"].values
        sc = df_cl[df_cl["pop_grupo"] == grupo]["sc_ha"].dropna().values if "sc_ha" in df_cl.columns else np.array([])
        sc = sc[sc > 0]
        resumo_cl.append({
            "grupo":   grupo,
            "n":       len(v),
            "media":   float(np.mean(v)) if len(v) else 0,
            "mediana": float(np.median(v)) if len(v) else 0,
            "dp":      float(np.std(v))  if len(v) else 0,
            "cv":      float(np.std(v)/np.mean(v)*100) if len(v) and np.mean(v) > 0 else 0,
            "vmin":    float(np.min(v))  if len(v) else 0,
            "vmax":    float(np.max(v))  if len(v) else 0,
            "q1":      float(np.percentile(v, 25)) if len(v) else 0,
            "q3":      float(np.percentile(v, 75)) if len(v) else 0,
            "sc_media":float(np.mean(sc)) if len(sc) else np.nan,
        })
    df_res = pd.DataFrame(resumo_cl)

    # Grupo com mais parcelas → destaque no ato 1
    grupo_maior   = df_res.loc[df_res["n"].idxmax(), "grupo"]
    pct_maior     = int(df_res.loc[df_res["n"].idxmax(), "n"] / df_res["n"].sum() * 100)
    total_parcelas = int(df_res["n"].sum())

    # Grupo com melhor produtividade → destaque no ato 3
    df_sc = df_res.dropna(subset=["sc_media"])
    grupo_melhor = df_sc.loc[df_sc["sc_media"].idxmax(), "grupo"] if not df_sc.empty else None
    sc_melhor    = round(df_sc["sc_media"].max(), 1) if not df_sc.empty else None

    # ── Cabeçalho da seção ────────────────────────────────────────────────────
    st.markdown(f"""
<div style="margin: 1.5rem 0 0.2rem;">
    <p style="font-size:12px;font-weight:600;color:{_CL_SUB};text-transform:uppercase;
              letter-spacing:0.07em;margin:0 0 4px;">Passo 1 de 4 — Definição dos Grupos de Densidade</p>
    <h2 style="font-size:1.9rem;font-weight:700;color:{_CL_TEXTO};margin:0;line-height:1.2;">
        {len(_ordem_grupos)} faixas de densidade identificadas no ensaio
    </h2>
    <p style="font-size:15px;color:{_CL_SUB};margin:6px 0 0;line-height:1.6;">
        Esta etapa <strong>não analisa nenhum cultivar específico</strong> —
        ela define as faixas de população que existem nos dados e serão usadas
        em todas as análises seguintes. {total_parcelas:,} parcelas no total.
    </p>
</div>
""", unsafe_allow_html=True)

    with st.popover("ℹ️ Como ler esta seção", use_container_width=False):
        st.markdown("""
⚠️ **Esta seção não analisa nenhum cultivar específico.**

Ela mostra como as **densidades de plantio do ensaio inteiro** foram organizadas em grupos —
uma etapa necessária antes de qualquer análise por cultivar.

O algoritmo leu as populações reais registradas em campo (ex: 287.000, 312.000 pl/ha)
e as agrupou automaticamente nas faixas mais comuns (ex: grupo **300k**).
Esses grupos são usados em todas as análises das seções seguintes.

**Os 3 gráficos respondem:**

**Ato 1 — Volume** → Quantas parcelas do ensaio caíram em cada faixa?

**Ato 2 — Separação** → Os grupos estão bem definidos ou há sobreposição?
Grupos sobrepostos indicam variação no plantio — o produtor não seguiu exatamente o planejado.

**Ato 3 — Desempenho médio** → Qual faixa produziu mais **na média geral do ensaio**?
Não é o resultado de um cultivar — é a tendência de todos os materiais juntos.

**Grupos com CV 0% ou poucas parcelas** indicam densidades testadas em poucos locais — interprete com cautela.

---

**⚙️ Metodologia — Como os grupos foram formados**

O algoritmo utilizado é o **K-Means**, uma técnica de agrupamento automático amplamente usada em ciência de dados.

Ele funciona assim:
1. Lê todas as populações reais registradas nas parcelas (ex: 145.000, 287.000, 320.000 pl/ha)
2. Identifica automaticamente quantos grupos distintos existem nos dados (até 6 grupos)
3. Agrupa as parcelas pela proximidade ao centro de cada grupo (**centróide**)
4. Nomeia cada grupo pela média arredondada em milhares (ex: média 317.000 → grupo **300k**)

A vantagem sobre faixas fixas (ex: 100k–200k, 200k–300k) é que o K-Means
**se adapta aos dados reais de cada safra** — se ninguém plantou entre 200k e 250k, esse intervalo
não vira um grupo vazio.

Grupos com **CV próximo de 0%** têm todas as parcelas com a mesma população —
geralmente porque aquele tratamento foi aplicado de forma muito uniforme ou há poucas observações.
""")


    # ── Cards de contexto ─────────────────────────────────────────────────────
    _ccols = st.columns(len(_ordem_grupos))
    for ci, row in df_res.iterrows():
        destaque = row["grupo"] == grupo_maior
        cor_borda = _CL_VERDE if destaque else "#E5E7EB"
        cor_num   = _CL_VERDE if destaque else _CL_TEXTO
        pct       = int(row["n"] / total_parcelas * 100)
        _ccols[ci].markdown(f"""
<div style="border:2px solid {cor_borda};border-radius:12px;padding:14px 12px;
            background:#FAFAFA;text-align:center;">
    <p style="font-size:22px;font-weight:800;color:{cor_num};margin:0;">{row['grupo']}</p>
    <p style="font-size:13px;color:{_CL_SUB};margin:2px 0 0;">{int(row['media']/1000):.0f}k pl/ha em média</p>
    <p style="font-size:12px;color:{_CL_SUB};margin:4px 0 0;"><b style="color:{_CL_TEXTO}">{row['n']}</b> parc. do ensaio · {pct}%</p>
</div>""", unsafe_allow_html=True)

    st.markdown("<div style='margin-top:2rem;'></div>", unsafe_allow_html=True)

    # ════════════════════════════════════════════════════════════════════════
    # ATO 1 — Volume: lollipop horizontal de n parcelas por grupo
    # ════════════════════════════════════════════════════════════════════════
    _titulo1 = (
        f"O grupo <b style='color:{_CL_VERDE}'>{grupo_maior}</b> concentra "
        f"{pct_maior}% das parcelas avaliadas"
    )
    st.markdown(f"""
<div style="margin:1rem 0 0.2rem;">
    <p style="font-size:12px;font-weight:600;color:{_CL_SUB};text-transform:uppercase;
              letter-spacing:0.07em;margin:0 0 2px;">Ato 1 — Volume</p>
    <p style="font-size:1.15rem;font-weight:700;color:{_CL_TEXTO};margin:0;">{_titulo1}</p>
    <p style="font-size:13px;color:{_CL_SUB};margin:3px 0 0;">
        Quantas parcelas foram plantadas em cada faixa de densidade?
    </p>
</div>""", unsafe_allow_html=True)

    fig_lol = _go_cl.Figure()
    grupos_rev1 = list(reversed(_ordem_grupos))
    for i, grupo in enumerate(grupos_rev1):
        row   = df_res[df_res["grupo"] == grupo].iloc[0]
        n_val = int(row["n"])
        pct_v = int(n_val / total_parcelas * 100)
        is_dest = grupo == grupo_maior
        cor = _CL_VERDE if is_dest else _CL_CINZA
        # Linha do lollipop
        fig_lol.add_shape(type="line",
            x0=0, x1=n_val, y0=i, y1=i,
            line=dict(color=cor, width=2.5))
        # Ponto
        fig_lol.add_trace(_go_cl.Scatter(
            x=[n_val], y=[i],
            mode="markers",
            marker=dict(color=cor, size=14,
                        line=dict(color="#FFFFFF", width=2)),
            hovertemplate=f"<b>{grupo}</b><br>{n_val} parcelas ({pct_v}%)<extra></extra>",
        ))
        # Label direto: n parcelas e %
        fig_lol.add_annotation(
            x=n_val, y=i,
            text=f"<b>{n_val}</b> parcelas <span style='color:{_CL_SUB}'>{pct_v}%</span>",
            xanchor="left", yanchor="middle",
            showarrow=False, xshift=14,
            font=dict(size=13, color=_CL_TEXTO if is_dest else _CL_SUB, weight="bold"),
        )
        # Label do grupo (eixo Y substituído)
        fig_lol.add_annotation(
            x=0, y=i,
            text=f"<b>{grupo}</b>",
            xanchor="right", yanchor="middle",
            showarrow=False, xshift=-10,
            font=dict(size=13, color=_CL_VERDE if is_dest else _CL_TEXTO, weight="bold"),
        )

    _x_max1 = int(df_res["n"].max() * 1.45)
    fig_lol.update_layout(
        height=max(240, len(_ordem_grupos) * 60 + 80),
        plot_bgcolor=_CL_BG, paper_bgcolor=_CL_BG,
        font=dict(family=_CL_FONTE),
        showlegend=False,
        margin=dict(t=20, b=20, l=60, r=20),
        xaxis=dict(range=[0, _x_max1], showgrid=True, gridcolor=_CL_GRID,
                   zeroline=False, showline=False,
                   tickfont=dict(size=12, color="#1A1A1A", weight="bold"),
                   title=dict(text="<b>Número de parcelas</b>", font=dict(size=14, color="#1A1A1A", weight="bold"))),
        yaxis=dict(showticklabels=False, showgrid=False, zeroline=False, showline=False,
                   range=[-0.7, len(_ordem_grupos) - 0.3]),
    )
    st.plotly_chart(fig_lol, use_container_width=True)

    st.markdown("<div style='margin-top:1.5rem;'></div>", unsafe_allow_html=True)

    # ════════════════════════════════════════════════════════════════════════
    # ATO 2 — Separação: dot-range mostrando dispersão e sobreposição
    # ════════════════════════════════════════════════════════════════════════
    # Detectar sobreposição entre grupos adjacentes
    _sobrepostos = []
    for gi in range(len(_ordem_grupos) - 1):
        ra = df_res[df_res["grupo"] == _ordem_grupos[gi]].iloc[0]
        rb = df_res[df_res["grupo"] == _ordem_grupos[gi+1]].iloc[0]
        if ra["q3"] >= rb["q1"]:
            _sobrepostos.append(f"{_ordem_grupos[gi]} e {_ordem_grupos[gi+1]}")

    _sep_txt = (
        "Os grupos estão bem separados — boa qualidade de agrupamento."
        if not _sobrepostos else
        f"Atenção: sobreposição entre {', '.join(_sobrepostos)}."
    )
    _sep_cor = _CL_VERDE if not _sobrepostos else "#E67E22"

    st.markdown(f"""
<div style="margin:0 0 0.2rem;">
    <p style="font-size:12px;font-weight:600;color:{_CL_SUB};text-transform:uppercase;
              letter-spacing:0.07em;margin:0 0 2px;">Ato 2 — Separação</p>
    <p style="font-size:1.15rem;font-weight:700;color:{_CL_TEXTO};margin:0;">
        Cada grupo cobre uma faixa distinta de população no campo
    </p>
    <p style="font-size:13px;color:{_sep_cor};margin:3px 0 0;font-weight:500;">
        {_sep_txt}
    </p>
</div>""", unsafe_allow_html=True)

    with st.popover("ℹ️ Como ler este gráfico", use_container_width=False):
        st.markdown("""
Este gráfico mostra **onde cada parcela do ensaio foi plantada** em termos de população real.
O objetivo é verificar se os grupos de densidade estão bem separados ou se há sobreposição.

**Como ler:**

- **Cada ponto** → uma parcela real; posição no eixo X = população registrada (pl/ha)
- **Faixa clara** → amplitude total do grupo (menor ao maior valor)
- **Faixa escura** → onde estão 50% das parcelas (IQR)
- **Linha sólida** → média do grupo
- **Linha tracejada** → mediana
- **Label à direita** → média em mil plantas · (n parcelas) · CV do grupo

**Subtítulo verde** → grupos bem separados, sem sobreposição — o agrupamento é confiável.
**Subtítulo laranja** → dois grupos se sobrepõem — algumas parcelas estão na fronteira entre densidades.

**O que é CV aqui?**

O CV (coeficiente de variação) mede a uniformidade das populações **dentro do grupo**:
- **CV 0%** → todas as parcelas foram plantadas com a mesma população exata
  (ensaio com população única nessa faixa, ou grupo com poucas observações)
- **CV alto** → as parcelas do grupo variaram bastante — o K-Means agrupou densidades distintas
  porque não havia uma faixa mais específica nos dados

**Dica:** passe o mouse sobre os pontos para ver fazenda, cidade e produtividade de cada parcela.
""")

    # ── Ato 2: strip chart horizontal — cada parcela como ponto ─────────────
    _grupos_rev2 = list(reversed(_ordem_grupos))

    fig_range = _go_cl.Figure()
    for i, grupo in enumerate(_grupos_rev2):
        row     = df_res[df_res["grupo"] == grupo].iloc[0]
        is_dest = grupo == grupo_maior
        cor     = _CL_VERDE if is_dest else "#94A3B8"
        cor_txt = _CL_VERDE if is_dest else _CL_TEXTO
        op_area = 0.18 if is_dest else 0.10
        op_iqr  = 0.30 if is_dest else 0.16
        n_g     = int(row["n"])

        # Faixa amplitude total (min–max)
        fig_range.add_shape(type="rect",
            x0=row["vmin"], x1=row["vmax"],
            y0=i - 0.32, y1=i + 0.32,
            fillcolor=cor, opacity=op_area, line=dict(width=0))

        # Faixa IQR
        fig_range.add_shape(type="rect",
            x0=row["q1"], x1=row["q3"],
            y0=i - 0.32, y1=i + 0.32,
            fillcolor=cor, opacity=op_iqr, line=dict(width=0))

        # Linha da média — sólida, mais grossa no destaque
        fig_range.add_shape(type="line",
            x0=row["media"], x1=row["media"],
            y0=i - 0.38, y1=i + 0.38,
            line=dict(color=cor, width=3 if is_dest else 2))

        # Linha da mediana — tracejada
        fig_range.add_shape(type="line",
            x0=row["mediana"], x1=row["mediana"],
            y0=i - 0.38, y1=i + 0.38,
            line=dict(color=cor, width=1.8, dash="dash"))

        # Pontos individuais com jitter + hover com localização
        vals_pop = df_cl[df_cl["pop_grupo"] == grupo]["populacao"].values
        _cd_cols_cl  = ["cod_fazenda", "nomeFazenda", "cidade_nome", "estado_sigla", "sc_ha"]
        _cd_avail_cl = [c for c in _cd_cols_cl if c in df_cl.columns]
        df_g_cl  = df_cl[df_cl["pop_grupo"] == grupo].reset_index(drop=True)

        if len(vals_pop) > 0:
            jitter_cl = np.random.uniform(-0.22, 0.22, size=len(vals_pop))
            if _cd_avail_cl and not df_g_cl.empty:
                cd_cl = df_g_cl[_cd_avail_cl].values
                _ht_cl = (
                    "<b>%{customdata[0]}</b> · %{customdata[2]}, %{customdata[3]}<br>"
                    "Pop: <b>%{x:,.0f}</b> pl/ha<br>"
                    "sc/ha: %{customdata[4]:.1f}"
                    "<extra></extra>"
                )
            else:
                cd_cl  = None
                _ht_cl = f"<b>{grupo}</b><br>Pop: %{{x:,.0f}} pl/ha<extra></extra>"

            fig_range.add_trace(_go_cl.Scatter(
                x=vals_pop,
                y=[i + j for j in jitter_cl],
                mode="markers",
                marker=dict(color=cor, size=6 if is_dest else 5,
                            opacity=0.75 if is_dest else 0.55,
                            line=dict(color="#FFFFFF", width=0.5)),
                customdata=cd_cl,
                hovertemplate=_ht_cl,
                showlegend=False,
            ))

        # Label: média + n parcelas + CV
        fig_range.add_annotation(
            x=row["vmax"], y=i,
            text=(
                f"<b style='color:{cor_txt}'>{row['media']/1000:.0f}k</b>"
                f"<span style='color:{_CL_SUB};font-size:10px'>"
                f" ({n_g}) · CV {row['cv']:.0f}%</span>"
            ),
            xanchor="left", yanchor="middle",
            showarrow=False, xshift=10,
            font=dict(size=13, color=_CL_TEXTO, weight="bold"),
        )

    _x_pop_min     = df_res["vmin"].min()
    _x_pop_max     = df_res["vmax"].max()
    _x_pad         = (_x_pop_max - _x_pop_min) * 0.05
    _x_label_space = (_x_pop_max - _x_pop_min) * 0.30

    fig_range.update_layout(
        height=max(220, len(_grupos_rev2) * 90 + 60),
        plot_bgcolor=_CL_BG, paper_bgcolor=_CL_BG,
        font=dict(family=_CL_FONTE),
        showlegend=False,
        margin=dict(t=20, b=50, l=70, r=20),
        xaxis=dict(
            range=[_x_pop_min - _x_pad, _x_pop_max + _x_label_space],
            title=dict(text="<b>População real (pl/ha)</b>", font=dict(size=14, color=_CL_SUB, weight="bold")),
            tickformat=",", tickfont=dict(size=12, color="#1A1A1A", weight="bold"),
            showgrid=True, gridcolor=_CL_GRID,
            zeroline=False, showline=False,
        ),
        yaxis=dict(
            tickmode="array",
            tickvals=list(range(len(_grupos_rev2))),
            ticktext=[
                f"<b style='color:{_CL_VERDE}'>{g}</b>" if g == grupo_maior
                else f"<b style='color:{_CL_SUB}'>{g}</b>"
                for g in _grupos_rev2
            ],
            tickfont=dict(size=13, weight="bold"),
            showgrid=False, zeroline=False, showline=False,
            range=[-0.7, len(_grupos_rev2) - 0.3],
        ),
    )
    st.plotly_chart(fig_range, use_container_width=True)
    st.caption(
        "Cada ponto = uma parcela · "
        "Linha sólida = média · Linha tracejada = mediana · "
        "Faixa escura = 50% das parcelas (IQR) · Faixa clara = amplitude total · "
        "Passe o mouse para ver fazenda, cidade e produtividade."
    )

    st.markdown("<div style='margin-top:1.5rem;'></div>", unsafe_allow_html=True)

    # ════════════════════════════════════════════════════════════════════════
    # ATO 3 — Desempenho: dot plot com erro padrão (Schwabish)
    # ════════════════════════════════════════════════════════════════════════
    if df_sc.empty or df_sc["sc_media"].isna().all():
        st.info("Dados de produtividade por cluster não disponíveis.")
    else:
        # Calcular EP (erro padrão) por grupo
        ep_por_grupo = {}
        for grupo in _ordem_grupos:
            sc_vals = df_cl[df_cl["pop_grupo"] == grupo]["sc_ha"].dropna().values
            sc_vals = sc_vals[sc_vals > 0]
            ep_por_grupo[grupo] = float(np.std(sc_vals) / np.sqrt(len(sc_vals))) if len(sc_vals) > 1 else 0.0

        _titulo3 = (
            f"O grupo <b style='color:{_CL_VERDE}'>{grupo_melhor}</b> registrou "
            f"a maior produtividade média: <b style='color:{_CL_VERDE}'>{sc_melhor:.1f} sc/ha</b>"
            if grupo_melhor else "Produtividade média por grupo de densidade"
        )
        st.markdown(f"""
<div style="margin:0 0 0.2rem;">
    <p style="font-size:12px;font-weight:600;color:{_CL_SUB};text-transform:uppercase;
              letter-spacing:0.07em;margin:0 0 2px;">Ato 3 — Desempenho</p>
    <p style="font-size:1.15rem;font-weight:700;color:{_CL_TEXTO};margin:0;">{_titulo3}</p>
    <p style="font-size:13px;color:{_CL_SUB};margin:3px 0 0;">
        Média geral do ensaio por grupo — não é o resultado de um cultivar específico.
        Barras de erro = ± erro padrão da média (quanto menor, mais consistente o grupo).
    </p>
</div>""", unsafe_allow_html=True)

        with st.popover("ℹ️ Como ler este gráfico", use_container_width=False):
            st.markdown("""
Este gráfico mostra a **produtividade média de todos os cultivares juntos** dentro de cada
grupo de densidade — é uma visão do ensaio como um todo, não de um material específico.

**Como ler:**

- **Ponto** → média de sc/ha de todas as parcelas do grupo
- **Barra horizontal** (`├───┤`) → erro padrão da média (EP)
- **Linha tracejada vertical** → média geral do ensaio (referência)

**O que é o erro padrão (EP)?**

O EP indica o quanto a média pode variar se o experimento fosse repetido.
Na prática:
- **Barra curta** → as parcelas desse grupo produziram de forma parecida entre si — resultado confiável
- **Barra longa** → alta variação entre parcelas — interprete com cautela

Se as barras de dois grupos **se sobrepõem**, a diferença entre eles provavelmente
não é significativa — pode ser variação normal do campo.

**⚠️ Atenção:** grupos com poucas parcelas (n pequeno) tendem a ter EP menor
apenas porque há menos dados — verifique o n na tabela abaixo antes de concluir.

---

**📊 Sobre a coluna IQR pop. na tabela**

O IQR mostra a faixa de população entre os 25% menores e os 25% maiores valores do grupo.

- **Ex: `280.000 – 320.000`** → as parcelas variaram dentro dessa faixa de densidade
- **`—`** → todas as parcelas têm exatamente a mesma população (CV = 0%) —
  o ensaio foi conduzido com população uniforme nesse grupo, ou há muito poucas observações
""")

        # Ordem: grupos rev para Y (maior no topo)
        _grupos_rev3 = list(reversed(_ordem_grupos))
        media_geral_sc = float(np.nanmean([
            df_cl[df_cl["pop_grupo"] == g]["sc_ha"].dropna().values[
                df_cl[df_cl["pop_grupo"] == g]["sc_ha"].dropna().values > 0
            ].mean()
            for g in _ordem_grupos
            if len(df_cl[df_cl["pop_grupo"] == g]["sc_ha"].dropna().values[
                df_cl[df_cl["pop_grupo"] == g]["sc_ha"].dropna().values > 0
            ]) > 0
        ]))

        fig_dot3 = _go_cl.Figure()

        # Linha de referência vertical (média geral)
        fig_dot3.add_vline(
            x=media_geral_sc,
            line=dict(color="#374151", width=1.5, dash="dot"),
        )
        fig_dot3.add_annotation(
            x=media_geral_sc, y=len(_grupos_rev3) - 0.1,
            text=f"<b style='color:#374151'>Média geral<br>{media_geral_sc:.1f} sc/ha</b>",
            showarrow=False, xanchor="left", yanchor="top",
            xshift=6, font=dict(size=11, color="#374151"),
        )

        for i, grupo in enumerate(_grupos_rev3):
            row     = df_res[df_res["grupo"] == grupo].iloc[0]
            ep      = ep_por_grupo.get(grupo, 0.0)
            is_dest = grupo == grupo_melhor
            cor     = _CL_VERDE if is_dest else "#94A3B8"
            cor_txt = _CL_VERDE if is_dest else _CL_TEXTO
            n_g     = int(row["n"])

            if pd.isna(row["sc_media"]):
                continue

            media_g = row["sc_media"]

            # Barra de erro padrão (faixa horizontal)
            if ep > 0:
                fig_dot3.add_shape(type="line",
                    x0=media_g - ep, x1=media_g + ep,
                    y0=i, y1=i,
                    line=dict(color=cor, width=3))
                # Terminais do EP
                for x_ep in [media_g - ep, media_g + ep]:
                    fig_dot3.add_shape(type="line",
                        x0=x_ep, x1=x_ep,
                        y0=i - 0.18, y1=i + 0.18,
                        line=dict(color=cor, width=2))

            # Ponto da média — maior e mais destacado no líder
            fig_dot3.add_trace(_go_cl.Scatter(
                x=[media_g], y=[i],
                mode="markers",
                marker=dict(
                    color=cor, size=16 if is_dest else 12,
                    symbol="circle",
                    line=dict(color="#FFFFFF", width=2.5),
                ),
                hovertemplate=(
                    f"<b>{grupo}</b><br>"
                    f"Média: {media_g:.1f} sc/ha<br>"
                    f"EP: ±{ep:.2f} sc/ha<br>"
                    f"Parcelas: {n_g}"
                    f"<extra></extra>"
                ),
                showlegend=False,
            ))

            # Label direto à direita: valor + EP
            fig_dot3.add_annotation(
                x=media_g + ep, y=i,
                text=(
                    f"<b style='color:{cor_txt}'>{media_g:.1f}</b>"
                    f"<span style='color:{_CL_SUB};font-size:10px'>"
                    f" ±{ep:.1f} ({n_g})</span>"
                ),
                xanchor="left", yanchor="middle",
                showarrow=False, xshift=10,
                font=dict(size=12, color=_CL_TEXTO),
            )

        # Eixo X: range perto dos dados (não começa no zero)
        _sc_vals_all = df_cl[df_cl["sc_ha"] > 0]["sc_ha"].dropna()
        _x_min3 = float(_sc_vals_all.min()) - 2 if not _sc_vals_all.empty else 0
        _x_max3 = float(_sc_vals_all.max()) + 12

        fig_dot3.update_layout(
            height=max(220, len(_grupos_rev3) * 72 + 80),
            plot_bgcolor=_CL_BG, paper_bgcolor=_CL_BG,
            font=dict(family=_CL_FONTE),
            showlegend=False,
            margin=dict(t=20, b=50, l=70, r=20),
            xaxis=dict(
                range=[_x_min3, _x_max3],
                title=dict(text="<b>sc/ha (média do grupo)</b>", font=dict(size=14, color=_CL_SUB, weight="bold")),
                tickfont=dict(size=12, color="#1A1A1A", weight="bold"),
                showgrid=True, gridcolor=_CL_GRID,
                zeroline=False, showline=False,
            ),
            yaxis=dict(
                tickmode="array",
                tickvals=list(range(len(_grupos_rev3))),
                ticktext=[
                    f"<b style='color:{_CL_VERDE}'>{g}</b>" if g == grupo_melhor
                    else f"<b style='color:{_CL_SUB}'>{g}</b>"
                    for g in _grupos_rev3
                ],
                tickfont=dict(size=13, weight="bold"),
                showgrid=False, zeroline=False, showline=False,
                range=[-0.7, len(_grupos_rev3) - 0.3],
            ),
        )
        st.plotly_chart(fig_dot3, use_container_width=True)
        st.caption(
            "Ponto = média de sc/ha do grupo · "
            "Barra horizontal = ± erro padrão da média · "
            "Barras menores = grupo mais consistente entre locais · "
            "Linha tracejada = média geral do ensaio."
        )

    # ── Tabela AgGrid sempre visível ──────────────────────────────────────────
    st.markdown(
        f'<p style="font-size:12px;font-weight:600;color:{_CL_SUB};text-transform:uppercase;'
        f'letter-spacing:0.07em;margin:1.2rem 0 0.4rem;">Resumo numérico por grupo</p>',
        unsafe_allow_html=True,
    )
    tbl_cl = []
    for _, row in df_res.iterrows():
        ep_g = ep_por_grupo.get(row["grupo"], 0.0) if df_sc is not None and not df_sc.empty else 0.0
        tbl_cl.append({
            "Grupo":              row["grupo"],
            "n (parcelas)":       int(row["n"]),
            "% do ensaio":        f"{int(row['n']/total_parcelas*100)}%",
            "Pop. média (pl/ha)": f"{int(row['media']):,}",
            "IQR pop.":           f"{int(row['q1']):,} – {int(row['q3']):,}" if int(row['q1']) != int(row['q3']) else "—",
            "CV pop. (%)":        f"{row['cv']:.1f}%",
            "sc/ha médio":        f"{row['sc_media']:.1f}" if not pd.isna(row['sc_media']) else "—",
            "EP (sc/ha)":         f"±{ep_g:.2f}" if ep_g > 0 else "—",
        })

    # AgGrid com linha do grupo líder destacada em verde claro
    df_tbl_cl = pd.DataFrame(tbl_cl)
    _gb3 = GridOptionsBuilder.from_dataframe(df_tbl_cl)
    _gb3.configure_default_column(
        resizable=True, sortable=True, filter=True,
        suppressMenu=False,
        menuTabs=["generalMenuTab", "filterMenuTab", "columnsMenuTab"],
        cellStyle={"fontSize": "13px", "color": "#000000",
                   "fontFamily": "Helvetica Neue, sans-serif"},
    )
    _gb3.configure_grid_options(
        headerHeight=36, rowHeight=32, domLayout="normal",
        suppressMenuHide=True, suppressColumnVirtualisation=True,
        getRowStyle=JsCode(f"""
            function(params) {{
                if (params.data["Grupo"] === "{grupo_melhor}") {{
                    return {{
                        'background': '#D5F5E3',
                        'fontWeight': '600',
                    }};
                }}
            }}
        """),
    )
    _go3 = _gb3.build()
    _go3["defaultColDef"]["headerClass"] = "ag-header-black"
    _go3["onFirstDataRendered"] = JsCode("function(params) { params.api.sizeColumnsToFit(); }")
    AgGrid(
        df_tbl_cl, gridOptions=_go3,
        height=min(320, 36 + 32 * len(df_tbl_cl) + 20),
        update_mode=GridUpdateMode.NO_UPDATE,
        fit_columns_on_grid_load=False,
        allow_unsafe_jscode=True,
        enable_enterprise_modules=True,
        custom_css={
            ".ag-header":                      {"background-color": "#4A4A4A !important"},
            ".ag-header-row":                  {"background-color": "#4A4A4A !important"},
            ".ag-header-cell":                 {"background-color": "#4A4A4A !important"},
            ".ag-header-cell-label":           {"color": "#FFFFFF !important", "font-weight": "700"},
            ".ag-header-cell-text":            {"color": "#FFFFFF !important", "font-size": "13px !important", "font-weight": "700 !important"},
            ".ag-icon":                        {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-header-icon":                 {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-header-cell-menu-button":     {"opacity": "1 !important", "visibility": "visible !important"},
            ".ag-header-cell-menu-button span":{"color": "#FFFFFF !important"},
            ".ag-cell":                        {"font-size": "13px !important", "color": "#000000 !important"},
        },
        theme="streamlit", use_container_width=True,
    )

st.divider()

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 2 — DISTRIBUIÇÃO POR DENSIDADE + LSD (Schwabish)
# ════════════════════════════════════════════════════════════════════════════════
import plotly.graph_objects as go_plt
from scipy import stats as _stats
from itertools import combinations as _comb

if not _tem_grupos:
    st.info("Dados de grupo de densidade não disponíveis.")
else:
    cultivares_anova = sorted(
        ta_filtrado[ta_filtrado["sc_ha"] > 0]["dePara"].dropna().unique().tolist()
    )

    # ── Label estático da seção ───────────────────────────────────
    st.markdown(
        f'<p style="font-size:12px;font-weight:600;color:{_CL_SUB};text-transform:uppercase;'
        f'letter-spacing:0.07em;margin:1.5rem 0 0.6rem;">Distribuição por Densidade</p>',
        unsafe_allow_html=True,
    )

    # ── Controles compactos ───────────────────────────────────────────────────
    col_a1, col_a2, col_a3 = st.columns([3, 2, 2])
    with col_a1:
        cult_anova = st.selectbox("Cultivar", options=cultivares_anova, key="den_anova_cult",
                                  label_visibility="collapsed",
                                  placeholder="Selecione o cultivar...")
    with col_a2:
        mostrar_lsd = st.checkbox("Mostrar corte LSD", value=True, key="den_lsd_chk")
    with col_a3:
        mostrar_pontos = st.checkbox("Mostrar parcelas individuais", value=True, key="den_pts_chk")

    _cols_loc  = ["sc_ha", "pop_grupo", "cod_fazenda", "nomeFazenda",
                  "cidade_nome", "estado_sigla", "populacao"]
    _cols_disp = [c for c in _cols_loc if c in ta_filtrado.columns]
    df_av      = ta_filtrado[
        (ta_filtrado["dePara"] == cult_anova) & (ta_filtrado["sc_ha"] > 0)
    ][_cols_disp].copy()
    grupos_cult = [g for g in _ordem_grupos if g in df_av["pop_grupo"].values]
    dados_pg    = {g: df_av[df_av["pop_grupo"] == g]["sc_ha"].values for g in grupos_cult}
    df_pg       = {g: df_av[df_av["pop_grupo"] == g].reset_index(drop=True) for g in grupos_cult}

    # ── Estatísticas ──────────────────────────────────────────────────────────
    amostras = [v for v in dados_pg.values() if len(v) >= 2]
    f_stat, p_valor = _stats.f_oneway(*amostras) if len(amostras) >= 2 else (np.nan, np.nan)

    def _calcular_lsd_dens(df, fator="pop_grupo", bloco="cod_fazenda", col="sc_ha", alpha=0.05):
        try:
            d = df[[col, fator, bloco]].dropna()
            d = d[d[col] > 0].reset_index(drop=True)
            if d.empty or d[fator].nunique() < 2 or d[bloco].nunique() < 2:
                return np.nan
            y      = d[col].values.astype(float)
            X_fat  = pd.get_dummies(d[fator],  drop_first=True).values.astype(float)
            X_blk  = pd.get_dummies(d[bloco],  drop_first=True).values.astype(float)
            X      = np.hstack([np.ones((len(y),1)), X_fat, X_blk])
            beta, _, rank, _ = np.linalg.lstsq(X, y, rcond=None)
            ss_res = np.sum((y - X @ beta)**2)
            gl_res = len(y) - rank
            if gl_res <= 0: return np.nan
            qmr    = ss_res / gl_res
            n_blk  = d[bloco].nunique()
            t_crit = _stats.t.ppf(1 - alpha/2, df=gl_res)
            return round(t_crit * np.sqrt(2 * qmr / n_blk), 1)
        except Exception:
            return np.nan

    lsd_val = _calcular_lsd_dens(df_av)

    # ── Derivar grupo líder ───────────────────────────────────────────────────
    medias_pg = {g: float(np.mean(v)) for g, v in dados_pg.items() if len(v) > 0}
    grupo_lider = max(medias_pg, key=medias_pg.get) if medias_pg else None
    media_lider = medias_pg[grupo_lider] if grupo_lider else None

    # ── Título dinâmico em linguagem natural ──────────────────────────────────
    p_significativo = not np.isnan(p_valor) and p_valor < 0.05
    if grupo_lider and media_lider:
        if p_significativo:
            _titulo_sec2 = (
                f"Em <b>{cult_anova}</b>, plantar em <b style='color:{_CL_VERDE}'>{grupo_lider}</b> "
                f"resultou na maior produtividade — diferença estatisticamente significativa"
            )
            _sub_sec2_cor = _CL_VERDE
            _sub_sec2 = (
                f"A densidade influenciou a produtividade (p={p_valor:.3f}). "
                f"Grupo líder: {grupo_lider} com {media_lider:.1f} sc/ha de média."
            )
        else:
            _titulo_sec2 = (
                f"Em <b>{cult_anova}</b>, a densidade <b>não alterou</b> "
                f"significativamente a produtividade"
            )
            _sub_sec2_cor = "#E67E22"
            _sub_sec2 = (
                f"Diferença entre grupos não é estatisticamente significativa (p={p_valor:.3f}). "
                f"Grupo com maior média: {grupo_lider} ({media_lider:.1f} sc/ha)."
            )
    else:
        _titulo_sec2 = f"Distribuição de produtividade por grupo de densidade · {cult_anova}"
        _sub_sec2_cor = _CL_SUB
        _sub_sec2 = "Dados insuficientes para análise estatística."

    st.markdown(f"""
<div style="margin: 1.5rem 0 0.8rem;">
        <p style="font-size:1.15rem;font-weight:700;color:{_CL_TEXTO};
              margin:0;line-height:1.4;">{_titulo_sec2}</p>
    <p style="font-size:13px;color:{_sub_sec2_cor};margin:5px 0 0;font-weight:500;">
        {_sub_sec2}
    </p>
</div>
""", unsafe_allow_html=True)

    with st.popover("ℹ️ Como interpretar", use_container_width=False):
        st.markdown("""
Cada **ponto** é uma parcela real colhida no campo. A posição horizontal indica a produtividade (sc/ha).

**O que as formas significam:**
- **Faixa clara** → amplitude total: do menor ao maior valor do grupo
- **Faixa escura** → onde estão 50% das parcelas (IQR)
- **Linha sólida** → média do grupo
- **Linha tracejada** → mediana (valor do meio)
- **Linha vermelha** → limite LSD (5%): grupos à esquerda produziram significativamente menos

**ANOVA** verifica se existe diferença real entre os grupos:
- Subtítulo **verde** → a densidade influenciou a produtividade (p < 0,05)
- Subtítulo **laranja** → diferença não comprovada — pode ser variação do ambiente

**Dica:** IQR largo indica alta variação entre locais — o cultivar pode ser instável nessa densidade.
""")

    # ── Construir gráfico ─────────────────────────────────────────────────────
    fig_strip = go_plt.Figure()
    grupos_rev = list(reversed(grupos_cult))

    for i, grupo in enumerate(grupos_rev):
        vals  = dados_pg.get(grupo, np.array([]))
        if len(vals) == 0: continue

        is_lider = grupo == grupo_lider
        cor_base = _CL_VERDE if is_lider else "#94A3B8"   # verde líder / cinza azulado resto
        cor_pt   = _CL_VERDE if is_lider else "#94A3B8"
        op_area  = 0.18 if is_lider else 0.10
        op_iqr   = 0.30 if is_lider else 0.16

        media   = float(np.mean(vals))
        mediana = float(np.median(vals))
        q1      = float(np.percentile(vals, 25))
        q3      = float(np.percentile(vals, 75))
        n       = len(vals)

        # Faixa amplitude total
        fig_strip.add_shape(type="rect",
            x0=float(np.min(vals)), x1=float(np.max(vals)),
            y0=i - 0.32, y1=i + 0.32,
            fillcolor=cor_base, opacity=op_area, line=dict(width=0))

        # Faixa IQR
        fig_strip.add_shape(type="rect",
            x0=q1, x1=q3, y0=i - 0.32, y1=i + 0.32,
            fillcolor=cor_base, opacity=op_iqr, line=dict(width=0))

        # Linha média — sólida, mais grossa no líder
        fig_strip.add_shape(type="line",
            x0=media, x1=media, y0=i - 0.38, y1=i + 0.38,
            line=dict(color=cor_base, width=3 if is_lider else 2))

        # Linha mediana — tracejada
        fig_strip.add_shape(type="line",
            x0=mediana, x1=mediana, y0=i - 0.38, y1=i + 0.38,
            line=dict(color=cor_base, width=1.8, dash="dash"))

        # Pontos individuais com hover rico (local, cidade, estado, pop)
        if mostrar_pontos:
            df_g   = df_pg.get(grupo, pd.DataFrame())
            jitter = np.random.uniform(-0.22, 0.22, size=len(vals))
            _cd_cols  = ["cod_fazenda", "nomeFazenda", "cidade_nome", "estado_sigla", "populacao"]
            _cd_avail = [c for c in _cd_cols if c in df_g.columns]
            if not df_g.empty and _cd_avail:
                cd  = df_g[_cd_avail].values
                _ht = (
                    "<b>%{customdata[0]}</b> · %{customdata[2]}, %{customdata[3]}<br>"
                    "sc/ha: <b>%{x:.1f}</b><br>"
                    "Pop: %{customdata[4]:,.0f} pl/ha"
                    "<extra></extra>"
                )
            else:
                cd  = None
                _ht = f"<b>{grupo}</b><br>sc/ha: %{{x:.1f}}<extra></extra>"

            fig_strip.add_trace(go_plt.Scatter(
                x=vals, y=[i + j for j in jitter],
                mode="markers",
                marker=dict(color=cor_pt, size=9 if is_lider else 8,
                            opacity=0.90 if is_lider else 0.75,
                            line=dict(color="#FFFFFF", width=0.5)),
                customdata=cd,
                hovertemplate=_ht,
                showlegend=False,
            ))

        # Label direto: média em destaque, n em cinza
        cor_label = _CL_VERDE if is_lider else _CL_SUB
        peso_label = "bold" if is_lider else "normal"
        fig_strip.add_annotation(
            x=media, y=i + 0.42,
            text=f"<b style='color:{cor_label};font-size:14px'>{media:.1f}</b>"
                 f"<b style='color:{_CL_SUB};font-size:14px'> ({n})</b>",
            showarrow=False,
            font=dict(size=14, color=_CL_TEXTO, weight="bold"),
            xanchor="center", yanchor="bottom",
        )

    # ── Linha de corte LSD ────────────────────────────────────────────────────
    if mostrar_lsd and not np.isnan(lsd_val) and grupo_lider:
        lsd_x = media_lider - lsd_val
        fig_strip.add_vline(
            x=lsd_x,
            line=dict(color="#E74C3C", width=1.8, dash="dot"),
        )
        # Label simples: só o valor do LSD
        fig_strip.add_annotation(
            x=lsd_x, y=len(grupos_rev) - 0.1,
            text=f"<b style='color:#E74C3C;font-size:13px'>LSD = {lsd_val:.1f} sc/ha</b>",
            showarrow=False,
            xanchor="right", yanchor="top",
            xshift=-6,
            font=dict(size=13, color="#E74C3C", weight="bold"),
            align="right",
        )

    # ── Layout limpo ──────────────────────────────────────────────────────────
    _all_sc = np.concatenate([v for v in dados_pg.values() if len(v) > 0])
    _x_min  = float(np.min(_all_sc)) - 2
    _x_max  = float(np.max(_all_sc)) + 8   # espaço para labels

    fig_strip.update_layout(
        height=max(280, len(grupos_rev) * 95 + 80),
        plot_bgcolor=_CL_BG, paper_bgcolor=_CL_BG,
        font=dict(family=_CL_FONTE),
        showlegend=False,
        margin=dict(t=20, b=50, l=70, r=20),
        xaxis=dict(
            range=[_x_min, _x_max],
            title=dict(text="<b>sc/ha</b>", font=dict(size=14, color="#1A1A1A", weight="bold")),
            tickfont=dict(size=12, color="#1A1A1A", weight="bold"),
            showgrid=True, gridcolor=_CL_GRID,
            zeroline=False, showline=False,
        ),
        yaxis=dict(
            tickmode="array",
            tickvals=list(range(len(grupos_rev))),
            ticktext=[
                f"<b style='color:{_CL_VERDE}'>{g}</b>" if g == grupo_lider
                else f"<b style='color:{_CL_SUB}'>{g}</b>"
                for g in grupos_rev
            ],
            tickfont=dict(size=14, weight="bold"),
            showgrid=False, zeroline=False, showline=False,
        ),
    )
    st.plotly_chart(fig_strip, use_container_width=True)

    # ── Legenda de leitura + dicionário de locais ───────────────────────────
    _col_cap, _col_dic = st.columns([3, 1])
    with _col_cap:
        st.caption(
            "Cada ponto = uma parcela · "
            "Linha sólida = média · Linha tracejada = mediana · "
            "Faixa escura = 50% das parcelas (IQR) · Faixa clara = amplitude total · "
            "Linha vermelha = limite mínimo para diferença real (LSD 5%)."
        )
    with _col_dic:
        _dic_cols = ["cod_fazenda", "nomeFazenda", "cidade_nome", "estado_sigla"]
        _dic_avail = [c for c in _dic_cols if c in df_av.columns]
        if _dic_avail:
            df_dic2 = (
                df_av[_dic_avail]
                .drop_duplicates()
                .sort_values("cod_fazenda")
                .rename(columns={
                    "cod_fazenda":  "Código",
                    "nomeFazenda":  "Local",
                    "cidade_nome":  "Cidade",
                    "estado_sigla": "Estado",
                })
                .reset_index(drop=True)
            )
            n_dic = len(df_dic2)
            with st.popover(f"📍 {n_dic} locais", use_container_width=True):
                st.markdown(
                    "Passe o mouse sobre os pontos do gráfico para identificar o local.",
                    help=None,
                )
                st.dataframe(df_dic2, hide_index=True, use_container_width=True)

st.divider()

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 3 — REGRESSÃO POLINOMIAL QUADRÁTICA
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Regressão Polinomial",
    "Qual é a densidade de máxima produtividade?",
    "Curva quadrática ajustada usando a população real de cada parcela como variável contínua.",
)

with st.popover("ℹ️ Como interpretar", use_container_width=False):
    st.markdown("""
Uma curva é ajustada para cada cultivar usando as populações **reais coletadas** nas parcelas
(não os tratamentos planejados).

**Como ler:**
- **Cada ponto** → uma parcela real observada
- **Curva** → tendência ajustada matematicamente
- **★ com caixa** → população de máxima produtividade estimada

**⚠️ Máximo técnico ≠ máximo econômico**

A estrela indica o ponto de **maior produtividade física** (sc/ha) —
não considera o custo das sementes extras.
Na prática, a densidade economicamente ótima costuma ser **menor**,
pois o ganho de produtividade diminui antes de cobrir o custo adicional das sementes.

**R² baixo** não invalida o ensaio — indica que ambiente e genótipo
influenciaram mais do que a densidade nessa safra.

Se a curva não forma uma "colina", não há ponto de máximo no intervalo —
a estrela não aparece para esse cultivar.

Consulte a tabela **Dados técnicos** abaixo do gráfico para os valores de R².
""")

if not _tem_grupos:
    st.info("Dados de grupo de densidade não disponíveis.")
else:
    cultivares_reg = sorted(
        ta_filtrado[ta_filtrado["sc_ha"] > 0]["dePara"].dropna().unique().tolist()
    )
    col_r, _ = st.columns([2, 3])
    with col_r:
        cults_reg = st.multiselect(
            "Cultivares",
            options=cultivares_reg,
            default=cultivares_reg[:min(4, len(cultivares_reg))],
            key="den_reg_cults",
        )

    CORES_REG = ["#2976B6", "#27AE60", "#F39C12", "#E74C3C", "#9B59B6", "#1ABC9C"]
    fig_reg   = go_plt.Figure()

    # Coletar resultados para título dinâmico
    _reg_resultados = []   # (cultivar, x_max, y_max, r2)

    for i, cultivar in enumerate(cults_reg):
        df_r = ta_filtrado[
            (ta_filtrado["dePara"] == cultivar) &
            (ta_filtrado["sc_ha"] > 0) &
            (ta_filtrado["populacao"].notna()) &
            (ta_filtrado["populacao"] > 0)
        ].copy()
        if len(df_r) < 5:
            continue
        cor = CORES_REG[i % len(CORES_REG)]
        x   = df_r["populacao"].values.astype(float)
        y   = df_r["sc_ha"].values.astype(float)

        # Pontos observados — customdata com localização
        _cd_reg_cols  = ["cod_fazenda", "cidade_nome", "estado_sigla"]
        _cd_reg_avail = [c for c in _cd_reg_cols if c in df_r.columns]
        if _cd_reg_avail:
            _cd_reg = df_r[_cd_reg_avail].values
            _ht_reg = (
                "<b>%{customdata[0]}</b> · %{customdata[1]}, %{customdata[2]}<br>"
                f"<i>{cultivar}</i><br>"
                "Pop: <b>%{x:,.0f}</b> pl/ha<br>"
                "sc/ha: <b>%{y:.1f}</b>"
                "<extra></extra>"
            )
        else:
            _cd_reg = None
            _ht_reg = f"<b>{cultivar}</b><br>Pop: %{{x:,.0f}}<br>sc/ha: %{{y:.1f}}<extra></extra>"

        fig_reg.add_trace(go_plt.Scatter(
            x=x, y=y, mode="markers",
            showlegend=False,
            marker=dict(color=cor, size=6, opacity=0.45,
                        line=dict(color="#FFFFFF", width=0.5)),
            customdata=_cd_reg,
            hovertemplate=_ht_reg,
        ))

        try:
            coefs  = np.polyfit(x, y, 2)
            p_fn   = np.poly1d(coefs)
            x_line = np.linspace(x.min(), x.max(), 200)
            y_pred = p_fn(x)
            ss_res = np.sum((y - y_pred) ** 2)
            ss_tot = np.sum((y - y.mean()) ** 2)
            r2     = max(0, 1 - ss_res / ss_tot) if ss_tot > 0 else 0

            # Curva — sem legenda lateral
            fig_reg.add_trace(go_plt.Scatter(
                x=x_line, y=p_fn(x_line), mode="lines",
                showlegend=False,
                line=dict(color=cor, width=2.5),
                hovertemplate=f"<b>{cultivar}</b><br>Pop: %{{x:,.0f}}<br>Estimado: %{{y:.1f}}<extra></extra>",
            ))

            # Estrela de máximo — maior, com caixa de destaque
            if coefs[0] < 0:
                x_max = -coefs[1] / (2 * coefs[0])
                if x.min() < x_max < x.max():
                    y_max = float(p_fn(x_max))
                    _y_curva_vals = p_fn(x_line)
                    _reg_resultados.append((cultivar, x_max, y_max, r2, coefs, float(x.mean()), cor,
                                            float(_y_curva_vals.min()), float(_y_curva_vals.max())))

                    # Linha vertical tracejada até o eixo X
                    fig_reg.add_shape(type="line",
                        x0=x_max, x1=x_max, y0=y.min() - 2, y1=y_max,
                        line=dict(color=cor, width=1, dash="dot"),
                    )

                    # Estrela visual — apenas símbolo no gráfico, sem legenda
                    fig_reg.add_trace(go_plt.Scatter(
                        x=[x_max], y=[y_max],
                        mode="markers",
                        name="", showlegend=False,
                        marker=dict(color=cor, size=18, symbol="star",
                                    line=dict(color="#FFFFFF", width=2)),
                        hovertemplate=(
                            f"<b>★ {cultivar}</b><br>"
                            f"Pop. ótima: {x_max/1000:.0f}k pl/ha<br>"
                            f"sc/ha estimado: {y_max:.1f}<br>"
                            f"R²={r2:.2f}"
                            f"<extra></extra>"
                        ),
                    ))
        except Exception:
            pass

    # ── Entradas de legenda — traces invisíveis após o loop ─────────────────
    for cultivar_l, x_max_l, y_max_l, r2_l, _, __, cor_l, *_rest in _reg_resultados:
        fig_reg.add_trace(go_plt.Scatter(
            x=[None], y=[None],
            mode="markers",
            name=f"{cultivar_l}  {x_max_l/1000:.0f}k pl/ha · {y_max_l:.1f} sc/ha",
            marker=dict(color=cor_l, size=14, symbol="star",
                        line=dict(color="#FFFFFF", width=1.5)),
            showlegend=True,
        ))

    # ── Título dinâmico com resultado real ───────────────────────────────────
    if _reg_resultados:
        # Cultivar com maior R² entre os que têm máximo
        _melhor_reg = max(_reg_resultados, key=lambda t: t[3])
        _titulo_reg = (
            f"<b style='color:{_CL_VERDE}'>{_melhor_reg[0]}</b> atinge máximo em "
            f"<b style='color:{_CL_VERDE}'>{_melhor_reg[1]/1000:.0f}k pl/ha</b> "
            f"({_melhor_reg[2]:.1f} sc/ha estimado)"
        )
        if len(_reg_resultados) > 1:
            _outros = [f"{t[0]}: {t[1]/1000:.0f}k" for t in _reg_resultados if t[0] != _melhor_reg[0]]
            _titulo_reg += f" · Outros: {' · '.join(_outros)}"
    elif cults_reg:
        _titulo_reg = "Nenhum ponto de máximo identificado no intervalo observado"
    else:
        _titulo_reg = "Selecione ao menos um cultivar"

    # Atualizar secao_titulo retroativamente via markdown acima do gráfico
    st.markdown(f"""
<div style="margin:0.5rem 0 0.8rem;">
    <p style="font-size:13px;color:{_CL_SUB};margin:0;">{_titulo_reg}</p>
</div>""", unsafe_allow_html=True)

    # ── Layout Schwabish ──────────────────────────────────────────────────────
    _x_all = ta_filtrado[ta_filtrado["populacao"] > 0]["populacao"].dropna()
    _x_reg_max = float(_x_all.max()) if not _x_all.empty else 400000
    _label_space = _x_reg_max * 0.05   # margem pequena à direita
    # Range Y baseado exclusivamente nas curvas ajustadas
    if _reg_resultados:
        _y_reg_min = min(t[7] for t in _reg_resultados)  # y_min de cada curva
        _y_reg_max = max(t[8] for t in _reg_resultados)  # y_max de cada curva
    else:
        _y_all_reg = ta_filtrado[
            ta_filtrado["dePara"].isin(cults_reg) & (ta_filtrado["sc_ha"] > 0)
        ]["sc_ha"].dropna()
        _y_reg_min = float(_y_all_reg.quantile(0.10)) if not _y_all_reg.empty else 40
        _y_reg_max = float(_y_all_reg.quantile(0.90)) if not _y_all_reg.empty else 90

    fig_reg.update_layout(
        height=520,
        plot_bgcolor=_CL_BG, paper_bgcolor=_CL_BG,
        font=dict(family=_CL_FONTE),
        showlegend=True,
        legend=dict(
            title=dict(text="Máximos estimados", font=dict(size=11, color=_CL_SUB)),
            orientation="v",
            x=1.01, y=1,
            xanchor="left", yanchor="top",
            bgcolor="rgba(255,255,255,0.92)",
            bordercolor="#E5E7EB", borderwidth=1,
            font=dict(size=12, color=_CL_TEXTO),
        ),
        margin=dict(t=20, b=60, l=70, r=20),
        xaxis=dict(
            title=dict(text="<b>População real (pl/ha)</b>", font=dict(size=14, color="#1A1A1A", weight="bold")),
            tickfont=dict(size=12, color="#1A1A1A", weight="bold"),
            tickvals=[100000, 150000, 200000, 250000, 300000, 350000, 400000, 450000, 500000],
            ticktext=["100k", "150k", "200k", "250k", "300k", "350k", "400k", "450k", "500k"],
            showgrid=True, gridcolor=_CL_GRID,
            zeroline=False, showline=False,
            range=[_x_all.min() * 0.92 if not _x_all.empty else 0,
                   _x_reg_max + _label_space],
        ),
        yaxis=dict(
            title=dict(text="<b>sc/ha</b>", font=dict(size=14, color="#1A1A1A", weight="bold")),
            tickfont=dict(size=12, color="#1A1A1A", weight="bold"),
            showgrid=True, gridcolor=_CL_GRID,
            zeroline=False, showline=False,
            range=[
                _y_reg_min - (_y_reg_max - _y_reg_min) * 0.06,
                _y_reg_max + (_y_reg_max - _y_reg_min) * 0.12,
            ],
        ),
    )
    st.plotly_chart(fig_reg, use_container_width=True)
    _col_cap_reg, _col_dic_reg = st.columns([3, 1])
    with _col_cap_reg:
        st.caption(
            "★ = ponto de máxima produtividade estimada (legenda: pop. ótima e sc/ha) · "
            "R² na tabela abaixo · Pontos = parcelas reais observadas."
        )
    with _col_dic_reg:
        _dic_reg_cols = ["cod_fazenda", "nomeFazenda", "cidade_nome", "estado_sigla"]
        _dic_reg_avail = [c for c in _dic_reg_cols if c in ta_filtrado.columns]
        if _dic_reg_avail and cults_reg:
            _df_dic_reg = (
                ta_filtrado[
                    ta_filtrado["dePara"].isin(cults_reg) &
                    (ta_filtrado["sc_ha"] > 0)
                ][_dic_reg_avail]
                .drop_duplicates()
                .sort_values("cod_fazenda")
                .rename(columns={
                    "cod_fazenda":  "Código",
                    "nomeFazenda":  "Local",
                    "cidade_nome":  "Cidade",
                    "estado_sigla": "Estado",
                })
                .reset_index(drop=True)
            )
            with st.popover(f"📍 {len(_df_dic_reg)} locais", use_container_width=True):
                st.markdown("Passe o mouse sobre os pontos para identificar o local.")
                st.dataframe(_df_dic_reg, hide_index=True, use_container_width=True)

    if _reg_resultados:
        st.markdown(
            f'<p style="font-size:12px;font-weight:600;color:{_CL_SUB};text-transform:uppercase;'
            f'letter-spacing:0.07em;margin:1.2rem 0 0.4rem;">Dados técnicos da regressão</p>',
            unsafe_allow_html=True,
        )
        _tbl_reg = []
        for cultivar_r, x_max_r, y_max_r, r2_r, coefs_r, x_mean_r, _cor_r, *_ in sorted(_reg_resultados, key=lambda t: -t[2]):
            df_rr = ta_filtrado[
                (ta_filtrado["dePara"] == cultivar_r) &
                (ta_filtrado["sc_ha"] > 0) &
                (ta_filtrado["populacao"] > 0)
            ]
            # Marginal: derivada em x_mean * 10000 (variacao por +10k plantas)
            _marginal = (2 * coefs_r[0] * x_mean_r + coefs_r[1]) * 10000
            _marginal_txt = (
                f"+{_marginal:.2f} sc/ha" if _marginal > 0
                else f"{_marginal:.2f} sc/ha"
            )
            # Funcao da curva formatada
            _a, _b, _c = coefs_r[0], coefs_r[1], coefs_r[2]
            _tbl_reg.append({
                "Cultivar":               cultivar_r,
                "Pop. ótima (técnica)":    f"{x_max_r/1000:.0f}k pl/ha",
                "sc/ha no máximo":         f"{y_max_r:.1f}",
                "Δsc/ha por +10k plantas":  _marginal_txt,
                "R²":                     f"{r2_r:.3f}",
                "Qualidade":              "Bom" if r2_r >= 0.50 else ("Moderado" if r2_r >= 0.25 else "Fraco"),
                "N parcelas":             len(df_rr),
                "Pop. média obs.":         f"{int(x_mean_r/1000):.0f}k",
            })
        ag_table(pd.DataFrame(_tbl_reg), height=min(280, 36 + 32 * len(_tbl_reg) + 20))
        st.caption(
            "Pop. ótima = máximo técnico da curva quadrática (não econômico). "
            "R² < 0.25 = outros fatores dominam sobre a densidade."
        )

st.divider()

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 4 — PERFIL VISUAL POR GRUPO DE DENSIDADE
# ════════════════════════════════════════════════════════════════════════════════
import streamlit.components.v1 as _components_dens
from pathlib import Path as _Path_dens

secao_titulo(
    "Perfil Visual por Densidade",
    "Como o perfil da planta muda com a população?",
    "Altura, inserção e acamamento médios por grupo de densidade para o cultivar selecionado.",
)

with st.popover("ℹ️ Como interpretar", use_container_width=False):
    st.markdown("""
Cada card representa um **grupo de densidade** — a média das parcelas daquele cultivar
plantadas naquela faixa de população.

**Métricas exibidas:**
- **sc/ha** → produtividade média das parcelas do grupo
- **ALP (cm)** → altura de planta — tende a aumentar com densidade por estiolamento
- **AIV (cm)** → inserção da 1ª vagem — sobe com densidade, afeta colheita mecânica
- **Acamamento** → nota 1–9 (maior = melhor); densidade alta pode reduzir a nota
- **n** → número de parcelas no grupo

**Dica:** compare o card de menor e maior densidade para ver a tendência do cultivar.
""")

if not _tem_grupos:
    st.info("Dados de grupo de densidade não disponíveis.")
else:
    cultivares_ep = sorted(ta_filtrado["dePara"].dropna().unique().tolist())
    col_e, _ = st.columns([2, 3])
    with col_e:
        cult_ep = st.selectbox("Cultivar", options=cultivares_ep, key="den_ep_cult")

    df_ep = ta_filtrado[ta_filtrado["dePara"] == cult_ep].copy()

    # Carregar imagem soja.png
    _img_path_ep = _Path_dens(__file__).parent.parent / "assets" / "soja.png"
    _img_b64_ep  = ""
    if _img_path_ep.exists():
        import base64 as _b64
        with open(_img_path_ep, "rb") as _f:
            _img_b64_ep = _b64.b64encode(_f.read()).decode()
    _img_src_ep = f"data:image/png;base64,{_img_b64_ep}" if _img_b64_ep else ""

    def _fmt_ep(v, suffix="", decimals=1):
        if v is None: return "—"
        if isinstance(v, str): return f"{v}{suffix}" if v else "—"
        try:
            if pd.isna(v): return "—"
        except (TypeError, ValueError): pass
        return f"{round(v, decimals)}{suffix}"

    def _moda_cultivar(col):
        """Moda calculada sobre todo o df_ep (cultivar inteiro, independente do grupo)."""
        if col not in df_ep.columns: return None
        v = df_ep[col].dropna().replace("", pd.NA).dropna()
        return v.mode().iloc[0] if len(v) > 0 else None

    # Características fixas do cultivar — calculadas uma vez para todos os cards
    _v_flor_cult = _moda_cultivar("corFlorNome")
    _v_hab_cult  = _moda_cultivar("habitoCrescimentoNome")
    _v_pub_cult  = _moda_cultivar("corPubNome")

    def _card_dens(grupo, df_grupo, cor_grupo, img_src):
        def _s(col):
            if col not in df_grupo.columns: return pd.Series(dtype=float)
            v = df_grupo[col].dropna()
            return v[v > 0]

        def _m(col): s = _s(col); return float(s.mean()) if len(s) > 0 else None

        n       = len(df_grupo)
        v_sc    = _m("sc_ha")
        v_alp   = _m("media_ALT")
        v_aiv   = _m("media_AIV")
        v_acam  = _m("notaAC")
        v_pmg   = _m("pesoMilGraos_corrigido")
        v_ciclo = _m("dias_ate_DMF")
        v_eng   = _m("media_RR") if "media_RR" in df_grupo.columns else None

        # Usa a moda do cultivar inteiro (não do grupo isolado)
        v_flor  = _v_flor_cult
        v_hab   = _v_hab_cult
        v_pub   = _v_pub_cult

        # % vagens por terço
        vts = _m("media_VTS"); vtm = _m("media_VTM"); vti = _m("media_VTI")
        _vv = [v for v in [vts, vtm, vti] if v is not None and v > 0]
        _tot = sum(_vv) if _vv else 0
        pct_sup = round(vts / _tot * 100, 1) if _tot > 0 and vts and vts > 0 else None
        pct_med = round(vtm / _tot * 100, 1) if _tot > 0 and vtm and vtm > 0 else None
        pct_inf = round(vti / _tot * 100, 1) if _tot > 0 and vti and vti > 0 else None
        sem_vagens = _tot == 0

        borda = cor_grupo

        img_tag = (
            f"<img src='{img_src}' alt='Soja' style='width:100%;height:auto;display:block;'/>"
            if img_src else
            "<div style='width:100%;height:400px;background:#f3f4f6;border-radius:8px;"
            "display:flex;align-items:center;justify-content:center;color:#9CA3AF;font-size:12px;'>soja.png</div>"
        )

        return f"""
<div style="font-family:'Helvetica Neue',sans-serif;width:100%;">

  <!-- Título -->
  <div style="text-align:center;font-size:13px;font-weight:700;color:#1A1A1A;
              letter-spacing:.06em;text-transform:uppercase;
              margin-bottom:10px;border-bottom:2px solid {borda};padding-bottom:6px;">
    {grupo} · {n} parcelas
  </div>

  <!-- Corpo: esquerda + imagem + direita -->
  <div style="display:flex;align-items:flex-start;justify-content:center;">

    <!-- Esquerda: % vagens por terço -->
    <div style="position:relative;width:150px;height:500px;flex-shrink:0;">
      {'<div style="position:absolute;right:10px;top:30px;text-align:right;"><div style="font-size:26px;font-weight:800;color:#1A1A1A;line-height:1;">' + _fmt_ep(pct_sup,"%") + '</div><div style="font-size:11px;color:#4B5563;margin-top:3px;">Terço Superior</div></div>' if not sem_vagens else '<div style="position:absolute;top:120px;right:10px;text-align:right;font-size:11px;color:#6B7280;line-height:1.5;">Avaliação de<br>vagens não<br>disponível</div>'}
      {'<div style="position:absolute;right:10px;top:200px;text-align:right;"><div style="font-size:26px;font-weight:800;color:#1A1A1A;line-height:1;">' + _fmt_ep(pct_med,"%") + '</div><div style="font-size:11px;color:#4B5563;margin-top:3px;">Terço Médio</div></div>' if not sem_vagens else ''}
      {'<div style="position:absolute;right:10px;top:370px;text-align:right;"><div style="font-size:26px;font-weight:800;color:#1A1A1A;line-height:1;">' + _fmt_ep(pct_inf,"%") + '</div><div style="font-size:11px;color:#4B5563;margin-top:3px;">Terço Inferior</div></div>' if not sem_vagens else ''}
    </div>

    <!-- Centro: imagem -->
    <div style="width:260px;flex-shrink:0;">
      {img_tag}
    </div>

    <!-- Direita: métricas -->
    <div style="position:relative;width:170px;height:500px;flex-shrink:0;">
      <div style="position:absolute;top:30px;left:14px;border-left:3px solid #D1D5DB;padding-left:10px;">
        <div style="font-size:10px;color:#374151;font-weight:600;text-transform:uppercase;letter-spacing:.04em;">Altura de Planta</div>
        <div style="font-size:16px;font-weight:700;color:#111827;">{_fmt_ep(v_alp," cm")}</div>
        <div style="font-size:10px;color:#374151;font-weight:600;text-transform:uppercase;letter-spacing:.04em;margin-top:8px;">Ciclo</div>
        <div style="font-size:16px;font-weight:700;color:#111827;">{_fmt_ep(v_ciclo," dias",0)}</div>
      </div>
      <div style="position:absolute;top:200px;left:14px;border-left:3px solid #D1D5DB;padding-left:10px;">
        <div style="font-size:10px;color:#374151;font-weight:600;text-transform:uppercase;letter-spacing:.04em;">Peso Mil Grãos</div>
        <div style="font-size:16px;font-weight:700;color:#111827;">{_fmt_ep(v_pmg," g")}</div>
        <div style="font-size:10px;color:#374151;font-weight:600;text-transform:uppercase;letter-spacing:.04em;margin-top:8px;">Inserção 1ª Vagem</div>
        <div style="font-size:16px;font-weight:700;color:#111827;">{_fmt_ep(v_aiv," cm")}</div>
      </div>
      <div style="position:absolute;top:360px;left:14px;border-left:3px solid #D1D5DB;padding-left:10px;">
        <div style="font-size:10px;color:#374151;font-weight:600;text-transform:uppercase;letter-spacing:.04em;">Cor da Flor</div>
        <div style="font-size:13px;font-weight:600;color:#1F2937;">{_fmt_ep(v_flor)}</div>
        <div style="font-size:10px;color:#374151;font-weight:600;text-transform:uppercase;letter-spacing:.04em;margin-top:5px;">Hábito</div>
        <div style="font-size:13px;font-weight:600;color:#1F2937;">{_fmt_ep(v_hab)}</div>
        <div style="font-size:10px;color:#374151;font-weight:600;text-transform:uppercase;letter-spacing:.04em;margin-top:5px;">Cor Pubescência</div>
        <div style="font-size:13px;font-weight:600;color:#1F2937;">{_fmt_ep(v_pub)}</div>
      </div>
    </div>

  </div>

  <!-- Rodapé: sc/ha · Acamamento · Engalhamento -->
  <div style="display:flex;justify-content:center;gap:32px;
              margin-top:12px;padding-top:10px;border-top:1px solid #E5E7EB;">
    <div style="text-align:center;">
      <div style="font-size:22px;font-weight:800;color:{borda};line-height:1;">{_fmt_ep(v_sc)}</div>
      <div style="font-size:10px;color:#4B5563;margin-top:3px;">sc/ha médio</div>
    </div>
    <div style="text-align:center;">
      <div style="font-size:22px;font-weight:800;color:#1A1A1A;line-height:1;">{_fmt_ep(v_acam)}</div>
      <div style="font-size:10px;color:#4B5563;margin-top:3px;">Acamamento</div>
    </div>
    <div style="text-align:center;">
      <div style="font-size:22px;font-weight:800;color:#1A1A1A;line-height:1;">{_fmt_ep(v_eng)}</div>
      <div style="font-size:10px;color:#4B5563;margin-top:3px;">Ramos Reprodutivos</div>
    </div>
  </div>

</div>"""

    # Definir grupo com maior sc/ha para destacar
    _sc_por_grupo = {}
    for _g in _ordem_grupos:
        _sc_g = df_ep[df_ep["pop_grupo"] == _g]["sc_ha"].dropna()
        _sc_g = _sc_g[_sc_g > 0]
        if len(_sc_g) > 0:
            _sc_por_grupo[_g] = float(_sc_g.mean())
    _grupo_lider_ep = max(_sc_por_grupo, key=_sc_por_grupo.get) if _sc_por_grupo else None

    # ── Renderizar cards com botão de download embutido ──────────────────────
    def _card_com_download(grupo, df_grupo, cor_grupo, img_src, cultivar_nome):
        card_inner = _card_dens(grupo, df_grupo, cor_grupo, img_src)
        _id = f"card-dens-{grupo}"
        _nome_arquivo = f"{cultivar_nome.replace(' ','_')}_{grupo}_perfil"
        return f"""
<script src="https://cdnjs.cloudflare.com/ajax/libs/html2canvas/1.4.1/html2canvas.min.js"></script>
<div id="{_id}" style="background:#FFFFFF;padding:16px 12px 12px;
     border-radius:12px;border:1px solid #E5E7EB;">
{card_inner}
</div>
<div style="text-align:center;margin-top:10px;">
  <button onclick="
    html2canvas(document.getElementById('{_id}'),{{
      backgroundColor:'#FFFFFF',scale:2,useCORS:true,allowTaint:true
    }}).then(function(c){{
      var a=document.createElement('a');
      a.download='{_nome_arquivo}.png';
      a.href=c.toDataURL('image/png');
      a.click();
    }});"
    style="background:#27AE60;color:#FFFFFF;border:none;border-radius:8px;
           padding:7px 18px;font-size:12px;font-weight:600;cursor:pointer;
           font-family:'Helvetica Neue',sans-serif;letter-spacing:.02em;">
    ⬇ Baixar imagem
  </button>
</div>"""

    _grupos_ep = [g for g in _ordem_grupos if g in df_ep["pop_grupo"].values]
    if not _grupos_ep:
        st.info("Nenhum grupo de densidade disponível para este cultivar.")
    else:
        _n_cols = min(len(_grupos_ep), 2)
        _linhas = [_grupos_ep[i:i+_n_cols] for i in range(0, len(_grupos_ep), _n_cols)]
        for _linha in _linhas:
            _cols_ep = st.columns(len(_linha))
            for _ci, _g in enumerate(_linha):
                _df_g  = df_ep[df_ep["pop_grupo"] == _g]
                _cor_g = _CL_VERDE if _g == _grupo_lider_ep else "#94A3B8"
                with _cols_ep[_ci]:
                    _components_dens.html(
                        _card_com_download(_g, _df_g, _cor_g, _img_src_ep, cult_ep),
                        height=760, scrolling=False,
                    )
        st.caption(
            f"Card verde = grupo com maior sc/ha médio ({_grupo_lider_ep}). "
            "Clique em ⬇ Baixar imagem para salvar o perfil como PNG."
        )

st.divider()

st.markdown(
    '<p style="font-size:13px;color:#374151;text-align:center;">Painel JAUM DTC · Stine Seed · '
    'Desenvolvido por <a href="https://www.linkedin.com/in/eng-agro-andre-ferreira/" '
    'target="_blank" style="color:#27AE60;text-decoration:none;">Andre Ferreira</a></p>',
    unsafe_allow_html=True,
)
